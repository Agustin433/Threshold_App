"""Shared report export, summary and narrative helpers."""

from __future__ import annotations

from datetime import datetime
from io import BytesIO
from pathlib import Path
import re
import textwrap
import unicodedata
import warnings

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from local_store import build_weekly_summaries
from modules.data_loader import prepare_raw_workouts_df
from modules.data_quality import compute_data_quality_report
from modules.jump_analysis import (
    _format_profile_source_date,
    _prepare_jump_df,
    build_neuromuscular_profile_result,
    build_composite_profile_metric_table,
    build_composite_profile_snapshot,
    build_jump_delta_display_table,
    build_jump_feedback_lines,
    build_jump_temporal_context,
    build_profile_radar_row,
    compute_baseline_delta,
    compute_swc_delta,
    resolve_zscore,
    semaphore_label,
)
from modules.metrics import calculate_completion_rate, summarize_completion_by_group
from modules.report_force_time import build_force_time_report_payload, draw_force_time_test_block


APP_ROOT = Path(__file__).resolve().parent.parent
BRAND_ASSET_DIR = APP_ROOT / "assets" / "brand"

REPORT_SHEET_ORDER = [
    "Resumen_Ejecutivo",
    "Interpretacion",
    "Contexto_Operativo",
    "Completion_Resumen",
    "ACWR_sRPE",
    "Monotonia_Strain",
    "Wellness",
    "Evaluaciones_Saltos",
    "Maximos_Ejercicios",
    "Volumen_Carga_Externa",
    "Volumen_RepLoad_Legacy",
    "Completion_Rate",
    "Session_Notes",
    "Reporte_Meta",
]

REPORT_SHEET_EXPORT_NAMES = {
    "Resumen_Ejecutivo": "01_Resumen",
    "Interpretacion": "02_Interpretacion",
    "Contexto_Operativo": "03_Contexto_Operativo",
    "Completion_Resumen": "03_Completion_Resumen",
    "ACWR_sRPE": "04_Carga_ACWR",
    "Monotonia_Strain": "05_Carga_Monotonia",
    "Wellness": "06_Wellness",
    "Evaluaciones_Saltos": "07_Evaluaciones",
    "Maximos_Ejercicios": "08_Maximos",
    "Volumen_Carga_Externa": "09_Carga_Externa",
    "Volumen_RepLoad_Legacy": "09_RepLoad_Legacy",
    "Completion_Rate": "10_Completion_Detalle",
    "Session_Notes": "11_Session_Notes",
    "Reporte_Meta": "99_Meta",
}

DATASET_LABELS = {
    "rpe_df": "RPE + Tiempo",
    "wellness_df": "Wellness",
    "completion_df": "Completion",
    "rep_load_df": "Rep/Load (legacy opcional)",
    "raw_df": "Raw Workouts",
    "session_notes_df": "Opt-outs / Notes",
    "maxes_df": "Maxes",
    "jump_df": "Evaluaciones",
}

MODERN_REPORT_DATASET_KEYS = [
    "rpe_df",
    "wellness_df",
    "completion_df",
    "raw_df",
    "session_notes_df",
    "maxes_df",
    "jump_df",
]
LEGACY_REPORT_DATASET_KEYS = ["rep_load_df"]

REPORT_AUDIENCE_OPTIONS = {
    "Atleta": "atleta",
    "Profe": "profe",
    "Cliente": "cliente",
}

REPORT_AUDIENCE_LABELS = {value: key for key, value in REPORT_AUDIENCE_OPTIONS.items()}
EUR_RATIO_LABEL = "EUR (ratio)"
PDF_MISSING_TEXT = "Faltan datos"
PROFESSIONAL_NO_CLEAR_LAGGING_TEXT = "No aparece una variable claramente rezagada según el umbral actual."
PROFESSIONAL_NO_CLEAR_LAGGING_NEXT_BLOCK = (
    "Mantener el perfil actual y priorizar consistencia técnica, sin abrir nuevos focos innecesarios."
)
PROFESSIONAL_NO_CLEAR_LAGGING_MONITOR = (
    "Monitorear si alguna variable empieza a separarse del perfil en próximas evaluaciones."
)
MOJIBAKE_MARKERS = tuple(chr(code) for code in (195, 194, 226))


def _repair_mojibake_text(value: object) -> str:
    """Repair UTF-8 text that was accidentally decoded as Latin-1/Windows-1252."""
    text = "" if value is None else str(value)
    for _ in range(3):
        if not any(marker in text for marker in MOJIBAKE_MARKERS):
            break
        previous = text
        for encoding in ("cp1252", "latin-1"):
            try:
                repaired = text.encode(encoding).decode("utf-8")
            except (UnicodeEncodeError, UnicodeDecodeError):
                continue
            if repaired != text:
                text = repaired
                break
        if text == previous:
            break
    manual_replacements = {
        "EstÃmulos": "Estímulos",
        "DÃas": "Días",
        "MonotonÃa": "Monotonía",
        "fisiolÃgica": "fisiológica",
        "biomecÃnica": "biomecánica",
        "evaluaciÃn": "evaluación",
        "prÃximo": "próximo",
        "seÃal": "señal",
        "SeÃal": "Señal",
    }
    manual_replacements.update(
        {
            "Est\u00c3mulos": "Est\u00edmulos",
            "Est\u00c3\u0192mulos": "Est\u00edmulos",
            "Est\u00c3\u0192\u00c2mulos": "Est\u00edmulos",
            "D\u00c3as": "D\u00edas",
            "Monoton\u00c3a": "Monoton\u00eda",
            "fisiol\u00c3gica": "fisiol\u00f3gica",
            "fisiol\u00c33gica": "fisiol\u00f3gica",
            "fisiol\u00c3\u0192\u00c23gica": "fisiol\u00f3gica",
            "biomec\u00c3nica": "biomec\u00e1nica",
            "biomec\u00c3\u0192\u00c2\u00a1nica": "biomec\u00e1nica",
            "evaluaci\u00c3n": "evaluaci\u00f3n",
            "pr\u00c3ximo": "pr\u00f3ximo",
            "pr\u00c33ximo": "pr\u00f3ximo",
            "pr\u00c3\u0192\u00c23ximo": "pr\u00f3ximo",
            "exposici\u00c33n": "exposici\u00f3n",
            "exposici\u00c3\u0192\u00c23n": "exposici\u00f3n",
            "se\u00c3al": "se\u00f1al",
            "Se\u00c3al": "Se\u00f1al",
        }
    )
    sequence_replacements = {
        "\u00c3\u00a1": "\u00e1",
        "\u00c3\u00a9": "\u00e9",
        "\u00c3\u00ad": "\u00ed",
        "\u00c3\u00b3": "\u00f3",
        "\u00c3\u00ba": "\u00fa",
        "\u00c3\u00b1": "\u00f1",
        "\u00c3\u0081": "\u00c1",
        "\u00c3\u0089": "\u00c9",
        "\u00c3\u008d": "\u00cd",
        "\u00c3\u0093": "\u00d3",
        "\u00c3\u009a": "\u00da",
        "\u00c3\u0091": "\u00d1",
        "\u00c2\u00b7": "\u00b7",
        "\u00c3\u0192\u00c2\u00a1": "\u00e1",
        "\u00c3\u0192\u00c2\u00a9": "\u00e9",
        "\u00c3\u0192\u00c2\u00ad": "\u00ed",
        "\u00c3\u0192\u00c2\u00b3": "\u00f3",
        "\u00c3\u0192\u00c2\u00ba": "\u00fa",
        "\u00c3\u0192\u00c2\u00b1": "\u00f1",
        "\u00c33": "\u00f3",
    }
    for broken, fixed in sequence_replacements.items():
        text = text.replace(broken, fixed)
    for broken, fixed in manual_replacements.items():
        text = text.replace(broken, fixed)
    return unicodedata.normalize("NFKC", text).replace("\u00a0", " ")


def _professional_visible_metric_text(value: object) -> str:
    text = _repair_mojibake_text(value)
    if text.strip() == "TC":
        return "Tiempo de contacto"
    replacements = {
        "TC inv": "Tiempo de contacto",
        "DJ TC": "Tiempo de contacto",
        "TC_inv_Z": "Tiempo de contacto (z-score invertido para análisis)",
    }
    for source, target in replacements.items():
        text = text.replace(source, target)
    accent_replacements = {
        "Pliometria": "Pliometría",
        "Isometricos": "Isométricos",
        "olimpicos": "olímpicos",
        "proximo": "próximo",
        "mas": "más",
        "evaluacion": "evaluación",
        "adquisicion": "adquisición",
        "produccion": "producción",
        "maxima": "máxima",
        "maximo": "máximo",
        "contraccion": "contracción",
        "familiarizacion": "familiarización",
        "tecnica": "técnica",
        "fisico": "físico",
        "fisiologico": "fisiológico",
        "biomecanico": "biomecánico",
        "indices": "índices",
        "deficits": "déficits",
        "mecanica": "mecánica",
        "expresion": "expresión",
        "progresion": "progresión",
        "concentrica": "concéntrica",
        "isometrica": "isométrica",
        "util": "útil",
        "limitacion": "limitación",
        "posicion": "posición",
        "medicion": "medición",
    }
    for source, target in accent_replacements.items():
        def _replace(match: re.Match[str], fixed: str = target) -> str:
            word = match.group(0)
            return fixed[:1].upper() + fixed[1:] if word[:1].isupper() else fixed

        text = re.sub(rf"\b{re.escape(source)}\b", _replace, text, flags=re.IGNORECASE)
    return _repair_mojibake_text(text)


PROFESSIONAL_COMPOSITE_PROFILE_NOTE = (
    "Perfil compuesto: usa el Ãºltimo dato vÃ¡lido disponible por variable; no todas las "
    "variables necesariamente provienen de la misma fecha."
)
PROFESSIONAL_CURRENT_PROFILE_SCOPE_NOTE = (
    "La lectura describe el perfil actual y conviene repetir la evaluación para confirmar si el patrón persiste."
)
PROFESSIONAL_FULL_REPORT_MIN_COMPOSITE_METRICS = 4
PROFESSIONAL_INTERVAL_WARNING = (
    "Intervalo entre evaluaciones menor al recomendado de 6-8 semanas. "
    "Interpretar cambios con cautela y evitar atribuirlos de forma directa a adaptación."
)
PROFESSIONAL_LARGE_CHANGE_WARNING = (
    "Mejora individual marcada. Verificar consistencia del protocolo, familiarización, "
    "condiciones del test y calidad del dato antes de atribuir todo el cambio a adaptación física."
)
PROFESSIONAL_RSI_METHOD_NOTE = (
    "RSI calculado como altura de salto dividida por tiempo de contacto. En este reporte se "
    "interpreta como índice de fuerza reactiva y debe leerse junto con DJ/DRI y Contact Time, "
    "no como velocidad lineal."
)
PROFESSIONAL_METRIC_DIRECTIONS = {
    "CMJ": "higher_is_better",
    "SJ": "higher_is_better",
    "DJ": "higher_is_better",
    "RSI": "higher_is_better",
    "Contact Time": "lower_is_better",
    "EUR": "context_dependent",
    "mRSI": "higher_is_better",
    "IMTP": "higher_is_better",
}
PROFESSIONAL_EVOLUTION_PRIORITY = ("CMJ", "SJ", "DJ", "IMTP", "EUR", "Contact Time", "RSI", "mRSI")
PROFESSIONAL_NO_EVALUATION_TEXT = (
    "Faltan datos de evaluación para este atleta en el período seleccionado.\n"
    "Este reporte se genera con la información disponible de entrenamiento, carga interna y wellness.\n"
    "Para completar el perfil físico profesional se recomienda cargar al menos una batería de evaluación con CMJ, SJ, DJ/RSI e IMTP."
)
PROFESSIONAL_NO_EVOLUTION_TEXT = (
    "Faltan datos para mostrar evolución entre evaluaciones.\n"
    "Se necesitan al menos dos fechas de evaluación para analizar cambios del perfil físico entre bloques."
)
PROFESSIONAL_NO_QUADRANTS_TEXT = (
    "Faltan datos para construir los cuadrantes de perfil físico.\n"
    "Para generar esta sección se necesitan combinaciones mínimas como IMTP relativa + CMJ, SJ + DRI y EUR + CMJ."
)
PROFESSIONAL_PDF_METRICS = (
    {
        "title": "CMJ",
        "value_cols": ("CMJ_cm",),
        "z_cols": ("CMJ_Z",),
        "unit": "cm",
        "digits": 1,
        "higher_is_better": True,
        "direction": PROFESSIONAL_METRIC_DIRECTIONS["CMJ"],
    },
    {
        "title": "SJ",
        "value_cols": ("SJ_cm",),
        "z_cols": ("SJ_Z",),
        "unit": "cm",
        "digits": 1,
        "higher_is_better": True,
        "direction": PROFESSIONAL_METRIC_DIRECTIONS["SJ"],
    },
    {
        "title": "DJ",
        "value_cols": ("DJ_cm",),
        "z_cols": ("DJ_height_Z",),
        "unit": "cm",
        "digits": 1,
        "higher_is_better": True,
        "direction": PROFESSIONAL_METRIC_DIRECTIONS["DJ"],
    },
    {
        "title": "RSI",
        "value_cols": ("DJ_RSI",),
        "z_cols": ("DJ_RSI_Z",),
        "unit": "rsi_index",
        "digits": 3,
        "higher_is_better": True,
        "direction": PROFESSIONAL_METRIC_DIRECTIONS["RSI"],
    },
    {
        "title": "Contact Time",
        "value_cols": ("DJ_tc_ms",),
        "z_cols": ("TC_inv_Z", "DJtc_Z"),
        "unit": "ms",
        "digits": 0,
        "higher_is_better": False,
        "direction": PROFESSIONAL_METRIC_DIRECTIONS["Contact Time"],
    },
    {
        "title": "EUR",
        "value_cols": ("EUR",),
        "z_cols": ("EUR_Z",),
        "unit": "ratio",
        "digits": 3,
        "higher_is_better": True,
        "direction": PROFESSIONAL_METRIC_DIRECTIONS["EUR"],
    },
    {
        "title": "mRSI",
        "value_cols": ("mRSI",),
        "z_cols": ("mRSI_Z",),
        "unit": "mrsi_index",
        "digits": 3,
        "higher_is_better": True,
        "direction": PROFESSIONAL_METRIC_DIRECTIONS["mRSI"],
    },
    {
        "title": "IMTP",
        "value_cols": ("IMTP_N", "IMTP_relPF"),
        "z_cols": ("IMTP_N_Z", "IMTP_Z", "IMTP_relPF_Z"),
        "unit": {"IMTP_N": "N", "IMTP_relPF": "N/kg"},
        "digits": {"IMTP_N": 0, "IMTP_relPF": 2},
        "higher_is_better": True,
        "direction": PROFESSIONAL_METRIC_DIRECTIONS["IMTP"],
    },
)
PROFESSIONAL_TE_REFERENCES = {
    "CMJ_cm": {"value": 1.5, "unit": "cm"},
    "SJ_cm": {"value": 1.5, "unit": "cm"},
    "DJ_cm": {"value": 0.4, "unit": "cm"},
    "DJ_RSI": {"value": 0.031, "unit": "m/s"},
    "DRI": {"value": 0.05, "unit": "dri_index"},
    "DJ_tc_ms": {"value": 4.0, "unit": "ms"},
    "mRSI": {"value": 0.05, "unit": "mRSI"},
    "EUR": {"value": 0.021, "unit": "ratio"},
    "IMTP_N": {"value": 30.0, "unit": "N"},
}
PROFESSIONAL_EXPOSURE_CATEGORY_SPECS = (
    {
        "title": "Fuerza con carga",
        "categories": ("strength_loaded",),
        "value_col": "Volume_Load_kg",
        "unit": "kg x rep",
    },
    {
        "title": "Pliometría y aterrizajes",
        "categories": ("plyo_jump", "landing_mechanics"),
        "value_col": "Contacts",
        "unit": "contactos",
    },
    {
        "title": "Derivados olímpicos",
        "categories": ("olympic_derivatives",),
        "value_col": "Exposures",
        "unit": "exposiciones",
    },
    {
        "title": "Isométricos y estabilidad",
        "categories": ("iso", "core_stability"),
        "value_col": "Exposures",
        "unit": "exposiciones",
    },
    {
        "title": "Movilidad y prehab",
        "categories": ("mobility_prehab",),
        "value_col": "Exposures",
        "unit": "exposiciones",
    },
    {
        "title": "Sprint / COD",
        "categories": ("sprint_cod",),
        "value_col": "Exposures",
        "unit": "esfuerzos",
    },
)


def _report_sample_suffix(
    n: int | None,
    *,
    singular: str = "día",
    plural: str = "días",
) -> str:
    if n is None or n <= 0:
        return ""
    unit = singular if int(n) == 1 else plural
    return f"(n={int(n)} {unit})"


def _report_sample_warning(n: int | None) -> str:
    if n is None or n <= 0:
        return ""
    if int(n) == 1:
        return "⚠ Interpretación limitada — basada en un solo registro."
    if int(n) < 3:
        return f"⚠ Datos insuficientes para tendencia confiable (n={int(n)})."
    return ""


def _report_wellness_score_label(value: float | None) -> dict[str, object]:
    if value is None or pd.isna(value):
        return {
            "score": float("nan"),
            "label": PDF_MISSING_TEXT,
            "interpretation": PDF_MISSING_TEXT,
        }
    score = max(1.0, min(5.0, float(value)))
    if score >= 4.0:
        return {
            "score": score,
            "label": "Óptimo",
            "interpretation": "Bienestar favorable para progresar.",
        }
    if score >= 3.0:
        return {
            "score": score,
            "label": "Aceptable",
            "interpretation": "Progresar con monitoreo.",
        }
    if score >= 2.0:
        return {
            "score": score,
            "label": "Atención",
            "interpretation": "Progresar conservadoramente — auditar sueño y estrés.",
        }
    return {
        "score": score,
        "label": "Crítico",
        "interpretation": "No aumentar carga hasta mejorar wellness.",
    }


def _report_wellness_source_column(frame: pd.DataFrame, *candidates: str) -> str | None:
    for candidate in candidates:
        if candidate in frame.columns:
            return candidate
    return None


def _normalize_report_wellness_component(series: pd.Series, *, source_col: str) -> pd.Series:
    numeric = pd.to_numeric(series, errors="coerce")
    valid = numeric.dropna()
    if valid.empty:
        return numeric

    # The PDF score is fixed to a 1-5 scale. Legacy exports may still arrive as
    # sleep hours or 0-10 questionnaire values, so we normalize only for the
    # report-level score while preserving raw component means elsewhere.
    if source_col == "Sueno_hs" and float(valid.max()) > 5.0:
        numeric = numeric / 2.0
    elif float(valid.max()) > 5.0:
        numeric = numeric / 2.0
    return numeric.clip(lower=1.0, upper=5.0)


def _report_wellness_score_series(frame: pd.DataFrame | None) -> pd.Series:
    if frame is None or frame.empty:
        return pd.Series(dtype="float64")

    score_components: dict[str, pd.Series] = {}
    sleep_col = _report_wellness_source_column(frame, "Sueno", "Sueno_hs")
    if sleep_col is not None:
        score_components["Sueno"] = _normalize_report_wellness_component(frame[sleep_col], source_col=sleep_col)
    if "Estres" in frame.columns:
        stress_raw = pd.to_numeric(frame["Estres"], errors="coerce")
        stress_norm = _normalize_report_wellness_component(frame["Estres"], source_col="Estres")
        if not stress_raw.dropna().empty and float(stress_raw.dropna().max()) > 5.0:
            stress_norm = ((11.0 - stress_raw) // 2.0).clip(lower=1.0, upper=5.0)
        else:
            stress_norm = (6.0 - stress_norm).clip(lower=1.0, upper=5.0)
        score_components["Estres"] = stress_norm
    if "Dolor" in frame.columns:
        pain_raw = pd.to_numeric(frame["Dolor"], errors="coerce")
        pain_norm = _normalize_report_wellness_component(frame["Dolor"], source_col="Dolor")
        if not pain_raw.dropna().empty and float(pain_raw.dropna().max()) > 5.0:
            pain_norm = ((11.0 - pain_raw) // 2.0).clip(lower=1.0, upper=5.0)
        else:
            pain_norm = (6.0 - pain_norm).clip(lower=1.0, upper=5.0)
        score_components["Dolor"] = pain_norm

    if not score_components:
        return pd.Series([float("nan")] * len(frame), index=frame.index, dtype="float64")

    score_df = pd.DataFrame(score_components, index=frame.index)
    return score_df.mean(axis=1).clip(lower=1.0, upper=5.0)


def _professional_metric_te_reference(spec: dict[str, object], value_col: str) -> dict[str, object] | None:
    title = str(spec.get("title", "")).strip()
    ref = PROFESSIONAL_TE_REFERENCES.get(value_col)
    if ref is not None:
        return ref
    title_fallback = {
        "CMJ": "CMJ_cm",
        "SJ": "SJ_cm",
        "DJ": "DJ_cm",
        "RSI": "DJ_RSI",
        "Contact Time": "DJ_tc_ms",
        "EUR": "EUR",
        "mRSI": "mRSI",
        "IMTP": "IMTP_N",
    }
    fallback_key = title_fallback.get(title)
    if fallback_key is None:
        return None
    return PROFESSIONAL_TE_REFERENCES.get(fallback_key)


def _professional_metric_te_caption(spec: dict[str, object], value_col: str) -> str:
    reference = _professional_metric_te_reference(spec, value_col)
    if reference is None:
        return PDF_MISSING_TEXT
    threshold_abs = _coerce_float(reference.get("value"))
    threshold_text = _professional_threshold_text(threshold_abs, "", spec, value_col)
    if threshold_text == PDF_MISSING_TEXT:
        return PDF_MISSING_TEXT
    return f"TE de referencia: {threshold_text}. Cambios dentro de este margen pueden deberse al error típico."

IMTP_FORCE_TIME_EXPORT_COLUMNS = [
    "IMTP_force_50_N",
    "IMTP_force_100_N",
    "IMTP_force_150_N",
    "IMTP_force_200_N",
    "IMTP_force_250_N",
    "IMTP_rfd_50_N_s",
    "IMTP_rfd_100_N_s",
    "IMTP_rfd_150_N_s",
    "IMTP_rfd_250_N_s",
    "IMTP_time_pull_s",
]
IMTP_LEGACY_RFD_EXPORT_COLUMNS = [
    "RFD_50",
    "RFD_100",
    "RFD_150",
    "RFD_250",
]


def normalize_report_audience(audience: str | None) -> str:
    clean = (audience or "profe").strip().lower()
    aliases = {
        "atleta": "atleta",
        "athlete": "atleta",
        "profe": "profe",
        "profesional": "profe",
        "staff": "profe",
        "coach": "profe",
        "cliente": "cliente",
        "client": "cliente",
    }
    return aliases.get(clean, "profe")


def report_audience_label(audience: str | None) -> str:
    return REPORT_AUDIENCE_LABELS.get(normalize_report_audience(audience), "Profe")


def _summary_eur_value(row: pd.Series | None) -> object:
    if row is None:
        return None
    return row.get(EUR_RATIO_LABEL, row.get("EUR"))


def collect_report_athletes(state: dict[str, pd.DataFrame | None]) -> list[str]:
    athletes: set[str] = set()
    for key in ["rpe_df", "wellness_df", "completion_df", "rep_load_df", "raw_df", "session_notes_df", "maxes_df", "jump_df"]:
        frame = state.get(key)
        if frame is None or frame.empty or "Athlete" not in frame.columns:
            continue
        athletes.update(frame["Athlete"].dropna().astype(str).str.strip().tolist())
    return sorted(athletes)


def _has_rows(frame: pd.DataFrame | None) -> bool:
    return frame is not None and not frame.empty


def build_report_source_checklist(state: dict[str, pd.DataFrame | None]) -> list[dict[str, object]]:
    """Checklist semantico: Raw Workouts es oficial; Rep/Load solo fallback legacy."""
    raw_loaded = _has_rows(state.get("raw_df"))
    rep_load_loaded = _has_rows(state.get("rep_load_df"))
    volume_ok = raw_loaded or rep_load_loaded
    if raw_loaded:
        volume_status = "Cubierto por Raw Workouts"
        volume_detail = "Fuente oficial para carga externa y analisis por estimulo."
    elif rep_load_loaded:
        volume_status = "Fallback legacy disponible"
        volume_detail = "Rep/Load no reemplaza Raw Workouts como fuente moderna."
    else:
        volume_status = "Faltante"
        volume_detail = "Cargar Raw Data Report - Workouts."

    rows = [
        {
            "label": "RPE + Tiempo (sRPE, ACWR EWMA)",
            "ok": _has_rows(state.get("rpe_df")),
            "status": "Activo" if _has_rows(state.get("rpe_df")) else "Faltante",
            "role": "principal",
        },
        {
            "label": "Wellness 3 preguntas",
            "ok": _has_rows(state.get("wellness_df")),
            "status": "Activo" if _has_rows(state.get("wellness_df")) else "Faltante",
            "role": "principal",
        },
        {
            "label": "Completion rate",
            "ok": _has_rows(state.get("completion_df")),
            "status": "Activo" if _has_rows(state.get("completion_df")) else "Faltante",
            "role": "principal",
        },
        {
            "label": "Volumen/carga externa (Raw Workouts)",
            "ok": volume_ok,
            "status": volume_status,
            "role": "principal" if raw_loaded else "legacy_fallback" if rep_load_loaded else "principal",
            "detail": volume_detail,
        },
        {
            "label": "Maximos ejercicios",
            "ok": _has_rows(state.get("maxes_df")),
            "status": "Activo" if _has_rows(state.get("maxes_df")) else "Faltante",
            "role": "principal",
        },
        {
            "label": "Evaluaciones saltos (CMJ/SJ/DJ/IMTP)",
            "ok": _has_rows(state.get("jump_df")),
            "status": "Activo" if _has_rows(state.get("jump_df")) else "Faltante",
            "role": "principal",
        },
    ]
    rows.append(
        {
            "label": "Rep/Load (legacy opcional)",
            "ok": rep_load_loaded,
            "status": "Legacy disponible" if rep_load_loaded else "Opcional no cargado",
            "role": "legacy_optional",
            "detail": "Fuente legacy. Usar Raw Data Report - Workouts como fuente oficial para carga externa.",
        }
    )
    return rows


def report_requires_individual(audience: str | None) -> bool:
    return normalize_report_audience(audience) in {"atleta", "cliente"}


def resolve_report_scope(
    state: dict[str, pd.DataFrame | None],
    report_athlete: str = "Todos",
    report_audience: str = "profe",
) -> str | None:
    audience = normalize_report_audience(report_audience)
    athletes = collect_report_athletes(state)
    clean_athlete = str(report_athlete or "Todos").strip() or "Todos"

    if report_requires_individual(audience):
        if not athletes:
            return None
        if clean_athlete == "Todos":
            return athletes[0]
        return clean_athlete

    if clean_athlete == "Todos":
        return "Todos"
    return clean_athlete


def _selected_athletes(state: dict[str, pd.DataFrame | None], report_athlete: str) -> list[str]:
    athletes = collect_report_athletes(state)
    if report_athlete != "Todos":
        return [report_athlete]
    return athletes


def _latest_acwr_row(state: dict[str, pd.DataFrame | None], athlete: str) -> pd.Series | None:
    acwr_dict = state.get("acwr_dict") or {}
    acwr_df = acwr_dict.get(athlete)
    if acwr_df is None or acwr_df.empty:
        return None
    latest = acwr_df[acwr_df["sRPE_diario"] > 0].tail(1)
    if latest.empty:
        return None
    return latest.iloc[-1]


def _latest_mono_row(state: dict[str, pd.DataFrame | None], athlete: str) -> pd.Series | None:
    mono_dict = state.get("mono_dict") or {}
    mono_df = mono_dict.get(athlete)
    if mono_df is None or mono_df.empty:
        return None
    latest = mono_df.tail(1)
    if latest.empty:
        return None
    return latest.iloc[-1]


def _recent_wellness_mean(state: dict[str, pd.DataFrame | None], athlete: str) -> float | None:
    wdf = state.get("wellness_df")
    if wdf is None or wdf.empty or "Athlete" not in wdf.columns:
        return None
    athlete_df = wdf[wdf["Athlete"] == athlete].sort_values("Date").tail(3)
    if athlete_df.empty:
        return None
    score_series = _report_wellness_score_series(athlete_df).dropna()
    if score_series.empty:
        return None
    return round(float(score_series.mean()), 1)


def _latest_jump_row(state: dict[str, pd.DataFrame | None], athlete: str) -> pd.Series | None:
    jdf = state.get("jump_df")
    if jdf is None or jdf.empty or "Athlete" not in jdf.columns:
        return None
    athlete_df = jdf[jdf["Athlete"] == athlete].sort_values("Date")
    if athlete_df.empty:
        return None
    return athlete_df.iloc[-1]


def _cmj_delta_vs_baseline(state: dict[str, pd.DataFrame | None], athlete: str) -> float | None:
    # Product v1 baseline: fixed per variable as the mean of the first
    # three valid measurements. Function name/column stay for report
    # compatibility, but the old full-history CMJ mean is deprecated.
    jdf = state.get("jump_df")
    if jdf is None or jdf.empty or "Athlete" not in jdf.columns or "CMJ_cm" not in jdf.columns:
        return None
    athlete_df = jdf[jdf["Athlete"] == athlete].sort_values("Date")
    if athlete_df.empty or "Date" not in athlete_df.columns:
        return None
    latest_date = pd.to_datetime(athlete_df["Date"], errors="coerce").dropna()
    if latest_date.empty:
        return None
    baseline_df = compute_baseline_delta(athlete_df, latest_date.max(), variables=["CMJ_cm"])
    if baseline_df.empty:
        return None
    cmj_row = baseline_df.iloc[0]
    if cmj_row.get("Signal") in {"baseline insuficiente", "sin dato actual"}:
        return None
    delta_pct = pd.to_numeric(pd.Series([cmj_row.get("Delta_pct")]), errors="coerce").iloc[0]
    if pd.isna(delta_pct):
        return None
    return round(float(delta_pct), 1)


def _team_completion_mean(state: dict[str, pd.DataFrame | None]) -> float | None:
    return _completion_snapshot(state, "Todos").get("numeric")


def _athlete_completion_mean(state: dict[str, pd.DataFrame | None], athlete: str) -> float | None:
    cdf = state.get("completion_df")
    if cdf is None or cdf.empty:
        return None
    if "Athlete" in cdf.columns:
        athlete_df = cdf[cdf["Athlete"] == athlete]
        if athlete_df.empty:
            return None
    return _completion_snapshot(state, athlete).get("numeric")


def _completion_plot_df(state: dict[str, pd.DataFrame | None], athlete: str = "Todos") -> pd.DataFrame:
    cdf = state.get("completion_df")
    if cdf is None or cdf.empty or "Date" not in cdf.columns:
        return pd.DataFrame(columns=["Date", "Pct"])

    result = cdf.copy()
    result["Date"] = pd.to_datetime(result["Date"], errors="coerce")
    result = result.dropna(subset=["Date"])
    if result.empty:
        return pd.DataFrame(columns=["Date", "Pct"])

    if athlete != "Todos" and "Athlete" in result.columns:
        result = result[result["Athlete"].astype(str).str.strip() == athlete]
        if result.empty:
            return pd.DataFrame(columns=["Date", "Pct"])

    grouped = summarize_completion_by_group(result, "Date", value_column="Pct")
    if grouped.empty:
        return pd.DataFrame(columns=["Date", "Pct"])
    return grouped[["Date", "Pct"]].sort_values("Date").reset_index(drop=True)


def _report_weekly_summaries(state: dict[str, pd.DataFrame | None]) -> dict[str, pd.DataFrame]:
    cached = state.get("weekly_summaries")
    if isinstance(cached, dict) and {"weekly_load", "weekly_wellness", "weekly_external", "weekly_team"}.issubset(cached.keys()):
        return cached
    return build_weekly_summaries(
        state.get("rpe_df"),
        state.get("wellness_df"),
        state.get("raw_df"),
        acwr_dict=state.get("acwr_dict"),
        prepared_raw_df=state.get("prepared_raw_df"),
    )


def _report_quality_report(state: dict[str, pd.DataFrame | None]) -> dict[str, object]:
    athletes = collect_report_athletes(state)
    return compute_data_quality_report(
        state.get("rpe_df"),
        state.get("wellness_df"),
        state.get("completion_df"),
        state.get("prepared_raw_df") if state.get("prepared_raw_df") is not None else state.get("raw_df"),
        state.get("maxes_df"),
        state.get("jump_df"),
        athletes,
    )


def _normalize_weekly_frame(frame: pd.DataFrame | None) -> pd.DataFrame:
    if frame is None or frame.empty:
        return pd.DataFrame()
    result = frame.copy()
    if "week_start" in result.columns:
        result["week_start"] = pd.to_datetime(result["week_start"], errors="coerce")
    if "is_current_week" in result.columns:
        result["is_current_week"] = result["is_current_week"].fillna(False).astype(bool)
    return result


def _current_week_slice(frame: pd.DataFrame | None, athlete: str = "Todos") -> pd.DataFrame:
    result = _normalize_weekly_frame(frame)
    if result.empty:
        return result
    if athlete != "Todos" and "Athlete" in result.columns:
        result = result[result["Athlete"].astype(str).str.strip() == athlete]
        if result.empty:
            return result
    if "is_current_week" in result.columns:
        current = result[result["is_current_week"].fillna(False)]
        if not current.empty:
            return current.copy()
    if "week_start" in result.columns and result["week_start"].notna().any():
        latest_week = result["week_start"].dropna().max()
        return result[result["week_start"] == latest_week].copy()
    return result.tail(1).copy()


def _completion_snapshot(
    state: dict[str, pd.DataFrame | None],
    athlete: str = "Todos",
    *,
    today: pd.Timestamp | None = None,
) -> dict[str, object]:
    cdf = state.get("completion_df")
    if cdf is None or cdf.empty or "Date" not in cdf.columns:
        return {"value": "Sin dato", "detail": "Sin completion cargado.", "numeric": None, "period_label": "Sin dato"}

    result = cdf.copy()
    result["Date"] = pd.to_datetime(result["Date"], errors="coerce")
    if athlete != "Todos" and "Athlete" in result.columns:
        result = result[result["Athlete"].astype(str).str.strip() == athlete]
    if result.empty:
        detail = "Sin completion para el atleta seleccionado." if athlete != "Todos" else "Sin completion visible."
        return {"value": "Sin dato", "detail": detail, "numeric": None, "period_label": "Sin dato"}

    dated_result = result.dropna(subset=["Date"]).copy()
    if dated_result.empty:
        completion_result = calculate_completion_rate(result)
        if completion_result.value is None:
            return {"value": "Sin dato", "detail": "Periodo cargado (sin fecha)", "numeric": None, "period_label": "Sin dato"}
        completion_mean = round(float(completion_result.value), 1)
        return {
            "value": f"{completion_mean:.1f}%",
            "detail": "Periodo cargado (sin fecha)",
            "numeric": completion_mean,
            "period_label": "Periodo cargado (sin fecha)",
        }
    result = dated_result

    today_ts = (pd.Timestamp(today) if today is not None else pd.Timestamp(datetime.now())).normalize()
    current_week = today_ts - pd.Timedelta(days=today_ts.weekday())
    current = result[
        (result["Date"].dt.normalize() >= current_week)
        & (result["Date"].dt.normalize() <= today_ts)
    ].copy()
    period_label = "Semana actual"
    if current.empty:
        latest_date = result["Date"].max().normalize()
        current = result[result["Date"].dt.normalize().eq(latest_date)].copy()
        period_label = f"Ultima fecha util ({latest_date:%d/%m})"

    completion_result = calculate_completion_rate(current)
    if completion_result.value is None:
        return {"value": "Sin dato", "detail": period_label, "numeric": None, "period_label": period_label}
    completion_mean = round(float(completion_result.value), 1)
    return {
        "value": f"{completion_mean:.1f}%",
        "detail": period_label,
        "numeric": completion_mean,
        "period_label": period_label,
    }


def _readiness_payload(quality_report: dict[str, object]) -> dict[str, object]:
    """Readiness real derivado del quality report, no inventario local de archivos."""
    dataset_summary = quality_report.get("dataset_summary", pd.DataFrame())
    alerts = quality_report.get("alerts", [])
    if dataset_summary is None or dataset_summary.empty:
        return {
            "label": "Limitado",
            "detail": "Sin fuentes validas para evaluar readiness.",
            "loaded_count": 0,
            "partial_count": 0,
            "total_count": 0,
            "alerts_count": len(alerts),
        }

    statuses = dataset_summary["Estado"].fillna("").astype(str).str.lower()
    loaded_count = int(statuses.str.contains("cargado").sum())
    partial_count = int(statuses.str.contains("parcial").sum())
    total_count = int(len(dataset_summary))
    alerts_count = int(len(alerts))

    if loaded_count >= 4 and alerts_count <= 1 and partial_count <= 1:
        label = "Listo"
    elif loaded_count >= 2:
        label = "Parcial"
    else:
        label = "Limitado"

    detail = f"{loaded_count}/{total_count} fuentes listas · {partial_count} parciales · {alerts_count} alerta(s)."
    return {
        "label": label,
        "detail": detail,
        "loaded_count": loaded_count,
        "partial_count": partial_count,
        "total_count": total_count,
        "alerts_count": alerts_count,
    }


def _quality_athlete_row(quality_report: dict[str, object], athlete: str) -> pd.Series | None:
    athlete_summary = quality_report.get("athlete_summary", pd.DataFrame())
    if athlete_summary is None or athlete_summary.empty or "Atleta" not in athlete_summary.columns:
        return None
    subset = athlete_summary[athlete_summary["Atleta"].astype(str).str.strip() == athlete]
    if subset.empty:
        return None
    return subset.iloc[0]


def _cmj_series(state: dict[str, pd.DataFrame | None], athlete: str) -> list[tuple[str, float]]:
    jdf = state.get("jump_df")
    if jdf is None or jdf.empty or "Athlete" not in jdf.columns or "CMJ_cm" not in jdf.columns:
        return []
    athlete_df = (
        jdf[jdf["Athlete"] == athlete]
        .dropna(subset=["CMJ_cm"])
        .sort_values("Date")
        .tail(8)
    )
    return [
        (pd.to_datetime(row["Date"]).strftime("%d/%m"), float(row["CMJ_cm"]))
        for _, row in athlete_df.iterrows()
    ]


def _acwr_series(state: dict[str, pd.DataFrame | None], athlete: str) -> list[tuple[str, float]]:
    acwr_dict = state.get("acwr_dict") or {}
    adf = acwr_dict.get(athlete)
    if adf is None or adf.empty or "ACWR_EWMA" not in adf.columns:
        return []
    filtered = adf.dropna(subset=["ACWR_EWMA"]).tail(10)
    return [
        (pd.to_datetime(row["Date"]).strftime("%d/%m"), float(row["ACWR_EWMA"]))
        for _, row in filtered.iterrows()
    ]


def _wellness_series(state: dict[str, pd.DataFrame | None], athlete: str) -> list[tuple[str, float]]:
    wdf = state.get("wellness_df")
    if wdf is None or wdf.empty or "Athlete" not in wdf.columns:
        return []
    athlete_df = (
        wdf[wdf["Athlete"] == athlete]
        .sort_values("Date")
        .tail(10)
    )
    athlete_df = athlete_df.copy()
    athlete_df["__report_wellness_score"] = _report_wellness_score_series(athlete_df)
    athlete_df = athlete_df.dropna(subset=["__report_wellness_score"])
    return [
        (pd.to_datetime(row["Date"]).strftime("%d/%m"), float(row["__report_wellness_score"]))
        for _, row in athlete_df.iterrows()
    ]


def _round_or_none(value, digits: int = 1):
    if value is None or pd.isna(value):
        return None
    return round(float(value), digits)


def _coerce_float(value) -> float | None:
    if value in [None, "", "—", "-"]:
        return None
    try:
        numeric = float(value)
    except (TypeError, ValueError):
        return None
    if pd.isna(numeric):
        return None
    return numeric


def safe_value(value: object, fallback: str = PDF_MISSING_TEXT) -> str:
    """Render scalar values safely for reports without leaking NaN/None."""
    if value is None:
        return fallback
    try:
        if pd.isna(value):
            return fallback
    except (TypeError, ValueError):
        pass
    text = str(value).strip()
    if text in {"", "-", "—", "â€”", "nan", "NaN", "None", "<NA>", "NaT", "Sin dato", "sin dato"}:
        return fallback
    return text


def _professional_name_key(value: object) -> str:
    text = unicodedata.normalize("NFKD", str(value or ""))
    text = "".join(char for char in text if not unicodedata.combining(char))
    return " ".join(text.casefold().strip().split())


def _professional_athlete_mask(series: pd.Series, athlete: str) -> pd.Series:
    target = _professional_name_key(athlete)
    return series.apply(_professional_name_key).eq(target)


def _professional_quadrant_display_name(value: object, *, max_chars: int = 26) -> str:
    text = " ".join(str(value or "").split()).strip()
    if not text:
        return "Atleta"
    if len(text) <= max_chars:
        return text
    parts = text.split()
    if len(parts) >= 2:
        simplified = f"{parts[0]} {parts[-1]}"
        if len(simplified) <= max_chars:
            return simplified
    return textwrap.shorten(text, width=max_chars, placeholder="…")


def _has_text(value: object) -> bool:
    return _ascii_text(value).strip() not in {"", "-", "—"}


def _row_has_eval_data(row: pd.Series | None) -> bool:
    if row is None:
        return False
    numeric_cols = ["CMJ cm", "CMJ vs BL %", EUR_RATIO_LABEL, "EUR", "DRI", "IMTP N"]
    return any(_coerce_float(row.get(col)) is not None for col in numeric_cols) or _has_text(row.get("Perfil NM"))


def _row_has_load_data(row: pd.Series | None) -> bool:
    if row is None:
        return False
    return any(
        [
            _coerce_float(row.get("ACWR EWMA")) is not None,
            _coerce_float(row.get("Monotonia")) is not None,
            _has_text(row.get("Zona")),
        ]
    )


def _row_has_wellness_data(row: pd.Series | None) -> bool:
    return row is not None and _coerce_float(row.get("Wellness 3d")) is not None


def _focus_completion_value(state: dict[str, pd.DataFrame | None], report_athlete: str) -> float | None:
    return (
        _athlete_completion_mean(state, report_athlete)
        if report_athlete != "Todos" else
        _team_completion_mean(state)
    )


def _profile_text(row: pd.Series, fallback: str = "-") -> str:
    text = _ascii_text(row.get("Perfil NM")).strip()
    return text or fallback


def _compact_lines(items: list[str | None]) -> list[str]:
    return [item for item in items if item and _ascii_text(item).strip()]


def _sentence_fragment(value: object, *, lowercase: bool = False) -> str:
    text = _professional_visible_metric_text(value).strip()
    text = re.sub(r"[\s\.;:!,?]+$", "", text)
    return text.lower() if lowercase else text


def _count_phrase(count: int, singular: str, plural: str | None = None) -> str:
    label = singular if count == 1 else (plural or f"{singular}s")
    return f"{count} {label}"


def _focus_metric_phrase(label: str, value: object, *, digits: int | None = None, suffix: str = "") -> str:
    numeric = _coerce_float(value)
    if numeric is None:
        return ""
    return _professional_visible_metric_text(
        f"{label} {_display_metric(numeric, digits=digits, suffix=suffix)}"
    ).strip()


def _focus_metric_context(row: pd.Series) -> dict[str, str]:
    profile = _professional_visible_metric_text(row.get("Perfil NM")).strip()
    zone = _display_zone(row.get("Zona")) if _has_text(row.get("Zona")) else ""
    acwr = _coerce_float(row.get("ACWR EWMA"))
    cmj_delta = _coerce_float(row.get("CMJ vs BL %"))
    return {
        "profile": profile,
        "profile_norm": _professional_normalized_text(profile),
        "zone": _professional_visible_metric_text(zone).strip() if zone else "",
        "load_phrase": _professional_visible_metric_text(
            f"ACWR EWMA {_display_metric(acwr, digits=2)} en zona {zone}"
        ).strip() if acwr is not None and zone else (
            _professional_visible_metric_text(f"ACWR EWMA {_display_metric(acwr, digits=2)}").strip()
            if acwr is not None else (
                _professional_visible_metric_text(f"zona {zone}").strip() if zone else ""
            )
        ),
        "cmj": _focus_metric_phrase("CMJ", row.get("CMJ cm"), suffix=" cm"),
        "cmj_delta": _focus_metric_phrase("CMJ vs BL", row.get("CMJ vs BL %"), digits=1, suffix="%"),
        "dri": _focus_metric_phrase("DRI", row.get("DRI"), digits=2),
        "eur": _focus_metric_phrase("EUR", _summary_eur_value(row), digits=2),
        "imtp": _focus_metric_phrase("IMTP", row.get("IMTP N"), suffix=" N"),
        "wellness": _focus_metric_phrase("Wellness 3d", row.get("Wellness 3d"), digits=1),
        "cmj_delta_value": _display_metric(cmj_delta, digits=1, suffix="%") if cmj_delta is not None else "",
    }


def _current_focus_text(row: pd.Series, *, audience: str) -> str:
    profile = _ascii_text(row.get("Perfil NM")).lower().strip()
    acwr = _coerce_float(row.get("ACWR EWMA"))
    cmj_delta = _coerce_float(row.get("CMJ vs BL %"))
    eval_available = _row_has_eval_data(row)

    if not eval_available:
        return "Hacer nueva evaluación" if audience == "cliente" else "Nueva evaluación"
    if "poca base" in profile:
        return "Construir fuerza base" if audience == "cliente" else "Fuerza base"
    if cmj_delta is not None and cmj_delta <= -5:
        return "Recuperar potencia" if audience == "cliente" else "Recuperar CMJ"
    if acwr is not None and acwr > 1.5:
        return "Ordenar carga" if audience == "cliente" else "Regular carga"
    if acwr is not None and acwr < 0.8:
        return "Ganar continuidad" if audience == "cliente" else "Subir estimulo"
    if audience == "cliente":
        return "Sostener progreso"
    if audience == "atleta":
        return "Sostener progreso"
    return "Sostener perfil"


def _objective_focuses_from_row(row: pd.Series, *, audience: str) -> list[str]:
    focus = _current_focus_text(row, audience=audience)
    profile = _ascii_text(row.get("Perfil NM")).lower().strip()
    acwr = _coerce_float(row.get("ACWR EWMA"))
    cmj_delta = _coerce_float(row.get("CMJ vs BL %"))
    eval_available = _row_has_eval_data(row)

    if not eval_available:
        return [
            "Necesitamos una referencia objetiva para ordenar mejor el siguiente bloque.",
            "La próxima evaluación va a permitir medir cambios reales y ajustar el plan.",
        ]
    if "poca base" in profile:
        return [
            "Esta prioridad busca mejorar la base de fuerza que sostiene el resto del perfil.",
            "Si responde bien, después va a tener más sentido pedir reactividad o velocidad.",
        ]
    if cmj_delta is not None and cmj_delta <= -5:
        return [
            "La prioridad es recuperar calidad de expresión antes de volver a escalar el trabajo.",
            "El próximo control debería mostrar una mejor respuesta de salto y disponibilidad.",
        ]
    if acwr is not None and acwr > 1.5:
        return [
            "El foco inmediato es bajar ruido y recuperar tolerancia para volver a producir bien.",
            "Si la carga se ordena, las próximas semanas deberían mostrar mejor disponibilidad.",
        ]
    if acwr is not None and acwr < 0.8:
        return [
            "La prioridad es recuperar continuidad de estímulo sin perder control del proceso.",
            "El próximo tramo debería dar una señal más clara sobre la adaptación real.",
        ]
    return [
        f"El foco inmediato es {focus.lower()} sin perder calidad en el resto de las variables.",
        "La próxima revisión debería confirmar si esta prioridad ya se está trasladando al rendimiento.",
    ]


def _current_focus_text(row: pd.Series, *, audience: str) -> str:
    audience = normalize_report_audience(audience)
    context = _focus_metric_context(row)
    profile = context["profile_norm"]
    acwr = _coerce_float(row.get("ACWR EWMA"))
    cmj_delta = _coerce_float(row.get("CMJ vs BL %"))
    eval_available = _row_has_eval_data(row)

    if not eval_available:
        if _has_text(context["load_phrase"]):
            if audience == "cliente":
                return _professional_visible_metric_text(f"Sostener carga actual con {context['load_phrase']}")
            if acwr is not None and acwr < 0.8:
                return _professional_visible_metric_text(f"Recuperar continuidad de carga con {context['load_phrase']}")
            return _professional_visible_metric_text(f"Operar con {context['load_phrase']}")
        return _professional_visible_metric_text("Completar evaluación")

    if cmj_delta is not None and cmj_delta <= -5:
        if audience == "cliente":
            metric = context["cmj"] or "CMJ"
            return _professional_visible_metric_text(f"Recuperar salto vertical con {metric}")
        metric = context["cmj"] or "CMJ"
        delta = context["cmj_delta"] or "CMJ vs BL"
        return _professional_visible_metric_text(f"Recuperar {metric} tras {delta}")

    if "poca base" in profile or "fuerza" in profile or "base" in profile:
        anchor = context["imtp"] or context["cmj"] or "fuerza base"
        if audience == "cliente":
            return _professional_visible_metric_text(f"Construir base con {anchor}")
        if audience == "atleta":
            return _professional_visible_metric_text(f"Construir fuerza base con {anchor}")
        return _professional_visible_metric_text(f"Construir fuerza base desde {anchor}")

    if "react" in profile:
        anchors = [metric for metric in (context["dri"], context["eur"], context["cmj"]) if _has_text(metric)]
        if audience == "cliente":
            anchor = context["cmj"] or context["dri"] or "reactividad"
            return _professional_visible_metric_text(f"Sostener lo mejor del perfil con {anchor}")
        if len(anchors) >= 2:
            return _professional_visible_metric_text(f"Sostener reactividad con {anchors[0]} y {anchors[1]}")
        if anchors:
            return _professional_visible_metric_text(f"Sostener reactividad con {anchors[0]}")

    anchors = [metric for metric in (context["cmj"], context["dri"], context["eur"], context["imtp"]) if _has_text(metric)]
    if audience == "cliente":
        if anchors:
            return _professional_visible_metric_text(f"Sostener progreso con {anchors[0]}")
        if _has_text(context["load_phrase"]):
            return _professional_visible_metric_text(f"Sostener continuidad con {context['load_phrase']}")
        return _professional_visible_metric_text("Completar evaluación")
    if audience == "atleta":
        if len(anchors) >= 2:
            return _professional_visible_metric_text(f"Sostener el perfil con {anchors[0]} y {anchors[1]}")
        if anchors:
            return _professional_visible_metric_text(f"Sostener el perfil con {anchors[0]}")
        return _professional_visible_metric_text("Completar evaluación")
    if len(anchors) >= 2:
        return _professional_visible_metric_text(f"Sostener el perfil actual con {anchors[0]} y {anchors[1]}")
    if anchors:
        return _professional_visible_metric_text(f"Sostener el perfil actual con {anchors[0]}")
    if _has_text(context["profile"]):
        return _professional_visible_metric_text(f"Sostener el perfil {context['profile']}")
    return _professional_visible_metric_text("Completar evaluación")


def _objective_focuses_from_row(row: pd.Series, *, audience: str) -> list[str]:
    audience = normalize_report_audience(audience)
    focus = _current_focus_text(row, audience=audience)
    context = _focus_metric_context(row)
    profile = context["profile_norm"]
    acwr = _coerce_float(row.get("ACWR EWMA"))
    cmj_delta = _coerce_float(row.get("CMJ vs BL %"))
    eval_available = _row_has_eval_data(row)

    if not eval_available:
        load_phrase = context["load_phrase"]
        lines = [
            (
                f"Completar CMJ, SJ, DJ e IMTP mientras sostenemos {load_phrase} para ordenar la siguiente decisión."
                if _has_text(load_phrase)
                else "Completar CMJ, SJ, DJ e IMTP antes de definir una prioridad física del siguiente bloque."
            ),
            (
                f"La señal de que funciona es llegar a la próxima evaluación con {load_phrase} y una referencia física comparable."
                if _has_text(load_phrase)
                else "La señal de que funciona es llegar a la próxima evaluación con una referencia física comparable."
            ),
        ]
        if audience == "cliente":
            lines = [
                "Completar una nueva evaluación y ordenar la carga actual antes de sacar conclusiones más firmes."
                if _has_text(load_phrase)
                else "Completar una nueva evaluación antes de sacar conclusiones más firmes.",
                f"La señal de que funciona es sostener {load_phrase} y sumar una referencia clara del proceso."
                if _has_text(load_phrase)
                else "La señal de que funciona es sumar una referencia clara del proceso.",
            ]
        return [_professional_visible_metric_text(line) for line in lines]
    if "poca base" in profile or "fuerza" in profile or "base" in profile:
        anchor = context["imtp"] or context["cmj"] or "fuerza base"
        lines = [
            f"La prioridad es empujar {anchor} para sostener mejor el resto del perfil neuromuscular.",
            f"La señal de que funciona es que {anchor} acompañe sin caída de CMJ ni pérdida de disponibilidad.",
        ]
        if audience == "cliente":
            lines = [
                f"La prioridad es construir mejor base desde {anchor}.",
                "La señal de que funciona es sentir más consistencia en el salto y en cómo tolerás la carga.",
            ]
        return [_professional_visible_metric_text(line) for line in lines]
    if cmj_delta is not None and cmj_delta <= -5:
        cmj_text = context["cmj"] or "CMJ"
        delta_text = context["cmj_delta"] or "CMJ vs BL"
        lines = [
            f"La prioridad es recuperar {cmj_text} tras {delta_text} antes de volver a escalar el trabajo reactivo.",
            "La señal de que funciona es que el próximo control muestre mejor respuesta de salto y una carga mejor tolerada.",
        ]
        if audience == "cliente":
            lines = [
                f"La prioridad es recuperar {cmj_text} después de {delta_text}.",
                "La señal de que funciona es volver a ver un salto más sólido y mejores sensaciones para entrenar.",
            ]
        return [_professional_visible_metric_text(line) for line in lines]
    if acwr is not None and acwr > 1.5:
        load_phrase = context["load_phrase"] or "la carga actual"
        lines = [
            f"La primera decisión es ordenar {load_phrase} antes de pedir más reactividad o potencia.",
            f"La señal de que funciona es sostener mejor disponibilidad y que {load_phrase} salga de la zona alta.",
        ]
        if audience == "cliente":
            lines = [
                f"La prioridad es ordenar {load_phrase} antes de apretar más el entrenamiento.",
                f"La señal de que funciona es sentirte mejor para entrenar y sostener {load_phrase} con más control.",
            ]
        return [_professional_visible_metric_text(line) for line in lines]
    if acwr is not None and acwr < 0.8:
        load_phrase = context["load_phrase"] or "la carga actual"
        lines = [
            f"La prioridad es recuperar continuidad de estímulo con {load_phrase} antes de progresar hacia más reactividad.",
            f"La señal de que funciona es que {load_phrase} salga de subcarga y aparezca una lectura más estable del perfil.",
        ]
        if audience == "cliente":
            lines = [
                f"La prioridad es recuperar continuidad con {load_phrase}.",
                "La señal de que funciona es sostener mejor el entrenamiento y salir de una carga demasiado baja.",
            ]
        return [_professional_visible_metric_text(line) for line in lines]

    named_metrics = [metric for metric in (context["dri"], context["eur"], context["cmj"], context["imtp"]) if _has_text(metric)]
    primary_metric = named_metrics[0] if named_metrics else focus
    secondary_metric = named_metrics[1] if len(named_metrics) > 1 else (context["load_phrase"] or primary_metric)
    if audience == "cliente":
        return [
            _professional_visible_metric_text(f"La prioridad es sostener {primary_metric} sin perder continuidad en el proceso."),
            _professional_visible_metric_text(f"La señal de que funciona es ver que {secondary_metric} acompaña mejor en la próxima revisión."),
        ]
    return [
        _professional_visible_metric_text(f"La primera decisión es sostener {primary_metric} y ordenar el siguiente bloque alrededor de esa referencia."),
        _professional_visible_metric_text(f"La señal de que funciona es que {secondary_metric} acompañe mejor en la próxima revisión."),
    ]


def _technical_planning_focuses(row: pd.Series, completion_value: float | None) -> list[str]:
    focuses: list[str] = []
    profile = _ascii_text(row.get("Perfil NM")).lower().strip()
    acwr = _coerce_float(row.get("ACWR EWMA"))
    monotony = _coerce_float(row.get("Monotonia"))
    wellness = _coerce_float(row.get("Wellness 3d"))
    cmj_delta = _coerce_float(row.get("CMJ vs BL %"))

    if "poca base" in profile:
        focuses.append("Priorizar fuerza base y fuerza máxima antes de aumentar la densidad reactiva.")
    if cmj_delta is not None and cmj_delta <= -5:
        focuses.append("Controlar fatiga reciente y repetir control corto de CMJ antes de progresar volumen reactivo.")
    elif cmj_delta is not None and cmj_delta >= 5:
        focuses.append("La salida vertical acompaña; se puede sostener o transferir sin abrir demasiados frentes.")
    if acwr is not None and acwr > 1.5:
        focuses.append("Reducir densidad o exposición del próximo microciclo para salir de zona alta.")
    elif acwr is not None and acwr < 0.8:
        focuses.append("Subir continuidad de estímulo para salir de subcarga y ganar lectura más estable.")
    if monotony is not None and monotony > 2.0:
        focuses.append("Aumentar variabilidad semanal para bajar monotonía antes de seguir acumulando strain.")
    if wellness is not None and wellness < 15:
        focuses.append("Seguir recuperación diaria y cruzar wellness con adherencia para evitar falsas lecturas.")
    if completion_value is not None and completion_value < 70:
        focuses.append("La adherencia actual limita la interpretación del bloque; conviene intervenir primero sobre cumplimiento.")
    if not focuses:
        focuses.append("Sostener el bloque actual y recontrolar en la próxima ventana para confirmar tendencia.")
    focuses.append("En la próxima medición, volver a controlar CMJ, perfil neuromuscular y disponibilidad general.")
    return list(dict.fromkeys(focuses))[:4]


def build_executive_summary_df(
    state: dict[str, pd.DataFrame | None],
    report_athlete: str = "Todos",
    report_audience: str | None = None,
) -> pd.DataFrame:
    effective_athlete = (
        resolve_report_scope(state, report_athlete, report_audience)
        if report_audience is not None else
        report_athlete
    )
    if effective_athlete is None:
        return pd.DataFrame()

    rows: list[dict[str, object]] = []
    athletes = _selected_athletes(state, effective_athlete)

    for athlete in athletes:
        load_row = _latest_acwr_row(state, athlete)
        mono_row = _latest_mono_row(state, athlete)
        jump_row = _latest_jump_row(state, athlete)

        row = {"Atleta": athlete}
        if load_row is not None:
            row["Fecha carga"] = pd.to_datetime(load_row["Date"]).strftime("%d/%m/%Y")
            row["sRPE"] = _round_or_none(load_row.get("sRPE_diario"), 0)
            row["ACWR EWMA"] = _round_or_none(load_row.get("ACWR_EWMA"), 2)
            row["Zona"] = load_row.get("Zona")
        if mono_row is not None:
            row["Monotonia"] = _round_or_none(mono_row.get("Monotonia"), 2)
            row["Strain"] = _round_or_none(mono_row.get("Strain"), 0)

        row["Wellness 3d"] = _recent_wellness_mean(state, athlete)

        if jump_row is not None:
            row["Fecha evaluación"] = pd.to_datetime(jump_row["Date"]).strftime("%d/%m/%Y")
            row["CMJ cm"] = _round_or_none(jump_row.get("CMJ_cm"), 1)
            row["CMJ vs BL %"] = _cmj_delta_vs_baseline(state, athlete)
            row[EUR_RATIO_LABEL] = _round_or_none(jump_row.get("EUR"), 3)
            row["DRI"] = _round_or_none(jump_row.get("DRI"), 3)
            row["IMTP N"] = _round_or_none(jump_row.get("IMTP_N"), 0)
            row["Perfil NM"] = jump_row.get("NM_Profile")

        rows.append(row)

    if not rows:
        completion_mean = _team_completion_mean(state)
        if completion_mean is None:
            return pd.DataFrame()
        summary_df = pd.DataFrame([{"Atleta": "Equipo", "Completion promedio": completion_mean}])
        summary_df.attrs["neuromuscular_profile_source"] = "unknown"
        summary_df.attrs["neuromuscular_profile_source_label"] = "Fuente no determinada"
        summary_df.attrs["neuromuscular_profile_source_note"] = (
            "No se pudo determinar una fuente de perfil NM para este resumen ejecutivo."
        )
        return summary_df

    summary_df = pd.DataFrame(rows)
    completion_mean = _team_completion_mean(state)
    if completion_mean is not None:
        summary_df["Completion equipo %"] = completion_mean
    
    # Convert all columns to string type to ensure Arrow serialization compatibility
    # This avoids mixing numeric and string types which causes Arrow errors
    for col in summary_df.columns:
        summary_df[col] = summary_df[col].apply(
            lambda x: "—" if pd.isna(x) or x is None else str(x) if not isinstance(x, str) else x
        )
    
    summary_df.attrs["neuromuscular_profile_source"] = "latest_valid_row"
    summary_df.attrs["neuromuscular_profile_source_label"] = "Ultima evaluacion valida"
    summary_df.attrs["neuromuscular_profile_source_note"] = (
        "El perfil NM del resumen ejecutivo usa la ultima fila de evaluacion disponible, "
        "no el snapshot compuesto."
    )
    return summary_df


def _quality_detail_text(athlete_quality: pd.Series | None) -> str | None:
    if athlete_quality is None:
        return None
    srpe_cov = _coerce_float(athlete_quality.get("% cobertura sRPE"))
    wellness_cov = _coerce_float(athlete_quality.get("% cobertura Wellness"))

    def coverage_text(label: str, value: float | None) -> str | None:
        if value is None:
            return None
        display_value = min(float(value), 100.0)
        suffix = " (registros adicionales)" if float(value) > 100.0 else ""
        return f"{label} {_display_metric(display_value, digits=0, suffix='%')}{suffix}"

    parts = _compact_lines(
        [
            coverage_text("sRPE", srpe_cov),
            coverage_text("Wellness", wellness_cov),
        ]
    )
    return " | ".join(parts) if parts else None


def _latest_eval_summary(summary_df: pd.DataFrame) -> tuple[str, str]:
    if summary_df.empty or "Fecha evaluación" not in summary_df.columns:
        return "Pendiente", "Sin evaluación útil dentro de la ventana visible."

    eval_rows = summary_df[summary_df["Fecha evaluación"].apply(_has_text)].copy()
    if eval_rows.empty:
        return "Pendiente", "Sin evaluación útil dentro de la ventana visible."

    eval_rows["_eval_date"] = pd.to_datetime(eval_rows["Fecha evaluación"], format="%d/%m/%Y", errors="coerce")
    eval_rows = eval_rows.dropna(subset=["_eval_date"])
    if eval_rows.empty:
        return "Pendiente", "Sin evaluación útil dentro de la ventana visible."

    latest_eval = eval_rows["_eval_date"].max()
    profiles = eval_rows["Perfil NM"].fillna("-").astype(str)
    profile_counts = profiles[profiles.apply(_has_text)].value_counts()
    profile_text = ", ".join(f"{name}: {count}" for name, count in profile_counts.head(2).items())
    detail_parts = [
        f"{len(eval_rows)} atleta(s) con evaluación visible" if "Atleta" in eval_rows.columns else None,
        profile_text if profile_text else None,
    ]
    return latest_eval.strftime("%d/%m/%Y"), " | ".join(part for part in detail_parts if part)


def _session_notes_source(state: dict[str, pd.DataFrame | None]) -> pd.DataFrame:
    df = state.get("session_notes_df")
    if df is None or df.empty or "Date" not in df.columns or "Athlete" not in df.columns:
        return pd.DataFrame()

    result = df.copy()
    result["Date"] = pd.to_datetime(result["Date"], errors="coerce").dt.normalize()
    if "Date_Assigned" in result.columns:
        result["Date_Assigned"] = pd.to_datetime(result["Date_Assigned"], errors="coerce").dt.normalize()
    result["Athlete"] = result["Athlete"].astype(str).str.strip().str.title()
    return result.dropna(subset=["Date", "Athlete"])


def _session_notes_for_scope(
    state: dict[str, pd.DataFrame | None],
    report_athlete: str,
    *,
    days: int | None = 42,
    max_rows: int | None = 6,
) -> pd.DataFrame:
    df = _session_notes_source(state)
    if df.empty:
        return df

    if report_athlete != "Todos":
        target = str(report_athlete).strip().casefold()
        df = df[df["Athlete"].astype(str).str.strip().str.casefold() == target]
        if df.empty:
            return df

    if days is not None and "Date" in df.columns:
        reference_date = pd.to_datetime(df["Date"], errors="coerce").max()
        if pd.notna(reference_date):
            window_start = pd.Timestamp(reference_date).normalize() - pd.Timedelta(days=max(days - 1, 0))
            df = df[df["Date"] >= window_start]

    sort_cols = [column for column in ["Date", "Athlete", "Source_Page"] if column in df.columns]
    ascending = [False, True, True][: len(sort_cols)]
    if sort_cols:
        df = df.sort_values(sort_cols, ascending=ascending, na_position="last")
    if max_rows is not None:
        df = df.head(max_rows)
    return df.reset_index(drop=True)


def _note_text(value: object, fallback: str = "-") -> str:
    if value is None or pd.isna(value):
        return fallback
    text = str(value).strip()
    return text if text else fallback


def _note_date_text(value: object) -> str:
    parsed = pd.to_datetime(value, errors="coerce")
    return parsed.strftime("%d/%m/%Y") if pd.notna(parsed) else "-"


def _session_note_summary(notes_df: pd.DataFrame) -> str:
    if notes_df.empty:
        return "Sin incidencias operativas recientes cargadas."
    categories = notes_df.get("Reason_Category", pd.Series(dtype=object)).dropna().astype(str)
    category_text = ", ".join(f"{name}: {count}" for name, count in categories.value_counts().head(2).items())
    latest = notes_df.iloc[0]
    latest_text = _compact_lines(
        [
            _note_text(latest.get("Opt_Out_Type")) if _has_text(latest.get("Opt_Out_Type")) else None,
            _note_text(latest.get("Assigned_Exercise")) if _has_text(latest.get("Assigned_Exercise")) else None,
        ]
    )
    return " | ".join(
        part for part in [
            f"{len(notes_df)} incidencia(s) recientes",
            category_text or None,
            f"Ultima: {', '.join(latest_text)}" if latest_text else None,
        ]
        if part
    )


def build_operational_context_sheet(
    state: dict[str, pd.DataFrame | None],
    report_athlete: str = "Todos",
    *,
    max_rows: int = 6,
) -> pd.DataFrame:
    notes_df = _session_notes_for_scope(state, report_athlete, days=42, max_rows=max_rows)
    if notes_df.empty:
        return pd.DataFrame()

    rows: list[dict[str, object]] = []
    for _, note in notes_df.iterrows():
        rows.append(
            {
                "Fecha": _note_date_text(note.get("Date")),
                "Atleta": _note_text(note.get("Athlete")),
                "Ejercicio asignado": _note_text(note.get("Assigned_Exercise")),
                "Tipo": _note_text(note.get("Opt_Out_Type")),
                "Categoria": _note_text(note.get("Reason_Category")),
                "Reemplazo": _note_text(note.get("Replacement_Exercise")),
                "Explicacion": _note_text(note.get("Explanation_Text")),
            }
        )
    return pd.DataFrame(rows)


def _build_session_notes_annex_sheet(
    state: dict[str, pd.DataFrame | None],
    report_athlete: str = "Todos",
) -> pd.DataFrame:
    notes_df = _session_notes_for_scope(state, report_athlete, days=None, max_rows=None)
    if notes_df.empty:
        return pd.DataFrame()
    ordered_columns = [
        "Date",
        "Athlete",
        "Date_Assigned",
        "Assigned_Exercise",
        "Opt_Out_Type",
        "Reason_Category",
        "Replacement_Exercise",
        "Explanation_Text",
        "Source",
        "Source_Page",
        "Raw_Text",
    ]
    return _preferred_columns(notes_df, ordered_columns)


def _build_raw_external_volume_sheet(
    state: dict[str, pd.DataFrame | None],
    report_athlete: str = "Todos",
) -> pd.DataFrame:
    prepared = state.get("prepared_raw_df")
    if prepared is None:
        prepared = prepare_raw_workouts_df(state.get("raw_df"))
    if prepared is None or prepared.empty:
        return pd.DataFrame()

    result = prepared.copy()
    if report_athlete != "Todos" and "Athlete" in result.columns:
        result = result[result["Athlete"].astype(str).str.strip() == report_athlete]
    if result.empty:
        return pd.DataFrame()

    result = result[
        ~result.get("is_invalid", pd.Series(False, index=result.index)).fillna(False)
        & ~result.get("is_untagged", pd.Series(False, index=result.index)).fillna(False)
    ].copy()
    if result.empty:
        return pd.DataFrame()

    if "Assigned Date" in result.columns:
        result["Date"] = pd.to_datetime(result["Assigned Date"], errors="coerce")
    elif "Date" in result.columns:
        result["Date"] = pd.to_datetime(result["Date"], errors="coerce")
    result["Source"] = "Raw Workouts"
    columns = [
        "Source",
        "Athlete",
        "Date",
        "Exercise",
        "Exercise Name",
        "Tags",
        "stimulus_category",
        "Category",
        "Result",
        "Reps",
        "Sets",
        "Volume_Load_kg",
        "Contacts",
        "Exposures",
        "Distance_m",
    ]
    available = [column for column in columns if column in result.columns]
    result = result[available].copy()
    sort_cols = [column for column in ["Athlete", "Date", "stimulus_category", "Exercise"] if column in result.columns]
    if sort_cols:
        result = result.sort_values(sort_cols, ascending=[True, False, True, True][: len(sort_cols)], na_position="last")
    return result.reset_index(drop=True)


def _build_rep_load_legacy_volume_sheet(
    state: dict[str, pd.DataFrame | None],
    report_athlete: str = "Todos",
) -> pd.DataFrame:
    df = state.get("rep_load_df")
    if df is None or df.empty:
        return pd.DataFrame()
    result = df.copy()
    if report_athlete != "Todos" and "Athlete" in result.columns:
        result = result[result["Athlete"].astype(str).str.strip() == report_athlete]
    if result.empty:
        return pd.DataFrame()
    result["Source"] = "Rep/Load legacy"
    return result.reset_index(drop=True)


def build_report_executive_sheet(
    state: dict[str, pd.DataFrame | None],
    report_athlete: str = "Todos",
    report_audience: str = "profe",
) -> pd.DataFrame:
    audience = normalize_report_audience(report_audience)
    effective_athlete = resolve_report_scope(state, report_athlete, audience)
    if effective_athlete is None:
        return pd.DataFrame()

    summary_df = build_executive_summary_df(state, effective_athlete, audience)
    quality_report = _report_quality_report(state)
    readiness = _readiness_payload(quality_report)
    alerts = quality_report.get("alerts", [])
    weekly_summaries = _report_weekly_summaries(state)
    completion_snapshot = _completion_snapshot(state, effective_athlete)

    rows: list[dict[str, object]] = []

    def add_row(block: str, indicator: str, value: object, detail: str) -> None:
        rows.append(
            {
                "Bloque": block,
                "Indicador": indicator,
                "Valor": value,
                "Detalle": detail,
            }
        )

    scope_label = effective_athlete if effective_athlete != "Todos" else "Equipo"
    add_row(
        "Alcance",
        "Reporte",
        scope_label,
        f"{report_audience_label(audience)} {'individual' if effective_athlete != 'Todos' else 'equipo'}".strip(),
    )
    add_row("Readiness", "Estado del sistema", readiness["label"], readiness["detail"])

    if effective_athlete == "Todos":
        current_team = _current_week_slice(weekly_summaries.get("weekly_team"))
        current_load = _current_week_slice(weekly_summaries.get("weekly_load"))
        team_row = current_team.tail(1).iloc[0] if not current_team.empty else pd.Series(dtype=object)
        risk_count = 0
        if not current_load.empty and "ACWR_EWMA_last" in current_load.columns:
            risk_series = pd.to_numeric(current_load["ACWR_EWMA_last"], errors="coerce")
            risk_count = int(risk_series.gt(1.3).sum())

        week_label = "Semana actual"
        athletes_active = _display_metric(team_row.get("athletes_active"), digits=0)
        load_value = _display_metric(team_row.get("team_sRPE_sum"), digits=0)
        load_detail = " | ".join(
            part for part in [
                f"{athletes_active} atleta(s) activos" if athletes_active != "Sin dato" else None,
                f"Monotonía media {_display_metric(team_row.get('team_monotony_mean'), digits=2)}" if _coerce_float(team_row.get("team_monotony_mean")) is not None else None,
                f"{risk_count} en precaución/riesgo" if risk_count > 0 else None,
            ]
            if part
        ) or "Sin lectura semanal consolidada."
        add_row("Semana actual", "Carga del equipo", load_value, f"{week_label} | {load_detail}")

        wellness_value = _display_metric(team_row.get("team_wellness_mean"), digits=1)
        strain_detail = _compact_lines(
            [
                completion_snapshot["detail"] if completion_snapshot["numeric"] is not None else None,
                f"Strain medio {_display_metric(team_row.get('team_strain_mean'), digits=0)}" if _coerce_float(team_row.get("team_strain_mean")) is not None else None,
            ]
        )
        add_row("Semana actual", "Wellness promedio", wellness_value, " | ".join(strain_detail) or "Sin wellness consolidado.")

        add_row("Adherencia", "Completion equipo", completion_snapshot["value"], completion_snapshot["detail"])
        eval_value, eval_detail = _latest_eval_summary(summary_df)
        add_row("Evaluación útil", "Última evaluación visible", eval_value, eval_detail)
    else:
        current_load = _current_week_slice(weekly_summaries.get("weekly_load"), effective_athlete)
        current_wellness = _current_week_slice(weekly_summaries.get("weekly_wellness"), effective_athlete)
        load_row = current_load.tail(1).iloc[0] if not current_load.empty else pd.Series(dtype=object)
        wellness_row = current_wellness.tail(1).iloc[0] if not current_wellness.empty else pd.Series(dtype=object)
        athlete_quality = _quality_athlete_row(quality_report, effective_athlete)
        quality_detail = _quality_detail_text(athlete_quality)
        if athlete_quality is not None:
            add_row(
                "Readiness",
                "Cobertura del atleta",
                athlete_quality.get("Semaforo", "Sin dato"),
                quality_detail or "Sin cobertura reciente para resumir.",
            )

        weekly_value = _display_metric(load_row.get("weekly_sRPE"), digits=0)
        weekly_detail = " | ".join(
            part for part in [
                f"{_display_metric(load_row.get('sessions_count'), digits=0)} sesión(es)" if _coerce_float(load_row.get("sessions_count")) is not None else None,
                f"ACWR EWMA {_display_metric(load_row.get('ACWR_EWMA_last'), digits=2)}" if _coerce_float(load_row.get("ACWR_EWMA_last")) is not None else None,
                f"Monotonía {_display_metric(load_row.get('monotony'), digits=2)}" if _coerce_float(load_row.get("monotony")) is not None else None,
            ]
            if part
        ) or "Sin carga semanal útil para el atleta."
        add_row("Semana actual", "Carga semanal", weekly_value, f"Semana actual | {weekly_detail}")

        wellness_value = _display_metric(wellness_row.get("Wellness_mean"), digits=1)
        wellness_detail = " | ".join(
            part for part in [
                f"{_display_metric(wellness_row.get('wellness_days'), digits=0)} registro(s)" if _coerce_float(wellness_row.get("wellness_days")) is not None else None,
                f"Cumplimiento {_display_metric(_coerce_float(wellness_row.get('wellness_compliance')) * 100, digits=0, suffix='%')}" if _coerce_float(wellness_row.get("wellness_compliance")) is not None else None,
            ]
            if part
        ) or "Sin wellness útil para la semana actual."
        add_row("Semana actual", "Wellness promedio", wellness_value, wellness_detail)

        add_row("Adherencia", "Completion", completion_snapshot["value"], completion_snapshot["detail"])

        focus_row = summary_df.iloc[0] if not summary_df.empty else pd.Series(dtype=object)
        eval_value = focus_row.get("Fecha evaluación", "Pendiente")
        eval_detail = " | ".join(
            part for part in [
                f"CMJ {_display_metric(focus_row.get('CMJ cm'), digits=1, suffix=' cm')}" if _coerce_float(focus_row.get("CMJ cm")) is not None else None,
                f"CMJ vs BL {_display_metric(focus_row.get('CMJ vs BL %'), digits=1, suffix='%')}" if _coerce_float(focus_row.get("CMJ vs BL %")) is not None else None,
                f"Perfil {_profile_text(focus_row)}" if _has_text(focus_row.get("Perfil NM")) else None,
            ]
            if part
        ) or "Sin evaluación útil dentro de la ventana visible."
        add_row("Evaluación útil", "Último test", eval_value, eval_detail)

    operational_notes = _session_notes_for_scope(state, effective_athlete, days=42, max_rows=3)
    if not operational_notes.empty:
        add_row(
            "Contexto operativo",
            "Opt-outs / notes recientes",
            str(len(operational_notes)),
            _session_note_summary(operational_notes),
        )

    top_alert = _ascii_text(alerts[0]) if alerts else "Sin alertas activas en la ventana visible."
    add_row("Calidad", "Alertas activas", str(len(alerts)), top_alert)

    return pd.DataFrame(rows, columns=["Bloque", "Indicador", "Valor", "Detalle"])


def build_interpretation_sheet(
    state: dict[str, pd.DataFrame | None],
    report_athlete: str = "Todos",
    report_audience: str = "profe",
) -> pd.DataFrame:
    insights = generate_module_insights(state, report_athlete, report_audience)
    rows = []
    for module_name, payload in insights.items():
        rows.append(
            {
                "Modulo": module_name.title(),
                "Lectura": payload.get("summary", ""),
                "Próximos focos": " | ".join(payload.get("focuses", [])),
            }
        )
    return pd.DataFrame(rows)


def _build_report_metadata_df(
    state: dict[str, pd.DataFrame | None],
    report_athlete: str,
    report_audience: str,
    included_sections: list[str],
    *,
    include_technical_annex: bool = False,
) -> pd.DataFrame:
    active_datasets = [
        DATASET_LABELS.get(key, key)
        for key in MODERN_REPORT_DATASET_KEYS
        if state.get(key) is not None and not state.get(key).empty
    ]
    legacy_datasets = [
        DATASET_LABELS.get(key, key)
        for key in LEGACY_REPORT_DATASET_KEYS
        if state.get(key) is not None and not state.get(key).empty
    ]
    visible_athletes = collect_report_athletes(state)
    return pd.DataFrame(
        [
            {"Campo": "Reporte", "Valor": "Threshold S&C - Reporte de rendimiento"},
            {"Campo": "Alcance", "Valor": report_athlete},
            {"Campo": "Destinatario", "Valor": report_audience_label(report_audience)},
            {"Campo": "Modo exportacion", "Valor": "Curado + anexo tecnico" if include_technical_annex else "Curado"},
            {"Campo": "Generado", "Valor": datetime.now().strftime("%d/%m/%Y %H:%M")},
            {"Campo": "Ventana operativa", "Valor": "Últimas 6 semanas visibles"},
            {"Campo": "Atletas visibles", "Valor": len(visible_athletes)},
            {"Campo": "Fuentes modernas activas", "Valor": ", ".join(active_datasets) if active_datasets else "Sin fuentes modernas activas"},
            {"Campo": "Fuentes legacy opcionales", "Valor": ", ".join(legacy_datasets) if legacy_datasets else "Sin fuentes legacy"},
            {"Campo": "Secciones incluidas", "Valor": ", ".join(included_sections) if included_sections else "Sin secciones"},
        ]
    )


def build_report_sheets(
    state: dict[str, pd.DataFrame | None],
    report_athlete: str = "Todos",
    report_audience: str = "profe",
    *,
    include_technical_annex: bool = False,
    include_acwr: bool = True,
    include_mono: bool = True,
    include_wellness: bool = True,
    include_jumps: bool = True,
    include_maxes: bool = True,
    include_volume: bool = True,
    include_completion: bool = True,
) -> dict[str, pd.DataFrame]:
    audience = normalize_report_audience(report_audience)
    effective_athlete = resolve_report_scope(state, report_athlete, audience)
    if effective_athlete is None:
        return {}

    include_technical_annex = bool(include_technical_annex and audience == "profe")
    sheets: dict[str, pd.DataFrame] = {}
    included_sections: list[str] = []

    executive_df = build_report_executive_sheet(state, effective_athlete, audience)
    if not executive_df.empty:
        sheets["Resumen_Ejecutivo"] = executive_df
        included_sections.append("Resumen ejecutivo")

    interpretation_df = build_interpretation_sheet(state, effective_athlete, audience)
    if not interpretation_df.empty:
        sheets["Interpretacion"] = interpretation_df
        included_sections.append("Interpretacion")

    operational_context_df = build_operational_context_sheet(state, effective_athlete, max_rows=6)
    if not operational_context_df.empty:
        sheets["Contexto_Operativo"] = operational_context_df
        included_sections.append("Contexto operativo")

    if state.get("completion_df") is not None:
        completion_df = state["completion_df"]
        if effective_athlete != "Todos" and "Athlete" in completion_df.columns:
            completion_df = completion_df[completion_df["Athlete"].astype(str).str.strip() == effective_athlete]
        summary_df = _build_completion_summary_sheet(completion_df, effective_athlete)
        if not summary_df.empty:
            sheets["Completion_Resumen"] = summary_df
            included_sections.append("Completion resumen")

    if include_technical_annex:
        acwr_dict = state.get("acwr_dict") or {}
        mono_dict = state.get("mono_dict") or {}

        if include_acwr and acwr_dict:
            acwr_rows = []
            for athlete, athlete_df in acwr_dict.items():
                if effective_athlete != "Todos" and athlete != effective_athlete:
                    continue
                tmp = athlete_df.copy()
                tmp["Athlete"] = athlete
                acwr_rows.append(tmp)
            if acwr_rows:
                sheets["ACWR_sRPE"] = pd.concat(acwr_rows, ignore_index=True).round(2)
                included_sections.append("ACWR EWMA + sRPE")

        if include_mono and mono_dict:
            mono_rows = []
            for athlete, athlete_df in mono_dict.items():
                if effective_athlete != "Todos" and athlete != effective_athlete:
                    continue
                tmp = athlete_df.copy()
                tmp["Athlete"] = athlete
                mono_rows.append(tmp)
            if mono_rows:
                sheets["Monotonia_Strain"] = pd.concat(mono_rows, ignore_index=True).round(2)
                included_sections.append("Monotonia + Strain")

        if include_wellness and state.get("wellness_df") is not None:
            df = state["wellness_df"]
            if effective_athlete != "Todos" and "Athlete" in df.columns:
                df = df[df["Athlete"].astype(str).str.strip() == effective_athlete]
            sheets["Wellness"] = df.round(2)
            if not df.empty:
                included_sections.append("Wellness")

        if include_jumps and state.get("jump_df") is not None:
            df = state["jump_df"]
            if effective_athlete != "Todos" and "Athlete" in df.columns:
                df = df[df["Athlete"].astype(str).str.strip() == effective_athlete]
            sheets["Evaluaciones_Saltos"] = df.round(2)
            if not df.empty:
                included_sections.append("Evaluaciones")

        if include_maxes and state.get("maxes_df") is not None:
            df = state["maxes_df"]
            if effective_athlete != "Todos" and "Athlete" in df.columns:
                df = df[df["Athlete"].astype(str).str.strip() == effective_athlete]
            sheets["Maximos_Ejercicios"] = df
            if not df.empty:
                included_sections.append("Maximos")

        if include_volume:
            raw_volume_df = _build_raw_external_volume_sheet(state, effective_athlete)
            if not raw_volume_df.empty:
                sheets["Volumen_Carga_Externa"] = raw_volume_df
                included_sections.append("Carga externa Raw Workouts")
            else:
                legacy_volume_df = _build_rep_load_legacy_volume_sheet(state, effective_athlete)
                if not legacy_volume_df.empty:
                    sheets["Volumen_RepLoad_Legacy"] = legacy_volume_df
                    included_sections.append("Volumen Rep/Load legacy")

        if include_completion and state.get("completion_df") is not None:
            df = state["completion_df"]
            if effective_athlete != "Todos" and "Athlete" in df.columns:
                df = df[df["Athlete"].astype(str).str.strip() == effective_athlete]
            if not df.empty:
                sheets["Completion_Rate"] = df
                included_sections.append("Completion detalle")

        session_notes_annex_df = _build_session_notes_annex_sheet(state, effective_athlete)
        if not session_notes_annex_df.empty:
            sheets["Session_Notes"] = session_notes_annex_df
            included_sections.append("Opt-outs / Notes detalle")

    if sheets:
        sheets["Reporte_Meta"] = _build_report_metadata_df(
            state,
            effective_athlete,
            audience,
            included_sections,
            include_technical_annex=include_technical_annex,
        )

    return {name: df for name, df in sheets.items() if df is not None and not df.empty}


def _ordered_sheet_items(data_dict: dict[str, pd.DataFrame]) -> list[tuple[str, pd.DataFrame]]:
    ordered_names = [name for name in REPORT_SHEET_ORDER if name in data_dict]
    ordered_names.extend(name for name in data_dict if name not in ordered_names)
    return [(name, data_dict[name]) for name in ordered_names]


def _preferred_columns(df: pd.DataFrame, ordered_columns: list[str]) -> pd.DataFrame:
    selected = [column for column in ordered_columns if column in df.columns]
    selected.extend(column for column in df.columns if column not in selected)
    return df[selected]


def _normalize_date_like_columns(df: pd.DataFrame) -> pd.DataFrame:
    result = df.copy()
    for column in result.columns:
        col_text = str(column).lower()
        if "date" in col_text or "fecha" in col_text or "semana" in col_text:
            parsed = pd.to_datetime(result[column], errors="coerce")
            if parsed.notna().any():
                result[column] = parsed
    return result


def _sort_export_frame(df: pd.DataFrame, sort_columns: list[str], ascending: list[bool]) -> pd.DataFrame:
    valid_columns = [column for column in sort_columns if column in df.columns]
    if not valid_columns:
        return df.reset_index(drop=True)
    valid_ascending = ascending[: len(valid_columns)]
    return df.sort_values(valid_columns, ascending=valid_ascending, na_position="last").reset_index(drop=True)


def _build_completion_summary_sheet(df: pd.DataFrame | None, report_athlete: str) -> pd.DataFrame:
    if df is None or df.empty or "Date" not in df.columns:
        return pd.DataFrame()

    result = df.copy()
    result["Date"] = pd.to_datetime(result["Date"], errors="coerce")
    result = result.dropna(subset=["Date"])
    if result.empty:
        return pd.DataFrame()

    scope_label = report_athlete if report_athlete != "Todos" else "Equipo"
    grouped = summarize_completion_by_group(result, "Date", value_column="Pct_Promedio")
    if grouped.empty:
        return pd.DataFrame()
    session_counts = result.groupby("Date").size().reset_index(name="Sesiones")
    grouped = grouped.merge(session_counts, on="Date", how="left")
    grouped = grouped.sort_values("Date", ascending=False).reset_index(drop=True)
    grouped.insert(0, "Alcance", scope_label)
    if "Athlete" in result.columns:
        athlete_counts = (
            result.groupby("Date")["Athlete"]
            .nunique(dropna=True)
            .reset_index(name="Atletas")
        )
        grouped = grouped.merge(athlete_counts, on="Date", how="left")
    return grouped


def _prepare_export_frame(sheet_name: str, df: pd.DataFrame) -> pd.DataFrame:
    result = _normalize_date_like_columns(df)

    if sheet_name == "ACWR_sRPE":
        result = _preferred_columns(
            result,
            ["Athlete", "Date", "sRPE_diario", "Aguda_7d", "Cronica_28d", "ACWR", "EWMA_Aguda", "EWMA_Cronica", "ACWR_EWMA", "Zona"],
        )
        return _sort_export_frame(result, ["Athlete", "Date"], [True, False])
    if sheet_name == "Monotonia_Strain":
        result = _preferred_columns(
            result,
            ["Athlete", "Semana", "Carga_Semanal", "Monotonia", "Monotony_Status", "Monotony_Warning", "Strain"],
        )
        return _sort_export_frame(result, ["Athlete", "Semana"], [True, False])
    if sheet_name == "Wellness":
        result = _preferred_columns(
            result,
            ["Athlete", "Date", "Sueno_hs", "Estres", "Dolor", "Wellness_Score"],
        )
        return _sort_export_frame(result, ["Athlete", "Date"], [True, False])
    if sheet_name == "Evaluaciones_Saltos":
        result = result.drop(columns=[col for col in IMTP_LEGACY_RFD_EXPORT_COLUMNS if col in result.columns])
        result = _preferred_columns(
            result,
            [
                "Athlete",
                "Date",
                "CMJ_cm",
                "SJ_cm",
                "DJ_cm",
                "DJ_tc_ms",
                "EUR",
                "DRI",
                "IMTP_N",
                "NM_Profile",
                "IMTP_avg_N",
                "IMTP_force_L_N",
                "IMTP_force_R_N",
                "IMTP_asym_pct",
                "IMTP_pretension",
                "IMTP_time_max_s",
                *IMTP_FORCE_TIME_EXPORT_COLUMNS,
            ],
        )
        result = result.rename(columns={"EUR": EUR_RATIO_LABEL})
        return _sort_export_frame(result, ["Athlete", "Date"], [True, False])
    if sheet_name == "Maximos_Ejercicios":
        result = _preferred_columns(
            result,
            ["Athlete", "Exercise Name", "Added Date", "Max Value"],
        )
        return _sort_export_frame(result, ["Athlete", "Added Date", "Exercise Name"], [True, False, True])
    if sheet_name == "Volumen_Carga_Externa":
        result = _preferred_columns(
            result,
            [
                "Source",
                "Athlete",
                "Date",
                "Exercise",
                "Exercise Name",
                "Tags",
                "stimulus_category",
                "Category",
                "Result",
                "Reps",
                "Sets",
                "Volume_Load_kg",
                "Contacts",
                "Exposures",
                "Distance_m",
            ],
        )
        return _sort_export_frame(result, ["Athlete", "Date", "stimulus_category", "Exercise"], [True, False, True, True])
    if sheet_name in {"Volumen_RepLoad_Legacy", "Volumen_Sesion"}:
        result = _preferred_columns(
            result,
            ["Source", "Athlete", "Date", "Exercise", "Reps_Assigned", "Reps_Completed", "Load_kg"],
        )
        if "Date" not in result.columns and "Assigned Date" in result.columns:
            result = _preferred_columns(
                result,
                ["Source", "Athlete", "Assigned Date", "Exercise", "Reps_Assigned", "Reps_Completed", "Load_kg", "Volume_Load", "Category"],
            )
            return _sort_export_frame(result, ["Athlete", "Assigned Date", "Exercise"], [True, False, True])
        return _sort_export_frame(result, ["Athlete", "Date", "Exercise"], [True, False, True])
    if sheet_name == "Completion_Resumen":
        result = _preferred_columns(
            result,
            ["Alcance", "Date", "Pct_Promedio", "Sesiones", "Atletas"],
        )
        return _sort_export_frame(result, ["Date"], [False])
    if sheet_name == "Completion_Rate":
        result = _preferred_columns(
            result,
            ["Athlete", "Date", "Assigned", "Completed", "Pct"],
        )
        return _sort_export_frame(result, ["Athlete", "Date"], [True, False])
    return result.reset_index(drop=True)


def _format_excel_sheet(worksheet, df: pd.DataFrame, sheet_name: str) -> None:
    header_fill = PatternFill(fill_type="solid", fgColor="0D3C5E")
    header_font = Font(color="FEFEFE", bold=True)
    body_font = Font(color="221F20")
    thin_side = Side(style="thin", color="D8DEE4")
    border = Border(bottom=thin_side)
    stripe_fill = PatternFill(fill_type="solid", fgColor="F7FAFC")

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    worksheet.sheet_view.showGridLines = False
    worksheet.row_dimensions[1].height = 22

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    for row_idx, row in enumerate(worksheet.iter_rows(min_row=2), start=2):
        for cell in row:
            cell.font = body_font
            cell.alignment = Alignment(vertical="top", wrap_text=sheet_name in {"Interpretacion", "Reporte_Meta"})
            cell.border = border
            if row_idx % 2 == 0:
                cell.fill = stripe_fill

    date_like_cols = [
        idx for idx, col in enumerate(df.columns, start=1)
        if "date" in str(col).lower() or "fecha" in str(col).lower() or "semana" in str(col).lower()
    ]
    percent_like_cols = [
        idx for idx, col in enumerate(df.columns, start=1)
        if "pct" in str(col).lower() or "%" in str(col) or "completion promedio" in str(col).lower()
    ]
    numeric_like_cols = [
        idx for idx, col in enumerate(df.columns, start=1)
        if pd.api.types.is_numeric_dtype(df[col])
    ]

    for col_idx in date_like_cols:
        for cell in worksheet[get_column_letter(col_idx)][1:]:
            cell.number_format = "DD/MM/YYYY"

    for col_idx in numeric_like_cols:
        for cell in worksheet[get_column_letter(col_idx)][1:]:
            cell.alignment = Alignment(horizontal="right", vertical="top")

    for col_idx in percent_like_cols:
        for cell in worksheet[get_column_letter(col_idx)][1:]:
            cell.number_format = '0.0"%"'

    for idx, column in enumerate(df.columns, start=1):
        values = [str(column)] + ["" if pd.isna(value) else str(value) for value in df[column].head(200)]
        max_len = min(max(len(value) for value in values) + 2, 42)
        worksheet.column_dimensions[get_column_letter(idx)].width = max(12, max_len)

    if sheet_name == "Reporte_Meta":
        worksheet.column_dimensions["A"].width = 22
        worksheet.column_dimensions["B"].width = 72


def export_excel(data_dict: dict[str, pd.DataFrame]) -> bytes:
    """Export selected dataframes to a multi-sheet Excel workbook."""
    def _append_full_executive_page(target: list[object]) -> None:
        target.extend(
            [
                _p("Reporte profesional", "ProfMuted"),
                _p(report_athlete, "ProfTitle"),
                _p(f"Generado el {datetime.now():%d/%m/%Y %H:%M}", "ProfMuted"),
                _box(
                    [
                        _p("Reporte madre profesional orientado a toma de decisiones.", "ProfBody"),
                        _p(
                            "Prioriza perfil actual, cambios relevantes, contexto de carga y decisiÃ³n sugerida para el prÃ³ximo bloque.",
                            "ProfMuted",
                        ),
                    ]
                ),
                Spacer(1, 6 * mm),
                _p(executive_payload.get("title", "Resumen ejecutivo profesional"), "ProfSection"),
                _executive_summary_table(executive_payload),
                Spacer(1, 4 * mm),
                _bullet_box("SeÃ±ales clave", executive_payload.get("signals", [])),
                Spacer(1, 4 * mm),
                _box(
                    [
                        _p("DecisiÃ³n sugerida", "ProfCardTitle"),
                        _p(executive_payload.get("decision_suggested", PDF_MISSING_TEXT), "ProfBody"),
                    ],
                    padding=6,
                ),
            ]
        )

    def _append_full_composite_profile_page(target: list[object]) -> None:
        target.append(_p(composite_profile.get("title", "Perfil actual compuesto"), "ProfSection"))
        if composite_profile.get("state") == "missing":
            target.append(_collapsed_box(str(composite_profile.get("message") or "Faltan datos para el perfil compuesto.")))
            return
        target.append(_box([_p(composite_profile.get("note", PROFESSIONAL_COMPOSITE_PROFILE_NOTE), "ProfMuted")], padding=5))
        radar_image = _composite_radar_image(composite_profile)
        if radar_image is not None:
            target.append(Spacer(1, 3 * mm))
            target.append(radar_image)
        target.append(Spacer(1, 3 * mm))
        target.append(_dataframe_table(composite_profile.get("metric_table"), col_widths_mm=[36, 32, 22, 84]))
        feedback = composite_profile.get("feedback", {}) if isinstance(composite_profile.get("feedback"), dict) else {}
        target.append(Spacer(1, 3 * mm))
        target.append(
            _bullet_box(
                "Lectura del perfil",
                [
                    f"Variable dominante: {feedback.get('high', PDF_MISSING_TEXT)}",
                    f"Variable rezagada: {feedback.get('low', PDF_MISSING_TEXT)}",
                    f"Lectura fisiolÃ³gica: {feedback.get('physiological', PDF_MISSING_TEXT)}",
                    f"Lectura biomecÃ¡nica: {feedback.get('biomechanical', PDF_MISSING_TEXT)}",
                    f"Implicancia para prÃ³ximo bloque: {feedback.get('next_block', PDF_MISSING_TEXT)}",
                ],
            )
        )

    def _append_full_change_page(target: list[object]) -> None:
        target.append(_p(change_payload.get("title", "Cambios vs evaluaciÃ³n anterior"), "ProfSection"))
        if change_payload.get("state") == "missing":
            target.append(_collapsed_box(str(change_payload.get("message") or PROFESSIONAL_NO_EVOLUTION_TEXT)))
            return
        target.append(_dataframe_table(change_payload.get("display_table"), col_widths_mm=[34, 18, 18, 18, 18, 32, 36]))
        target.append(Spacer(1, 3 * mm))
        target.append(_bullet_box("SÃ­ntesis de cambios", change_payload.get("summary_lines", [])))
        if not _professional_any_quadrant_ready(quadrant_sections):
            target.append(Spacer(1, 3 * mm))
            target.append(_box([_p("Relaciones de perfil / cuadrantes: datos insuficientes para una ubicaciÃ³n Ãºtil en esta exportaciÃ³n.", "ProfMuted")], padding=5))

    def _append_full_quadrants_page(target: list[object]) -> None:
        target.append(_p("Relaciones de perfil / cuadrantes", "ProfSection"))
        if not _professional_any_quadrant_ready(quadrant_sections):
            target.append(_collapsed_box(PROFESSIONAL_NO_QUADRANTS_TEXT))
            return
        for section in quadrant_sections[:3]:
            chart = _quadrant_chart(section)
            if chart is None:
                continue
            target.append(_chart_explanation_panel(chart, str(section.get("title", "Cuadrante")), section))
            target.append(Spacer(1, 4 * mm))

    def _append_full_isometrics_page(target: list[object]) -> None:
        target.append(_p(isometric_payload.get("title", "IsomÃ©tricos y force-time avanzado"), "ProfSection"))
        if isometric_payload.get("state") == "missing":
            target.append(_collapsed_box(str(isometric_payload.get("message") or "Faltan datos isomÃ©tricos.")))
            return
        imtp_priority = [row for row in isometric_payload.get("imtp_rows", []) if row[0] in {"Peak Force", "Cambio relevante", "Fuerza relativa", "AsimetrÃ­a"}]
        if imtp_priority:
            target.append(_p("IMTP principal", "ProfCardTitle"))
            target.append(_mini_cards_table(imtp_priority))
        if isometric_payload.get("iso_available"):
            target.append(Spacer(1, 3 * mm))
            target.append(_p("ISO Push Hip-Hamstring Bilateral", "ProfCardTitle"))
            iso_priority = [row for row in isometric_payload.get("iso_rows", []) if row[0] in {"Peak Force", "Force Avg", "Time to Peak", "AsimetrÃ­a"}]
            if iso_priority:
                target.append(_mini_cards_table(iso_priority))
            if isometric_payload.get("iso_notes"):
                target.append(Spacer(1, 2 * mm))
                target.append(_bullet_box("Lectura prÃ¡ctica del test complementario", isometric_payload.get("iso_notes", [])[:3], style_name="ProfMuted"))
        if isometric_payload.get("force_time_available"):
            target.append(Spacer(1, 3 * mm))
            draw_force_time_test_block(
                {
                    "story": target,
                    "p": _p,
                    "box": _box,
                    "Table": Table,
                    "TableStyle": TableStyle,
                    "Spacer": Spacer,
                    "mm": mm,
                    "palette": palette,
                },
                force_time_payload,
                report_type="professional",
            )

    def _append_full_load_page(target: list[object]) -> None:
        target.append(_p(load_tolerance_payload.get("title", "Carga interna y tolerancia"), "ProfSection"))
        if load_tolerance_payload.get("state") == "missing":
            target.append(_collapsed_box(str(load_tolerance_payload.get("message") or "Faltan datos de carga interna.")))
            return
        target.append(_key_value_table(load_tolerance_payload.get("rows", [])))
        target.append(Spacer(1, 3 * mm))
        weekly_chart = _weekly_ema_chart(load_tolerance_payload.get("weekly_points", []), height_mm=48)
        if weekly_chart is not None:
            target.append(weekly_chart)
            target.append(Spacer(1, 3 * mm))
        target.append(_box([_p(load_tolerance_payload.get("risk_line", PDF_MISSING_TEXT), "ProfBody")], padding=6))

    def _append_full_wellness_page(target: list[object]) -> None:
        target.append(_p(wellness_availability_payload.get("title", "Wellness, disponibilidad y adherencia"), "ProfSection"))
        if wellness_availability_payload.get("state") == "missing":
            target.append(_collapsed_box(str(wellness_availability_payload.get("message") or "Faltan datos de wellness.")))
            return
        target.append(_mini_cards_table(wellness_availability_payload.get("rows", [])))
        chart_points = (
            wellness_availability_payload.get("weekly_points", [])
            if wellness_availability_payload.get("trend_allowed")
            else wellness_availability_payload.get("daily_points", [])
        )
        chart = _wellness_chart(chart_points, width_mm=174, height_mm=44)
        if chart is not None:
            target.append(Spacer(1, 3 * mm))
            target.append(chart)
        target.append(Spacer(1, 3 * mm))
        note_lines = [str(wellness_availability_payload.get("compatibility") or "")]
        if str(wellness_availability_payload.get("quality_note") or "").strip():
            note_lines.append(str(wellness_availability_payload.get("quality_note")))
        target.append(_bullet_box("Lectura de disponibilidad", note_lines, style_name="ProfMuted"))

    def _append_full_exposure_page(target: list[object]) -> None:
        target.append(_p(exposure_payload.get("title", "ExposiciÃ³n del bloque / contenido entrenado"), "ProfSection"))
        if exposure_payload.get("state") == "missing":
            target.append(_collapsed_box(str(exposure_payload.get("message") or "Faltan datos de exposiciÃ³n.")))
            return
        chart_image = _exposure_chart_image()
        if chart_image is not None:
            target.append(chart_image)
            target.append(Spacer(1, 3 * mm))
        target.append(_dataframe_table(exposure_payload.get("table"), col_widths_mm=[38, 30, 26, 80]))
        target.append(Spacer(1, 3 * mm))
        target.append(
            _bullet_box(
                "Lectura del bloque",
                [
                    str(exposure_payload.get("summary_line") or ""),
                    str(exposure_payload.get("context_link") or ""),
                    f"EstÃ­mulos bajos o ausentes: {_professional_join_labels(exposure_payload.get('low_or_absent', [])[:3])}.",
                ],
                style_name="ProfMuted",
            )
        )

    def _append_full_integrated_page(target: list[object]) -> None:
        target.append(_p(integrated_decision_payload.get("title", "InterpretaciÃ³n integrada profesional"), "ProfSection"))
        target.append(_bullet_box("QuÃ© sabemos con buena confianza", integrated_decision_payload.get("good_confidence", [])))
        target.append(Spacer(1, 3 * mm))
        target.append(_bullet_box("QuÃ© parece probable", integrated_decision_payload.get("probable", []), style_name="ProfMuted"))
        target.append(Spacer(1, 3 * mm))
        target.append(_bullet_box("QuÃ© no podemos afirmar todavÃ­a", integrated_decision_payload.get("unknown", []), style_name="ProfMuted"))
        target.append(Spacer(1, 3 * mm))
        target.append(_bullet_box("DecisiÃ³n prÃ¡ctica", integrated_decision_payload.get("decision_practical", [])))
        target.append(Spacer(1, 3 * mm))
        target.append(_bullet_box("QuÃ© monitorear en el prÃ³ximo bloque", integrated_decision_payload.get("monitor", []), style_name="ProfMuted"))

    def _append_full_action_plan_page(target: list[object]) -> None:
        target.append(_p(action_plan_payload.get("title", "PrÃ³ximos pasos y limitaciones metodolÃ³gicas"), "ProfSection"))
        for label in ["Mantener", "Ajustar", "Monitorear", "Medir"]:
            target.append(_bullet_box(label, action_plan_payload.get("actions", {}).get(label, [])))
            target.append(Spacer(1, 3 * mm))
        target.append(_bullet_box("Limitaciones metodolÃ³gicas", action_plan_payload.get("limitations", []), style_name="ProfMuted"))

    def _append_full_executive_page(target: list[object]) -> None:
        target.extend(
            [
                _p("Reporte profesional", "ProfMuted"),
                _p(report_athlete, "ProfTitle"),
                _p(f"Generado el {datetime.now():%d/%m/%Y %H:%M}", "ProfMuted"),
                _box(
                    [
                        _p("Reporte madre profesional orientado a toma de decisiones.", "ProfBody"),
                        _p(
                            "Prioriza perfil actual, cambios relevantes, contexto de carga y decisiÃƒÂ³n sugerida para el prÃƒÂ³ximo bloque.",
                            "ProfMuted",
                        ),
                    ]
                ),
                Spacer(1, 6 * mm),
                _p(executive_payload.get("title", "Resumen ejecutivo profesional"), "ProfSection"),
                _executive_summary_table(executive_payload),
                Spacer(1, 4 * mm),
                _bullet_box("SeÃƒÂ±ales clave", executive_payload.get("signals", [])),
                Spacer(1, 4 * mm),
                _box(
                    [
                        _p("DecisiÃƒÂ³n sugerida", "ProfCardTitle"),
                        _p(executive_payload.get("decision_suggested", PDF_MISSING_TEXT), "ProfBody"),
                    ],
                    padding=6,
                ),
            ]
        )

    def _append_full_composite_profile_page(target: list[object]) -> None:
        target.append(_p(composite_profile.get("title", "Perfil actual compuesto"), "ProfSection"))
        if composite_profile.get("state") == "missing":
            target.append(_collapsed_box(str(composite_profile.get("message") or "Faltan datos para el perfil compuesto.")))
            return
        target.append(_box([_p(composite_profile.get("note", PROFESSIONAL_COMPOSITE_PROFILE_NOTE), "ProfMuted")], padding=5))
        radar_image = _composite_radar_image(composite_profile)
        if radar_image is not None:
            target.append(Spacer(1, 3 * mm))
            target.append(radar_image)
        target.append(Spacer(1, 3 * mm))
        target.append(_dataframe_table(composite_profile.get("metric_table"), col_widths_mm=[36, 32, 22, 84]))
        feedback = composite_profile.get("feedback", {}) if isinstance(composite_profile.get("feedback"), dict) else {}
        target.append(Spacer(1, 3 * mm))
        target.append(
            _bullet_box(
                "Lectura del perfil",
                [
                    f"Variable dominante: {feedback.get('high', PDF_MISSING_TEXT)}",
                    f"Variable rezagada: {feedback.get('low', PDF_MISSING_TEXT)}",
                    f"Lectura fisiolÃƒÂ³gica: {feedback.get('physiological', PDF_MISSING_TEXT)}",
                    f"Lectura biomecÃƒÂ¡nica: {feedback.get('biomechanical', PDF_MISSING_TEXT)}",
                    f"Implicancia para prÃƒÂ³ximo bloque: {feedback.get('next_block', PDF_MISSING_TEXT)}",
                ],
            )
        )

    def _append_full_change_page(target: list[object]) -> None:
        target.append(_p(change_payload.get("title", "Cambios vs evaluaciÃƒÂ³n anterior"), "ProfSection"))
        if change_payload.get("state") == "missing":
            target.append(_collapsed_box(str(change_payload.get("message") or PROFESSIONAL_NO_EVOLUTION_TEXT)))
            return
        target.append(_dataframe_table(change_payload.get("display_table"), col_widths_mm=[34, 18, 18, 18, 18, 32, 36]))
        target.append(Spacer(1, 3 * mm))
        target.append(_bullet_box("SÃƒÂ­ntesis de cambios", change_payload.get("summary_lines", [])))
        if not _professional_any_quadrant_ready(quadrant_sections):
            target.append(Spacer(1, 3 * mm))
            target.append(_box([_p("Relaciones de perfil / cuadrantes: datos insuficientes para una ubicaciÃƒÂ³n ÃƒÂºtil en esta exportaciÃƒÂ³n.", "ProfMuted")], padding=5))

    def _append_full_quadrants_page(target: list[object]) -> None:
        target.append(_p("Relaciones de perfil / cuadrantes", "ProfSection"))
        if not _professional_any_quadrant_ready(quadrant_sections):
            target.append(_collapsed_box(PROFESSIONAL_NO_QUADRANTS_TEXT))
            return
        for section in quadrant_sections[:3]:
            chart = _quadrant_chart(section)
            if chart is None:
                continue
            target.append(_chart_explanation_panel(chart, str(section.get("title", "Cuadrante")), section))
            target.append(Spacer(1, 4 * mm))

    def _append_full_isometrics_page(target: list[object]) -> None:
        target.append(_p(isometric_payload.get("title", "IsomÃƒÂ©tricos y force-time avanzado"), "ProfSection"))
        if isometric_payload.get("state") == "missing":
            target.append(_collapsed_box(str(isometric_payload.get("message") or "Faltan datos isomÃƒÂ©tricos.")))
            return
        imtp_priority = [row for row in isometric_payload.get("imtp_rows", []) if row[0] in {"Peak Force", "Cambio relevante", "Fuerza relativa", "AsimetrÃƒÂ­a"}]
        if imtp_priority:
            target.append(_p("IMTP principal", "ProfCardTitle"))
            target.append(_mini_cards_table(imtp_priority))
        if isometric_payload.get("iso_available"):
            target.append(Spacer(1, 3 * mm))
            target.append(_p("ISO Push Hip-Hamstring Bilateral", "ProfCardTitle"))
            iso_priority = [row for row in isometric_payload.get("iso_rows", []) if row[0] in {"Peak Force", "Force Avg", "Time to Peak", "AsimetrÃƒÂ­a"}]
            if iso_priority:
                target.append(_mini_cards_table(iso_priority))
            if isometric_payload.get("iso_notes"):
                target.append(Spacer(1, 2 * mm))
                target.append(_bullet_box("Lectura prÃƒÂ¡ctica del test complementario", isometric_payload.get("iso_notes", [])[:3], style_name="ProfMuted"))
        if isometric_payload.get("force_time_available"):
            target.append(Spacer(1, 3 * mm))
            draw_force_time_test_block(
                {
                    "story": target,
                    "p": _p,
                    "box": _box,
                    "Table": Table,
                    "TableStyle": TableStyle,
                    "Spacer": Spacer,
                    "mm": mm,
                    "palette": palette,
                },
                force_time_payload,
                report_type="professional",
            )

    def _append_full_load_page(target: list[object]) -> None:
        target.append(_p(load_tolerance_payload.get("title", "Carga interna y tolerancia"), "ProfSection"))
        if load_tolerance_payload.get("state") == "missing":
            target.append(_collapsed_box(str(load_tolerance_payload.get("message") or "Faltan datos de carga interna.")))
            return
        target.append(_key_value_table(load_tolerance_payload.get("rows", [])))
        target.append(Spacer(1, 3 * mm))
        weekly_chart = _weekly_ema_chart(load_tolerance_payload.get("weekly_points", []), height_mm=48)
        if weekly_chart is not None:
            target.append(weekly_chart)
            target.append(Spacer(1, 3 * mm))
        target.append(_box([_p(load_tolerance_payload.get("risk_line", PDF_MISSING_TEXT), "ProfBody")], padding=6))

    def _append_full_wellness_page(target: list[object]) -> None:
        target.append(_p(wellness_availability_payload.get("title", "Wellness, disponibilidad y adherencia"), "ProfSection"))
        if wellness_availability_payload.get("state") == "missing":
            target.append(_collapsed_box(str(wellness_availability_payload.get("message") or "Faltan datos de wellness.")))
            return
        target.append(_mini_cards_table(wellness_availability_payload.get("rows", [])))
        chart_points = (
            wellness_availability_payload.get("weekly_points", [])
            if wellness_availability_payload.get("trend_allowed")
            else wellness_availability_payload.get("daily_points", [])
        )
        chart = _wellness_chart(chart_points, width_mm=174, height_mm=44)
        if chart is not None:
            target.append(Spacer(1, 3 * mm))
            target.append(chart)
        target.append(Spacer(1, 3 * mm))
        note_lines = [str(wellness_availability_payload.get("compatibility") or "")]
        if str(wellness_availability_payload.get("quality_note") or "").strip():
            note_lines.append(str(wellness_availability_payload.get("quality_note")))
        target.append(_bullet_box("Lectura de disponibilidad", note_lines, style_name="ProfMuted"))

    def _append_full_exposure_page(target: list[object]) -> None:
        target.append(_p(exposure_payload.get("title", "ExposiciÃƒÂ³n del bloque / contenido entrenado"), "ProfSection"))
        if exposure_payload.get("state") == "missing":
            target.append(_collapsed_box(str(exposure_payload.get("message") or "Faltan datos de exposiciÃƒÂ³n.")))
            return
        chart_image = _exposure_chart_image()
        if chart_image is not None:
            target.append(chart_image)
            target.append(Spacer(1, 3 * mm))
        target.append(_dataframe_table(exposure_payload.get("table"), col_widths_mm=[38, 30, 26, 80]))
        target.append(Spacer(1, 3 * mm))
        exposure_lines = [
            str(exposure_payload.get("summary_line") or ""),
            str(exposure_payload.get("context_link") or ""),
        ]
        if exposure_payload.get("low_or_absent"):
            exposure_lines.append(
                f"EstÃƒÂ­mulos bajos o ausentes: {_professional_join_labels(exposure_payload.get('low_or_absent', [])[:3])}."
            )
        target.append(
            _bullet_box(
                "Lectura del bloque",
                exposure_lines,
                style_name="ProfMuted",
            )
        )

    def _append_full_integrated_page(target: list[object]) -> None:
        target.append(_p(integrated_decision_payload.get("title", "InterpretaciÃƒÂ³n integrada profesional"), "ProfSection"))
        target.append(_bullet_box("QuÃƒÂ© sabemos con buena confianza", integrated_decision_payload.get("good_confidence", [])))
        target.append(Spacer(1, 3 * mm))
        target.append(_bullet_box("QuÃƒÂ© parece probable", integrated_decision_payload.get("probable", []), style_name="ProfMuted"))
        target.append(Spacer(1, 3 * mm))
        target.append(_bullet_box("QuÃƒÂ© no podemos afirmar todavÃƒÂ­a", integrated_decision_payload.get("unknown", []), style_name="ProfMuted"))
        target.append(Spacer(1, 3 * mm))
        target.append(_bullet_box("DecisiÃƒÂ³n prÃƒÂ¡ctica", integrated_decision_payload.get("decision_practical", [])))
        target.append(Spacer(1, 3 * mm))
        target.append(_bullet_box("QuÃƒÂ© monitorear en el prÃƒÂ³ximo bloque", integrated_decision_payload.get("monitor", []), style_name="ProfMuted"))

    def _append_full_action_plan_page(target: list[object]) -> None:
        target.append(_p(action_plan_payload.get("title", "PrÃƒÂ³ximos pasos y limitaciones metodolÃƒÂ³gicas"), "ProfSection"))
        for label in ["Mantener", "Ajustar", "Monitorear", "Medir"]:
            target.append(_bullet_box(label, action_plan_payload.get("actions", {}).get(label, [])))
            target.append(Spacer(1, 3 * mm))
        target.append(_bullet_box("Limitaciones metodolÃƒÂ³gicas", action_plan_payload.get("limitations", []), style_name="ProfMuted"))

    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        for sheet_name, df in _ordered_sheet_items(data_dict):
            if df is not None and not df.empty:
                prepared_df = _prepare_export_frame(sheet_name, df)
                export_name = REPORT_SHEET_EXPORT_NAMES.get(sheet_name, sheet_name)[:31]
                prepared_df.to_excel(writer, sheet_name=export_name, index=False)
                worksheet = writer.book[export_name]
                _format_excel_sheet(worksheet, prepared_df, sheet_name)
    return buffer.getvalue()


def _ascii_text(value: object) -> str:
    text = "" if value is None else str(value)
    text = (
        text.replace("—", "-")
        .replace("±", "+/-")
        .replace("σ", "sd")
        .replace("·", "-")
        .replace("→", "->")
    )
    normalized = unicodedata.normalize("NFKC", text)
    return normalized.encode("latin-1", "ignore").decode("latin-1")


def _pdf_escape(text: str) -> str:
    return text.replace("\\", "\\\\").replace("(", "\\(").replace(")", "\\)")


def _wrap_lines(text: str, width: int) -> list[str]:
    clean = _ascii_text(text).strip()
    if not clean:
        return [""]
    return textwrap.wrap(clean, width=width) or [clean]


def _pdf_color(hex_color: str) -> tuple[float, float, float]:
    hex_clean = hex_color.lstrip("#")
    return tuple(int(hex_clean[i:i + 2], 16) / 255 for i in (0, 2, 4))


def _pdf_text(commands: list[str], x: int, y: int, text: str, *, font: str = "F1", size: int = 11, color: str = "#221F20") -> None:
    r, g, b = _pdf_color(color)
    commands.append("BT")
    commands.append(f"/{font} {size} Tf")
    commands.append(f"{r:.3f} {g:.3f} {b:.3f} rg")
    commands.append(f"{x} {y} Td")
    commands.append(f"({_pdf_escape(_ascii_text(text))}) Tj")
    commands.append("ET")


def _pdf_multiline(
    commands: list[str],
    x: int,
    y: int,
    lines: list[str],
    *,
    font: str = "F1",
    size: int = 11,
    color: str = "#221F20",
    leading: int = 14,
) -> int:
    current_y = y
    for line in lines:
        _pdf_text(commands, x, current_y, line, font=font, size=size, color=color)
        current_y -= leading
    return current_y


def _pdf_rect(commands: list[str], x: int, y: int, w: int, h: int, fill: str) -> None:
    r, g, b = _pdf_color(fill)
    commands.append(f"{r:.3f} {g:.3f} {b:.3f} rg")
    commands.append(f"{x} {y} {w} {h} re f")


def _pdf_line(commands: list[str], x1: int, y1: int, x2: int, y2: int, color: str = "#D8DEE4") -> None:
    r, g, b = _pdf_color(color)
    commands.append(f"{r:.3f} {g:.3f} {b:.3f} RG")
    commands.append("1 w")
    commands.append(f"{x1} {y1} m {x2} {y2} l S")


def _pdf_stroke_rect(
    commands: list[str],
    x: int,
    y: int,
    w: int,
    h: int,
    *,
    color: str = "#D8DEE4",
    width: float = 1.0,
) -> None:
    r, g, b = _pdf_color(color)
    commands.append(f"{r:.3f} {g:.3f} {b:.3f} RG")
    commands.append(f"{width:.1f} w")
    commands.append(f"{x} {y} {w} {h} re S")


def _zone_color(zone: object) -> str:
    zone_text = _ascii_text(zone).strip().lower()
    if "alto" in zone_text or "riesgo" in zone_text:
        return "#B85C38"
    if "precauc" in zone_text:
        return "#C58A2D"
    if "opt" in zone_text:
        return "#2F6B52"
    if "sub" in zone_text:
        return "#708C9F"
    return "#5A595B"


def _display_metric(value: object, *, digits: int | None = None, suffix: str = "") -> str:
    numeric = _coerce_float(value)
    if numeric is not None:
        if digits is None:
            rendered = f"{numeric:g}"
        else:
            rendered = f"{numeric:.{digits}f}"
        return f"{rendered}{suffix}"
    text = _ascii_text(value).strip()
    return text or "-"


def _display_zone(zone: object) -> str:
    clean = _ascii_text(zone).strip().lower()
    mapping = {
        "optimo": "Óptimo",
        "precaucion": "Precaución",
        "subcarga": "Subcarga",
        "alto riesgo": "Alto riesgo",
    }
    return mapping.get(clean, _ascii_text(zone).strip() or "-")


def _snapshot_value(column: str, raw_value: object) -> str:
    if column == "Atleta":
        return _ascii_text(raw_value).strip() or "-"
    if column in {"ACWR", "ACWR EWMA"}:
        return _display_metric(raw_value, digits=2)
    if column == "Zona":
        return _display_zone(raw_value)
    if column == "Wellness":
        return _display_metric(raw_value, digits=1)
    if column == "CMJ":
        return _display_metric(raw_value, digits=1, suffix=" cm")
    if column == "DRI":
        return _display_metric(raw_value, digits=2)
    if column == "IMTP":
        return _display_metric(raw_value, digits=0, suffix=" N")
    if column == "Perfil NM":
        return _ascii_text(raw_value).strip() or "-"
    return _display_metric(raw_value)


def _short_profile_label(profile: object) -> str:
    text = _ascii_text(profile).strip()
    if not text:
        return "-"
    return text.split("/")[0].strip()[:18]


def _strengths_from_row(row: pd.Series, *, audience: str) -> list[str]:
    strengths: list[str] = []
    acwr = _coerce_float(row.get("ACWR EWMA"))
    cmj_delta = _coerce_float(row.get("CMJ vs BL %"))
    wellness = _coerce_float(row.get("Wellness 3d"))
    profile = _ascii_text(row.get("Perfil NM")).strip()
    if cmj_delta is not None and cmj_delta >= 0:
        strengths.append(
            "Tu salto vertical se sostiene o mejora respecto al baseline inicial."
            if audience != "profe" else
            "El CMJ se sostiene o mejora respecto al baseline inicial."
        )
    if acwr is not None and 0.8 <= acwr <= 1.3:
        strengths.append(
            "La carga reciente está en una zona útil para seguir construyendo."
            if audience != "profe" else
            "La relación agudo-crónica está dentro de una zona operativa útil."
        )
    if wellness is not None and wellness >= 15:
        strengths.append(
            "La recuperación percibida acompaña bien el bloque actual."
            if audience != "profe" else
            "El wellness reciente acompaña sin señales claras de caída."
        )
    if profile:
        strengths.append(
            f"El perfil actual muestra una base de trabajo definida: {profile}."
            if audience == "cliente" else
            f"El perfil neuromuscular actual es {profile}."
        )
    return list(dict.fromkeys(strengths))[:3] or [
        "Hay una base mínima de información para seguir comparando la evolución."
        if audience != "cliente" else
        "Ya contamos con información suficiente para seguir tu progreso con más claridad."
    ]


def _gaps_from_row(row: pd.Series, *, audience: str) -> list[str]:
    gaps: list[str] = []
    acwr = _coerce_float(row.get("ACWR EWMA"))
    monotony = _coerce_float(row.get("Monotonia"))
    wellness = _coerce_float(row.get("Wellness 3d"))
    cmj_delta = _coerce_float(row.get("CMJ vs BL %"))
    profile = _ascii_text(row.get("Perfil NM")).lower().strip()

    if cmj_delta is not None and cmj_delta <= -5:
        gaps.append(
            "Tu salto vertical cayo frente a tu baseline inicial; conviene recuperar calidad."
            if audience != "profe" else
            "El CMJ cae respecto al baseline inicial; conviene revisar fatiga reciente y exposicion."
        )
    if acwr is not None and acwr > 1.5:
        gaps.append(
            "La carga reciente viene alta y puede pedir una semana más controlada."
            if audience != "profe" else
            "El ACWR EWMA queda alto y sugiere controlar densidad y exposición semanal."
        )
    elif acwr is not None and acwr < 0.8:
        gaps.append(
            "La carga reciente quedó baja; puede faltar continuidad de estímulo."
            if audience != "profe" else
            "El ACWR EWMA cae en subcarga; conviene revisar continuidad y volumen útil."
        )
    if monotony is not None and monotony > 2.0:
        gaps.append(
            "La semana se repite demasiado y conviene darle más variación."
            if audience != "profe" else
            "La monotonía semanal está alta y puede limitar tolerancia al bloque."
        )
    if wellness is not None and wellness < 15:
        gaps.append(
            "La recuperación percibida viene baja y requiere seguimiento cercano."
            if audience != "profe" else
            "El wellness reciente cae y sugiere revisar recuperación diaria."
        )
    if "poca base" in profile:
        gaps.append(
            "Todavía hay margen para construir una base de fuerza más sólida."
            if audience != "profe" else
            "El perfil actual sigue marcando necesidad de consolidar fuerza base."
        )
    return list(dict.fromkeys(gaps))[:3] or [
        "No aparece una alerta dominante; el foco está en sostener continuidad y calidad."
        if audience != "cliente" else
        "No aparece una alarma grande; la prioridad es sostener lo que ya está funcionando."
    ]


def _next_steps_from_row(row: pd.Series, *, audience: str) -> list[str]:
    steps: list[str] = []
    profile = _ascii_text(row.get("Perfil NM")).lower().strip()
    acwr = _coerce_float(row.get("ACWR EWMA"))
    monotony = _coerce_float(row.get("Monotonia"))
    cmj_delta = _coerce_float(row.get("CMJ vs BL %"))

    if "poca base" in profile:
        steps.append(
            "Construir fuerza base antes de pedir más reactividad."
            if audience != "cliente" else
            "Dedicar más trabajo a la base de fuerza para que el progreso sea más estable."
        )
    if cmj_delta is not None and cmj_delta <= -5:
        steps.append(
            "Bajar fatiga y recuperar calidad de salto en la próxima ventana."
            if audience != "cliente" else
            "Buscar que el salto vuelva a acercarse a tu mejor referencia reciente."
        )
    if acwr is not None and acwr > 1.5:
        steps.append(
            "Regular la densidad de trabajo del próximo microciclo."
            if audience != "cliente" else
            "Ordenar la carga de la semana para que el cuerpo la tolere mejor."
        )
    if monotony is not None and monotony > 2.0:
        steps.append(
            "Meter más variación semanal y ajustar recuperación."
            if audience != "cliente" else
            "Variar más el trabajo de la semana para que no todo se sienta igual."
        )
    return steps[:3] or [
        "Sostener la línea actual y volver a medir en la próxima ventana operativa."
        if audience != "cliente" else
        "Mantener la línea actual y volver a medir para confirmar el progreso."
    ]


ATHLETE_METRIC_EXPLANATIONS = [
    ("sRPE", "carga interna percibida del entrenamiento."),
    ("ACWR EWMA", "relación entre carga reciente y carga habitual, suavizada para ver tendencia."),
    ("DRI", "índice del Drop Jump que combina altura de caída, altura de salto y tiempo de contacto para interpretar el perfil reactivo."),
    ("EUR", "relación entre salto con contramovimiento y salto sin contramovimiento."),
    ("IMTP", "fuerza isométrica máxima; ayuda a entender tu base de fuerza."),
    ("Monotonía", "qué tan parecida fue la carga entre días de la semana."),
]


def _athlete_metric_explanation_rows(row: pd.Series | None) -> list[tuple[str, str]]:
    if row is None or row.empty:
        return ATHLETE_METRIC_EXPLANATIONS[:4]
    available = {
        "ACWR EWMA": _coerce_float(row.get("ACWR EWMA")) is not None,
        "DRI": _coerce_float(row.get("DRI")) is not None,
        "EUR": _coerce_float(_summary_eur_value(row)) is not None,
        "IMTP": _coerce_float(row.get("IMTP N")) is not None,
        "Monotonía": _coerce_float(row.get("Monotonia")) is not None,
    }
    rows = [ATHLETE_METRIC_EXPLANATIONS[0]]
    rows.extend((label, text) for label, text in ATHLETE_METRIC_EXPLANATIONS[1:] if available.get(label))
    return rows[:5] if len(rows) > 1 else ATHLETE_METRIC_EXPLANATIONS[:4]


def _neuromuscular_metric_fact(
    profile_payload: dict[str, object],
    key: str,
    *,
    digits: int = 1,
) -> str | None:
    metrics = profile_payload.get("metrics", {})
    metric = metrics.get(key, {}) if isinstance(metrics, dict) else {}
    if not isinstance(metric, dict) or not metric.get("available"):
        return None
    value = _coerce_float(metric.get("value"))
    if value is None:
        return None
    label = _professional_visible_metric_text(metric.get("label", NEUROMUSCULAR_KPI_LABELS.get(key, key)))
    unit = str(metric.get("unit") or "").strip()
    if unit == "ratio":
        unit = ""
    suffix = f" {unit}" if unit else ""
    return _professional_visible_metric_text(f"{label} {_display_metric(value, digits=digits, suffix=suffix)}")


def _athlete_profile_interpretation(
    row: pd.Series | None,
    *,
    neuromuscular_profile: dict[str, object] | None = None,
) -> dict[str, str]:
    if row is None or row.empty or not _row_has_eval_data(row):
        return {
            "what": "Faltan datos de evaluación física para construir el perfil neuromuscular.",
            "meaning": "Por ahora no conviene sacar conclusiones sobre fuerza, salto o reactividad.",
            "priority": "Completar una batería de evaluación con CMJ, SJ, Drop Jump/DRI e IMTP.",
        }
    if isinstance(neuromuscular_profile, dict) and neuromuscular_profile.get("source") == "core":
        metric_facts = _compact_lines(
            [
                _neuromuscular_metric_fact(neuromuscular_profile, "CMJ_cm", digits=1),
                _neuromuscular_metric_fact(neuromuscular_profile, "SJ_cm", digits=1),
                _neuromuscular_metric_fact(neuromuscular_profile, "DJ_RSI", digits=2),
                _neuromuscular_metric_fact(neuromuscular_profile, "EUR", digits=2),
                _neuromuscular_metric_fact(neuromuscular_profile, "IMTP_N", digits=0),
                _neuromuscular_metric_fact(neuromuscular_profile, "IMTP_relPF", digits=2),
            ]
        )
        profile_label = _professional_visible_metric_text(
            neuromuscular_profile.get("profile_label") or _profile_text(row, fallback="perfil parcial")
        )
        confidence_label = _professional_visible_metric_text(
            neuromuscular_profile.get("confidence_label")
            or NEUROMUSCULAR_CONFIDENCE_LABELS.get(str(neuromuscular_profile.get("confidence") or "low"), "Baja")
        )
        summary_athlete = _professional_visible_metric_text(neuromuscular_profile.get("summary_athlete") or "")
        flag_messages = [
            _professional_visible_metric_text(message)
            for message in list(neuromuscular_profile.get("flag_messages_athlete", []))
        ]
        meaning_parts = _compact_lines(
            [
                summary_athlete,
                None if summary_athlete else (flag_messages[0] if flag_messages else None),
            ]
        )
        priority_parts = _compact_lines(
            [
                _professional_visible_metric_text(neuromuscular_profile.get("training_priority_detailed") or ""),
            ]
        )
        return {
            "what": (
                "El radar resume tus principales cualidades neuromusculares: "
                + (" | ".join(metric_facts) if metric_facts else "hay métricas parciales disponibles.")
                + f" Perfil actual: {profile_label}. Confianza de lectura: {confidence_label}."
            ),
            "meaning": " ".join(meaning_parts) if meaning_parts else "Todavía falta una lectura más clara del perfil neuromuscular.",
            "priority": " ".join(priority_parts) if priority_parts else "Completar una batería de evaluación para definir mejor el próximo foco.",
        }
    profile = _profile_text(row, fallback="perfil parcial")
    profile_key = profile.lower()
    dri = _coerce_float(row.get("DRI"))
    imtp = _coerce_float(row.get("IMTP N"))
    eur = _coerce_float(_summary_eur_value(row))
    cmj = _coerce_float(row.get("CMJ cm"))

    what_parts = _compact_lines(
        [
            f"CMJ {_display_metric(cmj, digits=1, suffix=' cm')}" if cmj is not None else None,
            f"DRI {_display_metric(dri, digits=2)}" if dri is not None else None,
            f"IMTP {_display_metric(imtp, digits=0, suffix=' N')}" if imtp is not None else None,
            f"{EUR_RATIO_LABEL} {_display_metric(eur, digits=2)}" if eur is not None else None,
        ]
    )
    what = (
        "El radar resume tus principales cualidades neuromusculares: "
        + (" | ".join(what_parts) if what_parts else "hay métricas parciales disponibles.")
    )

    if "reactivo" in profile_key:
        meaning = "Tu perfil actual se orienta hacia un perfil reactivo: buena respuesta en acciones rápidas, que debe sostenerse con fuerza de base y calidad técnica."
        priority = "Priorizar reactividad con baja fatiga, buena técnica de contacto y mantenimiento de fuerza base."
    elif "base" in profile_key or "fuerza" in profile_key:
        meaning = "Tu perfil actual marca una base de fuerza relevante, pero la transferencia hacia acciones rápidas todavía debe seguir construyéndose."
        priority = "Sostener fuerza base y progresar potencia/reactividad de forma gradual."
    elif "poca" in profile_key:
        meaning = "Tu perfil todavía necesita una base más sólida antes de pedir grandes aumentos de velocidad o reactividad."
        priority = "Construir fuerza general, técnica de salto y continuidad de entrenamiento."
    else:
        meaning = f"Tu perfil actual se clasifica como {profile}; es una referencia útil, no una etiqueta fija."
        priority = "Usar el perfil para orientar el próximo bloque y confirmar cambios en la siguiente medición."
    return {"what": what, "meaning": meaning, "priority": priority}


def _athlete_load_status_lines(row: pd.Series | None, internal_load: dict[str, object]) -> list[str]:
    scope = str(internal_load.get("analysis_scope") or "")
    acwr = _coerce_float(row.get("ACWR EWMA")) if row is not None and not row.empty else None
    zone = _display_zone(row.get("Zona")) if row is not None and not row.empty and _has_text(row.get("Zona")) else PDF_MISSING_TEXT
    monotony = _coerce_float(row.get("Monotonia")) if row is not None and not row.empty else None
    lines: list[str] = [
        "Las barras muestran sRPE: carga interna percibida del entrenamiento. La línea muestra ACWR EWMA: relación entre carga reciente y carga habitual.",
        "Zona óptima sugiere relación estable; precaución o riesgo indican aumentos bruscos; subcarga indica baja continuidad relativa.",
    ]
    if scope == "current_week_partial":
        total = _coerce_float(internal_load.get("current_week_total"))
        sessions = int(_coerce_float(internal_load.get("current_week_sessions")) or 0)
        if total is not None:
            lines.append(
                f"Semana en curso / datos parciales: acumulás {total:.0f} UA con {sessions} sesiones registradas. No compararlo todavía contra una semana completa."
            )
        else:
            lines.append("Semana en curso / datos parciales: faltan datos suficientes para interpretar la carga acumulada.")
    elif scope == "last_complete_week":
        total = _coerce_float(internal_load.get("last_week_total"))
        change_pct = _coerce_float(internal_load.get("weekly_change_pct"))
        line = f"Última semana completa: {total:.0f} UA." if total is not None else "Faltan datos suficientes para analizar la última semana completa."
        if change_pct is not None:
            line = f"{line} Cambio vs semana previa: {change_pct:+.1f}%."
        lines.append(line)
    else:
        lines.append("Faltan datos de carga interna para este período.")

    if acwr is not None and zone != PDF_MISSING_TEXT:
        lines.append(f"Lectura actual: ACWR EWMA {_display_metric(acwr, digits=2)} en zona {zone}.")
    if monotony is not None:
        lines.append(f"Monotonía semanal: {_display_metric(monotony, digits=2)}. Si sube mucho, conviene variar mejor la carga entre días.")
    return lines[:5]


def _athlete_final_focus_blocks(
    row: pd.Series | None,
    completion_value: float | None,
    *,
    neuromuscular_profile: dict[str, object] | None = None,
) -> list[dict[str, str]]:
    if row is None or row.empty:
        return [
            {"title": "Fortaleza principal", "body": "Faltan datos suficientes para definir una fortaleza principal."},
            {"title": "Punto a vigilar", "body": "Completar carga, wellness y evaluación para mejorar la lectura."},
            {"title": "Foco del próximo bloque", "body": "Construir continuidad y registrar datos de forma consistente."},
            {"title": "Próxima medición o revisión", "body": "Realizar una evaluación física inicial y revisar carga semanal."},
        ]
    if isinstance(neuromuscular_profile, dict) and neuromuscular_profile.get("source") == "core":
        strengths = _strengths_from_row(row, audience="atleta")
        strength_line = _professional_visible_metric_text(
            neuromuscular_profile.get("summary_short")
            or neuromuscular_profile.get("summary_athlete")
            or (strengths[0] if strengths else PDF_MISSING_TEXT)
        )
        gap_lines = list(neuromuscular_profile.get("flag_messages_athlete", [])) or _gaps_from_row(row, audience="atleta")
        if completion_value is not None and completion_value < 70:
            gap_lines = [f"La adherencia actual ({completion_value:.1f}%) puede limitar la lectura del bloque."] + gap_lines
        review_bits = []
        if neuromuscular_profile.get("kpi_labels"):
            review_bits.append(
                _professional_visible_metric_text(
                    f"En la próxima medición conviene volver a mirar {', '.join(neuromuscular_profile.get('kpi_labels', [])[:3])}."
                )
            )
        review_bits.append(
            "Repetir evaluación física en 6-8 semanas y revisar carga/bienestar semanalmente."
            if _row_has_eval_data(row)
            else "Completar evaluación física y usarla como línea base del perfil."
        )
        return [
            {"title": "Fortaleza principal", "body": strength_line},
            {"title": "Punto a vigilar", "body": gap_lines[0] if gap_lines else PDF_MISSING_TEXT},
            {
                "title": "Foco del próximo bloque",
                "body": _professional_visible_metric_text(
                    neuromuscular_profile.get("training_priority_detailed")
                    or neuromuscular_profile.get("training_priority_short")
                    or PDF_MISSING_TEXT
                ),
            },
            {"title": "Próxima medición o revisión", "body": " ".join(review_bits)},
        ]
    strengths = _strengths_from_row(row, audience="atleta")
    gaps = _gaps_from_row(row, audience="atleta")
    objective = _current_focus_text(row, audience="atleta")
    next_steps = _next_steps_from_row(row, audience="atleta")
    if completion_value is not None and completion_value < 70:
        gaps = [f"La adherencia actual ({completion_value:.1f}%) puede limitar la lectura del bloque."] + gaps
    next_review = (
        "Repetir evaluación física en 6-8 semanas y revisar carga/bienestar semanalmente."
        if _row_has_eval_data(row)
        else "Completar evaluación física y usarla como línea base del perfil."
    )
    return [
        {"title": "Fortaleza principal", "body": strengths[0] if strengths else PDF_MISSING_TEXT},
        {"title": "Punto a vigilar", "body": gaps[0] if gaps else PDF_MISSING_TEXT},
        {"title": "Foco del próximo bloque", "body": f"{objective}. {next_steps[0] if next_steps else ''}".strip()},
        {"title": "Próxima medición o revisión", "body": next_review},
    ]


def _pdf_label_value_card(
    commands: list[str],
    x: int,
    y: int,
    w: int,
    h: int,
    *,
    label: str,
    value: str,
    accent: str = "#0D3C5E",
    fill: str = "#FEFEFE",
    border: str = "#D8DEE4",
    value_color: str = "#221F20",
) -> None:
    _pdf_rect(commands, x, y, w, h, fill)
    _pdf_stroke_rect(commands, x, y, w, h, color=border)
    _pdf_rect(commands, x, y + h - 4, w, 4, accent)
    _pdf_text(commands, x + 14, y + h - 24, label, size=9, color="#5A595B")
    _pdf_text(commands, x + 14, y + h - 50, value, font="F2", size=18, color=value_color)


def _pdf_module_block(
    commands: list[str],
    x: int,
    y: int,
    w: int,
    h: int,
    *,
    title: str,
    summary: str,
    focuses: list[str],
    accent: str = "#0D3C5E",
) -> None:
    _pdf_rect(commands, x, y, w, h, "#FEFEFE")
    _pdf_stroke_rect(commands, x, y, w, h, color="#D8DEE4")
    _pdf_rect(commands, x, y + h - 6, w, 6, accent)
    _pdf_text(commands, x + 14, y + h - 28, title, font="F2", size=12, color="#221F20")
    current_y = _pdf_multiline(
        commands,
        x + 14,
        y + h - 48,
        _wrap_lines(summary, 42),
        size=9,
        color="#5A595B",
        leading=12,
    ) - 8
    for focus in focuses[:2]:
        if current_y < y + 20:
            break
        current_y = _pdf_multiline(
            commands,
            x + 18,
            current_y,
            _wrap_lines(f"- {focus}", 40),
            size=8,
            color="#221F20",
            leading=11,
        ) - 4


def _pdf_chart_panel(
    commands: list[str],
    x: int,
    y: int,
    w: int,
    h: int,
    *,
    title: str,
    series: list[tuple[str, float]],
    line_color: str = "#0D3C5E",
) -> None:
    _pdf_rect(commands, x, y, w, h, "#FEFEFE")
    _pdf_stroke_rect(commands, x, y, w, h, color="#D8DEE4")
    _pdf_text(commands, x + 14, y + h - 24, title, font="F2", size=11, color="#221F20")

    if not series or len(series) < 2:
        _pdf_text(commands, x + 14, y + h - 54, "Sin historial suficiente para mostrar tendencia.", size=9, color="#708C9F")
        return

    plot_x = x + 16
    plot_y = y + 24
    plot_w = w - 32
    plot_h = h - 58
    _pdf_line(commands, plot_x, plot_y, plot_x + plot_w, plot_y, color="#D8DEE4")
    _pdf_line(commands, plot_x, plot_y, plot_x, plot_y + plot_h, color="#D8DEE4")

    values = [value for _, value in series]
    min_v = min(values)
    max_v = max(values)
    if max_v == min_v:
        max_v += 1
        min_v -= 1

    r, g, b = _pdf_color(line_color)
    commands.append(f"{r:.3f} {g:.3f} {b:.3f} RG")
    commands.append("1.6 w")
    for idx, (_, value) in enumerate(series):
        px = plot_x + (plot_w * idx / max(1, len(series) - 1))
        py = plot_y + ((value - min_v) / (max_v - min_v) * plot_h)
        if idx == 0:
            commands.append(f"{px:.1f} {py:.1f} m")
        else:
            commands.append(f"{px:.1f} {py:.1f} l")
    commands.append("S")

    first_label = series[0][0]
    last_label = series[-1][0]
    _pdf_text(commands, plot_x, y + 10, first_label, size=8, color="#708C9F")
    _pdf_text(commands, plot_x + plot_w - 28, y + 10, last_label, size=8, color="#708C9F")
    _pdf_text(commands, plot_x, plot_y + plot_h + 6, _display_metric(min_v, digits=1), size=8, color="#708C9F")
    _pdf_text(commands, plot_x + plot_w - 28, plot_y + plot_h + 6, _display_metric(max_v, digits=1), size=8, color="#708C9F")


def _build_cover_page(
    report_athlete: str,
    summary_df: pd.DataFrame,
    insights: dict[str, dict[str, object]],
    report_audience: str,
) -> str:
    audience = normalize_report_audience(report_audience)
    commands: list[str] = []
    _pdf_rect(commands, 0, 0, 595, 842, "#F4F6F8")
    _pdf_rect(commands, 0, 742, 595, 100, "#221F20")
    _pdf_rect(commands, 48, 730, 220, 3, "#0D3C5E")
    _pdf_text(commands, 48, 794, "THRESHOLD STRENGTH & CONDITIONING", font="F2", size=21, color="#FEFEFE")
    subtitle = {
        "atleta": "Versión atleta",
        "profe": "Versión técnica",
        "cliente": "Versión cliente",
    }[audience]
    _pdf_text(commands, 48, 770, subtitle, font="F2", size=14, color="#9AA2A9")

    _pdf_text(commands, 48, 698, "Alcance", size=10, color="#708C9F")
    _pdf_text(commands, 48, 674, report_athlete, font="F2", size=20, color="#0D3C5E")
    _pdf_text(commands, 250, 698, "Generado", size=10, color="#708C9F")
    _pdf_text(commands, 250, 676, f"{datetime.now():%d/%m/%Y %H:%M}", font="F2", size=14, color="#221F20")
    _pdf_text(commands, 430, 698, "Filas ejecutivas", size=10, color="#708C9F")
    _pdf_text(commands, 430, 676, str(len(summary_df)), font="F2", size=14, color="#221F20")

    _pdf_rect(commands, 48, 560, 499, 84, "#FEFEFE")
    _pdf_stroke_rect(commands, 48, 560, 499, 84, color="#D8DEE4")
    _pdf_text(commands, 64, 620, "Narrativa ejecutiva", font="F2", size=13, color="#221F20")
    intro = insights.get("report", {}).get("summary", "Reporte operativo listo para revisión técnica.")
    _pdf_multiline(commands, 64, 598, _wrap_lines(intro, 78), size=10, color="#5A595B", leading=14)

    overview_summary = insights.get("overview", {}).get("summary", "Todavía no hay una lectura general disponible.")
    focus_value = {
        "atleta": "Evaluación y perfil",
        "profe": "Lectura integrada",
        "cliente": "Progreso y claridad",
    }[audience]
    _pdf_label_value_card(commands, 48, 450, 156, 78, label="Foco del reporte", value=focus_value, accent="#0D3C5E")
    _pdf_label_value_card(commands, 220, 450, 156, 78, label="Ventana", value="Últimas 6 semanas", accent="#708C9F")
    _pdf_label_value_card(commands, 392, 450, 155, 78, label="Módulos", value=str(len(insights)), accent="#5A595B")

    _pdf_text(commands, 48, 410, "Contexto de lectura", font="F2", size=13, color="#221F20")
    _pdf_multiline(commands, 48, 388, _wrap_lines(overview_summary, 82), size=10, color="#5A595B", leading=14)

    _pdf_text(commands, 48, 318, "Focos prioritarios", font="F2", size=13, color="#221F20")
    current_y = 294
    for focus in insights.get("report", {}).get("focuses", []):
        current_y = _pdf_multiline(
            commands,
            60,
            current_y,
            _wrap_lines(f"- {focus}", 78),
            size=10,
            color="#221F20",
            leading=14,
        ) - 6

    _pdf_line(commands, 48, 112, 547, 112, color="#D8DEE4")
    footer = "Threshold S&C - Monitoreo de carga, evaluaciones y seguimiento de atletas."
    _pdf_multiline(commands, 48, 92, _wrap_lines(footer, 84), size=9, color="#708C9F", leading=13)
    return "\n".join(commands)


def _build_dashboard_page(
    state: dict[str, pd.DataFrame | None],
    report_athlete: str,
    summary_df: pd.DataFrame,
    report_audience: str,
) -> str | None:
    audience = normalize_report_audience(report_audience)
    if summary_df.empty:
        return None

    commands: list[str] = []
    _pdf_rect(commands, 0, 0, 595, 842, "#F4F6F8")
    page_title = {
        "atleta": "Resumen de rendimiento",
        "profe": "Dashboard ejecutivo",
        "cliente": "Resumen de progreso",
    }[audience]
    page_subtitle = {
        "atleta": "Lectura cuasi técnica para entender perfil actual, fortalezas y próximos pasos.",
        "profe": "Última foto integrada para revisión operativa.",
        "cliente": "Lectura clara del estado actual y del progreso reciente.",
    }[audience]
    _pdf_text(commands, 48, 800, page_title, font="F2", size=18, color="#0D3C5E")
    _pdf_text(commands, 48, 778, page_subtitle, size=10, color="#5A595B")
    _pdf_line(commands, 48, 766, 547, 766, color="#D8DEE4")

    if report_athlete != "Todos" and not summary_df.empty:
        focus_row = summary_df.iloc[0]
        title_text = _ascii_text(focus_row.get("Atleta", report_athlete))
    else:
        focus_row = summary_df.iloc[0]
        title_text = _count_phrase(len(summary_df), "atleta visible", "atletas visibles")

    _pdf_text(commands, 48, 730, title_text, font="F2", size=22, color="#221F20")
    dashboard_note = {
        "atleta": "En esta versión priorizamos evaluaciones, perfil y foco inmediato de trabajo.",
        "profe": "Tablero listo para revisión técnica y exportación profesional.",
        "cliente": "En esta versión priorizamos comprensión, progreso y próximos pasos.",
    }[audience]
    _pdf_text(commands, 48, 708, dashboard_note, size=10, color="#708C9F")

    cards = _audience_dashboard_cards(state, focus_row, report_athlete, audience)

    positions = [
        (48, 604), (220, 604), (392, 604),
        (48, 504), (220, 504), (392, 504),
    ]
    for (label, value, accent), (x, y) in zip(cards, positions):
        _pdf_label_value_card(commands, x, y, 155, 82, label=label, value=value, accent=accent)

    completion_value = (
        _athlete_completion_mean(state, report_athlete)
        if report_athlete != "Todos" else
        _team_completion_mean(state)
    )
    completion_text = _display_metric(completion_value, digits=1, suffix="%") if completion_value is not None else "-"
    athletes_text = str(len(summary_df))
    datasets_count = len(
        [
            key for key in MODERN_REPORT_DATASET_KEYS
            if state.get(key) is not None and not state.get(key).empty
        ]
    )
    athletes_label = "Atleta visible" if len(summary_df) == 1 else "Atletas visibles"
    datasets_label = "Fuente activa" if datasets_count == 1 else "Fuentes activas"
    adherence_label = "Adherencia del atleta" if report_athlete != "Todos" else "Adherencia promedio"

    _pdf_rect(commands, 48, 392, 499, 78, "#FEFEFE")
    _pdf_stroke_rect(commands, 48, 392, 499, 78, color="#D8DEE4")
    pulse_title = "Pulso del reporte" if audience == "profe" else "Lectura rápida"
    _pdf_text(commands, 64, 438, pulse_title, font="F2", size=12, color="#221F20")
    _pdf_text(commands, 64, 414, f"{athletes_label}: {athletes_text}", size=10, color="#5A595B")
    _pdf_text(commands, 220, 414, f"{datasets_label}: {datasets_count}", size=10, color="#5A595B")
    _pdf_text(commands, 392, 414, f"{adherence_label}: {completion_text}", size=10, color="#5A595B")

    _pdf_text(commands, 48, 346, "Nota de exportación", font="F2", size=12, color="#221F20")
    _pdf_multiline(
        commands,
        48,
        324,
        (
            [
                "Usa esta página como snapshot ejecutivo para cuerpo técnico y para toma de decisiones.",
                "La tabla integrada y la lectura por módulos continúan en las siguientes páginas.",
            ]
            if audience == "profe" else
            (
                [
                    "Esta página resume tu estado actual con foco en evaluaciones, perfil y próximos pasos.",
                    "Después vas a ver un perfil más claro y una lectura específica para vos.",
                ]
                if audience == "atleta" else
                [
                    "Esta página resume de forma simple dónde estás hoy y qué viene después.",
                    "Más abajo vas a ver un perfil resumido y los próximos pasos recomendados.",
                ]
            )
        ),
        size=10,
        color="#5A595B",
        leading=14,
    )
    return "\n".join(commands)


def _build_metric_profile_page(
    state: dict[str, pd.DataFrame | None],
    report_athlete: str,
    summary_df: pd.DataFrame,
    report_audience: str,
) -> str | None:
    audience = normalize_report_audience(report_audience)
    if summary_df.empty or report_athlete == "Todos":
        return None

    focus_row = summary_df.iloc[0]
    rows = _audience_metric_rows(state, focus_row, report_athlete, audience)
    title = "Perfil actual" if audience == "atleta" else "Dónde estás hoy"
    subtitle = (
        "Resumen cuasi técnico de tus evaluaciones, carga y recuperación."
        if audience == "atleta" else
        "Resumen simple del estado actual y de los indicadores más importantes."
    )

    commands: list[str] = []
    _pdf_rect(commands, 0, 0, 595, 842, "#FEFEFE")
    _pdf_text(commands, 48, 800, title, font="F2", size=18, color="#0D3C5E")
    _pdf_text(commands, 48, 778, subtitle, size=10, color="#5A595B")
    _pdf_line(commands, 48, 766, 547, 766, color="#D8DEE4")

    box_x = 48
    box_y = 176
    box_w = 499
    box_h = 552
    _pdf_rect(commands, box_x, box_y, box_w, box_h, "#F8FAFB")
    _pdf_stroke_rect(commands, box_x, box_y, box_w, box_h, color="#D8DEE4")

    left_rows = rows[::2]
    right_rows = rows[1::2]
    for col_idx, chunk in enumerate([left_rows, right_rows]):
        start_x = box_x + 20 + col_idx * 240
        current_y = box_y + box_h - 42
        for label, value in chunk:
            _pdf_text(commands, start_x, current_y, label, font="F2", size=10, color="#5A595B")
            _pdf_text(commands, start_x, current_y - 20, value, size=12, color="#221F20")
            _pdf_line(commands, start_x, current_y - 32, start_x + 205, current_y - 32, color="#E5E9ED")
            current_y -= 62
    return "\n".join(commands)


def _build_snapshot_pages(summary_df: pd.DataFrame) -> list[str]:
    if summary_df.empty:
        return []

    display_df = summary_df.copy().fillna("-")
    keep_cols = [
        col for col in ["Atleta", "ACWR EWMA", "Zona", "Wellness 3d", "CMJ cm", "DRI", "IMTP N", "Perfil NM"]
        if col in display_df.columns
    ]
    display_df = display_df[keep_cols]
    rename_map = {
        "Atleta": "Atleta",
        "ACWR EWMA": "ACWR EWMA",
        "Wellness 3d": "Wellness",
        "CMJ cm": "CMJ",
        "IMTP N": "IMTP",
        "Perfil NM": "Perfil NM",
    }
    display_df = display_df.rename(columns=rename_map)

    rows_per_page = 14
    pages: list[str] = []
    columns = display_df.columns.tolist()
    widths = [122, 52, 58, 56, 44, 44, 56, 67][: len(columns)]

    for start in range(0, len(display_df), rows_per_page):
        chunk = display_df.iloc[start:start + rows_per_page]
        commands: list[str] = []
        _pdf_rect(commands, 0, 0, 595, 842, "#FEFEFE")
        _pdf_text(commands, 48, 800, "Tabla ejecutiva integrada", font="F2", size=18, color="#0D3C5E")
        _pdf_text(commands, 48, 778, "KPIs más recientes para la ventana operativa visible.", size=10, color="#5A595B")
        _pdf_line(commands, 48, 766, 547, 766, color="#D8DEE4")

        table_x = 48
        header_y = 726
        row_h = 34
        current_x = table_x
        for col, width in zip(columns, widths):
            _pdf_rect(commands, current_x, header_y, width, row_h, "#0D3C5E")
            _pdf_text(commands, current_x + 8, header_y + 12, col, font="F2", size=8, color="#FEFEFE")
            current_x += width

        current_y = header_y - row_h
        for row_idx, (_, row) in enumerate(chunk.iterrows()):
            fill = "#F8FAFB" if row_idx % 2 == 0 else "#FEFEFE"
            current_x = table_x
            for col, width in zip(columns, widths):
                _pdf_rect(commands, current_x, current_y, width, row_h, fill)
                _pdf_stroke_rect(commands, current_x, current_y, width, row_h, color="#D8DEE4", width=0.8)
                raw_value = _snapshot_value(col, row.get(col, "-"))
                text_color = _zone_color(raw_value) if col == "Zona" else "#221F20"
                wrap_width = max(8, int(width / 5.8))
                if col in {"Atleta", "Perfil NM"}:
                    wrap_width = max(10, int(width / 6.4))
                wrapped = _wrap_lines(raw_value, wrap_width)
                _pdf_multiline(commands, current_x + 6, current_y + 20, wrapped[:2], size=8, color=text_color, leading=10)
                current_x += width
            current_y -= row_h

        pages.append("\n".join(commands))
    return pages


def _build_insight_pages(
    state: dict[str, pd.DataFrame | None],
    report_athlete: str,
    summary_df: pd.DataFrame,
    insights: dict[str, dict[str, object]],
    report_audience: str,
) -> list[str]:
    audience = normalize_report_audience(report_audience)
    payloads = _audience_blocks(state, report_athlete, summary_df, insights, audience)
    if not payloads:
        return []

    positions = [
        (48, 558), (306, 558),
        (48, 366), (306, 366),
        (48, 174), (306, 174),
    ]
    pages: list[str] = []
    blocks_per_page = len(positions)

    for start in range(0, len(payloads), blocks_per_page):
        commands: list[str] = []
        _pdf_rect(commands, 0, 0, 595, 842, "#F4F6F8")
        title = "Interpretación y focos" if audience == "profe" else ("Fortalezas y próximos pasos" if audience == "atleta" else "Lectura simple y próximos pasos")
        subtitle = (
            "Lectura editorial por módulo para seguimiento y reporte."
            if audience == "profe" else
            (
                "Bloques ordenados para entender qué está bien, qué mejorar y cómo seguir."
                if audience == "atleta" else
                "Resumen amigable para entender dónde estás hoy y cómo seguimos."
            )
        )
        _pdf_text(commands, 48, 800, title, font="F2", size=18, color="#0D3C5E")
        _pdf_text(commands, 48, 778, subtitle, size=10, color="#5A595B")
        _pdf_line(commands, 48, 766, 547, 766, color="#D8DEE4")

        for payload, (x, y) in zip(payloads[start:start + blocks_per_page], positions):
            _pdf_module_block(
                commands,
                x,
                y,
                241,
                152,
                title=payload.get("title", "Module"),
                summary=payload.get("summary", ""),
                focuses=payload.get("focuses", []),
            )
        pages.append("\n".join(commands))
    return pages


def generate_module_insights(
    state: dict[str, pd.DataFrame | None],
    report_athlete: str = "Todos",
    report_audience: str = "profe",
) -> dict[str, dict[str, object]]:
    audience = normalize_report_audience(report_audience)
    effective_athlete = resolve_report_scope(state, report_athlete, audience)
    if effective_athlete is None:
        return {
            "overview": {
                "title": "Lectura general",
                "summary": "Todavía no hay atletas válidos para preparar un reporte individual.",
                "focuses": ["Cargar datos y seleccionar un atleta antes de exportar."],
            }
        }

    summary_df = build_executive_summary_df(state, effective_athlete, audience)
    athletes = _selected_athletes(state, effective_athlete)
    active_datasets = [
        key for key in MODERN_REPORT_DATASET_KEYS
        if state.get(key) is not None and not state.get(key).empty
    ]
    quality_report = _report_quality_report(state)
    readiness = _readiness_payload(quality_report)
    alerts = quality_report.get("alerts", [])
    weekly_summaries = _report_weekly_summaries(state)
    current_team = _current_week_slice(weekly_summaries.get("weekly_team"))
    current_load = _current_week_slice(weekly_summaries.get("weekly_load"), effective_athlete)
    current_team_row = current_team.tail(1).iloc[0] if not current_team.empty else pd.Series(dtype=object)
    current_load_row = current_load.tail(1).iloc[0] if not current_load.empty else pd.Series(dtype=object)
    completion_snapshot = _completion_snapshot(state, effective_athlete)

    if effective_athlete == "Todos":
        overview_summary = (
            f"Readiness {readiness['label'].lower()}: {readiness['detail']} "
            f"Semana actual con {_display_metric(current_team_row.get('athletes_active'), digits=0)} atleta(s) activos."
            if active_datasets else
            "Todavía no hay información suficiente para construir una lectura ejecutiva."
        )
        overview_focuses = [
            "Priorizar la semana actual, adherencia y readiness antes de compartir el reporte.",
            alerts[0] if alerts else "La base actual ya permite una lectura integrada de carga, bienestar y rendimiento.",
        ]
    else:
        athlete_quality = _quality_athlete_row(quality_report, effective_athlete)
        quality_text = _quality_detail_text(athlete_quality)
        weekly_parts = _compact_lines(
            [
                f"sRPE semanal {_display_metric(current_load_row.get('weekly_sRPE'), digits=0)}" if _coerce_float(current_load_row.get("weekly_sRPE")) is not None else None,
                f"ACWR EWMA {_display_metric(current_load_row.get('ACWR_EWMA_last'), digits=2)}" if _coerce_float(current_load_row.get("ACWR_EWMA_last")) is not None else None,
                completion_snapshot["value"] if completion_snapshot["numeric"] is not None else None,
            ]
        )
        overview_summary = (
            f"Readiness {readiness['label'].lower()} para {effective_athlete}. "
            f"{' | '.join(weekly_parts)}."
            if active_datasets else
            f"Todavía no hay información suficiente para construir una lectura ejecutiva de {effective_athlete}."
        )
        overview_focuses = [
            quality_text or "Cuidar continuidad de carga, adherencia y evaluaciones dentro de la misma ventana operativa.",
            alerts[0] if alerts else "El reporte prioriza semana actual, última evaluación útil y readiness del atleta.",
        ]

    insights: dict[str, dict[str, object]] = {
        "overview": {
            "title": "Lectura general",
            "summary": overview_summary,
            "focuses": overview_focuses,
        }
    }

    operational_notes = _session_notes_for_scope(state, effective_athlete, days=42, max_rows=4)
    if not operational_notes.empty:
        focus_lines = []
        for _, note in operational_notes.head(3).iterrows():
            detail = _compact_lines(
                [
                    _note_text(note.get("Assigned_Exercise")) if _has_text(note.get("Assigned_Exercise")) else None,
                    _note_text(note.get("Opt_Out_Type")) if _has_text(note.get("Opt_Out_Type")) else None,
                    _note_text(note.get("Explanation_Text")) if _has_text(note.get("Explanation_Text")) else None,
                ]
            )
            label = _note_text(note.get("Athlete")) if effective_athlete == "Todos" else _note_date_text(note.get("Date"))
            focus_lines.append(f"{label}: {' | '.join(detail)}" if detail else f"{label}: revisar nota operativa.")
        insights["operational_context"] = {
            "title": "Contexto operativo",
            "summary": _session_note_summary(operational_notes),
            "focuses": focus_lines,
        }

    individual_eval_available = False
    if effective_athlete != "Todos" and not summary_df.empty:
        row = summary_df.iloc[0]
        eval_available = _row_has_eval_data(row)
        load_available = _row_has_load_data(row)
        wellness_available = _row_has_wellness_data(row)
        individual_eval_available = eval_available

        acwr = _coerce_float(row.get("ACWR EWMA"))
        monotony = _coerce_float(row.get("Monotonia"))
        wellness = _coerce_float(row.get("Wellness 3d"))
        cmj_delta = _coerce_float(row.get("CMJ vs BL %"))

        load_notes: list[str] = []
        if acwr is not None:
            if acwr > 1.5:
                load_notes.append("Bajar carga aguda y revisar tolerancia de la semana.")
            elif acwr < 0.8:
                load_notes.append("Verificar si la subcarga es planificada o si falta estímulo.")
            else:
                load_notes.append("La relación agudo-crónica se mantiene en una zona de trabajo útil.")
        if monotony is not None and monotony > 2.0:
            load_notes.append("Aumentar variabilidad del microciclo para reducir monotonía.")
        if wellness is not None and wellness < 15:
            load_notes.append("Seguir recuperación diaria porque el wellness reciente está deprimido.")

        eval_focus: list[str] = []
        if cmj_delta is not None:
            if cmj_delta <= -5:
                eval_focus.append("El CMJ cae respecto al baseline inicial; conviene mirar fatiga y exposicion reciente.")
            elif cmj_delta >= 5:
                eval_focus.append("La salida vertical esta por encima del baseline inicial.")
            else:
                eval_focus.append("La evaluación se mantiene cerca de la línea base del atleta.")
        if _has_text(row.get("Perfil NM")):
            eval_focus.append(f"Perfil neuromuscular actual: {_profile_text(row)}.")
        if not eval_available:
            eval_focus.append("Todavía no hay una evaluación reciente para construir una comparación objetiva.")

        load_summary_parts = _compact_lines(
            [
                f"ACWR EWMA {_display_metric(acwr, digits=2)}" if acwr is not None else None,
                f"Monotonía {_display_metric(monotony, digits=2)}" if monotony is not None else None,
                f"Wellness 3 días {_display_metric(wellness, digits=1)}" if wellness_available else None,
            ]
        )
        eval_summary_parts = _compact_lines(
            [
                f"CMJ {_display_metric(row.get('CMJ cm'), digits=1, suffix=' cm')}" if _coerce_float(row.get("CMJ cm")) is not None else None,
                f"{EUR_RATIO_LABEL} {_display_metric(_summary_eur_value(row), digits=2)}" if _coerce_float(_summary_eur_value(row)) is not None else None,
                f"DRI {_display_metric(row.get('DRI'), digits=2)}" if _coerce_float(row.get("DRI")) is not None else None,
                f"IMTP {_display_metric(row.get('IMTP N'), digits=0, suffix=' N')}" if _coerce_float(row.get("IMTP N")) is not None else None,
            ]
        )

        insights["load"] = {
            "title": "Lectura de carga",
            "summary": " | ".join(load_summary_parts) if load_summary_parts else "No hay datos recientes de carga y recuperación para una lectura estable.",
            "focuses": load_notes or ["Sin suficientes datos para una lectura de carga completa."],
        }
        insights["evaluations"] = {
            "title": "Lectura de evaluación",
            "summary": " | ".join(eval_summary_parts) if eval_summary_parts else "No hay una evaluación reciente cargada para este atleta.",
            "focuses": eval_focus,
        }

        if eval_available and (load_available or wellness_available):
            profile_summary = f"El perfil integrado de {effective_athlete} combina carga reciente, percepción de recuperación y última evaluación."
        elif eval_available:
            profile_summary = f"La lectura actual de {effective_athlete} se apoya principalmente en la última evaluación disponible."
        elif load_available or wellness_available:
            profile_summary = f"La lectura actual de {effective_athlete} se apoya en carga y bienestar recientes, a la espera de una nueva evaluación."
        else:
            profile_summary = f"Todavía no hay información suficiente para construir un perfil integrado de {effective_athlete}."

        insights["profile"] = {
            "title": "Foco del atleta",
            "summary": profile_summary,
            "focuses": list(dict.fromkeys(load_notes + eval_focus))[:3] or ["Seguir acumulando historial para una lectura individual más precisa."],
        }
    else:
        zone_counts = summary_df["Zona"].value_counts().to_dict() if "Zona" in summary_df.columns else {}
        profile_counts = summary_df["Perfil NM"].value_counts().to_dict() if "Perfil NM" in summary_df.columns else {}
        zone_text = ", ".join(f"{key}: {value}" for key, value in zone_counts.items()) if zone_counts else "sin zonas de carga disponibles"
        profile_text = ", ".join(f"{key}: {value}" for key, value in profile_counts.items()) if profile_counts else "sin perfiles neuromusculares disponibles"

        insights["load"] = {
            "title": "Lectura de carga",
            "summary": f"Distribución actual del equipo: {zone_text}.",
            "focuses": [
                "Revisar atletas en alto riesgo o precaución antes del siguiente microciclo." if any(key in zone_counts for key in ["Alto riesgo", "Precaucion", "Precaución"]) else "La distribución de carga no muestra acumulación marcada de riesgo.",
                "Monitorear la monotonía de quienes concentran más strain semanal.",
            ],
        }
        insights["evaluations"] = {
            "title": "Lectura de evaluación",
            "summary": f"Perfiles neuromusculares visibles: {profile_text}.",
            "focuses": [
                "Cruzar los perfiles reactivos con la carga reciente para decidir exposición pliométrica.",
                f"Usar IMTP, {EUR_RATIO_LABEL} y DRI para separar necesidades de fuerza base vs reactividad.",
            ],
        }
        insights["profile"] = {
            "title": "Lectura del equipo",
            "summary": "La lectura grupal resume la foto actual del plantel dentro de la ventana operativa visible.",
            "focuses": [
                "Identificar atletas fuera de rango antes de programar el siguiente bloque.",
                "Completar wellness y evaluaciones faltantes para cerrar la lectura por jugador.",
            ],
        }

    completion_mean = _team_completion_mean(state)
    individual_completion = _focus_completion_value(state, effective_athlete) if effective_athlete != "Todos" else None
    insights["team"] = {
        "title": "Adherencia del plan" if effective_athlete != "Todos" else "Adherencia del equipo",
        "summary": (
            f"Adherencia individual reciente: {individual_completion:.1f}%."
            if effective_athlete != "Todos" and individual_completion is not None else
            (
                f"Adherencia promedio del equipo: {completion_mean:.1f}%."
                if completion_mean is not None else
                "No hay información de adherencia cargada para evaluar ejecución."
            )
        ),
        "focuses": [
            (
                "Si la adherencia baja, revisar barreras de cumplimiento, disponibilidad y organización semanal."
                if effective_athlete != "Todos" else
                "Si la adherencia baja, revisar progresiones, disponibilidad y fricción operativa."
            ),
            (
                "Cruzar adherencia con carga y bienestar para entender si el plan realmente se está sosteniendo."
                if effective_athlete != "Todos" else
                "Alinear carga y adherencia para entender si el volumen planificado realmente se ejecuta."
            ),
        ],
    }

    if audience == "profe":
        target = effective_athlete if effective_athlete != "Todos" else "el equipo"
        report_summary = f"Versión técnica curada para {target}, priorizando semana actual, última evaluación útil, adherencia y readiness."
        report_focuses = [
            "El corazón del reporte resume semana actual, evaluación útil, adherencia y calidad del dato.",
            "El anexo técnico queda opt-in para no convertir el export en un dump de tablas.",
        ]
    elif audience == "atleta":
        report_summary = (
            f"Versión individual orientada al atleta, con foco en semana actual, evaluación útil y próximos pasos de trabajo para {effective_athlete}."
            if individual_eval_available else
            f"Versión individual orientada al atleta, centrada en estado actual, adherencia y próximos pasos de trabajo para {effective_athlete}."
        )
        report_focuses = [
            "Destacar fortalezas, oportunidades y foco inmediato en lenguaje cuasi técnico.",
            "Ordenar la información para que el atleta entienda qué sigue y por qué.",
        ]
    else:
        report_summary = (
            f"Versión individual amigable para cliente, enfocada en explicar semana actual, estado actual y próximos pasos de {effective_athlete}."
            if individual_eval_available else
            f"Versión individual amigable para cliente, enfocada en explicar seguimiento actual y próximos pasos de {effective_athlete}."
        )
        report_focuses = [
            "Simplificar el lenguaje y evitar tecnicismos innecesarios.",
            "Mostrar progreso, estado actual y próximos pasos con claridad.",
        ]

    insights["report"] = {
        "title": "Estado del reporte" if audience == "profe" else "Enfoque del reporte",
        "summary": report_summary,
        "focuses": report_focuses,
    }
    return insights


def _audience_dashboard_cards(
    state: dict[str, pd.DataFrame | None],
    focus_row: pd.Series,
    report_athlete: str,
    audience: str,
) -> list[tuple[str, str, str]]:
    completion_value = _focus_completion_value(state, report_athlete)
    eval_available = _row_has_eval_data(focus_row)
    load_available = _row_has_load_data(focus_row)
    wellness_available = _row_has_wellness_data(focus_row)
    cards: list[tuple[str, str, str]] = []

    if audience == "atleta":
        if eval_available:
            if _has_text(focus_row.get("Perfil NM")):
                cards.append(("Perfil actual", _short_profile_label(focus_row.get("Perfil NM")), "#0D3C5E"))
            if _coerce_float(focus_row.get("CMJ cm")) is not None:
                cards.append(("CMJ", _display_metric(focus_row.get("CMJ cm"), digits=1, suffix=" cm"), "#0D3C5E"))
        else:
            cards.append(("Evaluación reciente", "Pendiente", "#708C9F"))

        if load_available:
            cards.append(("ACWR EWMA", _display_metric(focus_row.get("ACWR EWMA"), digits=2), _zone_color(focus_row.get("Zona"))))
        if wellness_available:
            cards.append(("Wellness 3 días", _display_metric(focus_row.get("Wellness 3d"), digits=1), "#2F6B52"))
        if completion_value is not None:
            cards.append(("Adherencia", _display_metric(completion_value, digits=1, suffix="%"), "#708C9F"))
        cards.append(("Objetivo actual", _current_focus_text(focus_row, audience="atleta"), "#5A595B"))
        return cards[:6]

    if audience == "cliente":
        if load_available and _has_text(focus_row.get("Zona")):
            cards.append(("Estado actual", _display_zone(focus_row.get("Zona")), _zone_color(focus_row.get("Zona"))))
        else:
            cards.append(("Estado actual", "En seguimiento", "#708C9F"))

        if eval_available and _coerce_float(focus_row.get("CMJ cm")) is not None:
            cards.append(("Salto vertical", _display_metric(focus_row.get("CMJ cm"), digits=1, suffix=" cm"), "#0D3C5E"))
        if eval_available and _coerce_float(focus_row.get("CMJ vs BL %")) is not None:
            cards.append(("Cambio vs baseline", _display_metric(focus_row.get("CMJ vs BL %"), digits=1, suffix="%"), "#134263"))
        elif not eval_available:
            cards.append(("Evaluación física", "Pendiente", "#708C9F"))

        if wellness_available:
            cards.append(("Bienestar", _display_metric(focus_row.get("Wellness 3d"), digits=1), "#2F6B52"))
        if completion_value is not None:
            cards.append(("Constancia", _display_metric(completion_value, digits=1, suffix="%"), "#708C9F"))
        if eval_available and _has_text(focus_row.get("Perfil NM")):
            cards.append(("Perfil actual", _short_profile_label(focus_row.get("Perfil NM")), "#5A595B"))
        cards.append(("Próximo foco", _current_focus_text(focus_row, audience="cliente"), "#134263"))
        return cards[:6]

    if load_available and _coerce_float(focus_row.get("ACWR EWMA")) is not None:
        cards.append(("ACWR EWMA", _display_metric(focus_row.get("ACWR EWMA"), digits=2), _zone_color(focus_row.get("Zona"))))
    if load_available and _has_text(focus_row.get("Zona")):
        cards.append(("Zona de carga", _display_zone(focus_row.get("Zona")), "#708C9F"))
    if wellness_available:
        cards.append(("Wellness 3 días", _display_metric(focus_row.get("Wellness 3d"), digits=1), "#2F6B52"))
    if _coerce_float(focus_row.get("Monotonia")) is not None:
        cards.append(("Monotonía", _display_metric(focus_row.get("Monotonia"), digits=2), "#5A595B"))
    if eval_available and _coerce_float(focus_row.get("CMJ cm")) is not None:
        cards.append(("CMJ", _display_metric(focus_row.get("CMJ cm"), digits=1, suffix=" cm"), "#0D3C5E"))
    if eval_available and _coerce_float(focus_row.get("IMTP N")) is not None:
        cards.append(("IMTP", _display_metric(focus_row.get("IMTP N"), digits=0, suffix=" N"), "#134263"))
    if not eval_available:
        cards.append(("Evaluación", "Sin registro reciente", "#708C9F"))
    if completion_value is not None and len(cards) < 6:
        cards.append(("Adherencia", _display_metric(completion_value, digits=1, suffix="%"), "#708C9F"))
    return cards[:6]


def _audience_metric_rows(
    state: dict[str, pd.DataFrame | None],
    focus_row: pd.Series,
    report_athlete: str,
    audience: str,
) -> list[tuple[str, str]]:
    completion_value = _focus_completion_value(state, report_athlete)
    eval_available = _row_has_eval_data(focus_row)
    load_available = _row_has_load_data(focus_row)
    wellness_available = _row_has_wellness_data(focus_row)
    rows: list[tuple[str, str]] = []

    if audience == "cliente":
        if load_available and _has_text(focus_row.get("Zona")):
            rows.append(("Cómo venís hoy", _display_zone(focus_row.get("Zona"))))
        if eval_available and _coerce_float(focus_row.get("CMJ vs BL %")) is not None:
            rows.append(("Cambio vs baseline inicial", _display_metric(focus_row.get("CMJ vs BL %"), digits=1, suffix="%")))
        elif not eval_available:
            rows.append(("Chequeo físico", "Todavía no contamos con una medición física comparable."))
        if eval_available and _coerce_float(focus_row.get("CMJ cm")) is not None:
            rows.append(("Tu salto hoy", _display_metric(focus_row.get("CMJ cm"), digits=1, suffix=" cm")))
        if wellness_available:
            rows.append(("Bienestar reciente", _display_metric(focus_row.get("Wellness 3d"), digits=1)))
        if completion_value is not None:
            rows.append(("Constancia del plan", _display_metric(completion_value, digits=1, suffix="%")))
        if eval_available and _has_text(focus_row.get("Perfil NM")):
            rows.append(("Lectura actual", _profile_text(focus_row)))
        elif load_available or wellness_available:
            rows.append(("Lo que estamos mirando", "Ya hay información reciente para seguir tu proceso."))
        rows.append(("Objetivo actual", _current_focus_text(focus_row, audience="cliente")))
        return rows or [("Estado actual", "Falta información reciente para resumir tu progreso.")]

    if audience == "atleta":
        if eval_available:
            if _has_text(focus_row.get("Perfil NM")):
                rows.append(("Perfil neuromuscular", _profile_text(focus_row)))
            if _coerce_float(focus_row.get("CMJ cm")) is not None:
                rows.append(("CMJ", _display_metric(focus_row.get("CMJ cm"), digits=1, suffix=" cm")))
            if _coerce_float(focus_row.get("DRI")) is not None:
                rows.append(("DRI", _display_metric(focus_row.get("DRI"), digits=2)))
            if _coerce_float(focus_row.get("IMTP N")) is not None:
                rows.append(("IMTP", _display_metric(focus_row.get("IMTP N"), digits=0, suffix=" N")))
            focus_eur = _summary_eur_value(focus_row)
            if _coerce_float(focus_eur) is not None:
                rows.append((EUR_RATIO_LABEL, _display_metric(focus_eur, digits=2)))
            if _coerce_float(focus_row.get("CMJ vs BL %")) is not None:
                rows.append(("Cambio vs baseline inicial", _display_metric(focus_row.get("CMJ vs BL %"), digits=1, suffix="%")))
        else:
            rows.append(("Evaluación reciente", "Sin evaluación física reciente"))

        if load_available and _coerce_float(focus_row.get("ACWR EWMA")) is not None:
            rows.append(("ACWR EWMA", _display_metric(focus_row.get("ACWR EWMA"), digits=2)))
        if load_available and _coerce_float(focus_row.get("Monotonia")) is not None:
            rows.append(("Monotonía", _display_metric(focus_row.get("Monotonia"), digits=2)))
        if wellness_available:
            rows.append(("Wellness 3 días", _display_metric(focus_row.get("Wellness 3d"), digits=1)))
        if completion_value is not None:
            rows.append(("Adherencia al plan", _display_metric(completion_value, digits=1, suffix="%")))
        rows.append(("Objetivo actual", _current_focus_text(focus_row, audience="atleta")))
        if not eval_available:
            rows.append(("Próxima acción", "Programar una nueva batería de evaluaciones."))
        return rows

    if _has_text(focus_row.get("Perfil NM")):
        rows.append(("Perfil neuromuscular", _profile_text(focus_row)))
    if _coerce_float(focus_row.get("CMJ cm")) is not None:
        rows.append(("CMJ", _display_metric(focus_row.get("CMJ cm"), digits=1, suffix=" cm")))
    if _coerce_float(focus_row.get("DRI")) is not None:
        rows.append(("DRI", _display_metric(focus_row.get("DRI"), digits=2)))
    if _coerce_float(focus_row.get("IMTP N")) is not None:
        rows.append(("IMTP", _display_metric(focus_row.get("IMTP N"), digits=0, suffix=" N")))
    if _coerce_float(focus_row.get("ACWR EWMA")) is not None:
        rows.append(("ACWR EWMA", _display_metric(focus_row.get("ACWR EWMA"), digits=2)))
    if _coerce_float(focus_row.get("Monotonia")) is not None:
        rows.append(("Monotonía", _display_metric(focus_row.get("Monotonia"), digits=2)))
    if wellness_available:
        rows.append(("Wellness 3 días", _display_metric(focus_row.get("Wellness 3d"), digits=1)))
    if completion_value is not None:
        rows.append(("Adherencia", _display_metric(completion_value, digits=1, suffix="%")))
    return rows or [("Estado actual", "Sin datos suficientes para una lectura técnica estable.")]


def _audience_blocks(
    state: dict[str, pd.DataFrame | None],
    report_athlete: str,
    summary_df: pd.DataFrame,
    insights: dict[str, dict[str, object]],
    audience: str,
) -> list[dict[str, object]]:
    if report_athlete == "Todos" or summary_df.empty:
        ordered_keys = ["overview", "load", "operational_context", "evaluations", "profile", "team", "report"]
        return [insights[key] for key in ordered_keys if insights.get(key)]

    row = summary_df.iloc[0]
    eval_available = _row_has_eval_data(row)
    load_available = _row_has_load_data(row)
    wellness_available = _row_has_wellness_data(row)
    completion_value = _focus_completion_value(state, report_athlete)

    if audience == "profe":
        integrated_focuses = _compact_lines(
            [
                f"Perfil actual: {_profile_text(row)}." if _has_text(row.get("Perfil NM")) else None,
                f"ACWR EWMA {_display_metric(row.get('ACWR EWMA'), digits=2)} en zona {_display_zone(row.get('Zona'))}." if load_available and _coerce_float(row.get("ACWR EWMA")) is not None else None,
                f"CMJ {_display_metric(row.get('CMJ cm'), digits=1, suffix=' cm')} y DRI {_display_metric(row.get('DRI'), digits=2)}." if eval_available and _coerce_float(row.get("CMJ cm")) is not None and _coerce_float(row.get("DRI")) is not None else None,
                f"IMTP {_display_metric(row.get('IMTP N'), digits=0, suffix=' N')}." if eval_available and _coerce_float(row.get("IMTP N")) is not None else None,
            ]
        )
        planning_focuses = _technical_planning_focuses(row, completion_value)
        blocks = [
            {
                "title": "Carga actual",
                "summary": insights.get("load", {}).get("summary", "Sin lectura de carga disponible."),
                "focuses": insights.get("load", {}).get("focuses", []),
            },
            {
                "title": "Evaluación actual",
                "summary": insights.get("evaluations", {}).get("summary", "Sin evaluación reciente disponible."),
                "focuses": insights.get("evaluations", {}).get("focuses", []),
            },
            {
                "title": "Lectura integrada",
                "summary": "Cruce actual entre carga, evaluación y disponibilidad para orientar decisiones del próximo bloque.",
                "focuses": integrated_focuses or ["Todavía no hay datos suficientes para una síntesis técnica robusta."],
            },
            {
                "title": "Implicancias para la planificación",
                "summary": "Prioridades concretas para decidir qué cualidad empujar y qué controlar en la próxima ventana.",
                "focuses": planning_focuses,
            },
            {
                "title": insights.get("team", {}).get("title", "Adherencia del plan"),
                "summary": insights.get("team", {}).get("summary", "Sin lectura de adherencia disponible."),
                "focuses": insights.get("team", {}).get("focuses", []),
            },
        ]
        operational = insights.get("operational_context")
        if operational:
            blocks.insert(2, operational)
        return blocks

    strengths = _strengths_from_row(row, audience=audience)
    gaps = _gaps_from_row(row, audience=audience)
    next_steps = _next_steps_from_row(row, audience=audience)
    if not eval_available:
        next_steps = list(
            dict.fromkeys(
                ["Completar una nueva batería de evaluaciones para tener una referencia objetiva."] + next_steps
            )
        )[:3]

    if audience == "cliente":
        objective_focuses = _objective_focuses_from_row(row, audience="cliente")
        if eval_available:
            zone_text = (
                f" y tu estado general se lee como {_display_zone(row.get('Zona'))}"
                if load_available and _has_text(row.get("Zona")) else
                ""
            )
            progress_summary = (
                f"Tomamos como referencia tu baseline inicial. Hoy el cambio visible es {_display_metric(row.get('CMJ vs BL %'), digits=1, suffix='%')}{zone_text}."
                if _coerce_float(row.get("CMJ vs BL %")) is not None else
                "Ya contamos con una evaluación reciente para ubicar mejor tu punto actual y seguir el proceso con más claridad."
            )
        else:
            progress_summary = "Todavía no tenemos una comparación física completa, así que esta lectura se apoya en tu seguimiento reciente y en cómo viene respondiendo el proceso."

        current_focuses = _compact_lines(
            [
                f"Estado actual: {_display_zone(row.get('Zona'))}." if load_available and _has_text(row.get("Zona")) else None,
                f"Bienestar reciente: {_display_metric(row.get('Wellness 3d'), digits=1)}." if wellness_available else None,
                f"Constancia reciente: {_display_metric(completion_value, digits=1, suffix='%')}." if completion_value is not None else None,
                f"Salto vertical actual: {_display_metric(row.get('CMJ cm'), digits=1, suffix=' cm')}." if eval_available and _coerce_float(row.get("CMJ cm")) is not None else None,
            ]
        ) or ["Seguimos construyendo información útil para entender mejor tu evolución."]

        blocks = [
            {
                "title": "Tu progreso hasta hoy",
                "summary": progress_summary,
                "focuses": current_focuses,
            },
            {
                "title": "Qué está funcionando bien",
                "summary": "Estos puntos muestran dónde el proceso ya está dando señales favorables.",
                "focuses": strengths,
            },
            {
                "title": "Lo que vamos a seguir mejorando",
                "summary": "Todavía hay margen de mejora y eso orienta lo que conviene trabajar ahora.",
                "focuses": gaps,
            },
            {
                "title": "Objetivo actual",
                "summary": f"El foco inmediato del trabajo es {_current_focus_text(row, audience='cliente').lower()}.",
                "focuses": objective_focuses or ["Vamos a sostener el proceso y seguir ordenando la progresión."],
            },
            {
                "title": "Próximos pasos",
                "summary": "Plan de acción inmediato para seguir avanzando con claridad y continuidad.",
                "focuses": next_steps,
            },
        ]
        operational = insights.get("operational_context")
        if operational:
            blocks.insert(-1, operational)
        return blocks

    if eval_available:
        profile_summary = " | ".join(
            _compact_lines(
                [
                    f"Perfil neuromuscular: {_profile_text(row)}" if _has_text(row.get("Perfil NM")) else None,
                    f"CMJ {_display_metric(row.get('CMJ cm'), digits=1, suffix=' cm')}" if _coerce_float(row.get("CMJ cm")) is not None else None,
                    f"DRI {_display_metric(row.get('DRI'), digits=2)}" if _coerce_float(row.get("DRI")) is not None else None,
                    f"IMTP {_display_metric(row.get('IMTP N'), digits=0, suffix=' N')}" if _coerce_float(row.get("IMTP N")) is not None else None,
                ]
            )
        )
        profile_focuses = _compact_lines(
            [
                f"ACWR EWMA actual: {_display_metric(row.get('ACWR EWMA'), digits=2)} en zona {_display_zone(row.get('Zona'))}." if load_available and _coerce_float(row.get("ACWR EWMA")) is not None else None,
                f"Wellness reciente: {_display_metric(row.get('Wellness 3d'), digits=1)}." if wellness_available else None,
                f"Cambio frente al baseline inicial: {_display_metric(row.get('CMJ vs BL %'), digits=1, suffix='%')}." if _coerce_float(row.get("CMJ vs BL %")) is not None else None,
            ]
        ) or strengths[:2]
        first_block = {
            "title": "Perfil actual",
            "summary": profile_summary or "Hay evaluación reciente, pero faltan algunas métricas clave para completar la lectura.",
            "focuses": profile_focuses,
        }
    else:
        first_block = {
            "title": "Estado actual",
            "summary": "Todavía no hay una evaluación reciente cargada; por ahora la lectura se apoya en carga, bienestar y adherencia.",
            "focuses": _compact_lines(
                [
                    f"Estado de carga: {_display_zone(row.get('Zona'))}." if load_available and _has_text(row.get("Zona")) else None,
                    f"Wellness reciente: {_display_metric(row.get('Wellness 3d'), digits=1)}." if wellness_available else None,
                    f"Adherencia promedio: {_display_metric(completion_value, digits=1, suffix='%')}." if completion_value is not None else None,
                ]
            ) or ["La prioridad es completar una nueva evaluación para definir mejor el perfil actual."],
        }

    objective_focuses = _objective_focuses_from_row(row, audience="atleta")

    blocks = [
        first_block,
        {
            "title": "Fortalezas actuales",
            "summary": "Variables que hoy acompañan bien tu rendimiento y tu disponibilidad.",
            "focuses": strengths,
        },
        {
            "title": "Cosas a mejorar",
            "summary": "Variables a vigilar para que el próximo bloque sea más sólido.",
            "focuses": gaps,
        },
        {
            "title": "Objetivo actual",
            "summary": f"El foco inmediato del bloque es {_current_focus_text(row, audience='atleta').lower()}.",
            "focuses": objective_focuses or ["La prioridad es sostener el perfil actual sin perder disponibilidad."],
        },
        {
            "title": "Siguientes pasos",
            "summary": "Próximas prioridades de trabajo con un lenguaje cuasi técnico y accionable.",
            "focuses": next_steps,
        },
    ]
    operational = insights.get("operational_context")
    if operational:
        blocks.insert(-1, operational)
    return blocks


def _build_trend_page(
    state: dict[str, pd.DataFrame | None],
    report_athlete: str,
    report_audience: str,
) -> str | None:
    if report_athlete == "Todos":
        return None

    audience = normalize_report_audience(report_audience)
    candidates = {
        "profe": [
            ("CMJ", _cmj_series(state, report_athlete), "#0D3C5E"),
            ("ACWR EWMA", _acwr_series(state, report_athlete), "#708C9F"),
            ("Wellness 3 días", _wellness_series(state, report_athlete), "#2F6B52"),
        ],
        "atleta": [
            ("CMJ", _cmj_series(state, report_athlete), "#0D3C5E"),
            ("Wellness 3 días", _wellness_series(state, report_athlete), "#2F6B52"),
            ("ACWR EWMA", _acwr_series(state, report_athlete), "#708C9F"),
        ],
        "cliente": [
            ("Progreso de salto", _cmj_series(state, report_athlete), "#0D3C5E"),
            ("Bienestar reciente", _wellness_series(state, report_athlete), "#2F6B52"),
            ("Seguimiento del proceso", _acwr_series(state, report_athlete), "#708C9F"),
        ],
    }[audience]
    panels = [(title, series, color) for title, series, color in candidates if len(series) >= 2][:2]
    if not panels:
        return None

    commands: list[str] = []
    _pdf_rect(commands, 0, 0, 595, 842, "#F4F6F8")
    title = "Tendencias clave" if audience != "cliente" else "Evolución reciente"
    subtitle = (
        "Curvas mínimas para seguir cómo se movieron las variables con mejor valor de decisión."
        if audience == "profe" else
        (
            "Lectura simple para ver cómo cambió tu estado reciente."
            if audience == "cliente" else
            "Lectura visual mínima para seguir tu evolución sin sobrecargar el reporte."
        )
    )
    _pdf_text(commands, 48, 800, title, font="F2", size=18, color="#0D3C5E")
    _pdf_text(commands, 48, 778, subtitle, size=10, color="#5A595B")
    _pdf_line(commands, 48, 766, 547, 766, color="#D8DEE4")

    if len(panels) == 1:
        panel_title, series, color = panels[0]
        _pdf_chart_panel(commands, 48, 196, 499, 470, title=panel_title, series=series, line_color=color)
    else:
        first_title, first_series, first_color = panels[0]
        second_title, second_series, second_color = panels[1]
        _pdf_chart_panel(commands, 48, 408, 499, 280, title=first_title, series=first_series, line_color=first_color)
        _pdf_chart_panel(commands, 48, 96, 499, 280, title=second_title, series=second_series, line_color=second_color)
    return "\n".join(commands)


def report_plotly_export_ready() -> bool:
    try:
        import plotly.io as pio  # noqa: F401
    except Exception:
        return False
    return _import_kaleido_backend() is not None


def _import_kaleido_backend():
    try:
        with warnings.catch_warnings():
            warnings.filterwarnings(
                "ignore",
                category=UserWarning,
                module=r"kaleido\._sync_server",
            )
            import kaleido
    except Exception:
        return None
    return kaleido


def _get_kaleido_chrome_sync() -> str | None:
    kaleido = _import_kaleido_backend()
    if kaleido is None:
        return None
    try:
        return kaleido.get_chrome_sync()
    except Exception:
        return None


def _plotly_figure_to_png_bytes(
    pio_module,
    figure: object,
    *,
    width: int,
    height: int,
    scale: int,
) -> bytes | None:
    try:
        with warnings.catch_warnings():
            warnings.filterwarnings(
                "ignore",
                category=UserWarning,
                module=r"kaleido\._sync_server",
            )
            return pio_module.to_image(figure, format="png", width=width, height=height, scale=scale)
    except Exception:
        return None


def _build_report_chart_theme() -> dict:
    colors = {
        "bg": "#F4F6F8",
        "card": "#FEFEFE",
        "navy": "#0D3C5E",
        "steel": "#134263",
        "black": "#221F20",
        "white": "#221F20",
        "gray": "#5A595B",
        "muted": "#708C9F",
        "border": "#D8DEE4",
        "blue": "#0D3C5E",
        "green": "#2F6B52",
        "yellow": "#C4A464",
        "orange": "#B87445",
        "red": "#B56B73",
    }
    return {
        "colors": colors,
        "layout": dict(
            template="plotly_white",
            paper_bgcolor=colors["bg"],
            plot_bgcolor=colors["card"],
            font=dict(family="Helvetica", color=colors["black"], size=11),
            margin=dict(l=44, r=28, t=68, b=46),
        ),
        "grid": "rgba(34, 31, 32, 0.08)",
        "grid_soft": "rgba(34, 31, 32, 0.05)",
        "reference_line": "rgba(34, 31, 32, 0.18)",
        "reference_fill": "rgba(13, 60, 94, 0.06)",
        "legend": dict(
            orientation="h",
            y=-0.18,
            bgcolor="rgba(254, 254, 254, 0.92)",
            bordercolor=colors["border"],
            borderwidth=1,
            font=dict(size=9, color=colors["gray"]),
        ),
        "monotony_high": 2.0,
    }


def _team_mean_for_radar(jdf: pd.DataFrame) -> dict[str, float]:
    if jdf is None or jdf.empty:
        return {}

    prepared = _prepare_jump_df(jdf)
    z_keys = ("SJ_Z", "CMJ_Z", "DJ_height_Z", "DJ_RSI_Z", "TC_inv_Z", "IMTP_relPF_Z")
    team_means: dict[str, float] = {}
    for key in z_keys:
        values = prepared.apply(lambda row: resolve_zscore(row, key), axis=1)
        numeric = pd.to_numeric(values, errors="coerce").dropna()
        if not numeric.empty:
            team_means[key] = float(numeric.mean())
    return team_means


def collect_report_plotly_figures(
    state: dict[str, pd.DataFrame | None],
    report_athlete: str = "Todos",
    report_audience: str = "profe",
) -> list[dict[str, object]]:
    try:
        from charts.load_charts import chart_acwr, chart_completion, chart_wellness
        from charts.dashboard_charts import (
            chart_cmj_trend,
            chart_quadrant_cmj_imtp,
            chart_quadrant_dri_sj,
            chart_quadrant_rsi_sj,
            chart_radar,
        )
    except Exception:
        return []

    audience = normalize_report_audience(report_audience)
    effective_athlete = resolve_report_scope(state, report_athlete, audience)
    if effective_athlete is None:
        return []
    theme = _build_report_chart_theme()
    figures: list[dict[str, object]] = []
    jdf = state.get("jump_df")

    if effective_athlete == "Todos":
        completion_plot_df = _completion_plot_df(state, "Todos")
        if not completion_plot_df.empty:
            figures.append(
                {
                    "slug": "completion_team",
                    "title": "Adherencia del equipo",
                    "figure": chart_completion(completion_plot_df, theme=theme, athlete_label="Todos"),
                }
            )

        if jdf is not None and not jdf.empty and "Athlete" in jdf.columns:
            latest_team = jdf.sort_values("Date").groupby("Athlete").last().reset_index()
            if len(latest_team) > 1 and {"CMJ_cm", "IMTP_N"}.issubset(latest_team.columns):
                figures.append(
                    {
                        "slug": "quadrant_cmj_imtp",
                        "title": "Mapa de potencia y fuerza máxima",
                        "figure": chart_quadrant_cmj_imtp(latest_team, theme=theme),
                    }
                )
            if len(latest_team) > 1 and {"DJ_RSI", "SJ_cm"}.issubset(latest_team.columns):
                figures.append(
                    {
                        "slug": "quadrant_rsi_sj",
                        "title": "Mapa de RSI y fuerza concéntrica",
                        "figure": chart_quadrant_rsi_sj(latest_team, theme=theme),
                    }
                )
            if len(latest_team) > 1 and {"DRI", "SJ_cm"}.issubset(latest_team.columns):
                figures.append(
                    {
                        "slug": "quadrant_dri_sj_experimental",
                        "title": "Mapa experimental de DRI y fuerza concéntrica",
                        "figure": chart_quadrant_dri_sj(latest_team, theme=theme),
                    }
                )

        if audience == "cliente":
            preferred = {"completion_team", "quadrant_cmj_imtp"}
        elif audience == "atleta":
            preferred = {"completion_team", "quadrant_cmj_imtp"}
        else:
            preferred = {"completion_team", "quadrant_cmj_imtp", "quadrant_rsi_sj"}
        return [item for item in figures if item["slug"] in preferred][:2]

    if jdf is not None and not jdf.empty and "Athlete" in jdf.columns and effective_athlete in jdf["Athlete"].values:
        athlete_jdf = jdf[jdf["Athlete"] == effective_athlete].sort_values("Date")
        radar_row = build_profile_radar_row(athlete_jdf)
        if audience in {"atleta", "profe"}:
            figures.append(
                {
                    "slug": "radar_perfil",
                    "title": "Perfil neuromuscular",
                    "figure": chart_radar(radar_row, effective_athlete, _team_mean_for_radar(jdf), theme=theme),
                }
            )
        if len(_cmj_series(state, effective_athlete)) >= 2:
            figures.append(
                {
                    "slug": "cmj_trend",
                    "title": "Tendencia de CMJ",
                    "figure": chart_cmj_trend(jdf, effective_athlete, theme=theme),
                }
            )

    acwr_dict = state.get("acwr_dict") or {}
    acwr_df = acwr_dict.get(effective_athlete)
    if acwr_df is not None and not acwr_df.empty and len(_acwr_series(state, effective_athlete)) >= 2:
        figures.append(
            {
                "slug": "acwr",
                "title": "Carga reciente",
                "figure": chart_acwr(acwr_df, effective_athlete, theme=theme),
                }
            )

    wdf = state.get("wellness_df")
    if wdf is not None and not wdf.empty and "Athlete" in wdf.columns:
        athlete_wdf = wdf[wdf["Athlete"] == effective_athlete].sort_values("Date").copy()
        athlete_wdf["Wellness_Score"] = _report_wellness_score_series(athlete_wdf)
        if not athlete_wdf.empty and len(_wellness_series(state, effective_athlete)) >= 2:
            figures.append(
                {
                    "slug": "wellness",
                    "title": "Bienestar reciente",
                    "figure": chart_wellness(athlete_wdf, effective_athlete, theme=theme),
                }
            )

    completion_plot_df = _completion_plot_df(state, effective_athlete)
    if not completion_plot_df.empty and len(completion_plot_df) >= 2:
        figures.append(
            {
                "slug": "completion",
                "title": "Adherencia reciente",
                "figure": chart_completion(completion_plot_df, theme=theme, athlete_label=effective_athlete),
            }
        )

    if audience == "cliente":
        preferred = {"cmj_trend", "wellness", "acwr", "completion"}
        figures = [item for item in figures if item["slug"] in preferred][:2]
    elif audience == "atleta":
        preferred = {"radar_perfil", "cmj_trend", "wellness", "acwr", "completion"}
        figures = [item for item in figures if item["slug"] in preferred][:2]
    else:
        preferred = {"radar_perfil", "cmj_trend", "acwr", "wellness", "completion"}
        figures = [item for item in figures if item["slug"] in preferred][:2]

    return figures


def _collect_athlete_pdf_chart_payloads(state: dict[str, pd.DataFrame | None], athlete: str) -> dict[str, dict[str, object]]:
    try:
        from charts.load_charts import chart_acwr, chart_completion, chart_wellness
        from charts.dashboard_charts import chart_radar
    except Exception:
        return {}

    theme = _build_report_chart_theme()
    payloads: dict[str, dict[str, object]] = {}

    jump_team = _professional_jump_history(state, "Todos")
    jump_history = _professional_jump_history(state, athlete)
    if not jump_history.empty:
        radar_row = build_profile_radar_row(jump_history)
        try:
            payloads["radar_perfil"] = {
                "slug": "radar_perfil",
                "title": "Perfil neuromuscular",
                "figure": chart_radar(radar_row, athlete, _team_mean_for_radar(jump_team), theme=theme),
            }
        except Exception:
            pass

    acwr_df = (state.get("acwr_dict") or {}).get(athlete)
    if acwr_df is not None and not acwr_df.empty and {"Date", "sRPE_diario", "ACWR_EWMA"}.issubset(acwr_df.columns):
        try:
            payloads["acwr"] = {
                "slug": "acwr",
                "title": "Carga reciente: cómo venís tolerando el entrenamiento",
                "figure": chart_acwr(acwr_df, athlete, theme=theme),
            }
        except Exception:
            pass

    wellness_df = state.get("wellness_df")
    if wellness_df is not None and not wellness_df.empty and "Athlete" in wellness_df.columns:
        athlete_wellness = wellness_df[_professional_athlete_mask(wellness_df["Athlete"], athlete)].copy()
        if not athlete_wellness.empty and "Date" in athlete_wellness.columns:
            athlete_wellness = athlete_wellness.sort_values("Date")
            athlete_wellness["Wellness_Score"] = _report_wellness_score_series(athlete_wellness)
            try:
                payloads["wellness"] = {
                    "slug": "wellness",
                    "title": "Bienestar reciente",
                    "figure": chart_wellness(athlete_wellness, athlete, theme=theme),
                }
            except Exception:
                pass

    completion_plot_df = _completion_plot_df(state, athlete)
    if not completion_plot_df.empty and len(completion_plot_df) >= 2:
        try:
            payloads["completion"] = {
                "slug": "completion",
                "title": "Adherencia reciente",
                "figure": chart_completion(completion_plot_df, theme=theme, athlete_label=athlete),
            }
        except Exception:
            pass

    return payloads


def _collect_client_pdf_chart_payloads(state: dict[str, pd.DataFrame | None], athlete: str) -> dict[str, dict[str, object]]:
    try:
        from charts.load_charts import chart_acwr, chart_completion, chart_wellness
        from charts.dashboard_charts import chart_cmj_trend
    except Exception:
        return {}

    theme = _build_report_chart_theme()
    payloads: dict[str, dict[str, object]] = {}
    jump_history = _professional_jump_history(state, athlete)
    if not jump_history.empty and len(_cmj_series(state, athlete)) >= 2:
        try:
            payloads["cmj_trend"] = {
                "slug": "cmj_trend",
                "title": "Progreso reciente",
                "figure": chart_cmj_trend(state.get("jump_df"), athlete, theme=theme),
            }
        except Exception:
            pass

    acwr_df = (state.get("acwr_dict") or {}).get(athlete)
    if acwr_df is not None and not acwr_df.empty and len(_acwr_series(state, athlete)) >= 2:
        try:
            payloads["acwr"] = {
                "slug": "acwr",
                "title": "Cómo venís tolerando el entrenamiento",
                "figure": chart_acwr(acwr_df, athlete, theme=theme),
            }
        except Exception:
            pass

    wellness_df = state.get("wellness_df")
    if wellness_df is not None and not wellness_df.empty and "Athlete" in wellness_df.columns:
        athlete_wellness = wellness_df[_professional_athlete_mask(wellness_df["Athlete"], athlete)].copy()
        if not athlete_wellness.empty and "Date" in athlete_wellness.columns and len(_wellness_series(state, athlete)) >= 2:
            athlete_wellness = athlete_wellness.sort_values("Date")
            athlete_wellness["Wellness_Score"] = _report_wellness_score_series(athlete_wellness)
            try:
                payloads["wellness"] = {
                    "slug": "wellness",
                    "title": "Bienestar reciente",
                    "figure": chart_wellness(athlete_wellness, athlete, theme=theme),
                }
            except Exception:
                pass

    completion_plot_df = _completion_plot_df(state, athlete)
    if not completion_plot_df.empty and len(completion_plot_df) >= 2:
        try:
            payloads["completion"] = {
                "slug": "completion",
                "title": "Constancia reciente",
                "figure": chart_completion(completion_plot_df, theme=theme, athlete_label=athlete),
            }
        except Exception:
            pass

    return payloads


def export_plotly_figure_png(
    figure: object,
    *,
    width: int = 1200,
    height: int = 700,
    scale: int = 2,
) -> bytes | None:
    try:
        import plotly.io as pio
    except Exception:
        return None
    image_bytes = _plotly_figure_to_png_bytes(
        pio,
        figure,
        width=width,
        height=height,
        scale=scale,
    )
    if image_bytes is not None:
        return image_bytes
    chrome_path = _get_kaleido_chrome_sync()
    if chrome_path:
        return _plotly_figure_to_png_bytes(
            pio,
            figure,
            width=width,
            height=height,
            scale=scale,
        )
    return None


def _professional_jump_history(state: dict[str, pd.DataFrame | None], athlete: str) -> pd.DataFrame:
    jdf = state.get("jump_df")
    if jdf is None or jdf.empty or "Athlete" not in jdf.columns:
        return pd.DataFrame()
    try:
        data = _prepare_jump_df(jdf.copy())
    except Exception:
        data = jdf.copy()
    if data.empty or "Date" not in data.columns or "Athlete" not in data.columns:
        return pd.DataFrame()
    data = data.copy()
    data["Date"] = pd.to_datetime(data["Date"], errors="coerce").dt.normalize()
    data["Athlete"] = data["Athlete"].astype(str).str.strip()
    data = data.dropna(subset=["Date"])
    if athlete != "Todos":
        data = data[_professional_athlete_mask(data["Athlete"], athlete)]
    if data.empty:
        return pd.DataFrame()
    return data.sort_values(["Athlete", "Date"]).reset_index(drop=True)


def _professional_latest_team_jump_rows(state: dict[str, pd.DataFrame | None]) -> pd.DataFrame:
    data = _professional_jump_history(state, "Todos")
    if data.empty or "Athlete" not in data.columns:
        return pd.DataFrame()
    return data.sort_values("Date").groupby("Athlete", as_index=False).tail(1).reset_index(drop=True)


def _build_current_pdf_neuromuscular_profile_payload(
    state: dict[str, pd.DataFrame | None],
    athlete: str,
    *,
    audience: str,
) -> dict[str, object]:
    history = _professional_jump_history(state, athlete)
    if history.empty:
        return _build_pdf_neuromuscular_profile_payload(
            pd.Series(dtype=object),
            context={"audience": audience, "scope": "latest_evaluation"},
        )
    return _build_pdf_neuromuscular_profile_payload(
        history.iloc[-1],
        reference_df=history,
        context={"audience": audience, "scope": "latest_evaluation", "profile_source": "latest_valid_row"},
    )


def _professional_completion_snapshot(
    state: dict[str, pd.DataFrame | None],
    athlete: str,
) -> dict[str, object]:
    cdf = state.get("completion_df")
    if cdf is None or cdf.empty or "Date" not in cdf.columns:
        return {"value": PDF_MISSING_TEXT, "detail": PDF_MISSING_TEXT, "numeric": None}
    result = cdf.copy()
    result["Date"] = pd.to_datetime(result["Date"], errors="coerce")
    result = result.dropna(subset=["Date"])
    if "Athlete" in result.columns:
        result = result[_professional_athlete_mask(result["Athlete"], athlete)]
    if result.empty:
        return {"value": PDF_MISSING_TEXT, "detail": PDF_MISSING_TEXT, "numeric": None}

    today_ts = pd.Timestamp(datetime.now()).normalize()
    current_week = today_ts - pd.Timedelta(days=today_ts.weekday())
    current = result[
        (result["Date"].dt.normalize() >= current_week)
        & (result["Date"].dt.normalize() <= today_ts)
    ].copy()
    period_label = "Semana actual"
    if current.empty:
        latest_date = result["Date"].max().normalize()
        current = result[result["Date"].dt.normalize().eq(latest_date)].copy()
        period_label = f"Ultima fecha util ({latest_date:%d/%m})"

    completion_result = calculate_completion_rate(current)
    if completion_result.value is None:
        return {"value": PDF_MISSING_TEXT, "detail": period_label, "numeric": None}
    completion_mean = round(float(completion_result.value), 1)
    return {
        "value": f"{completion_mean:.1f}%",
        "detail": period_label,
        "numeric": completion_mean,
    }


def _professional_metric_value_col(frame: pd.DataFrame, spec: dict[str, object]) -> str:
    value_cols = tuple(spec.get("value_cols", ()))
    for column in value_cols:
        if column in frame.columns and pd.to_numeric(frame[column], errors="coerce").notna().any():
            return str(column)
    for column in value_cols:
        if column in frame.columns:
            return str(column)
    return str(value_cols[0]) if value_cols else ""


def _professional_metric_unit(spec: dict[str, object], value_col: str) -> str:
    unit = spec.get("unit", "")
    if isinstance(unit, dict):
        return str(unit.get(value_col, ""))
    return str(unit or "")


def _professional_metric_display_unit(spec: dict[str, object], value_col: str) -> str:
    unit = _professional_metric_unit(spec, value_col)
    title = str(spec.get("title", "")).strip()
    if unit == "rsi_index":
        return "Índice RSI"
    if unit == "dri_index":
        return "Índice DRI"
    if unit == "mrsi_index":
        return "Índice mRSI"
    if unit == "ratio":
        return "Ratio"
    if title == "IMTP" and value_col == "IMTP_relPF":
        return "N/kg (fuerza relativa)"
    return unit or PDF_MISSING_TEXT


def _professional_metric_delta_unit(spec: dict[str, object], value_col: str) -> str:
    unit = _professional_metric_unit(spec, value_col)
    if unit == "rsi_index":
        return "unidades RSI"
    if unit == "dri_index":
        return "unidades DRI"
    if unit == "mrsi_index":
        return "unidades mRSI"
    return unit


def _professional_metric_digits(spec: dict[str, object], value_col: str) -> int:
    digits = spec.get("digits", 1)
    if isinstance(digits, dict):
        return int(digits.get(value_col, 1))
    return int(digits)


def _format_professional_metric_value(value: object, spec: dict[str, object], value_col: str) -> str:
    numeric = _coerce_float(value)
    if numeric is None:
        return PDF_MISSING_TEXT
    digits = _professional_metric_digits(spec, value_col)
    unit = _professional_metric_unit(spec, value_col)
    rendered = f"{numeric:.{digits}f}"
    if unit and unit not in {"ratio", "rsi_index", "mrsi_index"}:
        return f"{rendered} {unit}"
    return rendered


def _professional_z_value(row: pd.Series | None, spec: dict[str, object], value_col: str = "") -> float | None:
    if row is None:
        return None
    if value_col == "IMTP_relPF":
        return _professional_snapshot_zscore(row, "IMTP_relPF_Z")
    if value_col == "IMTP_N":
        return _professional_snapshot_zscore(row, "IMTP_N_Z")
    for column in tuple(spec.get("z_cols", ())):
        if str(column).endswith("_Z") or str(column) in {"DJtc_Z", "DJ_Z", "IMTP_Z"}:
            numeric = _professional_snapshot_zscore(row, str(column))
        else:
            numeric = _coerce_float(row.get(column)) if column in row.index else None
        if numeric is not None:
            return numeric
    return None


def _professional_delta_text(
    current_value: float | None,
    previous_value: float | None,
    spec: dict[str, object],
    value_col: str,
) -> str:
    if current_value is None or previous_value is None:
        return PDF_MISSING_TEXT
    delta = current_value - previous_value
    digits = _professional_metric_digits(spec, value_col)
    unit = _professional_metric_delta_unit(spec, value_col)
    rendered_delta = f"{delta:+.{digits}f}"
    if unit and unit != "ratio":
        rendered_delta = f"{rendered_delta} {unit}"
    pct = None
    if previous_value != 0:
        pct = (delta / previous_value) * 100
    return f"{rendered_delta} ({pct:+.1f}%)" if pct is not None else rendered_delta


def _professional_best_text(series: pd.Series, spec: dict[str, object], value_col: str) -> str:
    valid = pd.to_numeric(series, errors="coerce").dropna()
    if valid.empty:
        return PDF_MISSING_TEXT
    best = valid.max() if bool(spec.get("higher_is_better", True)) else valid.min()
    return _format_professional_metric_value(best, spec, value_col)


def _professional_metric_direction(spec: dict[str, object]) -> str:
    direction = str(spec.get("direction") or "").strip()
    if direction:
        return direction
    return "higher_is_better" if bool(spec.get("higher_is_better", True)) else "lower_is_better"


def _professional_threshold_text(
    threshold_abs: float | None,
    threshold_method: object,
    spec: dict[str, object],
    value_col: str,
) -> str:
    if threshold_abs is None:
        return PDF_MISSING_TEXT
    method = safe_value(threshold_method, fallback="")
    rendered = f"+/- {_format_professional_metric_value(threshold_abs, spec, value_col)}"
    delta_unit = _professional_metric_delta_unit(spec, value_col)
    if delta_unit and delta_unit not in {"ratio", _professional_metric_unit(spec, value_col)}:
        rendered = f"{rendered} {delta_unit}"
    return f"{rendered} ({method})" if method else rendered


def _professional_delta_payload(
    history: pd.DataFrame,
    latest_row: pd.Series | None,
    value_col: str,
    spec: dict[str, object] | None = None,
) -> dict[str, object]:
    if latest_row is None or history.empty or value_col not in history.columns:
        return {}

    metric_rows = history[["Date", value_col]].copy()
    metric_rows["Date"] = pd.to_datetime(metric_rows["Date"], errors="coerce").dt.normalize()
    metric_rows[value_col] = pd.to_numeric(metric_rows[value_col], errors="coerce")
    metric_rows = metric_rows.dropna(subset=["Date", value_col]).sort_values("Date")
    if metric_rows.empty:
        return {}

    latest_date = pd.to_datetime(latest_row.get("Date"), errors="coerce")
    if pd.isna(latest_date):
        latest_date = metric_rows.iloc[-1]["Date"]
    latest_date = pd.Timestamp(latest_date).normalize()

    latest_matches = metric_rows[metric_rows["Date"].eq(latest_date)]
    current_value = _coerce_float(latest_matches.iloc[-1].get(value_col)) if not latest_matches.empty else _coerce_float(metric_rows.iloc[-1].get(value_col))
    previous_rows = metric_rows[metric_rows["Date"] < latest_date]
    previous_value = _coerce_float(previous_rows.iloc[-1].get(value_col)) if not previous_rows.empty else None
    delta_abs = (current_value - previous_value) if current_value is not None and previous_value is not None else None
    delta_pct = ((delta_abs / previous_value) * 100) if delta_abs is not None and previous_value not in [None, 0] else None
    te_reference = _professional_metric_te_reference(spec or {}, value_col) if spec is not None else None

    return {
        "threshold_abs": _coerce_float(te_reference.get("value")) if te_reference is not None else None,
        "threshold_method": "TE de referencia" if te_reference is not None else "",
        "temporal_signal": PDF_MISSING_TEXT,
        "delta_abs": delta_abs,
        "delta_pct": delta_pct,
    }


def _professional_large_change_warning(spec: dict[str, object], delta_pct: float | None) -> str:
    if delta_pct is None:
        return ""
    thresholds = {
        "CMJ": 10.0,
        "SJ": 10.0,
        "DJ": 10.0,
        "IMTP": 8.0,
    }
    threshold = thresholds.get(str(spec.get("title", "")))
    if threshold is None or delta_pct <= threshold:
        return ""
    return PROFESSIONAL_LARGE_CHANGE_WARNING


def _professional_metric_signal(
    current_value: float | None,
    previous_value: float | None,
    z_value: float | None,
    spec: dict[str, object],
    threshold_abs: float | None = None,
) -> tuple[str, str, str]:
    if current_value is None or previous_value is None:
        return "Sin dato", "#708C9F", "missing"
    delta = _coerce_float(current_value - previous_value)
    if delta is None:
        return "Sin dato", "#708C9F", "missing"
    threshold = _coerce_float(threshold_abs)
    if threshold is None or threshold <= 0:
        return PDF_MISSING_TEXT, "#708C9F", "fallback_no_te"
    if abs(delta) <= threshold:
        return "Amarillo", "#C4A464", "typical_error"
    direction = _professional_metric_direction(spec)
    favorable = delta > 0
    if direction == "lower_is_better":
        favorable = delta < 0
    if favorable:
        return "Verde", "#2F6B52", "typical_error"
    return "Rojo", "#B56B73", "typical_error"


def _professional_metric_interpretation(
    label: str,
    current_value: float | None,
    spec: dict[str, object] | None = None,
    basis: str = "",
) -> str:
    if current_value is None or label in {PDF_MISSING_TEXT, "Sin dato"}:
        return "Faltan datos para generar una interpretación confiable."
    if basis == "fallback_no_te":
        return "Sin TE de referencia; usar el dato como señal descriptiva y cruzarlo con el contexto."
    title = str((spec or {}).get("title", ""))
    base_message = ""
    if label == "Amarillo":
        base_message = "Cambio dentro del TE/error típico; no tomar decisiones fuertes con este dato aislado."
    elif label == "Verde":
        base_message = "Cambio favorable mayor al TE; interpretar junto con el contexto."
    elif label == "Rojo":
        base_message = "Cambio desfavorable mayor al TE; revisar contexto antes de modificar carga o prioridades."
    contextual_notes = {
        "EUR": (
            "EUR debe interpretarse junto con CMJ y SJ; un aumento puede reflejar mejor uso del "
            "contramovimiento o menor rendimiento relativo del SJ."
        ),
        "RSI": "RSI debe leerse junto con DJ/DRI y Contact Time; no usarlo como conclusión aislada.",
        "Contact Time": "Contact Time debe leerse junto con DJ/DRI y RSI; un cambio aislado no define por sí solo el perfil reactivo.",
        "mRSI": "mRSI aporta contexto reactivo y conviene leerlo junto con CMJ, DJ y el protocolo aplicado.",
    }
    note = contextual_notes.get(title, "")
    if base_message and note:
        return f"{base_message} {note}"
    if base_message:
        return base_message
    if note:
        return note
    return "Faltan datos para interpretar el cambio frente a la evaluación previa."


def _professional_quadrant_location(selected: dict[str, object] | None, spec: dict[str, object]) -> str:
    if selected is None:
        return PDF_MISSING_TEXT
    x = _coerce_float(selected.get("x"))
    y = _coerce_float(selected.get("y"))
    if x is None or y is None:
        return PDF_MISSING_TEXT
    x_label = str(spec.get("x_label", "eje X"))
    y_label = str(spec.get("y_label", "eje Y"))
    x_zone = "alto" if x >= 0.35 else "bajo" if x <= -0.35 else "intermedio"
    y_zone = "alto" if y >= 0.35 else "bajo" if y <= -0.35 else "intermedio"
    return f"{x_label} {x_zone} (z={x:+.2f}) y {y_label} {y_zone} (z={y:+.2f})."


def _professional_quadrant_athlete_meaning(selected: dict[str, object] | None) -> str:
    if selected is None:
        return "Faltan datos para interpretar la ubicación individual."
    x = _coerce_float(selected.get("x"))
    y = _coerce_float(selected.get("y"))
    if x is None or y is None:
        return "Faltan datos para interpretar la ubicación individual."
    if x >= 0 and y >= 0:
        return "El atleta se ubica en un perfil relativamente favorable para ambas dimensiones del cuadrante."
    if x >= 0 > y:
        return "El atleta muestra mejor perfil en el eje horizontal que en el vertical; conviene atacar la dimensión rezagada."
    if y >= 0 > x:
        return "El atleta muestra mejor perfil en el eje vertical que en el horizontal; conviene mejorar transferencia hacia la dimensión rezagada."
    return "El atleta queda por debajo de la referencia en ambas dimensiones; priorizar bases antes de complejizar el bloque."


def _professional_cmj_imtp_athlete_meaning(selected: dict[str, object] | None) -> str:
    if selected is None:
        return "Faltan datos para interpretar la ubicación individual."
    cmj = _coerce_float(selected.get("x"))
    imtp = _coerce_float(selected.get("y"))
    if cmj is None or imtp is None:
        return "Faltan datos para interpretar la ubicación individual."
    cmj_zone = "alto" if cmj >= 0.35 else "bajo" if cmj <= -0.35 else "intermedio"
    imtp_zone = "alto" if imtp >= 0.35 else "bajo" if imtp <= -0.35 else "intermedio"
    if cmj_zone == "alto" and imtp_zone in {"intermedio", "bajo"}:
        return "El perfil sugiere buena salida vertical relativa, con fuerza isométrica relativa menos destacada. Para el próximo bloque, conviene sostener potencia/salto y reforzar fuerza base sin perder expresión rápida."
    if cmj_zone in {"intermedio", "bajo"} and imtp_zone == "alto":
        return "La fuerza isométrica relativa aparece mejor que la expresión vertical. Conviene trabajar transferencia hacia salto, coordinación de impulso y velocidad de aplicación."
    if cmj_zone == "alto" and imtp_zone == "alto":
        return "El atleta combina buena salida vertical y fuerza isométrica relativa. El foco puede pasar por sostener capacidad y afinar transferencia específica del deporte."
    if cmj_zone == "bajo" and imtp_zone == "bajo":
        return "El perfil muestra limitación simultánea en fuerza base y salida vertical. Priorizar fuerza general, técnica de salto y progresión de potencia."
    return "El perfil no muestra un extremo dominante; conviene integrar fuerza base y expresión vertical según el objetivo del bloque."


def _professional_dri_sj_athlete_meaning(selected: dict[str, object] | None) -> str:
    if selected is None:
        return "Faltan datos para interpretar la ubicación individual."
    dri = _coerce_float(selected.get("x"))
    sj = _coerce_float(selected.get("y"))
    if dri is None or sj is None:
        return "Faltan datos para interpretar la ubicación individual."
    dri_zone = "alto" if dri >= 0.35 else "bajo" if dri <= -0.35 else "intermedio"
    sj_zone = "alto" if sj >= 0.35 else "bajo" if sj <= -0.35 else "intermedio"
    if sj_zone == "alto" and dri_zone == "alto":
        return "SJ alto + DRI alto: buen perfil concéntrico y buen comportamiento reactivo en DJ."
    if sj_zone == "alto" and dri_zone in {"intermedio", "bajo"}:
        return f"SJ alto + DRI {dri_zone}: buena capacidad concéntrica, pero DRI rezagado; priorizar estrategia reactiva, stiffness, contacto y tolerancia progresiva a DJ."
    if sj_zone in {"intermedio", "bajo"} and dri_zone == "alto":
        return "SJ bajo + DRI alto: buena respuesta reactiva relativa, pero falta base concéntrica."
    if sj_zone == "bajo" and dri_zone == "bajo":
        return "SJ bajo + DRI bajo: prioridad general de fuerza base y reactividad progresiva."
    return f"SJ {sj_zone} + DRI {dri_zone}: no aparece un limitante extremo; conviene integrar fuerza concéntrica y reactividad según el bloque."


def _professional_rsi_sj_athlete_meaning(selected: dict[str, object] | None) -> str:
    if selected is None:
        return "Faltan datos para interpretar la ubicación individual."
    rsi = _coerce_float(selected.get("x"))
    sj = _coerce_float(selected.get("y"))
    if rsi is None or sj is None:
        return "Faltan datos para interpretar la ubicación individual."
    rsi_zone = "alto" if rsi >= 0.35 else "bajo" if rsi <= -0.35 else "intermedio"
    sj_zone = "alto" if sj >= 0.35 else "bajo" if sj <= -0.35 else "intermedio"
    if sj_zone == "alto" and rsi_zone == "alto":
        return "SJ alto + DJ RSI alto: buen perfil concéntrico y buena eficiencia reactiva en DJ."
    if sj_zone == "alto" and rsi_zone in {"intermedio", "bajo"}:
        return f"SJ alto + DJ RSI {rsi_zone}: buena capacidad concéntrica, pero la eficiencia reactiva del DJ queda rezagada; priorizar stiffness, contacto y progresión reactiva."
    if sj_zone in {"intermedio", "bajo"} and rsi_zone == "alto":
        return "SJ bajo + DJ RSI alto: buena eficiencia reactiva relativa, pero falta base concéntrica."
    if sj_zone == "bajo" and rsi_zone == "bajo":
        return "SJ bajo + DJ RSI bajo: prioridad general de fuerza base y eficiencia reactiva progresiva."
    return f"SJ {sj_zone} + DJ RSI {rsi_zone}: no aparece un limitante extremo; conviene integrar fuerza concéntrica y eficiencia reactiva según el bloque."


def _professional_eur_cmj_athlete_meaning(selected: dict[str, object] | None) -> str:
    if selected is None:
        return "Faltan datos para interpretar la ubicación individual."
    eur = _coerce_float(selected.get("x"))
    cmj = _coerce_float(selected.get("y"))
    if eur is None or cmj is None:
        return "Faltan datos para interpretar la ubicación individual."
    eur_zone = "alto" if eur >= 0.35 else "bajo" if eur <= -0.35 else "intermedio"
    cmj_zone = "alto" if cmj >= 0.35 else "bajo" if cmj <= -0.35 else "intermedio"
    if eur_zone == "alto" and cmj_zone == "alto":
        return "El atleta presenta buen uso del contramovimiento y buen rendimiento vertical. Interpretar EUR junto con SJ para evitar sobrevalorar eficiencia elástica si el SJ cambia mucho."
    if eur_zone == "alto" and cmj_zone in {"intermedio", "bajo"}:
        return "EUR alto con CMJ menos destacado puede reflejar dependencia del contramovimiento o SJ bajo. Revisar fuerza concéntrica y técnica antes de concluir ventaja elástica real."
    if eur_zone in {"intermedio", "bajo"} and cmj_zone == "alto":
        return "CMJ alto con EUR menos destacado sugiere buena salida vertical, pero menor diferencia CMJ-SJ. Mantener potencia y revisar si la estrategia de contramovimiento puede optimizarse."
    if eur_zone == "bajo" and cmj_zone == "bajo":
        return "EUR y CMJ bajos sugieren prioridad en fuerza propulsiva, técnica de salto y progresión de capacidades elásticas."
    return "La relación EUR-CMJ es intermedia; conviene leerla junto con SJ, DRI y contexto del bloque antes de cambiar prioridades."


def _professional_quadrant_specific_meaning(selected: dict[str, object] | None, spec: dict[str, object]) -> str:
    meaning_type = spec.get("meaning_type")
    if meaning_type == "cmj_imtp":
        return _professional_cmj_imtp_athlete_meaning(selected)
    if meaning_type == "rsi_sj":
        return _professional_rsi_sj_athlete_meaning(selected)
    if meaning_type == "dri_sj":
        return _professional_dri_sj_athlete_meaning(selected)
    if meaning_type == "eur_cmj":
        return _professional_eur_cmj_athlete_meaning(selected)
    return _professional_quadrant_athlete_meaning(selected)


def _professional_wellness_high(value: float | None, scale: object) -> bool:
    if value is None:
        return False
    scale_text = str(scale or "")
    if scale_text == "/5":
        return value >= 4
    if scale_text == "/10":
        return value >= 7
    return False


def _professional_wellness_partial_message(payload: dict[str, object]) -> str:
    summary = payload.get("last_week_summary", {}) if isinstance(payload.get("last_week_summary"), dict) else {}
    scales = payload.get("scales", {}) if isinstance(payload.get("scales"), dict) else {}
    days = int(_coerce_float(summary.get("days")) or 0)
    if days <= 0:
        return "Wellness parcial: faltan registros suficientes para tendencia. La lectura debe tomarse como alerta contextual puntual."
    if days == 1:
        sleep_mean = _coerce_float(summary.get("sleep_mean"))
        stress_mean = _coerce_float(summary.get("stress_mean"))
        pain_mean = _coerce_float(summary.get("pain_mean"))
        signals: list[str] = []
        if sleep_mean is not None:
            signals.append("sueño bajo" if sleep_mean < 6.5 else f"sueño {sleep_mean:.1f} h")
        if stress_mean is not None:
            signals.append("estrés elevado" if _professional_wellness_high(stress_mean, scales.get("stress")) else f"estrés {stress_mean:.1f}{scales.get('stress', '')}")
        if pain_mean is not None:
            signals.append("dolor elevado" if _professional_wellness_high(pain_mean, scales.get("pain")) else f"dolor {pain_mean:.1f}{scales.get('pain', '')}")
        detail = ", ".join(signals) if signals else "variables disponibles insuficientes"
        return f"Wellness parcial: solo 1 día con registro. {detail.capitalize()} en ese registro; usar como alerta contextual, no como tendencia."
    return f"Wellness parcial: solo {days} días con registro. Faltan registros suficientes para tendencia. La lectura debe tomarse como alerta contextual puntual."


def _build_professional_metric_cards(
    state: dict[str, pd.DataFrame | None],
    athlete: str,
) -> list[dict[str, object]]:
    history = _professional_jump_history(state, athlete)
    cards: list[dict[str, object]] = []
    for spec in PROFESSIONAL_PDF_METRICS:
        value_col = _professional_metric_value_col(history, spec)
        metric_rows = pd.DataFrame()
        if not history.empty and value_col in history.columns:
            metric_rows = history[["Date", value_col] + [col for col in tuple(spec.get("z_cols", ())) if col in history.columns]].copy()
            metric_rows[value_col] = pd.to_numeric(metric_rows[value_col], errors="coerce")
            metric_rows = metric_rows.dropna(subset=[value_col]).sort_values("Date")

        latest_row = metric_rows.iloc[-1] if not metric_rows.empty else None
        previous_row = metric_rows.iloc[-2] if len(metric_rows) >= 2 else None
        current_value = _coerce_float(latest_row.get(value_col)) if latest_row is not None else None
        previous_value = _coerce_float(previous_row.get(value_col)) if previous_row is not None else None
        z_value = _professional_z_value(latest_row, spec, value_col)
        delta_payload = _professional_delta_payload(history, latest_row, value_col, spec)
        threshold_abs = _coerce_float(delta_payload.get("threshold_abs"))
        threshold_method = delta_payload.get("threshold_method")
        delta_pct = _coerce_float(delta_payload.get("delta_pct"))
        signal, color, signal_basis = _professional_metric_signal(current_value, previous_value, z_value, spec, threshold_abs)
        missing_fields = sum(
            [
                current_value is None,
                previous_value is None,
                z_value is None,
                metric_rows.empty,
            ]
        )
        if threshold_abs is None:
            missing_fields += 1
        cards.append(
            {
                "title": spec["title"],
                "state": "missing" if current_value is None else "partial" if missing_fields else "available",
                "value": _format_professional_metric_value(current_value, spec, value_col),
                "unit_label": _professional_metric_display_unit(spec, value_col),
                "delta": _professional_delta_text(current_value, previous_value, spec, value_col),
                "z_score": f"{z_value:+.2f}" if z_value is not None else PDF_MISSING_TEXT,
                "best": _professional_best_text(metric_rows[value_col] if not metric_rows.empty else pd.Series(dtype=float), spec, value_col),
                "threshold": _professional_threshold_text(threshold_abs, threshold_method, spec, value_col),
                "te_caption": _professional_metric_te_caption(spec, value_col),
                "large_change_warning": _professional_large_change_warning(spec, delta_pct),
                "signal": signal,
                "signal_color": color,
                "signal_basis": signal_basis,
                "direction": _professional_metric_direction(spec),
                "interpretation": _professional_metric_interpretation(signal, current_value, spec, signal_basis),
            }
        )
    return cards


def _build_professional_evolution_sections(
    state: dict[str, pd.DataFrame | None],
    athlete: str,
) -> list[dict[str, object]]:
    history = _professional_jump_history(state, athlete)
    sections: list[dict[str, object]] = []
    for spec in PROFESSIONAL_PDF_METRICS:
        value_col = _professional_metric_value_col(history, spec)
        points: list[dict[str, object]] = []
        if not history.empty and value_col in history.columns:
            metric_rows = history[["Date", value_col]].copy()
            metric_rows[value_col] = pd.to_numeric(metric_rows[value_col], errors="coerce")
            metric_rows = metric_rows.dropna(subset=["Date", value_col]).sort_values("Date")
            points = [
                {
                    "label": pd.Timestamp(row["Date"]).strftime("%d/%m"),
                    "date": row["Date"],
                    "value": float(row[value_col]),
                }
                for _, row in metric_rows.iterrows()
            ]

        delta_text = PDF_MISSING_TEXT
        threshold_text = PDF_MISSING_TEXT
        signal = PDF_MISSING_TEXT
        large_change_warning = ""
        if len(points) >= 2:
            current_value = _coerce_float(points[-1].get("value"))
            previous_value = _coerce_float(points[-2].get("value"))
            delta_payload = _professional_delta_payload(history, pd.Series({"Date": points[-1]["date"]}), value_col, spec)
            delta_text = _professional_delta_text(current_value, previous_value, spec, value_col)
            large_change_warning = _professional_large_change_warning(spec, _coerce_float(delta_payload.get("delta_pct")))
            threshold = _coerce_float(delta_payload.get("threshold_abs"))
            method = safe_value(delta_payload.get("threshold_method"), fallback="")
            if threshold is not None:
                threshold_text = _professional_threshold_text(threshold, method, spec, value_col)
            signal, _, _ = _professional_metric_signal(current_value, previous_value, None, spec, threshold)

        if len(points) < 2:
            state_label = "missing"
            message = "Faltan datos para mostrar la evolución de esta métrica."
        elif threshold_text == PDF_MISSING_TEXT:
            state_label = "partial"
            message = "Zona de ruido/CV/TE: Faltan datos"
        else:
            state_label = "available"
            message = ""

        sections.append(
            {
                "title": str(spec["title"]),
                "state": state_label,
                "points": points,
                "unit": _professional_metric_unit(spec, value_col),
                "unit_label": _professional_metric_display_unit(spec, value_col),
                "delta": delta_text,
                "threshold": threshold_text,
                "signal": signal,
                "large_change_warning": large_change_warning,
                "message": message,
                "what": "Evolución entre evaluaciones de perfil físico separadas por ventanas de 6-8 semanas.",
                "meaning": f"Cambio vs evaluación previa: {delta_text}. TE de referencia: {threshold_text}.",
                "decision": "Usar la dirección del cambio para ajustar prioridades del siguiente bloque, no como readiness semanal.",
            }
        )
    return sections


def _build_professional_quadrant_sections(
    state: dict[str, pd.DataFrame | None],
    athlete: str,
) -> list[dict[str, object]]:
    data = _professional_latest_team_jump_rows(state)
    specs = [
        {
            "title": "Cuadrante fuerza relativa vs salida vertical",
            "x_cols": ("CMJ_Z", "Jump_Momentum_Z"),
            "x_label": "CMJ z",
            "y_col": "IMTP_relPF_Z",
            "y_label": "IMTP relPF z",
            "what": "Cruza salida vertical con fuerza isométrica relativa.",
            "meaning": "Arriba/derecha sugiere perfil fuerte y expresivo; abajo/izquierda marca déficit global.",
            "decision": "Priorizar fuerza base, potencia o transferencia según el cuadrante del atleta.",
            "meaning_type": "cmj_imtp",
        },
        {
            "title": "Cuadrante fuerza concéntrica vs DJ RSI",
            "x_cols": ("DJ_RSI_Z",),
            "x_label": "DJ RSI z",
            "y_col": "SJ_Z",
            "y_label": "SJ z",
            "what": "Cruza la capacidad concéntrica expresada en SJ con el DJ RSI del Drop Jump.",
            "meaning": "Distingue si el limitante principal parece estar en fuerza concéntrica, eficiencia reactiva del DJ o integración entre ambas.",
            "decision": "Orientar el próximo bloque según si el limitante principal parece estar en fuerza concéntrica, DJ RSI/reactividad o integración entre ambas.",
            "missing_message": "Faltan datos para construir el cuadrante SJ vs DJ RSI.",
            "meaning_type": "rsi_sj",
            "required_value_cols": ("DJ_RSI",),
        },
        {
            "title": "Cuadrante experimental - fuerza concéntrica vs DRI",
            "x_cols": ("DRI_Z",),
            "x_label": "DRI z",
            "y_col": "SJ_Z",
            "y_label": "SJ z",
            "what": "Cruza la capacidad concéntrica expresada en SJ con el DRI 2026 del Drop Jump.",
            "meaning": "Aporta una lectura experimental para distinguir si el limitante principal parece estar en fuerza concéntrica, DRI/reactividad o integración entre ambas.",
            "decision": "Usarlo como apoyo experimental; confirmar la lectura junto con DJ RSI, tiempo de contacto y contexto del bloque.",
            "missing_message": "Faltan datos para construir el cuadrante SJ vs DRI.",
            "meaning_type": "dri_sj",
            "required_value_cols": ("DRI",),
        },
        {
            "title": "Cuadrante eficiencia SSC lenta vs CMJ",
            "x_cols": ("EUR_Z",),
            "x_label": "EUR z",
            "y_col": "CMJ_Z",
            "y_label": "CMJ z",
            "what": "Cruza el uso del contramovimiento con el rendimiento vertical actual.",
            "meaning": "Ayuda a separar atletas que saltan alto por capacidad general de quienes aprovechan mejor el SSC.",
            "decision": "Ajustar técnica, fuerza concéntrica o estrategia elástica según la ubicación.",
            "meaning_type": "eur_cmj",
        },
    ]
    sections: list[dict[str, object]] = []
    for spec in specs:
        required_value_cols = tuple(spec.get("required_value_cols", ()))
        required_values_ready = all(
            column in data.columns and pd.to_numeric(data[column], errors="coerce").notna().any()
            for column in required_value_cols
        )
        x_col = (
            next((col for col in spec["x_cols"] if col in data.columns), "")
            if not required_value_cols or required_values_ready
            else ""
        )
        y_col = str(spec["y_col"])
        points: list[dict[str, object]] = []
        selected = None
        if x_col and y_col in data.columns and "Athlete" in data.columns:
            plot_df = data[["Athlete", x_col, y_col]].copy()
            plot_df[x_col] = pd.to_numeric(plot_df[x_col], errors="coerce")
            plot_df[y_col] = pd.to_numeric(plot_df[y_col], errors="coerce")
            plot_df = plot_df.dropna(subset=[x_col, y_col])
            target = str(athlete).strip().casefold()
            for _, row in plot_df.iterrows():
                is_selected = str(row["Athlete"]).strip().casefold() == target
                point = {
                    "x": float(row[x_col]),
                    "y": float(row[y_col]),
                    "selected": is_selected,
                }
                if is_selected:
                    selected = point
                points.append(point)
        state_label = "available" if selected is not None and len(points) > 1 else "partial" if selected is not None else "missing"
        sections.append(
            {
                **spec,
                "x_col": x_col,
                "points": points,
                "selected": selected,
                "state": state_label,
                "message": "" if selected is not None else str(spec.get("missing_message") or "Faltan datos para ubicar al atleta en este cuadrante."),
                "location": _professional_quadrant_location(selected, spec),
                "athlete_meaning": _professional_quadrant_specific_meaning(selected, spec),
            }
        )
    return sections


def _professional_report_today() -> pd.Timestamp:
    return pd.Timestamp(datetime.now()).normalize()


def _professional_week_start(value: object) -> pd.Timestamp:
    date = pd.Timestamp(value).normalize()
    return date - pd.Timedelta(days=int(date.weekday()))


def _professional_week_label(week_start: pd.Timestamp) -> str:
    start = pd.Timestamp(week_start).normalize()
    end = start + pd.Timedelta(days=6)
    return f"{start:%d/%m/%Y} - {end:%d/%m/%Y}"


def _professional_week_end(week_start: pd.Timestamp) -> pd.Timestamp:
    return pd.Timestamp(week_start).normalize() + pd.Timedelta(days=6)


def _professional_last_complete_week_start(reference_date: pd.Timestamp) -> pd.Timestamp:
    reference = pd.Timestamp(reference_date).normalize()
    week_start = _professional_week_start(reference)
    if _professional_week_end(week_start) <= reference:
        return week_start
    return week_start - pd.Timedelta(days=7)


def _professional_latest_data_date(state: dict[str, pd.DataFrame | None], athlete: str) -> pd.Timestamp | None:
    direct_dates: list[pd.Timestamp] = []
    weekly_dates: list[pd.Timestamp] = []

    def add_dates(frame: pd.DataFrame | None, date_candidates: tuple[str, ...], *, athlete_col: str = "Athlete") -> None:
        if frame is None or frame.empty:
            return
        result = frame.copy()
        if athlete_col in result.columns:
            result = result[_professional_athlete_mask(result[athlete_col], athlete)]
        for column in date_candidates:
            if column not in result.columns:
                continue
            parsed = pd.to_datetime(result[column], errors="coerce").dropna()
            direct_dates.extend(pd.Timestamp(value).normalize() for value in parsed.tolist())
            break

    add_dates(state.get("jump_df"), ("Date",))
    add_dates(state.get("rpe_df"), ("Date",))
    add_dates(state.get("wellness_df"), ("Date",))
    add_dates(state.get("completion_df"), ("Date",))
    add_dates(state.get("raw_df"), ("Assigned Date", "Date"))
    add_dates(state.get("prepared_raw_df"), ("Assigned Date", "Date"))

    weekly_summaries = state.get("weekly_summaries")
    if isinstance(weekly_summaries, dict):
        for key, value_columns in {
            "weekly_load": ("weekly_sRPE", "sessions_count"),
            "weekly_wellness": ("wellness_days", "Wellness_mean", "Sueno_mean", "Estres_mean", "Dolor_mean"),
        }.items():
            frame = _normalize_weekly_frame(weekly_summaries.get(key))
            if frame.empty or "week_start" not in frame.columns:
                continue
            if "Athlete" in frame.columns:
                frame = frame[_professional_athlete_mask(frame["Athlete"], athlete)].copy()
            valid = pd.Series(False, index=frame.index)
            for column in value_columns:
                if column in frame.columns:
                    valid = valid | pd.to_numeric(frame[column], errors="coerce").fillna(0).gt(0)
            frame = frame[valid]
            if frame.empty:
                continue
            starts = pd.to_datetime(frame["week_start"], errors="coerce").dropna()
            weekly_dates.extend(_professional_week_end(pd.Timestamp(value)) for value in starts.tolist())

    dates = direct_dates or weekly_dates
    if not dates:
        return None
    return max(dates)


def _professional_week_points_from_daily(
    frame: pd.DataFrame,
    week_start: pd.Timestamp,
    value_col: str,
    *,
    end_date: pd.Timestamp | None = None,
) -> tuple[list[dict[str, object]], int, int]:
    week_start = pd.Timestamp(week_start).normalize()
    week_end = week_start + pd.Timedelta(days=6)
    if end_date is not None:
        week_end = min(week_end, pd.Timestamp(end_date).normalize())
    week_df = frame[(frame["Date"] >= week_start) & (frame["Date"] <= week_end)].copy()
    if week_df.empty:
        return [], 0, 0
    daily_sum = week_df.groupby("Date", as_index=False)[value_col].sum()
    daily_lookup = {
        pd.Timestamp(row["Date"]).normalize(): float(row[value_col])
        for _, row in daily_sum.iterrows()
    }
    points = [
        {
            "label": (week_start + pd.Timedelta(days=day)).strftime("%d/%m"),
            "value": daily_lookup.get(week_start + pd.Timedelta(days=day), 0.0),
        }
        for day in range(7)
    ]
    return points, int(len(week_df)), int(week_df["Date"].nunique())


def _build_professional_training_context(
    state: dict[str, pd.DataFrame | None],
    athlete: str,
) -> dict[str, object]:
    rows: list[tuple[str, str]] = []
    rpe_df = state.get("rpe_df")
    completion = _professional_completion_snapshot(state, athlete)
    weekly_summaries = _report_weekly_summaries(state)
    weekly_load = _normalize_weekly_frame(weekly_summaries.get("weekly_load"))
    if not weekly_load.empty and "Athlete" in weekly_load.columns:
        weekly_load = weekly_load[_professional_athlete_mask(weekly_load["Athlete"], athlete)].copy()
    weekly_load = weekly_load.sort_values("week_start") if "week_start" in weekly_load.columns else weekly_load
    if not weekly_load.empty and "week_start" in weekly_load.columns:
        current_week_start = _professional_week_start(_professional_report_today())
        complete_week_start = current_week_start - pd.Timedelta(days=7)
        completed_load = weekly_load[pd.to_datetime(weekly_load["week_start"], errors="coerce").dt.normalize() == complete_week_start].copy()
        current_load = completed_load if not completed_load.empty else weekly_load[pd.to_datetime(weekly_load["week_start"], errors="coerce").dt.normalize() < current_week_start].tail(1).copy()
    else:
        current_load = weekly_load.tail(1).copy()
    load_row = current_load.iloc[-1] if not current_load.empty else pd.Series(dtype=object)
    sessions_count = _coerce_float(load_row.get("sessions_count"))
    weekly_srpe = _coerce_float(load_row.get("weekly_sRPE"))
    sessions_text = str(int(sessions_count)) if sessions_count is not None else PDF_MISSING_TEXT
    weekly_srpe_days = 0
    if not weekly_load.empty and "week_start" in weekly_load.columns:
        target_week = pd.to_datetime(load_row.get("week_start"), errors="coerce")
        if pd.notna(target_week) and rpe_df is not None and not rpe_df.empty and {"Athlete", "Date", "sRPE"}.issubset(rpe_df.columns):
            rpe_source = rpe_df.copy()
            rpe_source["Date"] = pd.to_datetime(rpe_source["Date"], errors="coerce").dt.normalize()
            rpe_source["sRPE"] = pd.to_numeric(rpe_source["sRPE"], errors="coerce")
            rpe_source = rpe_source.dropna(subset=["Date", "sRPE"])
            rpe_source = rpe_source[_professional_athlete_mask(rpe_source["Athlete"], athlete)]
            week_start = pd.Timestamp(target_week).normalize()
            week_end = week_start + pd.Timedelta(days=6)
            weekly_srpe_days = int(rpe_source[(rpe_source["Date"] >= week_start) & (rpe_source["Date"] <= week_end)]["Date"].nunique())
    if completion.get("numeric") is not None:
        adherence_text = safe_value(completion.get("value"))
    elif sessions_count is not None:
        adherence_text = PDF_MISSING_TEXT
    else:
        adherence_text = PDF_MISSING_TEXT
    rows.append(("Adherencia formal", adherence_text))
    rows.append(("Sesiones registradas", sessions_text))
    weekly_srpe_value = f"{weekly_srpe:.0f} UA" if weekly_srpe is not None else PDF_MISSING_TEXT
    weekly_srpe_suffix = _report_sample_suffix(weekly_srpe_days)
    if weekly_srpe_value != PDF_MISSING_TEXT and weekly_srpe_suffix:
        weekly_srpe_value = f"{weekly_srpe_value}  {weekly_srpe_suffix}"
    rows.append(("sRPE semanal", weekly_srpe_value))

    load_evolution = PDF_MISSING_TEXT
    if not weekly_load.empty and "weekly_sRPE" in weekly_load.columns:
        valid_load = weekly_load.copy()
        valid_load["weekly_sRPE"] = pd.to_numeric(valid_load["weekly_sRPE"], errors="coerce")
        valid_load = valid_load.dropna(subset=["weekly_sRPE"]).sort_values("week_start" if "week_start" in valid_load.columns else "weekly_sRPE")
        if "week_start" in valid_load.columns:
            current_week_start = _professional_week_start(_professional_report_today())
            valid_load = valid_load[pd.to_datetime(valid_load["week_start"], errors="coerce").dt.normalize() < current_week_start]
        if len(valid_load) >= 2:
            current = float(valid_load.iloc[-1]["weekly_sRPE"])
            previous = float(valid_load.iloc[-2]["weekly_sRPE"])
            delta = current - previous
            pct = (delta / previous * 100) if previous else None
            load_evolution = f"{delta:+.0f} UA vs semana previa"
            if pct is not None:
                load_evolution = f"{load_evolution} ({pct:+.1f}%)"
    rows.append(("Evolución de carga", load_evolution))

    exercises_text = PDF_MISSING_TEXT
    prepared = state.get("prepared_raw_df")
    if prepared is None:
        prepared = prepare_raw_workouts_df(state.get("raw_df"))
    if prepared is not None and not prepared.empty and "Athlete" in prepared.columns:
        athlete_df = prepared[_professional_athlete_mask(prepared["Athlete"], athlete)].copy()
        if not athlete_df.empty:
            invalid = athlete_df.get("is_invalid", pd.Series(False, index=athlete_df.index)).fillna(False)
            untagged = athlete_df.get("is_untagged", pd.Series(False, index=athlete_df.index)).fillna(False)
            athlete_df = athlete_df[~invalid & ~untagged]
            exercise_col = "Exercise" if "Exercise" in athlete_df.columns else "Exercise Name" if "Exercise Name" in athlete_df.columns else None
            if exercise_col and not athlete_df.empty:
                top_exercises = athlete_df[exercise_col].dropna().astype(str).str.strip()
                top_exercises = top_exercises[top_exercises != ""].value_counts().head(4)
                if not top_exercises.empty:
                    exercises_text = ", ".join(top_exercises.index.tolist())
    rows.append(("Ejercicios principales", exercises_text))

    available_count = sum(value != PDF_MISSING_TEXT for _, value in rows)
    has_training = any(
        [
            completion.get("numeric") is not None,
            sessions_count is not None,
            weekly_srpe is not None,
            exercises_text != PDF_MISSING_TEXT,
        ]
    )
    state_label = "available" if available_count == len(rows) else "partial" if available_count else "missing"
    return {
        "state": state_label if has_training else "missing",
        "rows": rows,
        "message": "" if has_training else "Faltan datos de entrenamiento para este período.",
        "what": "Resumen del bloque de entrenamiento visible para contextualizar el perfil físico.",
        "meaning": "Ayuda a separar cambios por adaptación real de cambios explicados por exposición, asistencia o carga.",
        "decision": "Ajustar la siguiente progresión según adherencia, sesiones cumplidas y ejercicios dominantes.",
    }


def _build_professional_internal_load_context(
    state: dict[str, pd.DataFrame | None],
    athlete: str,
) -> dict[str, object]:
    reference_date = _professional_latest_data_date(state, athlete) or _professional_report_today()
    reference_date = pd.Timestamp(reference_date).normalize()
    reference_week_start = _professional_week_start(reference_date)
    rpe_df = state.get("rpe_df")
    rpe_data = pd.DataFrame()
    daily_points: list[dict[str, object]] = []
    current_week_points: list[dict[str, object]] = []
    current_week_sessions = 0
    current_week_days = 0
    if rpe_df is not None and not rpe_df.empty and {"Athlete", "Date", "sRPE"}.issubset(rpe_df.columns):
        rpe_data = rpe_df.copy()
        rpe_data["Date"] = pd.to_datetime(rpe_data["Date"], errors="coerce").dt.normalize()
        rpe_data["sRPE"] = pd.to_numeric(rpe_data["sRPE"], errors="coerce")
        rpe_data = rpe_data.dropna(subset=["Athlete", "Date", "sRPE"])
        rpe_data = rpe_data[_professional_athlete_mask(rpe_data["Athlete"], athlete)]
        rpe_data = rpe_data[(rpe_data["sRPE"] > 0) & (rpe_data["Date"] <= reference_date)].sort_values("Date")
        if not rpe_data.empty:
            rpe_data["week_start"] = rpe_data["Date"] - pd.to_timedelta(rpe_data["Date"].dt.weekday, unit="D")

    weekly_summaries = _report_weekly_summaries(state)
    weekly_load = _normalize_weekly_frame(weekly_summaries.get("weekly_load"))
    weekly_points: list[dict[str, object]] = []
    weekly_valid = pd.DataFrame()
    if not weekly_load.empty and {"Athlete", "week_start", "weekly_sRPE"}.issubset(weekly_load.columns):
        weekly_load = weekly_load[_professional_athlete_mask(weekly_load["Athlete"], athlete)].copy()
        weekly_load["week_start"] = pd.to_datetime(weekly_load["week_start"], errors="coerce").dt.normalize()
        weekly_load["weekly_sRPE"] = pd.to_numeric(weekly_load["weekly_sRPE"], errors="coerce")
        weekly_load = weekly_load.dropna(subset=["week_start", "weekly_sRPE"])
        valid_mask = weekly_load["weekly_sRPE"].fillna(0).gt(0)
        if "sessions_count" in weekly_load.columns:
            valid_mask = valid_mask | pd.to_numeric(weekly_load["sessions_count"], errors="coerce").fillna(0).gt(0)
        weekly_valid = weekly_load[valid_mask & (weekly_load["week_start"] <= reference_week_start)].copy()

    daily_weeks = set(pd.to_datetime(rpe_data.get("week_start", pd.Series(dtype="datetime64[ns]")), errors="coerce").dropna().tolist()) if not rpe_data.empty else set()
    weekly_weeks = set(pd.to_datetime(weekly_valid.get("week_start", pd.Series(dtype="datetime64[ns]")), errors="coerce").dropna().tolist()) if not weekly_valid.empty else set()
    complete_candidates = {
        pd.Timestamp(week).normalize()
        for week in [*daily_weeks, *weekly_weeks]
        if _professional_week_end(pd.Timestamp(week)) <= reference_date
    }
    partial_candidates = {
        pd.Timestamp(week).normalize()
        for week in [*daily_weeks, *weekly_weeks]
        if pd.Timestamp(week).normalize() <= reference_date
    }
    if complete_candidates:
        analysis_week_start = max(complete_candidates)
        analysis_scope = "last_complete_week"
    elif partial_candidates:
        analysis_week_start = max(partial_candidates)
        analysis_scope = "current_week_partial"
    else:
        analysis_week_start = _professional_last_complete_week_start(reference_date)
        analysis_scope = "missing"

    sessions_registered = 0
    days_with_data = 0
    analysis_week_weekly_value = None
    analysis_week_sessions = None
    current_week_weekly_value = None
    if not weekly_valid.empty:
        analysis_week_rows = weekly_valid[weekly_valid["week_start"].eq(analysis_week_start)].dropna(subset=["weekly_sRPE"])
        if not analysis_week_rows.empty:
            analysis_week_weekly_value = _coerce_float(analysis_week_rows.iloc[-1].get("weekly_sRPE"))
            analysis_week_sessions = _coerce_float(analysis_week_rows.iloc[-1].get("sessions_count"))
        reference_week_rows = weekly_valid[weekly_valid["week_start"].eq(reference_week_start)].dropna(subset=["weekly_sRPE"])
        if not reference_week_rows.empty:
            current_week_weekly_value = _coerce_float(reference_week_rows.iloc[-1].get("weekly_sRPE"))
        weekly_chart_source = weekly_valid[weekly_valid["week_start"] <= analysis_week_start].sort_values("week_start").tail(16)
        if not weekly_chart_source.empty:
            weekly_chart_source["EMA6"] = weekly_chart_source["weekly_sRPE"].ewm(span=6, adjust=False).mean()
            eval_weeks = set()
            jump_history = _professional_jump_history(state, athlete)
            if not jump_history.empty and "Date" in jump_history.columns:
                eval_dates = pd.to_datetime(jump_history["Date"], errors="coerce").dropna()
                eval_weeks = {
                    (pd.Timestamp(value).normalize() - pd.Timedelta(days=int(pd.Timestamp(value).weekday())))
                    for value in eval_dates.tolist()
                }
            weekly_points = [
                {
                    "label": pd.Timestamp(row["week_start"]).strftime("%d/%m"),
                    "week_start": pd.Timestamp(row["week_start"]).normalize(),
                    "value": float(row["weekly_sRPE"]),
                    "ema": float(row["EMA6"]),
                    "sessions": _coerce_float(row.get("sessions_count")),
                    "evaluation": pd.Timestamp(row["week_start"]).normalize() in eval_weeks,
                }
                for _, row in weekly_chart_source.iterrows()
            ]

    if analysis_scope != "missing" and not rpe_data.empty:
        daily_points, sessions_registered, days_with_data = _professional_week_points_from_daily(
            rpe_data,
            analysis_week_start,
            "sRPE",
            end_date=reference_date if analysis_scope == "current_week_partial" else None,
        )
    if analysis_scope == "last_complete_week" and reference_week_start > analysis_week_start and not rpe_data.empty:
        current_week_points, current_week_sessions, current_week_days = _professional_week_points_from_daily(
            rpe_data,
            reference_week_start,
            "sRPE",
            end_date=reference_date,
        )
    elif analysis_scope == "current_week_partial":
        current_week_points = daily_points
        current_week_sessions = sessions_registered
        current_week_days = days_with_data

    has_daily = len(daily_points) > 0 and analysis_scope == "last_complete_week"
    has_weekly = len(weekly_points) >= 2
    analysis_total = sum(float(point.get("value") or 0) for point in daily_points) if daily_points else analysis_week_weekly_value
    total_last_week = analysis_total if analysis_scope == "last_complete_week" else None
    if sessions_registered == 0 and analysis_week_sessions is not None:
        sessions_registered = int(analysis_week_sessions)
    current_week_total = (
        sum(float(point.get("value") or 0) for point in current_week_points)
        if current_week_points
        else (analysis_total if analysis_scope == "current_week_partial" else current_week_weekly_value)
    )
    last_week_daily_mean = (total_last_week / days_with_data) if total_last_week is not None and days_with_data > 0 else None
    current_week_daily_mean = (current_week_total / current_week_days) if current_week_total is not None and current_week_days > 0 else None
    weekly_change = None
    weekly_change_pct = None
    if analysis_scope == "last_complete_week" and len(weekly_points) >= 2:
        current_week_value = _coerce_float(weekly_points[-1].get("value"))
        previous_week_value = _coerce_float(weekly_points[-2].get("value"))
        if current_week_value is not None and previous_week_value is not None:
            weekly_change = current_week_value - previous_week_value
            if previous_week_value:
                weekly_change_pct = (weekly_change / previous_week_value) * 100
    missing_parts = []
    if not has_daily:
        missing_parts.append("sRPE última semana")
    if not has_weekly:
        missing_parts.append("sRPE semanal + EMA 6 semanas")
    state_label = "available" if total_last_week is not None and has_weekly else "partial" if total_last_week is not None or has_weekly else "missing"
    if analysis_scope == "current_week_partial":
        state_label = "partial"
    return {
        "state": state_label,
        "analysis_scope": analysis_scope,
        "analysis_title": "Carga interna - última semana completa" if analysis_scope == "last_complete_week" else "Carga interna - semana en curso",
        "analysis_week_label": _professional_week_label(analysis_week_start),
        "daily_points": daily_points,
        "current_week_points": current_week_points,
        "weekly_points": weekly_points,
        "message": "" if has_daily or has_weekly or current_week_total is not None else "Faltan datos suficientes para analizar la última semana completa.",
        "missing_parts": missing_parts,
        "last_week_total": total_last_week,
        "last_week_daily_mean": last_week_daily_mean,
        "last_week_days_with_data": days_with_data if has_daily else 0,
        "current_week_total": current_week_total,
        "current_week_daily_mean": current_week_daily_mean,
        "current_week_sessions": current_week_sessions,
        "current_week_days": current_week_days,
        "current_week_partial_message": (
            f"La semana analizada está incompleta. El total acumulado hasta el momento es {current_week_total:.0f} UA "
            f"con {current_week_sessions} sesión(es) registrada(s) y no debe compararse directamente con una semana completa."
            if current_week_total is not None
            else ""
        ),
        "sessions_registered": sessions_registered,
        "days_without_data": max(0, 7 - days_with_data) if has_daily else None,
        "weekly_change": weekly_change,
        "weekly_change_pct": weekly_change_pct,
        "daily_explanation": {
            "what": "sRPE diario de la última semana completa." if analysis_scope == "last_complete_week" else "sRPE acumulado de la semana en curso.",
            "meaning": "Contexto agudo de carga interna percibida y duración acumulada.",
            "decision": "Regular densidad inmediata si aparecen picos, días sin registro o acumulación brusca." if analysis_scope == "last_complete_week" else "No comparar el acumulado parcial contra una semana completa; esperar cierre semanal o usarlo solo como señal operativa.",
        },
        "weekly_explanation": {
            "what": "Barras de sRPE semanal de las últimas 16 semanas con línea EMA de 6 semanas.",
            "meaning": "Permite observar tendencia del bloque, acumulaciones, descargas o aumentos bruscos.",
            "decision": "Cruzar la tendencia con wellness, dolor, adherencia y calendario deportivo.",
        },
    }


def _professional_metric_display_groups(cards: list[dict[str, object]]) -> tuple[list[dict[str, object]], list[str]]:
    available = [card for card in cards if card.get("state") != "missing"]
    missing = [str(card.get("title")) for card in cards if card.get("state") == "missing"]
    return available, missing


def _professional_prioritized_evolution_sections(
    sections: list[dict[str, object]],
    *,
    max_charts: int = 6,
) -> tuple[list[dict[str, object]], list[str]]:
    priority = {title: idx for idx, title in enumerate(PROFESSIONAL_EVOLUTION_PRIORITY)}
    chartable = [section for section in sections if len(section.get("points", [])) >= 2]
    chartable = sorted(chartable, key=lambda section: priority.get(str(section.get("title")), 99))
    selected = chartable[:max_charts]
    omitted = [str(section.get("title", "Métrica")) for section in chartable[max_charts:]]
    return selected, omitted


def _professional_assessment_date_count(state: dict[str, pd.DataFrame | None], athlete: str) -> int:
    history = _professional_jump_history(state, athlete)
    if history.empty or "Date" not in history.columns:
        return 0
    return int(pd.to_datetime(history["Date"], errors="coerce").dropna().dt.normalize().nunique())


def _professional_short_assessment_interval_warning(state: dict[str, pd.DataFrame | None], athlete: str) -> str:
    history = _professional_jump_history(state, athlete)
    if history.empty or "Date" not in history.columns:
        return ""
    dates = pd.to_datetime(history["Date"], errors="coerce").dropna().dt.normalize().drop_duplicates().sort_values()
    if len(dates) < 2:
        return ""
    intervals = dates.diff().dropna().dt.days
    if not intervals.empty and int(intervals.min()) < 42:
        return PROFESSIONAL_INTERVAL_WARNING
    return ""


def _professional_quadrants_ready(sections: list[dict[str, object]]) -> bool:
    return len(sections) >= 3 and all(section.get("selected") is not None for section in sections[:3])


def _professional_any_quadrant_ready(sections: list[dict[str, object]]) -> bool:
    return any(section.get("selected") is not None for section in sections[:3])


def _professional_status_label(state_label: object) -> str:
    clean = str(state_label or "").strip().lower()
    if clean == "available":
        return "Disponible"
    if clean == "partial":
        return "Parcial"
    return "Faltan datos"


def _professional_infer_response_scale(values: pd.Series | list[object]) -> str:
    numeric = pd.to_numeric(pd.Series(values), errors="coerce").dropna()
    if numeric.empty:
        return "Escala no definida"
    max_value = float(numeric.max())
    min_value = float(numeric.min())
    if min_value >= 0 and max_value <= 5:
        return "/5"
    if min_value >= 0 and max_value <= 10:
        return "/10"
    return "Escala no definida"


def _professional_wellness_scales(frame: pd.DataFrame) -> dict[str, str]:
    return {
        "sleep": "h",
        "stress": _professional_infer_response_scale(frame["Estres"]) if "Estres" in frame.columns else "Escala no definida",
        "pain": _professional_infer_response_scale(frame["Dolor"]) if "Dolor" in frame.columns else "Escala no definida",
        "score": "/5.0",
    }


def _professional_wellness_context(state: dict[str, pd.DataFrame | None], athlete: str) -> dict[str, object]:
    wdf = state.get("wellness_df")
    if wdf is None or wdf.empty or "Athlete" not in wdf.columns:
        return {
            "state": "missing",
            "rows": 0,
            "latest": PDF_MISSING_TEXT,
            "daily_points": [],
            "weekly_points": [],
            "last_week_summary": {},
            "current_week_summary": {},
            "analysis_scope": "missing",
            "analysis_title": "Wellness - última semana completa",
            "trend_allowed": False,
            "scales": {"sleep": "h", "stress": "Escala no definida", "pain": "Escala no definida", "score": "/5.0"},
            "message": "Faltan datos de wellness para este período.",
        }
    result = wdf.copy()
    if "Date" in result.columns:
        result["Date"] = pd.to_datetime(result["Date"], errors="coerce").dt.normalize()
    result = result[_professional_athlete_mask(result["Athlete"], athlete)]
    sleep_col = _report_wellness_source_column(result, "Sueno", "Sueno_hs")
    numeric_cols = [column for column in [sleep_col, "Estres", "Dolor"] if column]
    for column in numeric_cols:
        if column in result.columns:
            result[column] = pd.to_numeric(result[column], errors="coerce")
    result["__report_wellness_score"] = _report_wellness_score_series(result)
    if result.empty:
        return {
            "state": "missing",
            "rows": 0,
            "latest": PDF_MISSING_TEXT,
            "daily_points": [],
            "weekly_points": [],
            "last_week_summary": {},
            "current_week_summary": {},
            "analysis_scope": "missing",
            "analysis_title": "Wellness - última semana completa",
            "trend_allowed": False,
            "scales": {"sleep": "h", "stress": "Escala no definida", "pain": "Escala no definida", "score": "/5.0"},
            "message": "Faltan datos de wellness para este período.",
        }
    rows_count = int(len(result))
    scales = _professional_wellness_scales(result)
    latest = PDF_MISSING_TEXT
    if "Date" in result.columns and result["Date"].notna().any():
        latest = result["Date"].max().strftime("%d/%m/%Y")

    daily_points: list[dict[str, object]] = []
    current_week_points: list[dict[str, object]] = []
    last_week_summary: dict[str, object] = {}
    current_week_summary: dict[str, object] = {}
    reference_date = _professional_latest_data_date(state, athlete) or _professional_report_today()
    reference_date = pd.Timestamp(reference_date).normalize()
    current_week_start = _professional_week_start(reference_date)
    complete_week_start = _professional_last_complete_week_start(reference_date)

    def build_week_snapshot(
        week_start: pd.Timestamp,
        *,
        end_date: pd.Timestamp | None = None,
    ) -> tuple[list[dict[str, object]], dict[str, object]]:
        week_start = pd.Timestamp(week_start).normalize()
        week_end = week_start + pd.Timedelta(days=6)
        if end_date is not None:
            week_end = min(week_end, pd.Timestamp(end_date).normalize())
        week_df = result[(result["Date"] >= week_start) & (result["Date"] <= week_end)].copy()
        points: list[dict[str, object]] = []
        for day in range(7):
            date = week_start + pd.Timedelta(days=day)
            day_df = week_df[week_df["Date"].eq(date)]
            point = {"label": date.strftime("%d/%m")}
            for source_col, output_col in [
                (sleep_col, "sleep"),
                ("Estres", "stress"),
                ("Dolor", "pain"),
                ("__report_wellness_score", "score"),
            ]:
                point[output_col] = (
                    float(day_df[source_col].mean())
                    if source_col and source_col in day_df.columns and day_df[source_col].notna().any()
                    else None
                )
            points.append(point)
        summary: dict[str, object] = {}
        for source_col, output_col in [
            (sleep_col, "sleep_mean"),
            ("Estres", "stress_mean"),
            ("Dolor", "pain_mean"),
            ("__report_wellness_score", "score_mean"),
        ]:
            summary[output_col] = (
                float(week_df[source_col].mean())
                if source_col and source_col in week_df.columns and week_df[source_col].notna().any()
                else None
            )
            summary[output_col.replace("_mean", "_n")] = (
                int(week_df[source_col].notna().sum())
                if source_col and source_col in week_df.columns
                else 0
            )
        summary["days"] = int(week_df["Date"].nunique()) if not week_df.empty else 0
        return points, summary

    if "Date" in result.columns and result["Date"].notna().any():
        daily_points, last_week_summary = build_week_snapshot(complete_week_start)
        current_week_points, current_week_summary = build_week_snapshot(current_week_start, end_date=reference_date)

    weekly_points: list[dict[str, object]] = []
    if "Date" in result.columns and result["Date"].notna().any():
        weekly_source = result.dropna(subset=["Date"]).copy()
        weekly_source["week_start"] = weekly_source["Date"] - pd.to_timedelta(weekly_source["Date"].dt.weekday, unit="D")
        weekly_source = weekly_source[weekly_source["week_start"] <= complete_week_start]
        weekly_point_map: dict[pd.Timestamp, dict[str, object]] = {}
        if not weekly_source.empty:
            aggregation_kwargs: dict[str, tuple[str, str | object]] = {
                "stress": ("Estres", "mean"),
                "stress_n": ("Estres", lambda s: int(pd.Series(s).notna().sum())),
                "pain": ("Dolor", "mean"),
                "pain_n": ("Dolor", lambda s: int(pd.Series(s).notna().sum())),
                "score": ("__report_wellness_score", "mean"),
                "score_n": ("__report_wellness_score", lambda s: int(pd.Series(s).notna().sum())),
                "days": ("Date", lambda s: int(pd.to_datetime(pd.Series(s), errors="coerce").dropna().dt.normalize().nunique())),
            }
            if sleep_col and sleep_col in weekly_source.columns:
                aggregation_kwargs["sleep"] = (sleep_col, "mean")
                aggregation_kwargs["sleep_n"] = (sleep_col, lambda s: int(pd.Series(s).notna().sum()))
            weekly_wellness = (
                weekly_source.groupby("week_start", as_index=False)
                .agg(**aggregation_kwargs)
                .sort_values("week_start")
                .tail(16)
            )
            weekly_points = []
            for _, row in weekly_wellness.iterrows():
                week_start = pd.Timestamp(row["week_start"]).normalize()
                point = {
                    "label": week_start.strftime("%d/%m"),
                    "week_start": week_start,
                    "sleep": _coerce_float(row.get("sleep")),
                    "sleep_n": _coerce_float(row.get("sleep_n")),
                    "stress": _coerce_float(row.get("stress")),
                    "stress_n": _coerce_float(row.get("stress_n")),
                    "pain": _coerce_float(row.get("pain")),
                    "pain_n": _coerce_float(row.get("pain_n")),
                    "score": _coerce_float(row.get("score")),
                    "score_n": _coerce_float(row.get("score_n")),
                    "days": _coerce_float(row.get("days")),
                }
                weekly_points.append(point)
                weekly_point_map[week_start] = point

        weekly_summaries = _report_weekly_summaries(state)
        weekly_summary_source = _normalize_weekly_frame(weekly_summaries.get("weekly_wellness"))
        if not weekly_summary_source.empty and "Athlete" in weekly_summary_source.columns:
            weekly_summary_source = weekly_summary_source[_professional_athlete_mask(weekly_summary_source["Athlete"], athlete)].copy()
        if not weekly_summary_source.empty and "week_start" in weekly_summary_source.columns:
            weekly_summary_source = weekly_summary_source.sort_values("week_start")
            weekly_summary_source = weekly_summary_source[
                pd.to_datetime(weekly_summary_source["week_start"], errors="coerce").dt.normalize() <= complete_week_start
            ].tail(16)
            summary_points: list[dict[str, object]] = []
            for _, row in weekly_summary_source.iterrows():
                week_start = pd.to_datetime(row.get("week_start"), errors="coerce")
                if pd.isna(week_start):
                    continue
                week_start = pd.Timestamp(week_start).normalize()
                days_value = _coerce_float(row.get("wellness_days"))
                fallback_frame = pd.DataFrame(
                    [
                        {
                            "Sueno_hs": row.get("Sueno_mean"),
                            "Estres": row.get("Estres_mean"),
                            "Dolor": row.get("Dolor_mean"),
                        }
                    ]
                )
                fallback_score_series = _report_wellness_score_series(fallback_frame).dropna()
                fallback_score = float(fallback_score_series.iloc[0]) if not fallback_score_series.empty else None
                summary_point = {
                    "label": week_start.strftime("%d/%m"),
                    "week_start": week_start,
                    "sleep": _coerce_float(row.get("Sueno_mean")),
                    "sleep_n": days_value,
                    "stress": _coerce_float(row.get("Estres_mean")),
                    "stress_n": days_value,
                    "pain": _coerce_float(row.get("Dolor_mean")),
                    "pain_n": days_value,
                    "score": fallback_score,
                    "score_n": days_value,
                    "days": days_value,
                }
                summary_points.append(weekly_point_map.get(week_start, summary_point))
            if summary_points:
                weekly_points = summary_points

    has_any_metric = any(
        result[column].notna().any()
        for column in [*numeric_cols, "__report_wellness_score"]
        if column in result.columns
    )
    missing_variables = [
        label
        for column, label in [
            (sleep_col, "sueño"),
            ("Estres", "estrés"),
            ("Dolor", "dolor"),
            ("__report_wellness_score", "wellness score"),
        ]
        if not column or column not in result.columns or not result[column].notna().any()
    ]
    last_week_days = int(_coerce_float(last_week_summary.get("days")) or 0)
    if last_week_days == 0 and weekly_points:
        completed_points = [
            point
            for point in weekly_points
            if point.get("week_start") is not None
            and _professional_week_end(pd.Timestamp(point["week_start"])) <= reference_date
            and int(_coerce_float(point.get("days")) or 0) > 0
        ]
        if completed_points:
            selected_week = pd.Timestamp(completed_points[-1]["week_start"]).normalize()
            selected_points, selected_summary = build_week_snapshot(selected_week)
            selected_days = int(_coerce_float(selected_summary.get("days")) or 0)
            if selected_days > 0:
                complete_week_start = selected_week
                daily_points = selected_points
                last_week_summary = selected_summary
                last_week_days = selected_days
            else:
                point = completed_points[-1]
                complete_week_start = selected_week
                last_week_summary = {
                    "sleep_mean": point.get("sleep"),
                    "sleep_n": point.get("sleep_n"),
                    "stress_mean": point.get("stress"),
                    "stress_n": point.get("stress_n"),
                    "pain_mean": point.get("pain"),
                    "pain_n": point.get("pain_n"),
                    "score_mean": point.get("score"),
                    "score_n": point.get("score_n"),
                    "days": int(_coerce_float(point.get("days")) or 0),
                }
                last_week_days = int(_coerce_float(last_week_summary.get("days")) or 0)
    current_week_days = int(_coerce_float(current_week_summary.get("days")) or 0)
    if last_week_days > 0:
        analysis_scope = "last_complete_week"
        analysis_title = "Wellness - última semana completa"
        analysis_week_label = _professional_week_label(complete_week_start)
    elif current_week_days > 0:
        analysis_scope = "current_week_partial"
        analysis_title = "Wellness - semana en curso"
        analysis_week_label = _professional_week_label(current_week_start)
        daily_points = current_week_points
        last_week_summary = current_week_summary
        last_week_days = current_week_days
    else:
        analysis_scope = "missing"
        analysis_title = "Wellness - última semana completa"
        analysis_week_label = _professional_week_label(complete_week_start)
    trend_allowed = analysis_scope == "last_complete_week" and last_week_days >= 3
    partial_message = ""
    if has_any_metric and not trend_allowed:
        if analysis_scope == "current_week_partial":
            day_label = "1 día" if last_week_days == 1 else f"{last_week_days} días"
            partial_message = (
                f"Wellness - semana en curso. Registro parcial: {day_label} con datos. "
                "La lectura debe tomarse como alerta contextual puntual, no como tendencia semanal cerrada."
            )
        else:
            partial_message = _professional_wellness_partial_message(
                {
                    "last_week_summary": last_week_summary,
                    "scales": scales,
                }
            )

    if not has_any_metric:
        state_label = "missing"
    elif trend_allowed and rows_count >= 3 and not missing_variables:
        state_label = "available"
    else:
        state_label = "partial"
    return {
        "state": state_label,
        "rows": rows_count,
        "latest": latest,
        "daily_points": daily_points,
        "current_week_points": current_week_points,
        "weekly_points": weekly_points,
        "last_week_summary": last_week_summary,
        "current_week_summary": current_week_summary,
        "analysis_scope": analysis_scope,
        "analysis_title": analysis_title,
        "analysis_week_label": analysis_week_label,
        "missing_variables": missing_variables,
        "trend_allowed": trend_allowed,
        "partial_message": partial_message,
        "scales": scales,
        "message": "" if state_label != "missing" else "Faltan datos de wellness para este período.",
    }


def _professional_number_text(value: object, *, digits: int = 1, unit: str = "") -> str:
    numeric = _coerce_float(value)
    if numeric is None:
        return PDF_MISSING_TEXT
    rendered = f"{numeric:.{digits}f}"
    if digits == 0:
        rendered = rendered.split(".")[0]
    return f"{rendered} {unit}".strip()


def _professional_force_side_label(side: object) -> str:
    clean = str(side or "").strip().lower()
    if clean == "left":
        return "izquierda"
    if clean == "right":
        return "derecha"
    return PDF_MISSING_TEXT


def _professional_feedback_map(lines: list[str]) -> dict[str, str]:
    mapping = {
        "alto": "high",
        "bajo": "low",
        "fisiologico": "physiological",
        "biomecanico": "biomechanical",
        "proximo bloque": "next_block",
    }
    result = {value: "" for value in mapping.values()}
    extras: list[str] = []
    for raw_line in lines:
        text = _professional_visible_metric_text(raw_line).strip()
        if not text:
            continue
        prefix, separator, detail = text.partition(":")
        key = mapping.get(prefix.strip().casefold())
        if key and separator:
            result[key] = detail.strip()
        else:
            extras.append(text)
    if extras:
        result["extras"] = " ".join(extras)
    return result


def _professional_join_labels(values: list[str], *, fallback: str = PDF_MISSING_TEXT) -> str:
    clean = [_professional_visible_metric_text(value).strip() for value in values if str(value or "").strip()]
    if not clean:
        return fallback
    return ", ".join(dict.fromkeys(clean))


def _professional_normalized_text(value: object) -> str:
    repaired = _professional_visible_metric_text(value)
    normalized = unicodedata.normalize("NFD", repaired).encode("ascii", "ignore").decode("ascii")
    return re.sub(r"\s+", " ", normalized).strip().casefold()


def _professional_strip_terminal_period(value: object) -> str:
    return _professional_visible_metric_text(value).strip().rstrip(".")


def _professional_strip_scope_note(value: object) -> str:
    text = _professional_visible_metric_text(value).strip()
    note = PROFESSIONAL_CURRENT_PROFILE_SCOPE_NOTE
    if note and note in text:
        text = text.replace(note, "").strip()
    return re.sub(r"\s{2,}", " ", text).strip().rstrip(".")


def _professional_contact_time_dominant_text(value: object) -> str:
    text = _professional_visible_metric_text(value)
    for pattern in (
        r"Tiempo de contacto\s*\(([-+]\d+(?:\.\d+)?)\)",
        r"Contact Time\s*\(([-+]\d+(?:\.\d+)?)\)",
    ):
        text = re.sub(
            pattern,
            lambda match: f"Tiempo de contacto en zona favorable según escala invertida ({match.group(1)})",
            text,
            flags=re.IGNORECASE,
        )
    return text


def _professional_snapshot_metric(row: pd.Series | dict[str, object], *columns: str) -> float | None:
    row_series = row if isinstance(row, pd.Series) else pd.Series(row)
    for column in columns:
        numeric = pd.to_numeric(pd.Series([row_series.get(column)]), errors="coerce").iloc[0]
        if pd.notna(numeric):
            return float(numeric)
    return None


def _professional_snapshot_zscore(row: pd.Series | dict[str, object], canonical_field: str) -> float | None:
    row_series = row if isinstance(row, pd.Series) else pd.Series(row)
    return resolve_zscore(row_series, canonical_field)


PROFESSIONAL_NEUROMUSCULAR_SIGNAL_ORDER = (
    "SJ_cm",
    "CMJ_cm",
    "DJ_cm",
    "DJ_RSI",
    "DJ_tc_ms",
    "IMTP_relPF",
)

PROFESSIONAL_NEUROMUSCULAR_SUPPORT_METRICS = (
    ("DJ_RSI", "DJ RSI", "m/s", ("DJ_RSI_Z",), "higher_is_better"),
    ("DSI", "DSI", "", ("DSI_Z",), "higher_is_better"),
    ("mRSI", "mRSI", "m/s", ("mRSI_Z",), "higher_is_better"),
    ("IMTP_N", "IMTP", "N", ("IMTP_N_Z",), "higher_is_better"),
)

NEUROMUSCULAR_KPI_LABELS = {
    "CMJ_cm": "CMJ",
    "SJ_cm": "SJ",
    "DJ_cm": "DJ",
    "DJ_RSI": "DJ RSI",
    "DJ_tc_ms": "Tiempo de contacto",
    "DRI": "DRI",
    "EUR": "EUR",
    "IMTP_relPF": "IMTP relPF",
    "IMTP_N": "IMTP",
    "DSI": "DSI",
    "mRSI": "mRSI",
}

NEUROMUSCULAR_PRIORITY_SHORT_TEXT = {
    "A": "Mejorar calidad de contacto y reacción rápida.",
    "B": "Construir más base de fuerza para sostener la reacción rápida.",
    "C": "Construir más base de fuerza máxima e isométrica.",
    "D": "Reconstruir la base general antes de complejizar el bloque.",
    "E": "Revisar el salto con contramovimiento y la calidad del gesto.",
}

NEUROMUSCULAR_PRIORITY_DETAILED_TEXT = {
    "A": "Priorizar la reactividad rápida y la calidad de contacto sin perder la base de fuerza actual.",
    "B": "Priorizar fuerza base y transferencia vertical para sostener mejor la reactividad ya visible.",
    "C": "Priorizar fuerza máxima e isométricos específicos para sostener mejor la salida vertical.",
    "D": "Priorizar base general, fuerza básica y una progresión reactiva conservadora antes de complejizar el bloque.",
    "E": "Revisar técnica de CMJ, contexto de fatiga y progresión del contramovimiento antes de escalar la exigencia reactiva.",
}

NEUROMUSCULAR_FLAG_MESSAGES = {
    "atleta": {
        "missing_imtp": "Todavía falta IMTP para leer mejor tu base de fuerza.",
        "missing_dj": "Todavía falta Drop Jump para leer mejor tu reactividad.",
        "cmj_lower_than_sj": "En esta medición el CMJ no superó al SJ; conviene revisar técnica, fatiga y calidad del gesto.",
        "insufficient_pattern_evidence": "La foto actual todavía es parcial; conviene completar mediciones antes de cambiar demasiado el foco.",
    },
    "cliente": {
        "missing_imtp": "Todavía falta una referencia de fuerza para entender mejor la base actual.",
        "missing_dj": "Todavía falta una referencia reactiva para entender mejor cómo responde el salto.",
        "cmj_lower_than_sj": "En esta medición el salto con contramovimiento no mejoró al salto base; conviene revisarlo en el próximo control.",
        "insufficient_pattern_evidence": "Todavía falta información para definir con más claridad qué conviene priorizar.",
    },
    "profe": {
        "missing_imtp": "Falta IMTP para cerrar mejor la lectura de fuerza base.",
        "missing_dj": "Falta Drop Jump para cerrar mejor la lectura reactiva.",
        "cmj_lower_than_sj": "CMJ < SJ: revisar técnica, fatiga y coherencia del test antes de cambiar prioridades.",
        "insufficient_pattern_evidence": "La evidencia todavía es parcial para cerrar un patrón dominante.",
    },
}

NEUROMUSCULAR_CONFIDENCE_LABELS = {
    "high": "Alta",
    "moderate": "Media",
    "low": "Baja",
}


def _professional_metric_band_label(z_value: object) -> str:
    numeric = _coerce_float(z_value)
    return _professional_visible_metric_text(semaphore_label(numeric))


def _professional_metric_payload(
    *,
    key: str,
    label: object,
    value: object,
    unit: object,
    z_value: object,
    direction: object,
    available: object = None,
    value_col: object = "",
    z_col: object = "",
    source_date: object = "-",
) -> dict[str, object]:
    numeric_value = _coerce_float(value)
    numeric_z = _coerce_float(z_value)
    is_available = bool(available) if available is not None else numeric_value is not None
    return {
        "key": str(key),
        "label": _professional_visible_metric_text(label),
        "value": numeric_value,
        "unit": str(unit or "").strip(),
        "z": numeric_z,
        "z_score": numeric_z,
        "band": _professional_metric_band_label(numeric_z),
        "direction": str(direction or ""),
        "available": is_available,
        "value_col": str(value_col or key),
        "z_col": str(z_col or ""),
        "source_date": _professional_visible_metric_text(source_date or "-"),
    }


def _professional_neuromuscular_metrics(
    row: pd.Series | dict[str, object] | None,
    core_metrics: dict[str, object] | None,
) -> dict[str, dict[str, object]]:
    row_series = row if isinstance(row, pd.Series) else pd.Series(row or {}, dtype=object)
    metrics: dict[str, dict[str, object]] = {}
    for key, payload in (core_metrics or {}).items():
        if not isinstance(payload, dict):
            continue
        metric_key = str(key)
        value_col = str(payload.get("value_col") or metric_key)
        metrics[metric_key] = _professional_metric_payload(
            key=metric_key,
            label=payload.get("label", metric_key),
            value=payload.get("value"),
            unit=payload.get("unit", ""),
            z_value=payload.get("z_score", payload.get("z")),
            direction=payload.get("direction", ""),
            available=payload.get("available"),
            value_col=value_col,
            z_col=payload.get("z_col", ""),
            source_date=payload.get("source_date") or row_series.get(f"{value_col}__source_date") or "-",
        )

    occupied_value_cols = {
        str(metric.get("value_col") or key)
        for key, metric in metrics.items()
        if isinstance(metric, dict) and metric.get("available") and str(metric.get("value_col") or key).strip()
    }

    for metric_key, label, unit, z_columns, direction in PROFESSIONAL_NEUROMUSCULAR_SUPPORT_METRICS:
        value = _professional_snapshot_metric(row_series, metric_key)
        z_value = None
        for z_col in z_columns:
            if str(z_col).endswith("_Z") or str(z_col) in {"DJtc_Z", "DJ_Z", "IMTP_Z"}:
                z_value = _professional_snapshot_zscore(row_series, str(z_col))
            else:
                z_value = _professional_snapshot_metric(row_series, str(z_col))
            if z_value is not None:
                break
        if value is None and z_value is None:
            continue
        if metric_key in metrics and metrics[metric_key].get("available"):
            continue
        if metric_key in occupied_value_cols:
            continue
        metrics[metric_key] = _professional_metric_payload(
            key=metric_key,
            label=label,
            value=value,
            unit=unit,
            z_value=z_value,
            direction=direction,
            available=value is not None,
            value_col=metric_key,
            z_col=z_columns[0],
            source_date=row_series.get(f"{metric_key}__source_date") or "-",
        )
        occupied_value_cols.add(metric_key)
    return metrics


def _professional_neuromuscular_signal_text(
    profile_payload: dict[str, object],
    *,
    threshold: float,
) -> str:
    metrics = profile_payload.get("metrics", {})
    if not isinstance(metrics, dict):
        return ""
    comparisons: list[str] = []
    for key in PROFESSIONAL_NEUROMUSCULAR_SIGNAL_ORDER:
        metric = metrics.get(key)
        if not isinstance(metric, dict):
            continue
        z_value = _coerce_float(metric.get("z_score", metric.get("z")))
        if z_value is None:
            continue
        if threshold > 0 and z_value > threshold:
            comparisons.append(f"{metric.get('label', key)} ({z_value:+.2f})")
        elif threshold < 0 and z_value < threshold:
            comparisons.append(f"{metric.get('label', key)} ({z_value:+.2f})")
    return ", ".join(_professional_visible_metric_text(value) for value in dict.fromkeys(comparisons))


def _neuromuscular_primary_code(profile_payload: dict[str, object]) -> str:
    tokens = [token.strip() for token in str(profile_payload.get("profile_code") or "").split("+") if token.strip()]
    if "E" in tokens:
        return "E"
    return tokens[0] if tokens else "UNCLASSIFIED"


def _neuromuscular_kpi_labels(profile_payload: dict[str, object]) -> list[str]:
    metrics = profile_payload.get("metrics", {})
    labels: list[str] = []
    for key in profile_payload.get("kpi_to_track", []):
        metric = metrics.get(key, {}) if isinstance(metrics, dict) else {}
        label = metric.get("label") if isinstance(metric, dict) else ""
        labels.append(_professional_visible_metric_text(label or NEUROMUSCULAR_KPI_LABELS.get(str(key), str(key))))
    return list(dict.fromkeys([label for label in labels if str(label).strip()]))


def _neuromuscular_priority_short(profile_payload: dict[str, object]) -> str:
    primary_code = _neuromuscular_primary_code(profile_payload)
    flags = set(profile_payload.get("flags", []))
    if primary_code in NEUROMUSCULAR_PRIORITY_SHORT_TEXT:
        return _professional_visible_metric_text(NEUROMUSCULAR_PRIORITY_SHORT_TEXT[primary_code])
    if {"missing_imtp", "missing_dj"}.issubset(flags):
        return "Completar las mediciones clave que faltan y sostener la calidad actual."
    if "missing_dj" in flags:
        return "Completar la referencia reactiva y sostener la calidad del salto actual."
    if "missing_imtp" in flags:
        return "Completar la referencia de fuerza y sostener la calidad del salto actual."
    if "insufficient_pattern_evidence" in flags:
        return "Confirmar la próxima medición antes de cambiar demasiado el foco."
    return "Sostener la calidad actual y confirmar la próxima medición."


def _neuromuscular_priority_detailed(profile_payload: dict[str, object]) -> str:
    primary_code = _neuromuscular_primary_code(profile_payload)
    detail = NEUROMUSCULAR_PRIORITY_DETAILED_TEXT.get(primary_code, "")
    short_text = _neuromuscular_priority_short(profile_payload)
    if not detail:
        flags = set(profile_payload.get("flags", []))
        if {"missing_imtp", "missing_dj"}.issubset(flags):
            detail = "Completar IMTP y Drop Jump antes de mover demasiado la prioridad física del bloque."
        elif "missing_dj" in flags:
            detail = "Completar el Drop Jump para leer mejor la reactividad y no ajustar el bloque con una foto parcial."
        elif "missing_imtp" in flags:
            detail = "Completar IMTP para leer mejor la base de fuerza y no ajustar el bloque con una foto parcial."
        elif "insufficient_pattern_evidence" in flags:
            detail = "La evidencia actual todavía es parcial; conviene completar o repetir mediciones clave antes de cambiar demasiado el foco."
        else:
            detail = short_text
    kpi_labels = _neuromuscular_kpi_labels(profile_payload)
    if kpi_labels:
        detail = f"{detail} Seguir de cerca {', '.join(kpi_labels[:3])}."
    return _professional_visible_metric_text(detail)


def _neuromuscular_flag_messages(profile_payload: dict[str, object], audience: str) -> list[str]:
    audience_messages = NEUROMUSCULAR_FLAG_MESSAGES.get(normalize_report_audience(audience), {})
    messages = [
        _professional_visible_metric_text(audience_messages.get(flag, ""))
        for flag in profile_payload.get("flags", [])
        if str(audience_messages.get(flag, "")).strip()
    ]
    return list(dict.fromkeys(messages))


def _professional_feedback_from_structured_profile(profile_payload: dict[str, object]) -> dict[str, str]:
    feedback = {
        "high": _professional_neuromuscular_signal_text(profile_payload, threshold=0.5) or "sin variables > 0.5.",
        "low": _professional_neuromuscular_signal_text(profile_payload, threshold=-0.5) or "sin variables < -0.5.",
        "physiological": _professional_visible_metric_text(profile_payload.get("phys", "")),
        "biomechanical": _professional_visible_metric_text(profile_payload.get("bio", "")),
        "next_block": _professional_visible_metric_text(profile_payload.get("train", "")),
    }
    if profile_payload.get("profile_code") == "E" and str(profile_payload.get("summary_short") or "").strip():
        feedback["extras"] = _professional_visible_metric_text(profile_payload.get("summary_short"))
    return feedback


def _build_pdf_neuromuscular_profile_payload(
    row,
    reference_df: pd.DataFrame | None = None,
    context: dict[str, object] | None = None,
) -> dict[str, object]:
    row_series = row if isinstance(row, pd.Series) else pd.Series(row or {}, dtype=object)
    core_context = dict(context or {})
    core_context.setdefault("audience", "profe")
    core_available = True
    try:
        core_payload = build_neuromuscular_profile_result(
            row_series,
            reference_df=reference_df,
            context=core_context,
        )
    except Exception:
        core_available = False
        core_payload = {}

    if not isinstance(core_payload, dict):
        core_available = False
        core_payload = {}

    metrics = _professional_neuromuscular_metrics(row_series, core_payload.get("metrics"))
    payload = {
        "source": "core" if core_available else "legacy_fallback",
        "profile_code": str(core_payload.get("profile_code") or "UNCLASSIFIED"),
        "profile_label": _professional_visible_metric_text(core_payload.get("profile_label") or "Sin patron dominante"),
        "confidence": str(core_payload.get("confidence") or "low"),
        "confidence_label": NEUROMUSCULAR_CONFIDENCE_LABELS.get(str(core_payload.get("confidence") or "low"), "Baja"),
        "profile_source": str(core_payload.get("profile_source") or "unknown"),
        "profile_source_label": _professional_visible_metric_text(core_payload.get("profile_source_label") or "Fuente no determinada"),
        "profile_source_note": _professional_visible_metric_text(core_payload.get("profile_source_note") or ""),
        "profile_source_dates": {
            str(key): str(value)
            for key, value in dict(core_payload.get("profile_source_dates") or {}).items()
            if str(key).strip() and str(value).strip()
        },
        "profile_source_is_composite": bool(core_payload.get("profile_source_is_composite")),
        "phys": _professional_visible_metric_text(core_payload.get("phys") or ""),
        "bio": _professional_visible_metric_text(core_payload.get("bio") or ""),
        "train": _professional_visible_metric_text(core_payload.get("train") or ""),
        "summary_short": _professional_visible_metric_text(core_payload.get("summary_short") or ""),
        "summary_athlete": _professional_visible_metric_text(core_payload.get("summary_athlete") or ""),
        "summary_client": _professional_visible_metric_text(core_payload.get("summary_client") or ""),
        "summary_professional": _professional_visible_metric_text(core_payload.get("summary_professional") or ""),
        "metrics": metrics,
        "flags": [str(flag).strip() for flag in core_payload.get("flags", []) if str(flag).strip()],
        "evidence": [
            _professional_visible_metric_text(item)
            for item in core_payload.get("evidence", [])
            if str(item).strip()
        ],
        "kpi_to_track": [
            _professional_visible_metric_text(item)
            for item in core_payload.get("kpi_to_track", [])
            if str(item).strip()
        ],
    }
    payload["kpi_labels"] = _neuromuscular_kpi_labels(payload)
    payload["training_priority_short"] = _neuromuscular_priority_short(payload)
    payload["training_priority_detailed"] = _neuromuscular_priority_detailed(payload)
    payload["flag_messages_athlete"] = _neuromuscular_flag_messages(payload, "atleta")
    payload["flag_messages_client"] = _neuromuscular_flag_messages(payload, "cliente")
    payload["flag_messages_professional"] = _neuromuscular_flag_messages(payload, "profe")
    payload["feedback"] = _professional_feedback_from_structured_profile(payload) if core_available else {}
    return payload


def _professional_force_time_note_text(value: object) -> str:
    text = _professional_visible_metric_text(value).strip()
    targeted_replacements = {
        "produccion maxima": "producción máxima",
        "fuerza isometrica": "fuerza isométrica",
        "en esta medicion": "En esta medición",
        "en esta Medicion": "En esta medición",
        "posicion de imtp": "posición de IMTP",
        "describe como expresa": "describe cómo se expresa",
        "capacidad maxima": "capacidad máxima",
        "como apoyo descriptivo": "como apoyo descriptivo",
    }
    for source, target in targeted_replacements.items():
        def _replace(match: re.Match[str], fixed: str = target) -> str:
            word = match.group(0)
            return fixed[:1].upper() + fixed[1:] if word[:1].isupper() else fixed

        text = re.sub(rf"\b{re.escape(source)}\b", _replace, text, flags=re.IGNORECASE)
    return text


def _professional_build_profile_feedback_fallback(
    snapshot_row: pd.Series | dict[str, object],
    *,
    assessment_count: int,
    has_clear_lagging: bool,
) -> dict[str, str]:
    row_series = snapshot_row if isinstance(snapshot_row, pd.Series) else pd.Series(snapshot_row)
    sj_z = _professional_snapshot_zscore(row_series, "SJ_Z")
    cmj_z = _professional_snapshot_zscore(row_series, "CMJ_Z")
    dj_z = _professional_snapshot_zscore(row_series, "DJ_height_Z")
    dri_z = _professional_snapshot_zscore(row_series, "DRI_Z")
    contact_z = _professional_snapshot_zscore(row_series, "TC_inv_Z")
    eur_z = _professional_snapshot_zscore(row_series, "EUR_Z")
    imtp_z = _professional_snapshot_zscore(row_series, "IMTP_relPF_Z")

    vertical_high = any(value is not None and value >= 0.5 for value in (sj_z, cmj_z))
    imtp_high = imtp_z is not None and imtp_z >= 0.5
    dj_low = dj_z is not None and dj_z <= -0.5
    reactive_high = any(value is not None and value >= 0.5 for value in (dri_z, contact_z))

    physiological = "La lectura disponible describe el perfil actual del atleta, sin evidenciar todavía una adaptación cerrada del bloque."
    if vertical_high and imtp_high:
        physiological = "El perfil actual muestra buena salida vertical con una base de fuerza relativa que puede sostener esa expresión."
    elif vertical_high:
        physiological = "El perfil actual sugiere buena salida vertical y una expresión concéntrica/vertical favorable en los saltos disponibles."
    elif imtp_high:
        physiological = "La fuerza relativa disponible aporta una base útil para sostener la producción de fuerza del perfil actual."
    elif reactive_high:
        physiological = "El perfil actual muestra una señal reactiva favorable, pero debe leerse junto con DJ/DRI y tiempo de contacto."
    elif eur_z is not None:
        physiological = "El perfil actual permite una lectura conservadora de la relación entre cualidades verticales, aunque todavía requiere continuidad para cerrar conclusiones más firmes."

    biomechanical = "La mecánica disponible no muestra un déficit único y conviene leerla junto con el contexto del test."
    if dj_low:
        biomechanical = "La expresión en DJ aparece por debajo de otras cualidades y conviene verificar técnica, contacto, familiarización y contexto antes de sacar una conclusión fuerte."
    elif contact_z is not None and contact_z >= 0.5:
        biomechanical = "El tiempo de contacto aparece en zona favorable según escala invertida, pero debe interpretarse junto con DJ/DRI y RSI."
    elif reactive_high:
        biomechanical = "La lectura reactiva es utilizable, aunque debe confirmarse con la consistencia entre DJ, RSI y tiempo de contacto."
    elif vertical_high:
        biomechanical = "La lectura biomecánica se apoya más en la salida vertical actual que en un patrón reactivo completamente definido."

    next_block = "Sostener salida vertical y fuerza base."
    if dj_low:
        next_block = (
            "Priorizar calidad de contacto y progresión reactiva si la expresión en DJ sigue baja, "
            "sin aumentar el volumen pliométrico de forma brusca."
        )
    elif has_clear_lagging and any(value is not None and value <= -0.5 for value in (dri_z, contact_z)):
        next_block = (
            "Priorizar calidad de contacto y progresión reactiva, verificando tolerancia antes de aumentar la densidad pliométrica."
        )
    elif not has_clear_lagging and (vertical_high or imtp_high):
        next_block = "Sostener salida vertical y fuerza base."

    return {
        "physiological": physiological.strip(),
        "biomechanical": biomechanical.strip(),
        "next_block": next_block.strip(),
    }


def _professional_is_no_clear_lagging_text(value: object) -> bool:
    normalized = _professional_normalized_text(value)
    if not normalized or normalized in {"-", "sin dato", "faltan datos"}:
        return True
    no_lagging_markers = (
        "sin variables < -0.5",
        "sin una variable claramente rezagada",
        "sin un rezago dominante",
        "no aparece una variable claramente rezagada",
    )
    return any(marker in normalized for marker in no_lagging_markers)


def _professional_has_clear_lagging_variable(feedback: dict[str, object]) -> bool:
    return not _professional_is_no_clear_lagging_text(feedback.get("low", ""))


def _professional_sanitize_profile_feedback(
    feedback: dict[str, object],
    snapshot_row: pd.Series | dict[str, object] | None = None,
    *,
    assessment_count: int = 0,
) -> dict[str, str]:
    cleaned = {
        str(key): _professional_visible_metric_text(value).strip()
        for key, value in feedback.items()
        if str(value or "").strip()
    }
    for key in ("physiological", "biomechanical", "next_block"):
        if key in cleaned:
            cleaned[key] = _professional_strip_scope_note(cleaned.get(key, ""))
    cleaned["high"] = _professional_contact_time_dominant_text(cleaned.get("high", ""))
    has_clear_lagging = _professional_has_clear_lagging_variable(cleaned)
    next_block_normalized = _professional_normalized_text(cleaned.get("next_block", ""))
    if not has_clear_lagging:
        cleaned["low"] = PROFESSIONAL_NO_CLEAR_LAGGING_TEXT
        cleaned["next_block"] = PROFESSIONAL_NO_CLEAR_LAGGING_NEXT_BLOCK
        if not cleaned.get("physiological"):
            cleaned["physiological"] = "Perfil equilibrado en los índices disponibles."
        if not cleaned.get("biomechanical"):
            cleaned["biomechanical"] = "Sin déficits biomecánicos marcados en los tests disponibles."
    elif "variable mas rezagada" in next_block_normalized:
        cleaned["next_block"] = "Sostener la cualidad dominante y ajustar el limitante principal en el próximo bloque."
    if snapshot_row is not None:
        fallback = _professional_build_profile_feedback_fallback(
            snapshot_row,
            assessment_count=assessment_count,
            has_clear_lagging=has_clear_lagging,
        )
        for key, value in fallback.items():
            normalized = _professional_normalized_text(cleaned.get(key, ""))
            if (
                not cleaned.get(key)
                or "faltan datos" in normalized
                or (key == "next_block" and "variable mas rezagada" in normalized)
            ):
                cleaned[key] = value
        if has_clear_lagging and any(token in _professional_normalized_text(cleaned.get("low", "")) for token in ("dj height", "dj")):
            cleaned["next_block"] = fallback["next_block"]
    return cleaned


def _professional_delta_signal_labels(delta_df: pd.DataFrame, signal: str) -> list[str]:
    if delta_df is None or delta_df.empty or "Signal" not in delta_df.columns:
        return []
    mask = delta_df["Signal"].fillna("").astype(str).str.casefold().eq(signal.casefold())
    if not mask.any():
        return []
    return (
        delta_df.loc[mask, "Label"]
        .dropna()
        .astype(str)
        .map(_professional_visible_metric_text)
        .drop_duplicates()
        .tolist()
    )


def _professional_change_pattern_lines(delta_df: pd.DataFrame, assessment_interval_warning: str = "") -> list[str]:
    if delta_df is None or delta_df.empty or "Variable" not in delta_df.columns:
        return []
    frame = delta_df.copy()
    frame["Variable"] = frame["Variable"].astype(str)
    frame["Signal"] = frame.get("Signal", pd.Series("", index=frame.index)).fillna("").astype(str)

    def signal_for(variable: str) -> str:
        rows = frame[frame["Variable"].eq(variable)]
        return str(rows.iloc[-1].get("Signal", "")) if not rows.empty else ""

    vertical_vars = ["CMJ_cm", "SJ_cm", "DJ_cm"]
    vertical_improved = [
        _professional_visible_metric_text(row.get("Label", row.get("Variable")))
        for _, row in frame[frame["Variable"].isin(vertical_vars) & frame["Signal"].eq("mejora relevante")].iterrows()
    ]
    vertical_declined = [
        _professional_visible_metric_text(row.get("Label", row.get("Variable")))
        for _, row in frame[frame["Variable"].isin(vertical_vars) & frame["Signal"].eq("caida relevante")].iterrows()
    ]
    lines: list[str] = []
    if vertical_improved:
        lines.append(
            f"Patrón integrado: el output vertical mejoró en {_professional_join_labels(vertical_improved)}."
        )
    elif vertical_declined:
        lines.append(
            f"Patrón integrado: el output vertical cayó en {_professional_join_labels(vertical_declined)}; revisar protocolo, fatiga y contexto de carga."
        )

    rsi_signal = signal_for("DJ_RSI")
    tc_signal = signal_for("DJ_tc_ms")
    if rsi_signal == "caida relevante" and tc_signal == "caida relevante":
        lines.append(
            "Eficiencia reactiva: DJ RSI cayó y el tiempo de contacto aumentó; sugiere una estrategia más lenta o mayor costo de contacto, no necesariamente mejor reactividad rápida."
        )
    elif rsi_signal == "caida relevante" or tc_signal == "caida relevante":
        lines.append(
            "Eficiencia reactiva: aparece una señal desfavorable en DJ RSI/tiempo de contacto; verificar protocolo, familiarización, fatiga y exposición pliométrica antes de concluir adaptación."
        )
    elif rsi_signal == "mejora relevante" or tc_signal == "mejora relevante":
        lines.append(
            "Eficiencia reactiva: la señal de DJ RSI/tiempo de contacto es favorable, pero debe leerse junto con DJ, DRI y calidad técnica."
        )

    isometric_improved = signal_for("IMTP_N") == "mejora relevante" or signal_for("IMTP_relPF") == "mejora relevante"
    isometric_declined = signal_for("IMTP_N") == "caida relevante" or signal_for("IMTP_relPF") == "caida relevante"
    if isometric_improved:
        lines.append("Fuerza isométrica: mejora relevante en IMTP/fuerza relativa; interpretarla junto con masa corporal, ángulo y consistencia del protocolo.")
    elif isometric_declined:
        lines.append("Fuerza isométrica: caída relevante en IMTP/fuerza relativa; confirmar calidad del test y estado de fatiga antes de ajustar prioridades.")

    if assessment_interval_warning:
        lines.append("Cautela temporal: intervalo corto entre evaluaciones; evitar atribuir el patrón completo a adaptación sin seguimiento.")
    return lines


def _professional_composite_metric_count(metric_table: pd.DataFrame) -> int:
    if metric_table is None or metric_table.empty or "Valor" not in metric_table.columns:
        return 0
    values = metric_table["Valor"].fillna("-").astype(str).str.strip()
    return int(values.ne("-").sum())


def _build_professional_composite_profile_payload(
    state: dict[str, pd.DataFrame | None],
    athlete: str,
) -> dict[str, object]:
    history = _professional_jump_history(state, athlete)
    assessment_count = _professional_assessment_date_count(state, athlete)
    empty_neuromuscular_profile = _build_pdf_neuromuscular_profile_payload(
        pd.Series(dtype=object),
        context={"assessment_count": assessment_count, "scope": "professional_composite_profile"},
    )
    if history.empty:
        return {
            "title": "Perfil actual compuesto",
            "state": "missing",
            "message": "Faltan evaluaciones suficientes para construir el perfil actual compuesto.",
            "metric_table": pd.DataFrame(columns=["Variable", "Valor", "Z-score", "Origen / referencia"]),
            "available_metric_count": 0,
            "feedback": {},
            "neuromuscular_profile": empty_neuromuscular_profile,
            "profile_code": empty_neuromuscular_profile["profile_code"],
            "profile_label": empty_neuromuscular_profile["profile_label"],
            "confidence": empty_neuromuscular_profile["confidence"],
            "profile_source": empty_neuromuscular_profile["profile_source"],
            "profile_source_label": empty_neuromuscular_profile["profile_source_label"],
            "profile_source_note": empty_neuromuscular_profile["profile_source_note"],
            "profile_source_dates": empty_neuromuscular_profile["profile_source_dates"],
            "profile_source_is_composite": empty_neuromuscular_profile["profile_source_is_composite"],
            "summary_short": empty_neuromuscular_profile["summary_short"],
            "summary_athlete": empty_neuromuscular_profile["summary_athlete"],
            "summary_client": empty_neuromuscular_profile["summary_client"],
            "summary_professional": empty_neuromuscular_profile["summary_professional"],
            "phys": empty_neuromuscular_profile["phys"],
            "bio": empty_neuromuscular_profile["bio"],
            "train": empty_neuromuscular_profile["train"],
            "metrics": empty_neuromuscular_profile["metrics"],
            "flags": empty_neuromuscular_profile["flags"],
            "evidence": empty_neuromuscular_profile["evidence"],
            "kpi_to_track": empty_neuromuscular_profile["kpi_to_track"],
            "note": PROFESSIONAL_COMPOSITE_PROFILE_NOTE,
        }

    snapshot_row, source_df = build_composite_profile_snapshot(history)
    if snapshot_row is None:
        return {
            "title": "Perfil actual compuesto",
            "state": "missing",
            "message": "Faltan datos suficientes para construir el perfil actual compuesto.",
            "metric_table": pd.DataFrame(columns=["Variable", "Valor", "Z-score", "Origen / referencia"]),
            "available_metric_count": 0,
            "feedback": {},
            "neuromuscular_profile": empty_neuromuscular_profile,
            "profile_code": empty_neuromuscular_profile["profile_code"],
            "profile_label": empty_neuromuscular_profile["profile_label"],
            "confidence": empty_neuromuscular_profile["confidence"],
            "profile_source": empty_neuromuscular_profile["profile_source"],
            "profile_source_label": empty_neuromuscular_profile["profile_source_label"],
            "profile_source_note": empty_neuromuscular_profile["profile_source_note"],
            "profile_source_dates": empty_neuromuscular_profile["profile_source_dates"],
            "profile_source_is_composite": empty_neuromuscular_profile["profile_source_is_composite"],
            "summary_short": empty_neuromuscular_profile["summary_short"],
            "summary_athlete": empty_neuromuscular_profile["summary_athlete"],
            "summary_client": empty_neuromuscular_profile["summary_client"],
            "summary_professional": empty_neuromuscular_profile["summary_professional"],
            "phys": empty_neuromuscular_profile["phys"],
            "bio": empty_neuromuscular_profile["bio"],
            "train": empty_neuromuscular_profile["train"],
            "metrics": empty_neuromuscular_profile["metrics"],
            "flags": empty_neuromuscular_profile["flags"],
            "evidence": empty_neuromuscular_profile["evidence"],
            "kpi_to_track": empty_neuromuscular_profile["kpi_to_track"],
            "note": PROFESSIONAL_COMPOSITE_PROFILE_NOTE,
        }

    metric_table = build_composite_profile_metric_table(snapshot_row)
    if not metric_table.empty:
        metric_table = metric_table.apply(lambda column: column.map(_professional_visible_metric_text))
        if "Z-score" in metric_table.columns:
            metric_table["Z-score"] = metric_table["Z-score"].map(
                lambda value: "\u2014"
                if _professional_visible_metric_text(value).strip()
                in {"", "-", "\u2014", PDF_MISSING_TEXT, "Sin dato"}
                else value
            )
    neuromuscular_profile = _build_pdf_neuromuscular_profile_payload(
        snapshot_row,
        reference_df=history,
        context={
            "assessment_count": assessment_count,
            "scope": "professional_composite_profile",
            "profile_source": "composite_snapshot",
        },
    )
    feedback_seed = neuromuscular_profile.get("feedback", {})
    if neuromuscular_profile.get("source") != "core":
        feedback_seed = {}
    if not isinstance(feedback_seed, dict) or not any(str(value or "").strip() for value in feedback_seed.values()):
        feedback_seed = {}
    if not feedback_seed:
        feedback_seed = {
            key: _professional_visible_metric_text(value)
            for key, value in _professional_feedback_map(build_jump_feedback_lines(snapshot_row)).items()
        }
    feedback = _professional_sanitize_profile_feedback(
        feedback_seed,
        snapshot_row,
        assessment_count=assessment_count,
    )
    available_metric_count = _professional_composite_metric_count(metric_table)
    state_label = "available" if available_metric_count >= PROFESSIONAL_FULL_REPORT_MIN_COMPOSITE_METRICS else "partial"
    dominant_text = _professional_strip_terminal_period(
        feedback.get("high") or "sin variables claramente por encima de la referencia"
    )
    has_clear_lagging = _professional_has_clear_lagging_variable(feedback)
    lagging_text = _professional_strip_terminal_period(feedback.get("low") or PROFESSIONAL_NO_CLEAR_LAGGING_TEXT)
    if available_metric_count and has_clear_lagging:
        summary_line = f"Predominan hoy {dominant_text}. El limitante principal aparece en {lagging_text}."
    elif available_metric_count:
        summary_line = f"Predominan hoy {dominant_text}. {PROFESSIONAL_NO_CLEAR_LAGGING_TEXT}"
    else:
        summary_line = "No hay suficientes variables válidas para describir el perfil compuesto."
    scope_note = PROFESSIONAL_CURRENT_PROFILE_SCOPE_NOTE if assessment_count < 2 and available_metric_count else ""
    return {
        "title": "Perfil actual compuesto",
        "state": state_label,
        "message": "" if available_metric_count else "Faltan datos suficientes para construir el perfil actual compuesto.",
        "profile_row": snapshot_row,
        "source_rows": source_df,
        "metric_table": metric_table,
        "available_metric_count": available_metric_count,
        "feedback": feedback,
        "neuromuscular_profile": neuromuscular_profile,
        "profile_code": neuromuscular_profile.get("profile_code", "UNCLASSIFIED"),
        "profile_label": neuromuscular_profile.get("profile_label", "Sin patron dominante"),
        "confidence": neuromuscular_profile.get("confidence", "low"),
        "profile_source": neuromuscular_profile.get("profile_source", "unknown"),
        "profile_source_label": neuromuscular_profile.get("profile_source_label", "Fuente no determinada"),
        "profile_source_note": neuromuscular_profile.get("profile_source_note", ""),
        "profile_source_dates": neuromuscular_profile.get("profile_source_dates", {}),
        "profile_source_is_composite": neuromuscular_profile.get("profile_source_is_composite", False),
        "summary_short": neuromuscular_profile.get("summary_short", ""),
        "summary_athlete": neuromuscular_profile.get("summary_athlete", ""),
        "summary_client": neuromuscular_profile.get("summary_client", ""),
        "summary_professional": neuromuscular_profile.get("summary_professional", ""),
        "phys": neuromuscular_profile.get("phys", ""),
        "bio": neuromuscular_profile.get("bio", ""),
        "train": neuromuscular_profile.get("train", ""),
        "metrics": neuromuscular_profile.get("metrics", {}),
        "flags": neuromuscular_profile.get("flags", []),
        "evidence": neuromuscular_profile.get("evidence", []),
        "kpi_to_track": neuromuscular_profile.get("kpi_to_track", []),
        "has_clear_lagging": has_clear_lagging,
        "summary_line": summary_line.strip(),
        "latest_profile_date": _format_profile_source_date(history["Date"].max()),
        "note": PROFESSIONAL_COMPOSITE_PROFILE_NOTE,
        "scope_note": scope_note,
    }


def _build_professional_change_payload(
    state: dict[str, pd.DataFrame | None],
    athlete: str,
) -> dict[str, object]:
    history = _professional_jump_history(state, athlete)
    assessment_count = _professional_assessment_date_count(state, athlete)
    empty_table = pd.DataFrame(columns=["Variable", "Actual", "Anterior", "Delta abs", "Delta %", "Threshold", "Senal"])
    if history.empty or assessment_count < 2:
        return {
            "title": "Cambios vs evaluaciÃ³n anterior",
            "state": "missing",
            "message": PROFESSIONAL_NO_EVOLUTION_TEXT,
            "delta_df": pd.DataFrame(),
            "display_table": empty_table,
            "summary_lines": [],
            "row_count": 0,
            "improvements": [],
            "declines": [],
            "no_change": [],
            "no_previous": [],
            "to_verify": [],
        }

    latest_date = pd.to_datetime(history["Date"], errors="coerce").dropna().max()
    delta_df = compute_swc_delta(history, latest_date)
    display_table = build_jump_delta_display_table(delta_df)
    if not display_table.empty:
        display_table = display_table.apply(lambda column: column.map(_professional_visible_metric_text))
    improvements = _professional_delta_signal_labels(delta_df, "mejora relevante")
    declines = _professional_delta_signal_labels(delta_df, "caida relevante")
    no_change = _professional_delta_signal_labels(delta_df, "sin cambio relevante")
    no_previous = _professional_delta_signal_labels(delta_df, "sin dato anterior")
    to_verify = []
    interval_warning = _professional_short_assessment_interval_warning(state, athlete)
    if interval_warning:
        to_verify.append("intervalo corto entre evaluaciones")
    if no_previous:
        to_verify.append(f"sin dato anterior en {_professional_join_labels(no_previous)}")

    summary_lines = [_professional_visible_metric_text(line) for line in build_jump_temporal_context(delta_df)]
    summary_lines.extend(_professional_change_pattern_lines(delta_df, interval_warning))
    if improvements:
        summary_lines.append(f"Mejoras relevantes: {_professional_join_labels(improvements)}.")
    if declines:
        summary_lines.append(f"CaÃ­das relevantes: {_professional_join_labels(declines)}.")
    if no_change:
        summary_lines.append(f"Sin cambio relevante: {_professional_join_labels(no_change)}.")
    if no_previous:
        summary_lines.append(f"Sin dato anterior: {_professional_join_labels(no_previous)}.")
    if to_verify:
        summary_lines.append(f"Datos a verificar: {_professional_join_labels(to_verify)}.")

    state_label = "available" if not display_table.empty else "partial"
    return {
        "title": "Cambios vs evaluaciÃ³n anterior",
        "state": state_label,
        "message": "" if not display_table.empty else PROFESSIONAL_NO_EVOLUTION_TEXT,
        "delta_df": delta_df,
        "display_table": display_table,
        "summary_lines": list(dict.fromkeys(summary_lines)),
        "row_count": int(len(display_table)),
        "improvements": improvements,
        "declines": declines,
        "no_change": no_change,
        "no_previous": no_previous,
        "to_verify": to_verify,
    }


def _professional_full_report_ready(
    composite_payload: dict[str, object],
    change_payload: dict[str, object],
    assessment_count: int,
) -> bool:
    if assessment_count < 2:
        return False
    if int(composite_payload.get("available_metric_count") or 0) < PROFESSIONAL_FULL_REPORT_MIN_COMPOSITE_METRICS:
        return False
    return int(change_payload.get("row_count") or 0) >= 3


def _professional_latest_force_time_payload(
    state: dict[str, pd.DataFrame | None],
    athlete: str,
    *,
    test_id: str,
) -> tuple[pd.Series | None, dict[str, object]]:
    history = _professional_jump_history(state, athlete)
    if not history.empty:
        for _, row in history.sort_values("Date", ascending=False).iterrows():
            payload = build_force_time_report_payload(row, test_id=test_id, report_type="professional")
            summary = payload.get("summary", {})
            if payload.get("has_valid_force_time") or _coerce_float(summary.get("peak_force_n")) is not None:
                return row, payload
        latest_row = history.sort_values("Date").iloc[-1]
        return latest_row, build_force_time_report_payload(latest_row, test_id=test_id, report_type="professional")
    fallback_row = _latest_jump_row(state, athlete)
    return fallback_row, build_force_time_report_payload(
        fallback_row if fallback_row is not None else {},
        test_id=test_id,
        report_type="professional",
    )


def _build_professional_isometric_payload(
    state: dict[str, pd.DataFrame | None],
    athlete: str,
    cards: list[dict[str, object]],
) -> dict[str, object]:
    imtp_card = next((card for card in cards if card.get("title") == "IMTP"), {})
    imtp_row, imtp_payload = _professional_latest_force_time_payload(state, athlete, test_id="imtp")
    _, iso_payload = _professional_latest_force_time_payload(state, athlete, test_id="iso_push_hamstring")

    imtp_summary = imtp_payload.get("summary", {}) if isinstance(imtp_payload.get("summary"), dict) else {}
    imtp_asymmetry = imtp_payload.get("asymmetry", {}) if isinstance(imtp_payload.get("asymmetry"), dict) else {}
    iso_summary = iso_payload.get("summary", {}) if isinstance(iso_payload.get("summary"), dict) else {}
    iso_asymmetry = iso_payload.get("asymmetry", {}) if isinstance(iso_payload.get("asymmetry"), dict) else {}

    imtp_rows = [
        ("Peak Force", safe_value(imtp_card.get("value"))),
        ("Cambio relevante", safe_value(imtp_card.get("delta"))),
        (
            "Fuerza relativa",
            _professional_number_text(imtp_row.get("IMTP_relPF"), digits=2, unit="N/kg") if imtp_row is not None else PDF_MISSING_TEXT,
        ),
        ("Force Avg", _professional_number_text(imtp_summary.get("avg_force_n"), digits=0, unit="N")),
        ("Time to Peak", _professional_number_text(imtp_summary.get("time_to_peak_s"), digits=2, unit="s")),
        ("AsimetrÃ­a", _professional_number_text(imtp_summary.get("absolute_asymmetry_pct"), digits=1, unit="%")),
        ("Lado dominante", _professional_force_side_label(imtp_asymmetry.get("stronger_side"))),
    ]
    imtp_rows = [(label, value) for label, value in imtp_rows if value != PDF_MISSING_TEXT]

    iso_rows = [
        ("Peak Force", _professional_number_text(iso_summary.get("peak_force_n"), digits=0, unit="N")),
        ("Force Avg", _professional_number_text(iso_summary.get("avg_force_n"), digits=0, unit="N")),
        ("Time to Peak", _professional_number_text(iso_summary.get("time_to_peak_s"), digits=2, unit="s")),
        ("AsimetrÃ­a", _professional_number_text(iso_summary.get("absolute_asymmetry_pct"), digits=1, unit="%")),
        ("Lado dominante", _professional_force_side_label(iso_asymmetry.get("stronger_side"))),
    ]
    iso_rows = [(label, value) for label, value in iso_rows if value != PDF_MISSING_TEXT]

    imtp_notes = []
    for key in ["peak_force_text", "force_time_text", "rfd_text", "decision_note"]:
        text = safe_value((imtp_payload.get("interpretation") or {}).get(key) if isinstance(imtp_payload.get("interpretation"), dict) else None)
        if text != PDF_MISSING_TEXT:
            imtp_notes.append(_professional_force_time_note_text(text))
    iso_notes = []
    for key in ["peak_force_text", "force_time_text", "asymmetry_text", "decision_note"]:
        text = safe_value((iso_payload.get("interpretation") or {}).get(key) if isinstance(iso_payload.get("interpretation"), dict) else None)
        if text != PDF_MISSING_TEXT:
            iso_notes.append(_professional_force_time_note_text(text))

    has_isometric_data = bool(imtp_rows or iso_rows or imtp_payload.get("has_valid_force_time") or iso_payload.get("has_valid_force_time"))
    return {
        "title": "IsomÃ©tricos y force-time avanzado",
        "state": "available" if has_isometric_data else "missing",
        "message": "" if has_isometric_data else "Faltan datos isomÃ©tricos vÃ¡lidos para esta secciÃ³n.",
        "imtp_rows": imtp_rows,
        "iso_rows": iso_rows,
        "imtp_notes": imtp_notes[:4],
        "iso_notes": iso_notes[:4],
        "imtp_payload": imtp_payload,
        "iso_payload": iso_payload,
        "force_time_available": bool(imtp_payload.get("has_valid_force_time")),
        "iso_available": bool(iso_rows or iso_payload.get("has_valid_force_time")),
    }


def _draw_compact_professional_force_time_block(pdf: dict[str, object], payload: dict[str, object] | None, *_, **__) -> None:
    if not payload or not payload.get("has_valid_force_time"):
        return
    story = pdf.get("story")
    p = pdf.get("p")
    box = pdf.get("box")
    TableClass = pdf.get("Table")
    TableStyleClass = pdf.get("TableStyle")
    SpacerClass = pdf.get("Spacer")
    mm_unit = pdf.get("mm")
    palette = pdf.get("palette", {})
    if not isinstance(story, list) or p is None or box is None or TableClass is None or TableStyleClass is None or SpacerClass is None or mm_unit is None:
        return

    force_points = list(payload.get("force_time_points", []))
    rfd_points = list(payload.get("rfd_points", []))
    force_rows = [
        (f"Force@{str(point.get('label') or '').replace(' ms', '')}", _professional_number_text(point.get("value_n"), digits=0, unit="N"))
        for point in force_points
        if _coerce_float(point.get("value_n")) is not None
    ]
    rfd_rows = [
        (str(point.get("label") or ""), _professional_number_text(point.get("value_n_s"), digits=0, unit="N/s"))
        for point in rfd_points
        if _coerce_float(point.get("value_n_s")) is not None
    ]
    story.append(
        box(
            [
                p("Force-time descriptivo", "ProfCardTitle"),
                p(
                    "Puntos force-time para contextualizar cómo se expresa la fuerza en ventanas tempranas e intermedias antes del pico.",
                    "ProfMuted",
                ),
                p("RFD exportada descriptiva; interpretar con cautela si no hay TE propio.", "ProfMuted"),
            ],
            background=palette.get("panel"),
            border_color=palette.get("line_dark"),
            padding=5,
            accent_color=palette.get("steel"),
        )
    )
    if not force_rows and not rfd_rows:
        return

    rows = [[p("Fuerza", "ProfTableHeader"), p("Valor", "ProfTableHeader"), p("RFD", "ProfTableHeader"), p("Valor RFD", "ProfTableHeader")]]
    for idx in range(max(len(force_rows), len(rfd_rows))):
        force_label, force_value = force_rows[idx] if idx < len(force_rows) else ("", "")
        rfd_label, rfd_value = rfd_rows[idx] if idx < len(rfd_rows) else ("", "")
        rows.append([p(force_label, "ProfTableCell"), p(force_value, "ProfTableCell"), p(rfd_label, "ProfTableCell"), p(rfd_value, "ProfTableCell")])
    table = TableClass(rows, colWidths=[34 * mm_unit, 42 * mm_unit, 34 * mm_unit, 64 * mm_unit], hAlign="LEFT")
    table.setStyle(
        TableStyleClass(
            [
                ("BOX", (0, 0), (-1, -1), 0.8, palette["line_dark"]),
                ("INNERGRID", (0, 0), (-1, -1), 0.35, palette["line"]),
                ("BACKGROUND", (0, 0), (-1, 0), palette["navy"]),
                ("TEXTCOLOR", (0, 0), (-1, 0), palette.get("card")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [palette["card"], palette.get("panel", palette["card"])]),
                ("LEFTPADDING", (0, 0), (-1, -1), 6),
                ("RIGHTPADDING", (0, 0), (-1, -1), 6),
                ("TOPPADDING", (0, 0), (-1, -1), 5),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ]
        )
    )
    story.append(SpacerClass(1, 2 * mm_unit))
    story.append(p("Detalle técnico compacto", "ProfCardTitle"))
    story.append(table)


def _build_professional_load_tolerance_payload(
    state: dict[str, pd.DataFrame | None],
    athlete: str,
    internal_load: dict[str, object],
) -> dict[str, object]:
    acwr_row = _latest_acwr_row(state, athlete)
    mono_row = _latest_mono_row(state, athlete)
    acwr_value = _coerce_float(acwr_row.get("ACWR_EWMA")) if acwr_row is not None else None
    acwr_zone = safe_value(acwr_row.get("Zona")) if acwr_row is not None else PDF_MISSING_TEXT
    monotony_value = _coerce_float(mono_row.get("Monotonia")) if mono_row is not None else None
    strain_value = _coerce_float(mono_row.get("Strain")) if mono_row is not None else None
    weekly_total = _coerce_float(internal_load.get("last_week_total"))
    weekly_change_pct = _coerce_float(internal_load.get("weekly_change_pct"))
    change_text = PDF_MISSING_TEXT
    if str(internal_load.get("analysis_scope") or "") == "current_week_partial":
        change_text = "Semana en curso incompleta"
    elif weekly_change_pct is not None:
        change_text = f"{weekly_change_pct:+.1f}%"
    elif _coerce_float(internal_load.get("weekly_change")) is not None:
        change_text = _professional_number_text(internal_load.get("weekly_change"), digits=0, unit="UA")

    risk_line = "Faltan datos para valorar la tolerancia de carga."
    if str(internal_load.get("analysis_scope") or "") == "current_week_partial":
        risk_line = "La semana actual sigue abierta; usar el acumulado parcial solo como seÃ±al operativa."
    elif acwr_value is not None and acwr_value > 1.5:
        risk_line = "La carga reciente luce alta y con riesgo de acumulaciÃ³n; revisar densidad, calendario y tolerancia."
    elif monotony_value is not None and monotony_value > 2.0:
        risk_line = "La carga reciente parece homogÃ©nea y con riesgo de acumulaciÃ³n por monotonÃ­a."
    elif acwr_value is not None and acwr_value < 0.8:
        risk_line = "La carga reciente luce baja/subdosificada respecto a la carga crÃ³nica."
    elif weekly_change_pct is not None and weekly_change_pct > 10:
        risk_line = "La carga reciente estÃ¡ en ascenso y conviene corroborar su tolerancia con wellness y disponibilidad."
    elif any(value is not None for value in (acwr_value, monotony_value, strain_value)):
        risk_line = "ACWR, monotonía y strain disponibles: lectura conservadora compatible con control de carga si el wellness y la calidad de sesión acompañan."
    elif weekly_total is not None:
        risk_line = "La carga reciente parece estable y compatible con seguimiento normal del bloque."

    rows = [
        ("Semana analizada", safe_value(internal_load.get("analysis_week_label"))),
        ("sRPE semanal", _professional_number_text(weekly_total, digits=0, unit="UA")),
        ("Cambio vs semana previa", change_text),
        ("Sesiones registradas", safe_value(internal_load.get("sessions_registered"))),
        ("ACWR EWMA", _professional_number_text(acwr_value, digits=2)),
        ("Zona ACWR", acwr_zone),
        ("MonotonÃ­a", _professional_number_text(monotony_value, digits=2)),
        ("Strain", _professional_number_text(strain_value, digits=0)),
    ]
    available_rows = [(label, value) for label, value in rows if value != PDF_MISSING_TEXT]
    state_label = "available" if available_rows else "missing"
    if state_label == "available" and str(internal_load.get("analysis_scope") or "") == "current_week_partial":
        state_label = "partial"
    return {
        "title": "Carga interna y tolerancia",
        "state": state_label,
        "message": "" if available_rows else "Faltan datos suficientes para consolidar la carga interna reciente.",
        "rows": available_rows,
        "risk_line": risk_line,
        "weekly_points": internal_load.get("weekly_points", []),
        "analysis_scope": internal_load.get("analysis_scope"),
    }


def _build_professional_wellness_availability_payload(
    state: dict[str, pd.DataFrame | None],
    athlete: str,
    training_context: dict[str, object],
    internal_load: dict[str, object],
    wellness_context: dict[str, object],
) -> dict[str, object]:
    completion = _professional_completion_snapshot(state, athlete)
    summary = wellness_context.get("last_week_summary", {}) if isinstance(wellness_context.get("last_week_summary"), dict) else {}
    scales = wellness_context.get("scales", {}) if isinstance(wellness_context.get("scales"), dict) else {}

    def _with_scale(value: object, key: str, *, digits: int = 1) -> str:
        numeric = _coerce_float(value)
        if numeric is None:
            return PDF_MISSING_TEXT
        scale = str(scales.get(key, ""))
        if scale == "h":
            return f"{numeric:.{digits}f} h"
        if scale.startswith("/"):
            return f"{numeric:.{digits}f}{scale}"
        return f"{numeric:.{digits}f}"

    score_value = _coerce_float(summary.get("score_mean"))
    score_label = PDF_MISSING_TEXT
    if score_value is not None:
        score_meta = _report_wellness_score_label(score_value)
        score_label = f"{float(score_meta['score']):.1f} / 5.0 ({score_meta['label']})"
    day_candidates = [
        _coerce_float(summary.get("days")),
        _coerce_float(summary.get("score_n")),
        _coerce_float(summary.get("sleep_n")),
        _coerce_float(summary.get("stress_n")),
        _coerce_float(summary.get("pain_n")),
    ]
    days_count = max([int(value) for value in day_candidates if value is not None] or [0])
    days_with_record = safe_value(days_count if days_count > 0 else None)

    rows = [
        ("Wellness score", score_label),
        ("SueÃ±o", _with_scale(summary.get("sleep_mean"), "sleep")),
        ("EstrÃ©s", _with_scale(summary.get("stress_mean"), "stress")),
        ("Dolor", _with_scale(summary.get("pain_mean"), "pain")),
        ("DÃ­as con registro", days_with_record),
        ("Adherencia formal", safe_value(completion.get("value"))),
    ]
    available_rows = [(label, value) for label, value in rows if value != PDF_MISSING_TEXT]

    weekly_change_pct = _coerce_float(internal_load.get("weekly_change_pct"))
    load_state = str(internal_load.get("state") or "").strip().lower()
    stress_mean = _coerce_float(summary.get("stress_mean"))
    pain_mean = _coerce_float(summary.get("pain_mean"))
    sleep_mean = _coerce_float(summary.get("sleep_mean"))
    compatibility = "Faltan datos suficientes para cruzar wellness, disponibilidad y carga."
    if wellness_context.get("state") == "partial":
        compatibility = "Datos parciales de wellness/disponibilidad: usar la lectura como seÃ±al preliminar y no como tendencia cerrada."
    elif wellness_context.get("state") == "available" and load_state == "missing":
        compatibility = "Wellness/disponibilidad disponible, pero sin carga interna reciente conviene usarlo solo como contexto parcial."
    elif weekly_change_pct is not None and weekly_change_pct > 10 and (
        _professional_wellness_high(stress_mean, scales.get("stress"))
        or _professional_wellness_high(pain_mean, scales.get("pain"))
        or (sleep_mean is not None and sleep_mean < 6.5)
    ):
        compatibility = "Carga alta con wellness/disponibilidad menos favorables: conviene aumentar la vigilancia del prÃ³ximo microciclo."
    elif wellness_context.get("state") == "available":
        compatibility = "Carga estable + wellness/disponibilidad relativamente estables: lectura compatible con tolerancia del bloque."

    quality_note = wellness_context.get("partial_message") if wellness_context.get("state") == "partial" else ""
    return {
        "title": "Wellness, disponibilidad y adherencia",
        "state": "available" if available_rows else "missing",
        "message": "" if available_rows else "Faltan datos de wellness/disponibilidad para este perÃ­odo.",
        "rows": available_rows,
        "compatibility": compatibility,
        "quality_note": str(quality_note or ""),
        "daily_points": wellness_context.get("daily_points", []),
        "weekly_points": wellness_context.get("weekly_points", []),
        "analysis_scope": wellness_context.get("analysis_scope"),
        "trend_allowed": bool(wellness_context.get("trend_allowed")),
    }


def _build_professional_exposure_payload(
    state: dict[str, pd.DataFrame | None],
    athlete: str,
    change_payload: dict[str, object] | None = None,
) -> dict[str, object]:
    prepared = state.get("prepared_raw_df")
    if prepared is None:
        prepared = prepare_raw_workouts_df(state.get("raw_df"))
    if prepared is None or prepared.empty:
        return {
            "title": "ExposiciÃ³n del bloque / contenido entrenado",
            "state": "missing",
            "message": "Faltan raw workouts suficientes para resumir la exposiciÃ³n del bloque.",
            "table": pd.DataFrame(columns=["EstÃ­mulo", "Dosis", "Sesiones", "Ejercicios clave"]),
            "active_groups": [],
            "dominant": [],
            "secondary": [],
            "low_or_absent": [],
            "summary_line": "",
            "context_link": "",
        }

    athlete_col = "Athlete" if "Athlete" in prepared.columns else "Name" if "Name" in prepared.columns else None
    athlete_df = prepared[_professional_athlete_mask(prepared[athlete_col], athlete)].copy() if athlete_col is not None else prepared.copy()
    if athlete_df.empty:
        return {
            "title": "ExposiciÃ³n del bloque / contenido entrenado",
            "state": "missing",
            "message": "No hay raw workouts visibles para este atleta.",
            "table": pd.DataFrame(columns=["EstÃ­mulo", "Dosis", "Sesiones", "Ejercicios clave"]),
            "active_groups": [],
            "dominant": [],
            "secondary": [],
            "low_or_absent": [],
            "summary_line": "",
            "context_link": "",
        }

    invalid = athlete_df.get("is_invalid", pd.Series(False, index=athlete_df.index)).fillna(False)
    untagged = athlete_df.get("is_untagged", pd.Series(False, index=athlete_df.index)).fillna(False)
    athlete_df = athlete_df[~invalid & ~untagged].copy()
    if athlete_df.empty:
        return {
            "title": "ExposiciÃ³n del bloque / contenido entrenado",
            "state": "missing",
            "message": "No hay raw workouts clasificados para resumir la exposiciÃ³n del bloque.",
            "table": pd.DataFrame(columns=["EstÃ­mulo", "Dosis", "Sesiones", "Ejercicios clave"]),
            "active_groups": [],
            "dominant": [],
            "secondary": [],
            "low_or_absent": [],
            "summary_line": "",
            "context_link": "",
        }

    exercise_col = "Exercise" if "Exercise" in athlete_df.columns else "Exercise Name" if "Exercise Name" in athlete_df.columns else None
    rows: list[dict[str, object]] = []
    ranking_rows: list[dict[str, object]] = []
    for spec in PROFESSIONAL_EXPOSURE_CATEGORY_SPECS:
        subset = athlete_df[athlete_df["stimulus_category"].isin(spec["categories"])].copy()
        sessions = 0
        if not subset.empty and "Assigned Date" in subset.columns:
            sessions = int(pd.to_datetime(subset["Assigned Date"], errors="coerce").dropna().dt.normalize().nunique())
        value_col = str(spec["value_col"])
        value_sum = None
        if not subset.empty and value_col in subset.columns:
            value_sum = _coerce_float(pd.to_numeric(subset[value_col], errors="coerce").sum(min_count=1))
        top_exercises = []
        if not subset.empty and exercise_col:
            top_exercises = (
                subset[exercise_col]
                .dropna()
                .astype(str)
                .str.strip()
                .replace("", pd.NA)
                .dropna()
                .value_counts()
                .head(3)
                .index
                .tolist()
            )
        if sessions > 0 or value_sum is not None:
            dose_text = _professional_number_text(value_sum, digits=0, unit=str(spec["unit"])) if value_sum is not None else PDF_MISSING_TEXT
            rows.append(
                {
                    "EstÃ­mulo": str(spec["title"]),
                    "Dosis": dose_text,
                    "Sesiones": str(sessions) if sessions > 0 else PDF_MISSING_TEXT,
                    "Ejercicios clave": _professional_join_labels(top_exercises),
                }
            )
            ranking_rows.append(
                {
                    "title": str(spec["title"]),
                    "sessions": sessions,
                    "value": value_sum or 0.0,
                }
            )

    active_titles = [row["title"] for row in sorted(ranking_rows, key=lambda item: (-int(item["sessions"]), -float(item["value"]), item["title"]))]
    active_titles = [_professional_visible_metric_text(title) for title in active_titles]
    dominant = active_titles[:2]
    secondary = active_titles[2:4]
    low_or_absent: list[str] = []
    if len(ranking_rows) >= 2:
        lowest_row = sorted(ranking_rows, key=lambda item: (int(item["sessions"]), float(item["value"]), item["title"]))[0]
        low_or_absent = [str(lowest_row["title"])]
    low_or_absent = [_professional_visible_metric_text(title) for title in low_or_absent]

    change_payload = change_payload or {}
    improved_labels = " ".join(change_payload.get("improvements", []))
    context_link = "La exposiciÃ³n visible ayuda a contextualizar el perfil actual, pero no reemplaza la lectura de calidad del dato."
    dominant_text = _professional_join_labels(dominant, fallback="")
    if dominant_text and "Fuerza con carga" in dominant_text and any(token in improved_labels for token in ("CMJ", "SJ", "IMTP")):
        context_link = "El bloque tuvo predominio de fuerza con carga y es compatible con la seÃ±al actual de fuerza/salida vertical."
    elif dominant_text and "Pliometría y aterrizajes" in dominant_text and any(token in improved_labels for token in ("DRI", "RSI", "Contact")):
        context_link = "El bloque tuvo predominio reactivo y es compatible con la seÃ±al actual de reactividad/contacto."
    elif dominant_text and "Movilidad y prehab" in dominant_text:
        context_link = "El bloque visible parece orientado a soporte/prehab; conviene no sobreinterpretar cambios como respuesta a un bloque principal de rendimiento."

    summary_line = ""
    if dominant:
        summary_line = f"EstÃ­mulos dominantes: {_professional_join_labels(dominant)}."
        if secondary:
            summary_line = f"{summary_line} EstÃ­mulos secundarios: {_professional_join_labels(secondary)}."

    table = pd.DataFrame(rows, columns=["EstÃ­mulo", "Dosis", "Sesiones", "Ejercicios clave"])
    if not table.empty:
        table = table.apply(lambda column: column.map(_professional_visible_metric_text))
    if not table.empty:
        table.columns = [_professional_visible_metric_text(column) for column in table.columns]
    summary_line = _professional_visible_metric_text(summary_line)
    context_link = _professional_visible_metric_text(context_link)
    state_label = "available" if len(rows) >= 2 else "partial" if rows else "missing"
    return {
        "title": "ExposiciÃ³n del bloque / contenido entrenado",
        "state": state_label,
        "message": "" if rows else "Faltan raw workouts suficientes para resumir la exposiciÃ³n del bloque.",
        "table": table,
        "active_groups": active_titles,
        "dominant": dominant,
        "secondary": secondary,
        "low_or_absent": low_or_absent,
        "summary_line": summary_line,
        "context_link": context_link,
    }


def _professional_data_confidence_label(
    evaluation_state: str,
    training_context: dict[str, object],
    wellness_payload: dict[str, object],
    assessment_interval_warning: str,
) -> str:
    if (
        str(evaluation_state).strip().lower() == "available"
        and training_context.get("state") == "available"
        and wellness_payload.get("state") == "available"
        and not assessment_interval_warning
    ):
        return "Alta"
    if str(evaluation_state).strip().lower() in {"available", "partial"}:
        return "Media"
    return "Baja"


def _build_professional_integrated_decision_payload(
    state: dict[str, pd.DataFrame | None],
    athlete: str,
    *,
    evaluation_state: str,
    assessment_interval_warning: str,
    composite_payload: dict[str, object],
    change_payload: dict[str, object],
    load_payload: dict[str, object],
    wellness_payload: dict[str, object],
    exposure_payload: dict[str, object],
    training_context: dict[str, object],
) -> dict[str, object]:
    good_confidence: list[str] = []
    probable: list[str] = []
    unknown: list[str] = []
    decision_practical: list[str] = []
    monitor: list[str] = []

    feedback = composite_payload.get("feedback", {}) if isinstance(composite_payload.get("feedback"), dict) else {}
    has_clear_lagging = _professional_has_clear_lagging_variable(feedback)
    high_text = _professional_strip_terminal_period(feedback.get("high") or "sin una cualidad claramente dominante")
    low_text = _professional_strip_terminal_period(feedback.get("low") or "sin un rezago dominante")
    high_norm = _professional_normalized_text(high_text)
    load_rows = _professional_rows_map(load_payload.get("rows"))
    wellness_rows = _professional_rows_map(wellness_payload.get("rows"))
    dominant = exposure_payload.get("dominant", []) if isinstance(exposure_payload.get("dominant"), list) else []
    low_or_absent = exposure_payload.get("low_or_absent", []) if isinstance(exposure_payload.get("low_or_absent"), list) else []
    improvements = change_payload.get("improvements", []) if isinstance(change_payload.get("improvements"), list) else []
    declines = change_payload.get("declines", []) if isinstance(change_payload.get("declines"), list) else []
    no_previous = change_payload.get("no_previous", []) if isinstance(change_payload.get("no_previous"), list) else []
    zone_text = load_rows.get("Zona ACWR", "")
    acwr_text = load_rows.get("ACWR EWMA", "")
    monotony_text = load_rows.get("Monotonía", "")
    load_fact = _professional_join_labels(
        [
            f"ACWR {acwr_text}" if _has_text(acwr_text) else "",
            f"zona {zone_text}" if _has_text(zone_text) else "",
            f"monotonía {monotony_text}" if _has_text(monotony_text) else "",
        ],
        fallback="",
    )
    wellness_score = wellness_rows.get("Wellness score", "")
    wellness_days = wellness_rows.get("Días con registro", "")
    load_risk_norm = _professional_normalized_text(load_payload.get("risk_line", ""))
    wellness_compatibility_norm = _professional_normalized_text(wellness_payload.get("compatibility", ""))

    if composite_payload.get("state") != "missing":
        if dominant and _has_text(high_text):
            good_confidence.append(
                f"El perfil actual concentra la mejor señal en {high_text} y el bloque mostró más exposición en {_professional_join_labels(dominant[:2])}; ese cruce merece sostenerse."
            )
        elif _has_text(high_text) and _has_text(load_fact):
            good_confidence.append(
                f"La mejor expresión hoy aparece en {high_text} con {load_fact}; el contexto acompaña una lectura utilizable del perfil actual."
            )
        elif _has_text(high_text):
            good_confidence.append(
                f"El perfil actual concentra la mejor señal en {high_text} y no muestra un dato aislado sin apoyo del resto de la lectura."
            )
    if improvements:
        if _has_text(load_fact):
            good_confidence.append(
                f"Las mejoras en {_professional_join_labels(improvements[:3])} aparecen con {load_fact}; hoy conviene consolidarlas antes de cambiar el foco."
            )
        else:
            good_confidence.append(
                f"Las mejoras en {_professional_join_labels(improvements[:3])} ya dan una señal de progreso que merece continuidad."
            )
    if declines:
        if str(wellness_payload.get("state") or "").strip().lower() == "partial":
            good_confidence.append(
                f"Las caídas en {_professional_join_labels(declines[:3])} aparecen con wellness parcial; pesan más como señal a confirmar que como conclusión cerrada."
            )
        elif _has_text(load_fact) and not any(token in load_risk_norm for token in ("riesgo", "subdosificada", "alta")):
            good_confidence.append(
                f"Las caídas en {_professional_join_labels(declines[:3])} no coinciden con una carga claramente fuera de rango; conviene recontrolarlas antes de sobrerreaccionar."
            )
        elif _has_text(wellness_score):
            good_confidence.append(
                f"Las caídas en {_professional_join_labels(declines[:3])} conviven con {wellness_score}; la interpretación todavía necesita confirmación."
            )
    if composite_payload.get("state") != "missing" and not has_clear_lagging and _has_text(load_fact):
        good_confidence.append(
            f"No aparece un rezago dominante y la carga reciente se resume en {load_fact}; el contexto favorece sostener una prioridad acotada."
        )

    if dominant:
        dominant_text = _professional_join_labels(dominant[:2])
        if "fuerza con carga" in _professional_normalized_text(dominant_text) and any(token in f"{high_norm} {_professional_normalized_text(_professional_join_labels(improvements))}" for token in ("sj", "cmj", "imtp", "fuerza")):
            probable.append("El predominio de fuerza con carga es compatible con la señal actual de salida vertical y fuerza base.")
        elif any(token in _professional_normalized_text(dominant_text) for token in ("pliometr", "aterriz", "react")) and any(token in f"{high_norm} {_professional_normalized_text(_professional_join_labels(improvements))}" for token in ("dj", "dri", "rsi", "contact", "eur")):
            probable.append("El predominio reactivo del bloque es compatible con la señal actual de reactividad y contacto.")
        elif _has_text(high_text):
            probable.append(
                f"La exposición dominante en {dominant_text} probablemente esté empujando la señal visible en {high_text}."
            )
    if low_or_absent:
        probable.append(
            f"La baja exposición en {_professional_join_labels(low_or_absent[:2])} no muestra impacto visible todavía, pero debe vigilarse si empieza a limitar el perfil."
        )
    elif training_context.get("state") != "missing" and exposure_payload.get("state") == "missing":
        probable.append(
            "La asistencia y la carga visibles ayudan a ubicar el contexto, pero no reemplazan la exposición por estímulos del bloque."
        )

    if assessment_interval_warning:
        unknown.append("Con menos de 6-8 semanas entre evaluaciones no se puede atribuir el cambio con seguridad al bloque.")
    if str(evaluation_state).strip().lower() == "partial":
        unknown.append("La batería parcial impide cerrar el perfil completo y puede dejar cualidades subrepresentadas.")
    if load_payload.get("state") == "missing":
        unknown.append("Sin carga interna reciente no se puede definir si la tolerancia del bloque fue realmente estable.")
    if wellness_payload.get("state") == "missing":
        unknown.append("Falta wellness/disponibilidad reciente; evitar conclusiones fuertes sobre tolerancia del bloque.")
    if wellness_payload.get("state") == "partial":
        if _has_text(wellness_days):
            unknown.append(f"No se puede cerrar tolerancia real porque wellness tiene solo {wellness_days.lower()}.")
        else:
            unknown.append("No se puede cerrar tolerancia real porque wellness todavía tiene pocos registros.")
    if no_previous:
        unknown.append(
            f"Sin referencia previa en {_professional_join_labels(no_previous[:3])} no se puede afirmar dirección de cambio en esas variables."
        )
    if exposure_payload.get("state") == "missing":
        unknown.append("Falta exposición por estímulos suficiente para vincular el resultado a un contenido concreto del bloque.")

    low_text = _professional_normalized_text(feedback.get("low", ""))
    load_risk_text = str(load_payload.get("risk_line") or "").casefold()
    wellness_risk_text = str(wellness_payload.get("compatibility") or "").casefold()
    if "riesgo" in load_risk_text or "vigilancia" in wellness_risk_text or "menos favorables" in wellness_risk_text:
        decision_practical.append("Ajustar densidad/volumen del prÃ³ximo microciclo y sostener solo el estÃ­mulo prioritario hasta confirmar tolerancia.")
    elif has_clear_lagging and "fuerza" in low_text:
        decision_practical.append("Mantener la cualidad dominante y priorizar fuerza base/transferencia en el prÃ³ximo bloque.")
    elif has_clear_lagging and any(token in low_text for token in ("dj height", "dj", "react", "dri", "contact")):
        decision_practical.append("Sostener la base actual y priorizar la progresión reactiva para mejorar la expresión en DJ, con foco en calidad de contacto, stiffness y tolerancia progresiva.")
    else:
        decision_practical.append("Sostener la cualidad mejor expresada y traducir la variable rezagada a una prioridad concreta de entrenamiento.")

    if change_payload.get("declines"):
        decision_practical.append("Confirmar la seÃ±al en la siguiente ventana antes de atribuirla por completo a adaptaciÃ³n o fatiga.")
    if not has_clear_lagging:
        decision_practical = [
            line
            for line in decision_practical
            if "variable mas rezagada" not in _professional_normalized_text(line)
        ]
        decision_practical.append("Sostener el perfil actual, consolidar calidad técnica y monitorear evolución sin abrir nuevos focos innecesarios.")

    if exposure_payload.get("dominant"):
        decision_practical.append(f"Usar como base del siguiente bloque los estÃ­mulos dominantes ya visibles: {_professional_join_labels(exposure_payload.get('dominant', []))}.")

    if declines:
        monitor.append(
            f"Monitorear {_professional_join_labels(declines[:3])} para distinguir si la caída persiste o se normaliza con el siguiente microciclo."
        )
    if str(wellness_payload.get("state") or "").strip().lower() == "partial":
        monitor.append("Completar registros de wellness suficientes antes de mover la carga; con pocos días la tolerancia real queda abierta.")
    elif str(wellness_payload.get("state") or "").strip().lower() == "available" and any(token in wellness_compatibility_norm for token in ("menos favorables", "vigilancia")):
        monitor.append("Seguir de cerca sueño, estrés y dolor porque hoy condicionan la lectura de tolerancia del bloque.")
    if _has_text(load_fact) and any(token in load_risk_norm for token in ("riesgo", "homogenea", "ascenso", "subdosificada")):
        metric_watch = _professional_join_labels(declines[:2], fallback=low_text if has_clear_lagging else high_text)
        monitor.append(
            f"Monitorear {metric_watch} si la carga sigue en {load_fact}, para ver si cambia la respuesta neuromuscular."
        )
    if not has_clear_lagging:
        monitor.append(PROFESSIONAL_NO_CLEAR_LAGGING_MONITOR)
    if low_or_absent:
        monitor.append(
            f"Monitorear si la baja exposición en {_professional_join_labels(low_or_absent[:2])} empieza a generar una limitación futura del perfil."
        )
    metric_recheck = _professional_join_labels(declines[:2], fallback=low_text if has_clear_lagging else high_text)
    if _has_text(metric_recheck):
        monitor.append(f"Recontrolar {metric_recheck} antes de 6-8 semanas si el contexto competitivo necesita una verificación puntual.")

    return {
        "title": "InterpretaciÃ³n integrada profesional",
        "state": "available",
        "good_confidence": list(dict.fromkeys([line for line in good_confidence if line.strip()]))[:4],
        "probable": list(dict.fromkeys([line for line in probable if line.strip()]))[:3],
        "unknown": list(dict.fromkeys([line for line in unknown if line.strip()]))[:4],
        "decision_practical": list(dict.fromkeys([line for line in decision_practical if line.strip()]))[:3],
        "monitor": list(dict.fromkeys([line for line in monitor if line.strip()]))[:4],
    }


def _professional_decision_practical_lines(
    state: dict[str, pd.DataFrame | None],
    athlete: str,
    *,
    evaluation_state: str,
    composite_payload: dict[str, object],
    change_payload: dict[str, object],
    load_payload: dict[str, object],
    wellness_payload: dict[str, object],
    exposure_payload: dict[str, object],
    has_clear_lagging: bool,
    feedback: dict[str, object],
) -> list[str]:
    summary_df = build_executive_summary_df(state, athlete, "profe")
    row = summary_df.iloc[0] if not summary_df.empty else pd.Series(dtype=object)
    context = _focus_metric_context(row)
    eval_available = _row_has_eval_data(row)
    has_evaluation_reference = eval_available or bool(feedback) or str(evaluation_state).strip().lower() in {"available", "partial"}
    profile = context["profile_norm"]
    load_phrase = context["load_phrase"]
    low_label = _professional_strip_terminal_period(feedback.get("low", ""))
    low_norm = _professional_normalized_text(low_label)
    high_label = _professional_strip_terminal_period(feedback.get("high") or composite_payload.get("summary_line") or "")
    declines = change_payload.get("declines", []) if isinstance(change_payload.get("declines"), list) else []
    improvements = change_payload.get("improvements", []) if isinstance(change_payload.get("improvements"), list) else []
    dominant = exposure_payload.get("dominant", []) if isinstance(exposure_payload.get("dominant"), list) else []
    low_or_absent = exposure_payload.get("low_or_absent", []) if isinstance(exposure_payload.get("low_or_absent"), list) else []
    load_risk_text = _professional_normalized_text(load_payload.get("risk_line", ""))
    wellness_risk_text = _professional_normalized_text(wellness_payload.get("compatibility", ""))
    named_metrics = [metric for metric in (context["dri"], context["eur"], context["cmj"], context["imtp"]) if _has_text(metric)]
    lead_metric = named_metrics[0] if named_metrics else (_professional_visible_metric_text(high_label) if _has_text(high_label) else "")
    support_metric = named_metrics[1] if len(named_metrics) > 1 else (context["cmj"] or context["imtp"] or lead_metric)
    lines: list[str] = []

    if not has_evaluation_reference:
        if _has_text(load_phrase):
            lines.append(
                f"Operar con {load_phrase} y completar CMJ, SJ, DJ e IMTP antes de orientar la cualidad física principal del bloque."
            )
        else:
            lines.append("Completar CMJ, SJ, DJ e IMTP antes de orientar la cualidad física principal del bloque.")
        if _has_text(context["wellness"]):
            lines.append(
                f"Cruzar {context['wellness']} con la carga diaria para operar con seguridad hasta la próxima evaluación."
            )
        else:
            lines.append("La próxima evaluación debe confirmar qué cualidad física conviene priorizar.")
        if dominant:
            lines.append(
                f"Mantener visibles los estímulos del bloque ({_professional_join_labels(dominant)}) hasta contar con la próxima evaluación."
            )
        return [_professional_visible_metric_text(line) for line in lines if _has_text(line)]

    if has_clear_lagging:
        if any(token in low_norm for token in ("dj height", "dj", "react", "dri", "contact")):
            lines.append(
                f"Priorizar {low_label} traduciéndolo a progresión reactiva para mejorar la expresión en DJ, con foco en calidad de contacto, stiffness y tolerancia progresiva."
            )
            lines.append(
                f"Confirmar en la próxima evaluación que la expresión en DJ mejora sin perder {support_metric or high_label or 'la cualidad dominante actual'}."
            )
        elif any(token in low_norm for token in ("fuerza", "imtp")):
            lines.append(
                f"Priorizar {low_label} con trabajo de fuerza base y transferencia, sosteniendo {lead_metric or high_label or context['cmj'] or 'la salida vertical actual'}."
            )
            lines.append(
                f"Confirmar en la próxima evaluación que {low_label} mejora y que {support_metric or 'CMJ'} acompaña sin caer."
            )
        else:
            lines.append(
                f"Priorizar {low_label} y ordenar el bloque para recuperarlo sin perder {lead_metric or high_label or context['cmj'] or 'la cualidad dominante actual'}."
            )
            lines.append(f"Confirmar en la próxima evaluación si {low_label} se recupera de forma consistente.")
        if _has_text(load_phrase) and any(token in load_risk_text for token in ("riesgo", "precaucion", "alto")):
            lines.append(f"No escalar densidad ni volumen mientras {load_phrase} siga condicionando la tolerancia del bloque.")
        elif dominant:
            lines.append(
                f"Cruzar esta prioridad con los estímulos dominantes ya visibles ({_professional_join_labels(dominant)})."
            )
        return [_professional_visible_metric_text(line) for line in lines if _has_text(line)]

    if declines:
        decline_text = _professional_join_labels(declines)
        if _has_text(load_phrase):
            lines.append(f"Priorizar recuperar {decline_text} mientras {load_phrase} vuelve a una zona más estable.")
        else:
            lines.append(f"Priorizar recuperar {decline_text} antes de progresar hacia nuevas demandas del bloque.")
        lines.append(f"La señal práctica es que {decline_text} deje de caer en la próxima evaluación.")
    elif "react" in profile:
        reactive_metrics = [metric for metric in (context["dri"], context["eur"], context["cmj"]) if _has_text(metric)]
        if reactive_metrics:
            if _has_text(load_phrase) and any(token in load_risk_text for token in ("riesgo", "precaucion", "alto")):
                lines.append(
                    f"Sostener {reactive_metrics[0]}{f' y {reactive_metrics[1]}' if len(reactive_metrics) > 1 else ''} sin subir el volumen pliométrico mientras {load_phrase} siga alta."
                )
            else:
                lines.append(
                    f"Sostener {reactive_metrics[0]}{f' y {reactive_metrics[1]}' if len(reactive_metrics) > 1 else ''} como base del siguiente bloque."
                )
            lines.append(
                f"La señal de que funciona es que {reactive_metrics[-1]} y {context['cmj'] or reactive_metrics[0]} sigan acompañando en la próxima evaluación."
            )
    elif "poca base" in profile or "fuerza" in profile or "base" in profile:
        anchor = context["imtp"] or context["cmj"] or "fuerza base"
        if _has_text(load_phrase) and (_professional_normalized_text(load_phrase).find("subcarga") >= 0 or (_coerce_float(row.get("ACWR EWMA")) or 0) < 0.8):
            lines.append(f"Recuperar continuidad de carga y sostener {anchor} antes de progresar hacia más reactividad.")
        else:
            lines.append(f"Sostener {anchor} y transferirlo con cuidado hacia {context['cmj'] or 'la salida vertical actual'}.")
        lines.append(f"La señal de que funciona es que {anchor} acompañe mejor en la próxima evaluación.")
    elif dominant:
        dominant_text = _professional_join_labels(dominant)
        anchor_metric = lead_metric or context["cmj"] or context["imtp"] or "el perfil actual"
        lines.append(f"Sostener los estímulos dominantes del bloque ({dominant_text}) y verificar si siguen acompañando {anchor_metric}.")
        lines.append(
            f"La señal de que funciona es que {anchor_metric} se mantenga mientras rotamos o compensamos lo menos expuesto del bloque."
            if low_or_absent else
            f"La señal de que funciona es que {anchor_metric} se mantenga en la próxima evaluación."
        )
    else:
        lines.append(f"Sostener {lead_metric or context['cmj'] or context['imtp'] or 'la referencia física actual'} y usarlo como guía del siguiente bloque.")
        lines.append(
            f"La señal de que funciona es que {support_metric or lead_metric or 'la referencia física actual'} acompañe mejor en la próxima evaluación."
        )

    if low_or_absent:
        lines.append(f"Revisar si conviene rotar o compensar la baja exposición en {_professional_join_labels(low_or_absent[:2])}.")
    elif improvements:
        lines.append(f"Usar como apoyo del siguiente bloque las mejoras recientes en {_professional_join_labels(improvements[:2])}.")
    elif _has_text(context["wellness"]) and any(token in wellness_risk_text for token in ("vigilancia", "menos favorables", "baja")):
        lines.append(f"Monitorear de cerca {context['wellness']} antes de subir la demanda del próximo microciclo.")

    return [_professional_visible_metric_text(line) for line in lines if _has_text(line)]


def _clip_signal(text: str, max_chars: int = 120) -> str:
    clean = re.sub(r"\s+", " ", _professional_visible_metric_text(text).strip())
    if not clean:
        return ""
    if len(clean) <= max_chars:
        return clean
    return textwrap.shorten(clean, width=max_chars, placeholder="…")


def _professional_rows_map(rows: object) -> dict[str, str]:
    mapped: dict[str, str] = {}
    if not isinstance(rows, list):
        return mapped
    for item in rows:
        if not isinstance(item, (list, tuple)) or len(item) < 2:
            continue
        label = _professional_visible_metric_text(item[0]).strip()
        value = _professional_visible_metric_text(item[1]).strip()
        if _has_text(label) and _has_text(value):
            mapped[label] = value
    return mapped


def _build_professional_integrated_decision_payload(
    state: dict[str, pd.DataFrame | None],
    athlete: str,
    *,
    evaluation_state: str,
    assessment_interval_warning: str,
    composite_payload: dict[str, object],
    change_payload: dict[str, object],
    load_payload: dict[str, object],
    wellness_payload: dict[str, object],
    exposure_payload: dict[str, object],
    training_context: dict[str, object],
) -> dict[str, object]:
    good_confidence: list[str] = []
    probable: list[str] = []
    unknown: list[str] = []
    decision_practical: list[str] = []
    monitor: list[str] = []

    feedback = composite_payload.get("feedback", {}) if isinstance(composite_payload.get("feedback"), dict) else {}
    has_clear_lagging = _professional_has_clear_lagging_variable(feedback)
    high_text = _professional_strip_terminal_period(feedback.get("high") or "sin una cualidad claramente dominante")
    low_text = _professional_strip_terminal_period(feedback.get("low") or "sin un rezago dominante")
    high_norm = _professional_normalized_text(high_text)
    load_rows = _professional_rows_map(load_payload.get("rows"))
    wellness_rows = _professional_rows_map(wellness_payload.get("rows"))
    dominant = exposure_payload.get("dominant", []) if isinstance(exposure_payload.get("dominant"), list) else []
    low_or_absent = exposure_payload.get("low_or_absent", []) if isinstance(exposure_payload.get("low_or_absent"), list) else []
    improvements = change_payload.get("improvements", []) if isinstance(change_payload.get("improvements"), list) else []
    declines = change_payload.get("declines", []) if isinstance(change_payload.get("declines"), list) else []
    no_previous = change_payload.get("no_previous", []) if isinstance(change_payload.get("no_previous"), list) else []
    zone_text = load_rows.get("Zona ACWR", "")
    acwr_text = load_rows.get("ACWR EWMA", "")
    monotony_text = load_rows.get("Monotonía", "")
    load_fact = _professional_join_labels(
        [
            f"ACWR {acwr_text}" if _has_text(acwr_text) else "",
            f"zona {zone_text}" if _has_text(zone_text) else "",
            f"monotonía {monotony_text}" if _has_text(monotony_text) else "",
        ],
        fallback="",
    )
    wellness_score = wellness_rows.get("Wellness score", "")
    wellness_days = wellness_rows.get("Días con registro", "")
    load_risk_norm = _professional_normalized_text(load_payload.get("risk_line", ""))
    wellness_compatibility_norm = _professional_normalized_text(wellness_payload.get("compatibility", ""))

    if composite_payload.get("state") != "missing":
        if dominant and _has_text(high_text):
            good_confidence.append(
                f"El perfil actual concentra la mejor señal en {high_text} y el bloque mostró más exposición en {_professional_join_labels(dominant[:2])}; ese cruce merece sostenerse."
            )
        elif _has_text(high_text) and _has_text(load_fact):
            good_confidence.append(
                f"La mejor expresión hoy aparece en {high_text} con {load_fact}; el contexto acompaña una lectura utilizable del perfil actual."
            )
        elif _has_text(high_text):
            good_confidence.append(
                f"El perfil actual concentra la mejor señal en {high_text} y no muestra un dato aislado sin apoyo del resto de la lectura."
            )
    if improvements:
        if _has_text(load_fact):
            good_confidence.append(
                f"Las mejoras en {_professional_join_labels(improvements[:3])} aparecen con {load_fact}; hoy conviene consolidarlas antes de cambiar el foco."
            )
        else:
            good_confidence.append(
                f"Las mejoras en {_professional_join_labels(improvements[:3])} ya dan una señal de progreso que merece continuidad."
            )
    if declines:
        if str(wellness_payload.get("state") or "").strip().lower() == "partial":
            good_confidence.append(
                f"Las caídas en {_professional_join_labels(declines[:3])} aparecen con wellness parcial; pesan más como señal a confirmar que como conclusión cerrada."
            )
        elif _has_text(load_fact) and not any(token in load_risk_norm for token in ("riesgo", "subdosificada", "alta")):
            good_confidence.append(
                f"Las caídas en {_professional_join_labels(declines[:3])} no coinciden con una carga claramente fuera de rango; conviene recontrolarlas antes de sobrerreaccionar."
            )
        elif _has_text(wellness_score):
            good_confidence.append(
                f"Las caídas en {_professional_join_labels(declines[:3])} conviven con {wellness_score}; la interpretación todavía necesita confirmación."
            )
    if composite_payload.get("state") != "missing" and not has_clear_lagging and _has_text(load_fact):
        good_confidence.append(
            f"No aparece un rezago dominante y la carga reciente se resume en {load_fact}; el contexto favorece sostener una prioridad acotada."
        )

    if dominant:
        dominant_text = _professional_join_labels(dominant[:2])
        if "fuerza con carga" in _professional_normalized_text(dominant_text) and any(token in f"{high_norm} {_professional_normalized_text(_professional_join_labels(improvements))}" for token in ("sj", "cmj", "imtp", "fuerza")):
            probable.append("El predominio de fuerza con carga es compatible con la señal actual de salida vertical y fuerza base.")
        elif any(token in _professional_normalized_text(dominant_text) for token in ("pliometr", "aterriz", "react")) and any(token in f"{high_norm} {_professional_normalized_text(_professional_join_labels(improvements))}" for token in ("dj", "dri", "rsi", "contact", "eur")):
            probable.append("El predominio reactivo del bloque es compatible con la señal actual de reactividad y contacto.")
        elif _has_text(high_text):
            probable.append(
                f"La exposición dominante en {dominant_text} probablemente esté empujando la señal visible en {high_text}."
            )
    if low_or_absent:
        probable.append(
            f"La baja exposición en {_professional_join_labels(low_or_absent[:2])} no muestra impacto visible todavía, pero debe vigilarse si empieza a limitar el perfil."
        )
    elif training_context.get("state") != "missing" and exposure_payload.get("state") == "missing":
        probable.append(
            "La asistencia y la carga visibles ayudan a ubicar el contexto, pero no reemplazan la exposición por estímulos del bloque."
        )

    if assessment_interval_warning:
        unknown.append("Con menos de 6-8 semanas entre evaluaciones no se puede atribuir el cambio con seguridad al bloque.")
    if str(evaluation_state).strip().lower() == "partial":
        unknown.append("La batería parcial impide cerrar el perfil completo y puede dejar cualidades subrepresentadas.")
    if load_payload.get("state") == "missing":
        unknown.append("Sin carga interna reciente no se puede definir si la tolerancia del bloque fue realmente estable.")
    if wellness_payload.get("state") == "missing":
        unknown.append("Falta wellness/disponibilidad reciente; evitar conclusiones fuertes sobre tolerancia del bloque.")
    if wellness_payload.get("state") == "partial":
        if _has_text(wellness_days):
            unknown.append(f"No se puede cerrar tolerancia real porque wellness tiene solo {wellness_days.lower()}.")
        else:
            unknown.append("No se puede cerrar tolerancia real porque wellness todavía tiene pocos registros.")
    if no_previous:
        unknown.append(
            f"Sin referencia previa en {_professional_join_labels(no_previous[:3])} no se puede afirmar dirección de cambio en esas variables."
        )
    if exposure_payload.get("state") == "missing":
        unknown.append("Falta exposición por estímulos suficiente para vincular el resultado a un contenido concreto del bloque.")

    decision_practical.extend(
        _professional_decision_practical_lines(
            state,
            athlete,
            evaluation_state=evaluation_state,
            composite_payload=composite_payload,
            change_payload=change_payload,
            load_payload=load_payload,
            wellness_payload=wellness_payload,
            exposure_payload=exposure_payload,
            has_clear_lagging=has_clear_lagging,
            feedback=feedback,
        )
    )

    if declines:
        monitor.append(
            f"Monitorear {_professional_join_labels(declines[:3])} para distinguir si la caída persiste o se normaliza con el siguiente microciclo."
        )
    if str(wellness_payload.get("state") or "").strip().lower() == "partial":
        monitor.append("Completar registros de wellness suficientes antes de mover la carga; con pocos días la tolerancia real queda abierta.")
    elif str(wellness_payload.get("state") or "").strip().lower() == "available" and any(token in wellness_compatibility_norm for token in ("menos favorables", "vigilancia")):
        monitor.append("Seguir de cerca sueño, estrés y dolor porque hoy condicionan la lectura de tolerancia del bloque.")
    if _has_text(load_fact) and any(token in load_risk_norm for token in ("riesgo", "homogenea", "ascenso", "subdosificada")):
        metric_watch = _professional_join_labels(declines[:2], fallback=low_text if has_clear_lagging else high_text)
        monitor.append(
            f"Monitorear {metric_watch} si la carga sigue en {load_fact}, para ver si cambia la respuesta neuromuscular."
        )
    if not has_clear_lagging:
        monitor.append(PROFESSIONAL_NO_CLEAR_LAGGING_MONITOR)
    if low_or_absent:
        monitor.append(
            f"Monitorear si la baja exposición en {_professional_join_labels(low_or_absent[:2])} empieza a generar una limitación futura del perfil."
        )
    metric_recheck = _professional_join_labels(declines[:2], fallback=low_text if has_clear_lagging else high_text)
    if _has_text(metric_recheck):
        monitor.append(f"Recontrolar {metric_recheck} antes de 6-8 semanas si el contexto competitivo necesita una verificación puntual.")

    return {
        "title": "Interpretación integrada profesional",
        "state": "available",
        "good_confidence": list(dict.fromkeys([_professional_visible_metric_text(line) for line in good_confidence if line.strip()]))[:4],
        "probable": list(dict.fromkeys([_professional_visible_metric_text(line) for line in probable if line.strip()]))[:3],
        "unknown": list(dict.fromkeys([_professional_visible_metric_text(line) for line in unknown if line.strip()]))[:4],
        "decision_practical": list(dict.fromkeys([_professional_visible_metric_text(line) for line in decision_practical if line.strip()]))[:3],
        "monitor": list(dict.fromkeys([_professional_visible_metric_text(line) for line in monitor if line.strip()]))[:4],
    }


def _build_professional_action_plan_payload(
    state: dict[str, pd.DataFrame | None],
    athlete: str,
    *,
    evaluation_state: str,
    composite_payload: dict[str, object],
    change_payload: dict[str, object],
    integrated_payload: dict[str, object],
) -> dict[str, object]:
    feedback = composite_payload.get("feedback", {}) if isinstance(composite_payload.get("feedback"), dict) else {}
    has_clear_lagging = _professional_has_clear_lagging_variable(feedback)
    high_text = _professional_strip_terminal_period(feedback.get("high", ""))
    high_norm = _professional_normalized_text(high_text)
    low_text = _professional_normalized_text(feedback.get("low", ""))
    decline_text = _professional_join_labels(change_payload.get("declines", []), fallback="").casefold()
    maintain: list[str] = []
    if any(token in high_norm for token in ("sj", "dj", "dri", "react", "tiempo de contacto")):
        maintain.append("Mantener volumen y calidad de trabajo concéntrico/reactivo sin subir densidad pliométrica.")
    elif any(token in high_norm for token in ("imtp", "fuerza")):
        maintain.append("Mantener frecuencia de fuerza máxima/isométricos, mínimo 2 exposiciones semanales si el microciclo lo permite.")
    elif any(token in high_norm for token in ("cmj", "eur")):
        maintain.append("Mantener potencia vertical y estrategia elástica, controlando que CMJ/EUR no caigan.")
    elif not has_clear_lagging:
        maintain.append("Mantener el perfil actual y reforzar consistencia técnica con control de carga.")
    else:
        maintain.append("Mantener la estructura actual y reforzar consistencia técnica con control de carga.")
    load_context_available = _latest_acwr_row(state, athlete) is not None or _latest_mono_row(state, athlete) is not None
    if load_context_available:
        maintain.append("Sostener variabilidad del microciclo y evitar monotonía excesiva para proteger la tolerancia actual.")
    else:
        maintain.append("Sostener variabilidad del microciclo y proteger la tolerancia actual.")

    adjust: list[str] = []
    if has_clear_lagging and any(token in f"{low_text} {decline_text}" for token in ("dj height", "dj", "tiempo de contacto", "rsi", "react", "dri")):
        adjust.append("Ajustar la progresión reactiva para mejorar la expresión en DJ, priorizando calidad de contacto, stiffness y tolerancia progresiva.")
    elif has_clear_lagging and feedback.get("low"):
        adjust.append("Ajustar solo una prioridad principal del bloque y traducirla a una tarea concreta de entrenamiento, sin perseguir la métrica de forma aislada.")
    if change_payload.get("declines"):
        adjust.append(f"Revisar carga, protocolo y prioridades si persisten caídas en {_professional_join_labels(change_payload.get('declines', []))}.")
    if not adjust and has_clear_lagging:
        adjust.append("Ajustar solo una prioridad principal del bloque para evitar dispersar el estímulo.")

    elif not adjust:
        adjust.append("Sostener el perfil actual, consolidar calidad técnica y reforzar la base solo si el contexto de carga o wellness lo pide.")

    monitor = [
        "Monitorear DJ RSI, tiempo de contacto, EUR y DSI/DRI si están disponibles en la próxima evaluación.",
        "Cruzar calidad de sesión, dolor, sueño, estrés y adherencia antes de subir volumen o densidad.",
    ]
    if change_payload.get("declines"):
        monitor.append(f"Confirmar si la señal desfavorable en {_professional_join_labels(change_payload.get('declines', []))} persiste o se normaliza.")

    if not has_clear_lagging:
        monitor.append(PROFESSIONAL_NO_CLEAR_LAGGING_MONITOR)

    measure = [
        "Medir nuevamente en 6-8 semanas o antes si el calendario competitivo exige verificar una señal puntual.",
        "Medir las variables clave del objetivo principal y completar datos faltantes si los hubiera.",
    ]
    if str(evaluation_state).strip().lower() == "partial":
        measure.append("Completar la batería faltante antes de cerrar conclusiones más fuertes.")

    limitations = [
        "Las evaluaciones se interpretan como perfilado físico cada 6-8 semanas, no como readiness semanal.",
        "El sRPE es una estimación práctica de carga interna y debe cruzarse con contexto, wellness y criterio profesional.",
        "Si el intervalo entre evaluaciones es corto o faltan registros, la lectura debe ser más conservadora.",
        "El RSI se interpreta aquí como índice reactivo y no como velocidad lineal.",
    ]
    return {
        "title": "Próximos pasos y limitaciones metodológicas",
        "actions": {
            "Mantener": list(dict.fromkeys([line for line in maintain if line.strip()]))[:2],
            "Ajustar": list(dict.fromkeys([line for line in adjust if line.strip()]))[:3],
            "Monitorear": list(dict.fromkeys([line for line in monitor if line.strip()]))[:3],
            "Medir": list(dict.fromkeys([line for line in measure if line.strip()]))[:3],
        },
        "limitations": limitations,
    }
    maintain = [
        "Mantener la cualidad dominante actual sin perder calidad tÃ©cnica ni variabilidad del microciclo.",
    ]
    if feedback.get("next_block"):
        maintain.append(f"Mantener como referencia del bloque: {feedback.get('next_block')}")

    adjust = []
    if feedback.get("low"):
        adjust.append(f"Ajustar el prÃ³ximo bloque para priorizar {feedback.get('low')}.")
    adjust.extend(integrated_payload.get("decision_practical", [])[:2])
    if change_payload.get("declines"):
        adjust.append(f"Ajustar carga y prioridades si persisten las caÃ­das en {_professional_join_labels(change_payload.get('declines', []))}.")

    monitor = list(integrated_payload.get("monitor", [])[:3])
    measure = [
        "Medir nuevamente el perfil fÃ­sico en 6-8 semanas para comparar cambios de bloque.",
        "Medir en la prÃ³xima ventana las variables clave del objetivo principal y completar datos faltantes si los hubiera.",
    ]
    if str(evaluation_state).strip().lower() == "partial":
        measure.append("Medir y completar la baterÃ­a faltante antes de cerrar conclusiones mÃ¡s fuertes.")

    limitations = [
        "Las evaluaciones se interpretan como perfilado fÃ­sico cada 6-8 semanas, no como readiness semanal.",
        "El sRPE es una estimaciÃ³n prÃ¡ctica de carga interna y debe cruzarse con contexto, wellness y criterio profesional.",
        "Si el intervalo entre evaluaciones es corto o faltan registros, la lectura debe ser mÃ¡s conservadora.",
        "El RSI se interpreta aquÃ­ como Ã­ndice reactivo y no como velocidad lineal.",
    ]
    return {
        "title": "PrÃ³ximos pasos y limitaciones metodolÃ³gicas",
        "actions": {
            "Mantener": list(dict.fromkeys([line for line in maintain if line.strip()]))[:2],
            "Ajustar": list(dict.fromkeys([line for line in adjust if line.strip()]))[:3],
            "Monitorear": list(dict.fromkeys([line for line in monitor if line.strip()]))[:3],
            "Medir": list(dict.fromkeys([line for line in measure if line.strip()]))[:3],
        },
        "limitations": limitations,
    }


def _build_professional_executive_payload(
    state: dict[str, pd.DataFrame | None],
    athlete: str,
    *,
    evaluation_state: str,
    assessment_interval_warning: str,
    composite_payload: dict[str, object],
    change_payload: dict[str, object],
    load_payload: dict[str, object],
    wellness_payload: dict[str, object],
    exposure_payload: dict[str, object],
    training_context: dict[str, object],
    integrated_payload: dict[str, object],
) -> dict[str, object]:
    quality_report = _report_quality_report(state)
    athlete_quality = _quality_athlete_row(quality_report, athlete)
    quality_detail = _quality_detail_text(athlete_quality)
    coverage_parts = []
    if athlete_quality is not None:
        coverage_parts.append(f"Cobertura: {safe_value(athlete_quality.get('Semaforo'))}")
    if quality_detail:
        coverage_parts.append(quality_detail)
    if not coverage_parts:
        coverage_parts.append(f"Evaluaciones: {_professional_status_label(evaluation_state)}")
        coverage_parts.append(f"Entrenamiento: {_professional_status_label(training_context.get('state'))}")
        coverage_parts.append(f"Wellness: {_professional_status_label(wellness_payload.get('state'))}")

    signals: list[str] = []
    feedback = composite_payload.get("feedback", {}) if isinstance(composite_payload.get("feedback"), dict) else {}
    has_clear_lagging = _professional_has_clear_lagging_variable(feedback)
    high_text = _professional_strip_terminal_period(feedback.get("high", ""))
    low_text = _professional_strip_terminal_period(feedback.get("low", ""))
    load_rows = _professional_rows_map(load_payload.get("rows"))
    wellness_rows = _professional_rows_map(wellness_payload.get("rows"))
    improvements = change_payload.get("improvements", []) if isinstance(change_payload.get("improvements"), list) else []
    declines = change_payload.get("declines", []) if isinstance(change_payload.get("declines"), list) else []
    dominant = exposure_payload.get("dominant", []) if isinstance(exposure_payload.get("dominant"), list) else []
    low_or_absent = exposure_payload.get("low_or_absent", []) if isinstance(exposure_payload.get("low_or_absent"), list) else []

    if _has_text(high_text):
        if has_clear_lagging and _has_text(low_text):
            signals.append(_clip_signal(f"Predominan {high_text}. Rezago principal: {low_text}."))
        else:
            signals.append(_clip_signal(f"Predominan {high_text}. Sin rezago claro."))
    if improvements or declines:
        change_parts: list[str] = []
        if improvements:
            change_parts.append(f"Mejoran {_professional_join_labels(improvements[:3])}.")
        if declines:
            change_parts.append(f"Caen {_professional_join_labels(declines[:3])}.")
        signals.append(_clip_signal(" ".join(change_parts)))
    if load_rows:
        load_parts = []
        if _has_text(load_rows.get("ACWR EWMA")):
            load_parts.append(f"ACWR {load_rows['ACWR EWMA']}")
        if _has_text(load_rows.get("Zona ACWR")):
            load_parts.append(f"zona {load_rows['Zona ACWR']}")
        if _has_text(load_rows.get("Monotonía")):
            load_parts.append(f"Monotonía {load_rows['Monotonía']}")
        if load_parts:
            signals.append(_clip_signal(". ".join(load_parts) + "."))
    if wellness_rows:
        days_text = wellness_rows.get("Días con registro", "")
        score_text = wellness_rows.get("Wellness score", "")
        if str(wellness_payload.get("state") or "").strip().lower() == "partial":
            signals.append(_clip_signal(f"Wellness parcial: {days_text}." if _has_text(days_text) else "Wellness parcial."))
        elif _has_text(score_text):
            details = [f"Wellness {score_text}"]
            if _has_text(days_text):
                details.append(f"{days_text} con registro")
            signals.append(_clip_signal(". ".join(details) + "."))
    if dominant:
        signals.append(_clip_signal(f"Dominan {_professional_join_labels(dominant[:2])}."))
    elif low_or_absent:
        signals.append(_clip_signal(f"Baja exposición en {_professional_join_labels(low_or_absent[:2])}."))

    if len([signal for signal in signals if _has_text(signal)]) < 3:
        fallback_signals = [
            f"Evaluación: {_professional_status_label(evaluation_state)}.",
            f"Carga: {_professional_status_label(load_payload.get('state'))}.",
            f"Wellness: {_professional_status_label(wellness_payload.get('state'))}.",
            f"Entrenamiento: {_professional_status_label(training_context.get('state'))}.",
        ]
        for fallback in fallback_signals:
            clipped = _clip_signal(fallback)
            if _has_text(clipped):
                signals.append(clipped)
            if len(list(dict.fromkeys([signal for signal in signals if _has_text(signal)]))) >= 5:
                break
    confidence = _professional_data_confidence_label(
        evaluation_state,
        training_context,
        wellness_payload,
        assessment_interval_warning,
    )
    confidence_detail = "Sin seÃ±ales metodolÃ³gicas fuertes de cautela." if confidence == "Alta" else (
        "Lectura Ãºtil, pero todavÃ­a condicionada por cobertura parcial o por el contexto de mediciÃ³n."
        if confidence == "Media"
        else "La lectura requiere mucha prudencia por cobertura parcial o falta de continuidad."
    )
    if assessment_interval_warning:
        confidence_detail = assessment_interval_warning

    decision_suggested = _professional_join_labels(
        integrated_payload.get("decision_practical", [])[:1],
        fallback="Completar datos y sostener decisiones conservadoras hasta confirmar la tendencia.",
    )
    return {
        "title": "Resumen ejecutivo profesional",
        "athlete": athlete,
        "date": f"{datetime.now():%d/%m/%Y}",
        "period": _professional_period_label(state, athlete),
        "coverage": " | ".join(coverage_parts),
        "confidence": confidence,
        "confidence_detail": confidence_detail,
        "signals": list(dict.fromkeys([signal for signal in signals if _has_text(signal)]))[:5],
        "decision_suggested": decision_suggested,
    }


def _professional_period_label(state: dict[str, pd.DataFrame | None], athlete: str) -> str:
    dates: list[pd.Timestamp] = []

    def add_dates(frame: pd.DataFrame | None, date_candidates: tuple[str, ...], *, athlete_col: str = "Athlete") -> None:
        if frame is None or frame.empty:
            return
        result = frame.copy()
        if athlete_col in result.columns:
            result = result[_professional_athlete_mask(result[athlete_col], athlete)]
        for column in date_candidates:
            if column not in result.columns:
                continue
            parsed = pd.to_datetime(result[column], errors="coerce").dropna()
            dates.extend(pd.Timestamp(value).normalize() for value in parsed.tolist())
            break

    add_dates(state.get("jump_df"), ("Date",))
    add_dates(state.get("rpe_df"), ("Date",))
    add_dates(state.get("wellness_df"), ("Date",))
    add_dates(state.get("completion_df"), ("Date",))
    add_dates(state.get("raw_df"), ("Assigned Date", "Date"))

    if not dates:
        return PDF_MISSING_TEXT
    start = min(dates)
    end = max(dates)
    if start == end:
        return start.strftime("%d/%m/%Y")
    return f"{start:%d/%m/%Y} - {end:%d/%m/%Y}"


def _build_professional_report_overview(
    state: dict[str, pd.DataFrame | None],
    athlete: str,
    cards: list[dict[str, object]],
    training_context: dict[str, object],
    internal_load: dict[str, object],
) -> dict[str, object]:
    available_cards, _ = _professional_metric_display_groups(cards)
    wellness = _professional_wellness_context(state, athlete)
    assessment_interval_warning = _professional_short_assessment_interval_warning(state, athlete)
    evaluation_state = "missing"
    if len(available_cards) == len(PROFESSIONAL_PDF_METRICS):
        evaluation_state = "available"
    elif available_cards:
        evaluation_state = "partial"

    if evaluation_state == "available":
        reading = "Perfil físico disponible con batería de evaluación útil para orientar prioridades del bloque."
        decision = "Usar las métricas junto con carga interna, wellness y contexto deportivo para orientar el siguiente bloque."
    elif evaluation_state == "partial":
        reading = "Perfil físico parcial: hay señales útiles, pero la cobertura de evaluación todavía es incompleta."
        decision = "Tomar esta lectura como preliminar, completar la batería faltante y evitar conclusiones fuertes con métricas aisladas."
    elif internal_load.get("state") != "missing" or training_context.get("state") != "missing":
        reading = "Sin perfil físico suficiente; el reporte se apoya en entrenamiento, carga interna y contexto del bloque."
        decision = "Completar una batería de evaluación y usar la carga disponible solo para regular decisiones de corto plazo."
    else:
        reading = "Información insuficiente para una lectura profesional completa."
        decision = "Cargar evaluaciones, entrenamiento y sRPE antes de tomar decisiones de bloque."
    if assessment_interval_warning and evaluation_state != "missing":
        reading = f"{reading} Intervalo corto entre evaluaciones: interpretar los cambios como señal preliminar."

    return {
        "athlete": athlete,
        "date": f"{datetime.now():%d/%m/%Y}",
        "period": _professional_period_label(state, athlete),
        "statuses": {
            "Evaluaciones físicas": _professional_status_label(evaluation_state),
            "Entrenamiento": _professional_status_label(training_context.get("state")),
            "Carga interna": _professional_status_label(internal_load.get("state")),
            "Wellness": _professional_status_label(wellness.get("state")),
        },
        "reading": reading,
        "decision": decision,
    }


def _professional_trend_delta(points: list[dict[str, object]], key: str) -> float | None:
    valid = [_coerce_float(point.get(key)) for point in points]
    valid = [value for value in valid if value is not None]
    if len(valid) < 2:
        return None
    return float(valid[-1] - valid[-2])

def _build_professional_integrated_interpretation(
    internal_load: dict[str, object],
    wellness_context: dict[str, object],
    *,
    evaluation_state: str | None = None,
    assessment_interval_warning: str = "",
) -> list[str]:
    lines: list[str] = []
    clean_evaluation_state = str(evaluation_state or "").strip().lower()
    if clean_evaluation_state == "partial":
        lines.append(
            "La cobertura de evaluaciones físicas es parcial; esta lectura debe entenderse como señal preliminar y no como cierre completo del perfil."
        )
    elif clean_evaluation_state == "missing":
        lines.append(
            "No hay batería física suficiente para integrar cambios neuromusculares; la lectura se apoya sobre todo en carga interna y wellness."
        )
    if assessment_interval_warning:
        lines.append(
            "El intervalo entre evaluaciones fue menor a 6-8 semanas; evitar atribuir cambios de forma directa a adaptación sin seguimiento adicional."
        )
    load_scope = str(internal_load.get("analysis_scope") or "")
    weekly_change_pct = _coerce_float(internal_load.get("weekly_change_pct"))
    weekly_change = _coerce_float(internal_load.get("weekly_change"))
    if load_scope == "current_week_partial":
        partial_total = _coerce_float(internal_load.get("current_week_total"))
        partial_sessions = int(_coerce_float(internal_load.get("current_week_sessions")) or 0)
        if partial_total is not None:
            lines.append(
                f"El acumulado parcial de la semana actual es {partial_total:.0f} UA con {partial_sessions} sesiones registradas. "
                "Al tratarse de una semana en curso, no debe interpretarse todavía como una caída real frente a la semana previa."
            )
        else:
            lines.append("La semana actual está incompleta y faltan datos suficientes para interpretar la carga acumulada.")
    elif weekly_change_pct is not None:
        direction = "aumentó" if weekly_change_pct > 0 else "disminuyó" if weekly_change_pct < 0 else "se mantuvo estable"
        lines.append(
            f"La carga interna semanal {direction} {weekly_change_pct:+.1f}% respecto a la semana previa; "
            "es una señal de exposición reciente y no una conclusión aislada por sí sola."
        )
    elif weekly_change is not None:
        direction = "aumentó" if weekly_change > 0 else "disminuyó" if weekly_change < 0 else "se mantuvo estable"
        lines.append(
            f"La carga interna semanal {direction} {weekly_change:+.0f} UA respecto a la semana previa; "
            "es una señal operativa y debe leerse con el resto del contexto."
        )
    else:
        lines.append("Faltan datos para comparar la carga interna contra la semana previa.")

    weekly_wellness = wellness_context.get("weekly_points", [])
    sleep_delta = _professional_trend_delta(weekly_wellness, "sleep")
    stress_delta = _professional_trend_delta(weekly_wellness, "stress")
    pain_delta = _professional_trend_delta(weekly_wellness, "pain")
    score_delta = _professional_trend_delta(weekly_wellness, "score")
    summary = wellness_context.get("last_week_summary", {}) if isinstance(wellness_context.get("last_week_summary"), dict) else {}
    stress_mean = _coerce_float(summary.get("stress_mean"))
    pain_mean = _coerce_float(summary.get("pain_mean"))
    sleep_mean = _coerce_float(summary.get("sleep_mean"))
    scales = wellness_context.get("scales", {}) if isinstance(wellness_context.get("scales"), dict) else {}

    wellness_comments: list[str] = []
    if wellness_context.get("state") == "partial" and wellness_context.get("partial_message"):
        wellness_comments.append("La lectura del wellness es limitada por baja cantidad de registros y requiere seguimiento.")
    if sleep_delta is not None and sleep_delta < -0.2:
        wellness_comments.append(
            f"El sueño cambió {sleep_delta:+.1f} h respecto a la semana previa; si baja, puede reducir el margen de recuperación."
        )
    if stress_delta is not None and stress_delta > 0.2:
        wellness_comments.append(
            f"El estrés cambió {stress_delta:+.1f} respecto a la semana previa y es una señal compatible con mayor carga contextual del bloque."
        )
    if pain_delta is not None and pain_delta > 0.2:
        wellness_comments.append(
            f"El dolor cambió {pain_delta:+.1f} respecto a la semana previa; conviene controlar volumen, densidad y selección de ejercicios."
        )
    score_label = "wellness score"
    if score_delta is not None:
        magnitude = abs(score_delta)
        if abs(score_delta) <= 0.5:
            wellness_comments.append(f"El {score_label} cambió {score_delta:+.1f} puntos respecto a la semana previa y se mantiene relativamente estable.")
        elif score_delta < -0.5:
            wellness_comments.append(
                f"El {score_label} disminuyó {magnitude:.1f} puntos respecto a la semana previa; señal compatible con menor tolerancia o mayor carga contextual."
            )
        else:
            wellness_comments.append(
                f"El {score_label} aumentó {magnitude:.1f} puntos respecto a la semana previa, señal compatible con buena tolerancia si el contexto acompaña."
            )
    if load_scope != "current_week_partial" and weekly_change_pct is not None and weekly_change_pct > 10 and _professional_wellness_high(stress_mean, scales.get("stress")):
        wellness_comments.append(
            f"El estrés promedio fue elevado ({stress_mean:.1f}{scales.get('stress', '')}), por lo que conviene progresar de manera conservadora y controlar volumen/densidad."
        )
    if pain_mean is not None and _professional_wellness_high(pain_mean, scales.get("pain")):
        wellness_comments.append(f"El dolor promedio fue elevado ({pain_mean:.1f}{scales.get('pain', '')}); revisar tolerancia antes de aumentar carga.")
    if sleep_mean is not None and sleep_mean < 6.5:
        wellness_comments.append(f"El sueño promedio fue bajo ({sleep_mean:.1f} h); priorizar recuperación antes de escalar volumen.")
    if not wellness_comments and wellness_context.get("state") == "missing":
        wellness_comments.append("Faltan datos de wellness para cruzar la carga interna con recuperación percibida.")
    elif not wellness_comments:
        wellness_comments.append("No aparece una señal dominante de wellness; interpretar junto con sueño, estrés, dolor y asistencia.")
    score_comment_idx = next((idx for idx, comment in enumerate(wellness_comments) if score_label in comment), None)
    stress_comment_idx = next((idx for idx, comment in enumerate(wellness_comments) if "estrés promedio fue elevado" in comment), None)
    if (
        load_scope != "current_week_partial"
        and weekly_change_pct is not None
        and weekly_change_pct > 10
        and score_comment_idx is not None
        and stress_comment_idx is not None
    ):
        lines.append(
            f"{wellness_comments[score_comment_idx]} {wellness_comments[stress_comment_idx]} Esto no implica necesariamente mala tolerancia, pero sí sugiere que la siguiente progresión debería priorizar control de volumen, densidad y calidad de ejecución antes que aumentar carga de forma agresiva."
        )
        lines.extend(
            comment
            for idx, comment in enumerate(wellness_comments)
            if idx not in {score_comment_idx, stress_comment_idx}
        )
    else:
        lines.extend(wellness_comments[:3])
    lines.append(
        "Integrar siempre carga interna, sueño, estrés, dolor, adherencia, calendario deportivo y criterio profesional antes de modificar carga o prioridades."
    )
    return lines


def _build_professional_next_steps(evaluation_state: str) -> list[str]:
    clean = str(evaluation_state or "").strip().lower()
    if clean == "missing":
        return [
            "Completar batería de evaluación: CMJ, SJ, DJ/RSI e IMTP.",
            "Mantener registro de sRPE y wellness durante las próximas semanas.",
            "Usar la próxima evaluación como línea base del perfil físico.",
            "No interpretar carga interna como perfil físico.",
        ]
    if clean == "partial":
        return [
            "Completar métricas faltantes, especialmente mRSI o DRI si corresponde.",
            "Repetir evaluación en 6-8 semanas antes de cerrar conclusiones más fuertes.",
            "Mantener misma entrada en calor, protocolo y condiciones para comparar mejor.",
            "Usar carga interna y wellness solo para regular el corto plazo; no reemplazan el perfil físico.",
        ]
    return [
        "Usar el perfil físico para orientar el próximo bloque junto con carga interna, wellness y contexto deportivo.",
        "Usar carga interna y wellness para regular el corto plazo, no como sustituto de la evaluación física.",
        "Repetir evaluación en 6-8 semanas.",
    ]


def _pdf_theme_threshold(colors_module, *, variant: str = "general") -> dict[str, object]:
    variant_key = str(variant or "general").strip().lower()
    shared = {
        "page_width_mm": 174,
        "page_gap_mm": 4,
        "box_padding": 8,
        "table_padding": 6,
        "footer_label": "Threshold S&C",
        "footer_font_name": "Helvetica",
        "footer_font_size": 7.5,
        "footer_y_mm": 8,
        "footer_rule_gap_mm": 3.2,
        "footer_rule_width": 0.45,
        "white": colors_module.white,
    }
    variants = {
        "general": {
            "bg": colors_module.HexColor("#F4F6F8"),
            "card": colors_module.HexColor("#FEFEFE"),
            "panel": colors_module.HexColor("#F6F9FB"),
            "panel_alt": colors_module.HexColor("#EEF3F8"),
            "navy": colors_module.HexColor("#0D3C5E"),
            "steel": colors_module.HexColor("#134263"),
            "ink": colors_module.HexColor("#221F20"),
            "muted": colors_module.HexColor("#708C9F"),
            "gray": colors_module.HexColor("#5A595B"),
            "line": colors_module.HexColor("#D8DEE4"),
            "line_dark": colors_module.HexColor("#C5D0D9"),
            "green": colors_module.HexColor("#2F6B52"),
            "yellow": colors_module.HexColor("#C4A464"),
            "orange": colors_module.HexColor("#B87445"),
            "red": colors_module.HexColor("#B56B73"),
        },
        "professional": {
            "bg": colors_module.HexColor("#F3F5F7"),
            "card": colors_module.HexColor("#FFFFFF"),
            "panel": colors_module.HexColor("#F7FAFC"),
            "panel_alt": colors_module.HexColor("#EEF3F8"),
            "navy": colors_module.HexColor("#102C44"),
            "steel": colors_module.HexColor("#29485F"),
            "ink": colors_module.HexColor("#161A1D"),
            "muted": colors_module.HexColor("#617383"),
            "gray": colors_module.HexColor("#4E5B67"),
            "line": colors_module.HexColor("#D6DEE5"),
            "line_dark": colors_module.HexColor("#BCC9D3"),
            "green": colors_module.HexColor("#2F6B52"),
            "yellow": colors_module.HexColor("#C4A464"),
            "orange": colors_module.HexColor("#B87445"),
            "red": colors_module.HexColor("#B56B73"),
        },
    }
    selected = variants.get(variant_key, variants["general"]).copy()
    selected.update(shared)
    return selected


def _register_threshold_pdf_styles(
    styles,
    ParagraphStyleClass,
    palette: dict[str, object],
    *,
    variant: str = "general",
):
    def add(name: str, parent_name: str, **kwargs) -> None:
        if name in styles.byName:
            return
        styles.add(ParagraphStyleClass(name=name, parent=styles[parent_name], **kwargs))

    add("ReportTitle", "Heading1", fontName="Helvetica-Bold", fontSize=23, leading=28, textColor=palette["navy"], spaceAfter=8)
    add("ReportSection", "Heading2", fontName="Helvetica-Bold", fontSize=16, leading=20, textColor=palette["navy"], spaceAfter=8, spaceBefore=4)
    add("ReportBody", "BodyText", fontName="Helvetica", fontSize=10, leading=14, textColor=palette["ink"], spaceAfter=6)
    add("ReportMuted", "BodyText", fontName="Helvetica", fontSize=9, leading=13, textColor=palette["gray"], spaceAfter=4)
    add("ReportMutedItalic", "BodyText", fontName="Helvetica-Oblique", fontSize=9, leading=13, textColor=palette["gray"], spaceAfter=3)
    add("CardLabel", "BodyText", fontName="Helvetica-Bold", fontSize=8, leading=10, textColor=palette["gray"], spaceAfter=4)
    add("CardValue", "BodyText", fontName="Helvetica-Bold", fontSize=15, leading=18, textColor=palette["ink"])
    add("BlockTitle", "BodyText", fontName="Helvetica-Bold", fontSize=11, leading=14, textColor=palette["ink"], spaceAfter=5)
    add("ReportDecisionTitle", "BodyText", fontName="Helvetica-Bold", fontSize=10, leading=13, textColor=palette["white"], spaceAfter=3)
    add("ReportBodyWhite", "BodyText", fontName="Helvetica", fontSize=9.4, leading=12.8, textColor=palette["white"], spaceAfter=4)
    add("ReportMutedWhite", "BodyText", fontName="Helvetica", fontSize=8.5, leading=11.6, textColor=palette["white"], spaceAfter=3)
    add("ReportTableHeader", "BodyText", fontName="Helvetica-Bold", fontSize=8.4, leading=10.5, textColor=palette["white"], spaceAfter=0)
    add("ReportTableCell", "BodyText", fontName="Helvetica", fontSize=8.7, leading=11.4, textColor=palette["ink"], spaceAfter=0)

    if str(variant or "").strip().lower() != "professional":
        return styles

    add("ProfTitle", "Heading1", fontName="Helvetica-Bold", fontSize=24, leading=28, textColor=palette["navy"], spaceAfter=6)
    add("ProfHeroTitle", "Heading1", fontName="Helvetica-Bold", fontSize=24, leading=29, textColor=palette["white"], spaceAfter=4)
    add("ProfHeroMeta", "BodyText", fontName="Helvetica", fontSize=9, leading=12, textColor=palette["line_dark"], spaceAfter=3)
    add("ProfEyebrow", "BodyText", fontName="Helvetica-Bold", fontSize=8, leading=10, textColor=palette["muted"], spaceAfter=2)
    add("ProfSection", "Heading2", fontName="Helvetica-Bold", fontSize=18, leading=22, textColor=palette["navy"], spaceBefore=0, spaceAfter=3)
    add("ProfSectionSubtitle", "BodyText", fontName="Helvetica", fontSize=9, leading=12, textColor=palette["muted"], spaceAfter=3)
    add("ProfBody", "BodyText", fontName="Helvetica", fontSize=9.2, leading=12.6, textColor=palette["ink"], spaceAfter=4)
    add("ProfMuted", "BodyText", fontName="Helvetica", fontSize=8.4, leading=11.2, textColor=palette["gray"], spaceAfter=4)
    add("ProfMutedItalic", "BodyText", fontName="Helvetica-Oblique", fontSize=8.5, leading=11.5, textColor=palette["gray"], spaceAfter=3)
    add("ProfCardTitle", "BodyText", fontName="Helvetica-Bold", fontSize=9.6, leading=12, textColor=palette["navy"], spaceAfter=3)
    add("ProfCardValue", "BodyText", fontName="Helvetica-Bold", fontSize=14.2, leading=17, textColor=palette["ink"], spaceAfter=4)
    add("ProfDecisionTitle", "BodyText", fontName="Helvetica-Bold", fontSize=9.5, leading=12, textColor=palette["white"], spaceAfter=3)
    add("ProfBodyWhite", "BodyText", fontName="Helvetica", fontSize=9.3, leading=12.8, textColor=palette["white"], spaceAfter=4)
    add("ProfMutedWhite", "BodyText", fontName="Helvetica", fontSize=8.3, leading=11.2, textColor=palette["line_dark"], spaceAfter=3)
    add("ProfTableHeader", "BodyText", fontName="Helvetica-Bold", fontSize=8.5, leading=10.5, textColor=palette["white"], spaceAfter=0)
    add("ProfTableCell", "BodyText", fontName="Helvetica", fontSize=8.5, leading=11.2, textColor=palette["ink"], spaceAfter=0)
    return styles


def _fit_pdf_image(
    path: Path | None,
    *,
    max_width_mm: float,
    max_height_mm: float,
    mm_unit,
    ImageClass,
    ImageReaderClass,
):
    if path is None:
        return None
    try:
        reader = ImageReaderClass(str(path))
        width, height = reader.getSize()
    except Exception:
        return None
    max_width = max_width_mm * mm_unit
    max_height = max_height_mm * mm_unit
    scale = min(max_width / width, max_height / height)
    return ImageClass(str(path), width=width * scale, height=height * scale)


def _build_threshold_box(
    flowables: list[object],
    *,
    TableClass,
    TableStyleClass,
    mm_unit,
    palette: dict[str, object],
    width_mm: float | None = None,
    padding: int | None = None,
    background=None,
    border_color=None,
    accent_color=None,
    h_align: str = "LEFT",
    border_width: float = 0.8,
):
    cell_padding = padding if padding is not None else int(palette.get("box_padding", 8))
    table = TableClass(
        [[flowables]],
        colWidths=[(width_mm or float(palette.get("page_width_mm", 174))) * mm_unit],
        hAlign=h_align,
    )
    style_commands = [
        ("BOX", (0, 0), (-1, -1), border_width, border_color or palette.get("line")),
        ("BACKGROUND", (0, 0), (-1, -1), background or palette.get("card")),
        ("LEFTPADDING", (0, 0), (-1, -1), cell_padding),
        ("RIGHTPADDING", (0, 0), (-1, -1), cell_padding),
        ("TOPPADDING", (0, 0), (-1, -1), cell_padding),
        ("BOTTOMPADDING", (0, 0), (-1, -1), cell_padding),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
    ]
    if accent_color is not None:
        style_commands.append(("LINEABOVE", (0, 0), (-1, 0), 2.2, accent_color))
    table.setStyle(TableStyleClass(style_commands))
    return table


def _build_threshold_table(
    rows: list[list[object]],
    *,
    TableClass,
    TableStyleClass,
    mm_unit,
    palette: dict[str, object],
    col_widths_mm: list[float] | None = None,
    col_widths: list[object] | None = None,
    header_rows: int = 0,
    repeat_rows: int = 0,
    h_align: str = "LEFT",
    background=None,
    header_background=None,
    header_text_color=None,
    row_backgrounds: list[object] | None = None,
    border_color=None,
    inner_grid_color=None,
    show_inner_grid: bool = True,
    border_width: float = 0.8,
    inner_grid_width: float = 0.35,
    left_padding: int | None = None,
    right_padding: int | None = None,
    top_padding: int | None = None,
    bottom_padding: int | None = None,
    valign: str = "TOP",
):
    resolved_widths = col_widths if col_widths is not None else ([width * mm_unit for width in col_widths_mm] if col_widths_mm is not None else None)
    table = TableClass(rows, colWidths=resolved_widths, repeatRows=repeat_rows, hAlign=h_align)
    body_background = background or palette.get("card")
    style_commands = [
        ("BOX", (0, 0), (-1, -1), border_width, border_color or palette.get("line")),
        ("LEFTPADDING", (0, 0), (-1, -1), left_padding if left_padding is not None else int(palette.get("table_padding", 6))),
        ("RIGHTPADDING", (0, 0), (-1, -1), right_padding if right_padding is not None else int(palette.get("table_padding", 6))),
        ("TOPPADDING", (0, 0), (-1, -1), top_padding if top_padding is not None else int(palette.get("table_padding", 6))),
        ("BOTTOMPADDING", (0, 0), (-1, -1), bottom_padding if bottom_padding is not None else int(palette.get("table_padding", 6))),
        ("VALIGN", (0, 0), (-1, -1), valign),
    ]
    if show_inner_grid and rows and (len(rows) > 1 or len(rows[0]) > 1):
        style_commands.append(("INNERGRID", (0, 0), (-1, -1), inner_grid_width, inner_grid_color or palette.get("line")))
    if header_rows > 0:
        style_commands.append(("BACKGROUND", (0, 0), (-1, header_rows - 1), header_background or palette.get("navy")))
        style_commands.append(("TEXTCOLOR", (0, 0), (-1, header_rows - 1), header_text_color or palette.get("white")))
        if row_backgrounds:
            style_commands.append(("ROWBACKGROUNDS", (0, header_rows), (-1, -1), row_backgrounds))
        else:
            style_commands.append(("BACKGROUND", (0, header_rows), (-1, -1), body_background))
    elif row_backgrounds:
        style_commands.append(("ROWBACKGROUNDS", (0, 0), (-1, -1), row_backgrounds))
    else:
        style_commands.append(("BACKGROUND", (0, 0), (-1, -1), body_background))
    table.setStyle(TableStyleClass(style_commands))
    return table


def _build_metric_card(
    label: object,
    value: object,
    *,
    paragraph_builder,
    TableClass,
    TableStyleClass,
    mm_unit,
    palette: dict[str, object],
    width_mm: float = 54,
    label_style: str = "CardLabel",
    value_style: str = "CardValue",
):
    table = TableClass(
        [[paragraph_builder(label, label_style)], [paragraph_builder(value, value_style)]],
        colWidths=[width_mm * mm_unit],
    )
    table.setStyle(
        TableStyleClass(
            [
                ("BOX", (0, 0), (-1, -1), 0.7, palette.get("line")),
                ("BACKGROUND", (0, 0), (-1, -1), palette.get("card")),
                ("LEFTPADDING", (0, 0), (-1, -1), 8),
                ("RIGHTPADDING", (0, 0), (-1, -1), 8),
                ("TOPPADDING", (0, 0), (-1, -1), 7),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
            ]
        )
    )
    return table


def _build_threshold_bullets(
    lines: list[object],
    *,
    paragraph_builder,
    style_name: str = "ReportBody",
) -> list[object]:
    return [
        paragraph_builder(f"- {text}", style_name)
        for text in [str(line or "").strip() for line in lines]
        if text
    ]


def _build_note_box(
    title: str,
    notes: object,
    *,
    paragraph_builder,
    TableClass,
    TableStyleClass,
    mm_unit,
    palette: dict[str, object],
    width_mm: float | None = None,
    title_style: str = "ProfCardTitle",
    body_style: str = "ProfMuted",
    background=None,
    border_color=None,
    accent_color=None,
):
    items = notes if isinstance(notes, list) else [notes]
    flowables = [paragraph_builder(title, title_style)] if title else []
    flowables.extend(paragraph_builder(note, body_style) for note in items if str(note or "").strip())
    return _build_threshold_box(
        flowables,
        TableClass=TableClass,
        TableStyleClass=TableStyleClass,
        mm_unit=mm_unit,
        palette=palette,
        width_mm=width_mm,
        padding=6,
        background=background,
        border_color=border_color,
        accent_color=accent_color,
    )


def _build_chart_container(
    chart_flowable: object | None,
    *,
    paragraph_builder,
    TableClass,
    TableStyleClass,
    mm_unit,
    palette: dict[str, object],
    title: str = "",
    note: str = "",
    empty_text: str = "",
    width_mm: float | None = None,
    title_style: str = "ProfCardTitle",
    note_style: str = "ProfMuted",
    background=None,
    border_color=None,
    accent_color=None,
):
    flowables: list[object] = []
    if title:
        flowables.append(paragraph_builder(title, title_style))
    if chart_flowable is not None:
        flowables.append(chart_flowable)
    elif empty_text:
        flowables.append(paragraph_builder(empty_text, note_style))
    if note:
        flowables.append(paragraph_builder(note, note_style))
    return _build_threshold_box(
        flowables,
        TableClass=TableClass,
        TableStyleClass=TableStyleClass,
        mm_unit=mm_unit,
        palette=palette,
        width_mm=width_mm,
        padding=6,
        background=background or palette.get("panel"),
        border_color=border_color or palette.get("line_dark", palette.get("line")),
        accent_color=accent_color,
    )


def _build_pdf_page_title(
    title: str,
    *,
    paragraph_builder,
    TableClass,
    TableStyleClass,
    mm_unit,
    palette: dict[str, object],
    subtitle: str = "",
    eyebrow: str = "Threshold S&C",
    width_mm: float | None = None,
    title_style: str = "ProfSection",
    subtitle_style: str = "ProfSectionSubtitle",
    eyebrow_style: str = "ProfEyebrow",
):
    flowables: list[object] = []
    if eyebrow:
        flowables.append(paragraph_builder(str(eyebrow).upper(), eyebrow_style))
    flowables.append(paragraph_builder(title, title_style))
    if subtitle:
        flowables.append(paragraph_builder(subtitle, subtitle_style))
    return _build_threshold_box(
        flowables,
        TableClass=TableClass,
        TableStyleClass=TableStyleClass,
        mm_unit=mm_unit,
        palette=palette,
        width_mm=width_mm,
        padding=7,
        background=palette.get("panel"),
        border_color=palette.get("line_dark", palette.get("line")),
        accent_color=palette.get("navy"),
    )


def _build_decision_box(
    title: str,
    text: object,
    *,
    paragraph_builder,
    TableClass,
    TableStyleClass,
    mm_unit,
    palette: dict[str, object],
    width_mm: float | None = None,
    note: str = "",
    inverted: bool = True,
    title_style: str | None = None,
    body_style: str | None = None,
    note_style: str | None = None,
):
    resolved_title_style = title_style or ("ProfDecisionTitle" if inverted else "ProfCardTitle")
    resolved_body_style = body_style or ("ProfBodyWhite" if inverted else "ProfBody")
    resolved_note_style = note_style or ("ProfMutedWhite" if inverted else "ProfMuted")
    flowables = [
        paragraph_builder(title, resolved_title_style),
        paragraph_builder(text, resolved_body_style),
    ]
    if note:
        flowables.append(paragraph_builder(note, resolved_note_style))
    return _build_threshold_box(
        flowables,
        TableClass=TableClass,
        TableStyleClass=TableStyleClass,
        mm_unit=mm_unit,
        palette=palette,
        width_mm=width_mm,
        padding=7,
        background=palette.get("navy") if inverted else palette.get("panel_alt", palette.get("panel")),
        border_color=palette.get("navy") if inverted else palette.get("line_dark", palette.get("line")),
        accent_color=None if inverted else palette.get("navy"),
    )


def _build_threshold_separator(
    *,
    TableClass,
    TableStyleClass,
    mm_unit,
    palette: dict[str, object],
    width_mm: float | None = None,
    color=None,
    thickness: float = 0.6,
):
    table = TableClass([[""]], colWidths=[(width_mm or float(palette.get("page_width_mm", 174))) * mm_unit], rowHeights=[1], hAlign="LEFT")
    table.setStyle(
        TableStyleClass(
            [
                ("LINEABOVE", (0, 0), (-1, 0), thickness, color or palette.get("line_dark", palette.get("line"))),
                ("LEFTPADDING", (0, 0), (-1, -1), 0),
                ("RIGHTPADDING", (0, 0), (-1, -1), 0),
                ("TOPPADDING", (0, 0), (-1, -1), 0),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
            ]
        )
    )
    return table


def _draw_threshold_header(canvas_obj, doc_obj, *, palette: dict[str, object], mm_unit, label: str = "Threshold S&C") -> None:
    canvas_obj.saveState()
    top_y = float(doc_obj.pagesize[1]) - (10 * mm_unit)
    page_width = float(doc_obj.pagesize[0])
    canvas_obj.setStrokeColor(palette.get("line_dark", palette.get("line")))
    canvas_obj.setLineWidth(0.35)
    canvas_obj.line(doc_obj.leftMargin, top_y, page_width - doc_obj.rightMargin, top_y)
    canvas_obj.setFillColor(palette.get("muted", palette.get("gray")))
    canvas_obj.setFont(str(palette.get("footer_font_name", "Helvetica")), 7.0)
    canvas_obj.drawString(doc_obj.leftMargin, top_y + (1.4 * mm_unit), label)
    canvas_obj.restoreState()


def _draw_threshold_footer(canvas_obj, doc_obj, *, palette: dict[str, object], mm_unit, label: str | None = None) -> None:
    canvas_obj.saveState()
    footer_y = float(palette.get("footer_y_mm", 8)) * mm_unit
    rule_gap = float(palette.get("footer_rule_gap_mm", 3.2)) * mm_unit
    page_width = float(doc_obj.pagesize[0])
    canvas_obj.setStrokeColor(palette.get("line_dark", palette.get("line")))
    canvas_obj.setLineWidth(float(palette.get("footer_rule_width", 0.45)))
    canvas_obj.line(doc_obj.leftMargin, footer_y + rule_gap, page_width - doc_obj.rightMargin, footer_y + rule_gap)
    canvas_obj.setFillColor(palette.get("muted", palette.get("gray")))
    canvas_obj.setFont(str(palette.get("footer_font_name", "Helvetica")), float(palette.get("footer_font_size", 7.5)))
    canvas_obj.drawString(doc_obj.leftMargin, footer_y, label or str(palette.get("footer_label", "Threshold S&C")))
    canvas_obj.drawRightString(page_width - doc_obj.rightMargin, footer_y, f"Página {doc_obj.page}")
    canvas_obj.restoreState()


def _resolve_brand_asset_path(kind: str = "wordmark") -> Path | None:
    patterns = {
        "wordmark": [
            "threshold_logo_horizontal.*",
            "threshold_wordmark.*",
            "threshold-wordmark.*",
            "threshold-horizontal.*",
            "wordmark.*",
            "Untitled-2.*",
            "untitled-2.*",
        ],
        "icon": [
            "threshold_isotipo.*",
            "threshold_icon.*",
            "threshold-isotype.*",
            "threshold-icon.*",
            "isotype.*",
            "Untitled-1.*",
            "untitled-1.*",
        ],
    }
    for pattern in patterns.get(kind, []):
        matches = sorted(BRAND_ASSET_DIR.glob(pattern))
        if matches:
            return matches[0]
    return None


def _generate_professional_profile_pdf_reportlab(
    state: dict[str, pd.DataFrame | None],
    report_athlete: str,
) -> bytes | None:
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
        from reportlab.lib.units import mm
        from reportlab.lib.utils import ImageReader
        from reportlab.graphics.shapes import Circle, Drawing, Line, PolyLine, Rect, String
        from reportlab.platypus import Image, PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
        from xml.sax.saxutils import escape
    except Exception:
        return None

    palette = _pdf_theme_threshold(colors, variant="professional")
    styles = _register_threshold_pdf_styles(
        getSampleStyleSheet(),
        ParagraphStyle,
        palette,
        variant="professional",
    )

    def _p(text: object, style_name: str = "ProfBody") -> Paragraph:
        safe = escape(_repair_mojibake_text(safe_value(text, fallback=""))).replace("\n", "<br/>")
        return Paragraph(safe, styles[style_name])

    def _pdf_label(text: object, fallback: str = "") -> str:
        return _repair_mojibake_text(safe_value(text, fallback=fallback))

    def _fit_image(path: Path | None, max_width_mm: float, max_height_mm: float) -> Image | None:
        return _fit_pdf_image(
            path,
            max_width_mm=max_width_mm,
            max_height_mm=max_height_mm,
            mm_unit=mm,
            ImageClass=Image,
            ImageReaderClass=ImageReader,
        )

    def _box(
        flowables: list[object],
        *,
        background=None,
        padding: int = 8,
        width_mm: float = 174,
        border_color=None,
        accent_color=None,
    ) -> Table:
        return _build_threshold_box(
            flowables,
            TableClass=Table,
            TableStyleClass=TableStyle,
            mm_unit=mm,
            palette=palette,
            width_mm=width_mm,
            padding=padding,
            background=background,
            border_color=border_color,
            accent_color=accent_color,
        )

    def _page_header(title: str, subtitle: str = "", *, eyebrow: str = "Threshold S&C") -> Table:
        return _build_pdf_page_title(
            title,
            paragraph_builder=_p,
            TableClass=Table,
            TableStyleClass=TableStyle,
            mm_unit=mm,
            palette=palette,
            subtitle=subtitle,
            eyebrow=eyebrow,
        )

    def _decision_box(
        title: str,
        text: object,
        *,
        width_mm: float = 174,
        note: str = "",
        inverted: bool = True,
    ) -> Table:
        return _build_decision_box(
            title,
            text,
            paragraph_builder=_p,
            TableClass=Table,
            TableStyleClass=TableStyle,
            mm_unit=mm,
            palette=palette,
            width_mm=width_mm,
            note=note,
            inverted=inverted,
        )

    def _two_column(
        left: object,
        right: object,
        *,
        left_width_mm: float = 86,
        right_width_mm: float = 88,
        gap_mm: float = 4,
        padding: int = 0,
    ) -> Table:
        table = Table(
            [[left or "", right or ""]],
            colWidths=[left_width_mm * mm, right_width_mm * mm],
            spaceBefore=0,
            spaceAfter=0,
            hAlign="LEFT",
        )
        table.setStyle(
            TableStyle(
                [
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("LEFTPADDING", (0, 0), (-1, -1), padding),
                    ("RIGHTPADDING", (0, 0), (-1, -1), padding),
                    ("TOPPADDING", (0, 0), (-1, -1), 0),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
                    ("RIGHTPADDING", (0, 0), (0, 0), gap_mm * mm / 2),
                    ("LEFTPADDING", (1, 0), (1, 0), gap_mm * mm / 2),
                ]
            )
        )
        return table

    def _chart_panel(chart_flowable: object, *, title: str = "", note: str = "", width_mm: float = 174) -> Table:
        return _build_chart_container(
            chart_flowable,
            paragraph_builder=_p,
            TableClass=Table,
            TableStyleClass=TableStyle,
            mm_unit=mm,
            palette=palette,
            title=title,
            note=note,
            width_mm=width_mm,
            accent_color=palette["steel"],
        )

    def _brand_lockup() -> object | None:
        wordmark = _resolve_brand_asset_path("wordmark")
        icon = _resolve_brand_asset_path("icon")
        wordmark_logo = _fit_image(wordmark, 140, 18) if wordmark is not None else None
        icon_logo = _fit_image(icon, 14, 14) if icon is not None else None
        if icon_logo is not None and wordmark_logo is not None:
            return Table(
                [[icon_logo, wordmark_logo]],
                colWidths=[18 * mm, 128 * mm],
                style=TableStyle(
                    [
                        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                        ("LEFTPADDING", (0, 0), (-1, -1), 0),
                        ("RIGHTPADDING", (0, 0), (-1, -1), 0),
                        ("TOPPADDING", (0, 0), (-1, -1), 0),
                        ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
                    ]
                ),
            )
        return wordmark_logo or icon_logo

    def _missing_box(message: str) -> Table:
        return _box([_p(message, "ProfBody"), _p("Faltan datos para generar una interpretación confiable.", "ProfMuted")])

    def _collapsed_box(message: str) -> Table:
        return _box([_p(line, "ProfBody") for line in str(message).splitlines() if line.strip()])

    def _overview_table(overview: dict[str, object]) -> Table:
        statuses = overview.get("statuses", {})
        data = [
            [_p("Atleta", "ProfCardTitle"), _p(overview.get("athlete", PDF_MISSING_TEXT), "ProfBody")],
            [_p("Fecha del reporte", "ProfCardTitle"), _p(overview.get("date", PDF_MISSING_TEXT), "ProfBody")],
            [_p("Período analizado", "ProfCardTitle"), _p(overview.get("period", PDF_MISSING_TEXT), "ProfBody")],
        ]
        for label, value in statuses.items():
            data.append([_p(label, "ProfCardTitle"), _p(value, "ProfBody")])
        data.extend(
            [
                [_p("Lectura principal", "ProfCardTitle"), _p(overview.get("reading", PDF_MISSING_TEXT), "ProfBody")],
                [_p("Decisión sugerida", "ProfCardTitle"), _p(overview.get("decision", PDF_MISSING_TEXT), "ProfBody")],
            ]
        )
        table = Table(data, colWidths=[54 * mm, 120 * mm], hAlign="LEFT")
        table.setStyle(
            TableStyle(
                [
                    ("BOX", (0, 0), (-1, -1), 0.7, palette["line"]),
                    ("INNERGRID", (0, 0), (-1, -1), 0.35, palette["line"]),
                    ("BACKGROUND", (0, 0), (-1, -1), palette["card"]),
                    ("LEFTPADDING", (0, 0), (-1, -1), 8),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 8),
                    ("TOPPADDING", (0, 0), (-1, -1), 6),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ]
            )
        )
        return table

    def _metric_cards_table(cards: list[dict[str, object]]) -> Table:
        rows: list[list[object]] = []
        row: list[object] = []
        for idx, card in enumerate(cards, start=1):
            color = colors.HexColor(str(card.get("signal_color", "#708C9F")))
            cell_flow = [
                _p(card.get("title", "-"), "ProfCardTitle"),
                _p(card.get("value", PDF_MISSING_TEXT), "ProfCardValue"),
                _p(f"Unidad: {card.get('unit_label', PDF_MISSING_TEXT)}", "ProfMuted"),
                _p(f"Delta vs última: {card.get('delta', PDF_MISSING_TEXT)}", "ProfMuted"),
                _p(f"Z-score: {card.get('z_score', PDF_MISSING_TEXT)}", "ProfMuted"),
                _p(f"Mejor valor histórico: {card.get('best', PDF_MISSING_TEXT)}", "ProfMuted"),
                _p(f"Zona ruido/CV/TE: {card.get('threshold', PDF_MISSING_TEXT)}", "ProfMuted"),
                _p(f"Semáforo: {card.get('signal', PDF_MISSING_TEXT)}", "ProfMuted"),
                _p(card.get("te_caption", PDF_MISSING_TEXT), "ProfMuted"),
                _p(card.get("interpretation", ""), "ProfBody"),
            ]
            if card.get("large_change_warning"):
                cell_flow.append(_p(f"Advertencia: {card.get('large_change_warning')}", "ProfMuted"))
            cell = Table([[cell_flow]], colWidths=[84 * mm])
            cell.setStyle(
                TableStyle(
                    [
                        ("BOX", (0, 0), (-1, -1), 0.7, palette["line"]),
                        ("LINEABOVE", (0, 0), (-1, 0), 4, color),
                        ("BACKGROUND", (0, 0), (-1, -1), palette["card"]),
                        ("LEFTPADDING", (0, 0), (-1, -1), 8),
                        ("RIGHTPADDING", (0, 0), (-1, -1), 8),
                        ("TOPPADDING", (0, 0), (-1, -1), 8),
                        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
                        ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ]
                )
            )
            row.append(cell)
            if idx % 2 == 0:
                rows.append(row)
                row = []
        if row:
            row.append("")
            rows.append(row)
        table = Table(rows, colWidths=[87 * mm, 87 * mm], hAlign="LEFT")
        table.setStyle(TableStyle([("VALIGN", (0, 0), (-1, -1), "TOP")]))
        return table

    def _scale(value: float, min_value: float, max_value: float, start: float, length: float) -> float:
        if max_value == min_value:
            return start + (length / 2)
        return start + ((value - min_value) / (max_value - min_value) * length)

    def _line_chart(points: list[dict[str, object]], *, width_mm: float = 82, height_mm: float = 42, color=None) -> Drawing | None:
        values = [_coerce_float(point.get("value")) for point in points]
        values = [value for value in values if value is not None]
        if len(values) < 2:
            return None
        color = color or palette["steel"]
        width = width_mm * mm
        height = height_mm * mm
        left = 20
        bottom = 15
        plot_w = width - 30
        plot_h = height - 24
        min_v = min(values)
        max_v = max(values)
        if min_v == max_v:
            min_v -= 1
            max_v += 1
        drawing = Drawing(width, height)
        drawing.add(Rect(0, 0, width, height, fillColor=palette["card"], strokeColor=palette["line"], strokeWidth=0.5))
        drawing.add(Line(left, bottom, left + plot_w, bottom, strokeColor=palette["line"], strokeWidth=0.7))
        drawing.add(Line(left, bottom, left, bottom + plot_h, strokeColor=palette["line"], strokeWidth=0.7))
        coords: list[float] = []
        valid_points = [point for point in points if _coerce_float(point.get("value")) is not None]
        for idx, point in enumerate(valid_points):
            value = float(_coerce_float(point.get("value")) or 0)
            x = left + (plot_w * idx / max(1, len(valid_points) - 1))
            y = _scale(value, min_v, max_v, bottom, plot_h)
            coords.extend([x, y])
        drawing.add(PolyLine(coords, strokeColor=color, strokeWidth=1.7))
        for idx, point in enumerate(valid_points):
            value = float(_coerce_float(point.get("value")) or 0)
            x = left + (plot_w * idx / max(1, len(valid_points) - 1))
            y = _scale(value, min_v, max_v, bottom, plot_h)
            drawing.add(Circle(x, y, 2.2, fillColor=color, strokeColor=colors.white, strokeWidth=0.4))
        drawing.add(String(left, 4, _pdf_label(valid_points[0].get("label")), fontSize=6.5, fillColor=palette["gray"]))
        drawing.add(String(left + plot_w - 20, 4, _pdf_label(valid_points[-1].get("label")), fontSize=6.5, fillColor=palette["gray"]))
        drawing.add(String(left + 2, bottom + plot_h + 3, f"{max_v:.1f}", fontSize=6.5, fillColor=palette["muted"]))
        return drawing

    def _bar_chart(
        points: list[dict[str, object]],
        *,
        width_mm: float = 82,
        height_mm: float = 42,
        title: str = "",
        y_unit: str = "UA",
        show_values: bool = True,
    ) -> Drawing | None:
        values = [_coerce_float(point.get("value")) for point in points]
        values = [value for value in values if value is not None]
        if not values:
            return None
        width = width_mm * mm
        height = height_mm * mm
        left = 20
        bottom = 15
        top_pad = 13 if title else 5
        plot_w = width - 32
        plot_h = height - bottom - top_pad
        max_v = max(values) if max(values) > 0 else 1
        drawing = Drawing(width, height)
        drawing.add(Rect(0, 0, width, height, fillColor=palette["card"], strokeColor=palette["line"], strokeWidth=0.5))
        drawing.add(Line(left, bottom, left + plot_w, bottom, strokeColor=palette["line"], strokeWidth=0.7))
        drawing.add(Line(left, bottom, left, bottom + plot_h, strokeColor=palette["line"], strokeWidth=0.7))
        if title:
            drawing.add(String(left, height - 9, _pdf_label(title), fontSize=7.5, fillColor=palette["navy"]))
        if y_unit:
            drawing.add(String(4, bottom + plot_h - 2, _pdf_label(y_unit), fontSize=6.5, fillColor=palette["muted"]))
        bar_gap = 3
        bar_w = max(4, (plot_w - (bar_gap * max(0, len(points) - 1))) / max(1, len(points)))
        for idx, point in enumerate(points):
            value = _coerce_float(point.get("value"))
            if value is None:
                continue
            x = left + idx * (bar_w + bar_gap)
            bar_h = (value / max_v) * plot_h
            drawing.add(Rect(x, bottom, bar_w, bar_h, fillColor=palette["steel"], strokeColor=None))
            if show_values and value > 0:
                drawing.add(String(x, bottom + bar_h + 2, f"{value:.0f}", fontSize=5.5, fillColor=palette["gray"]))
            drawing.add(String(x, 4, _pdf_label(point.get("label")), fontSize=6, fillColor=palette["gray"]))
        drawing.add(String(left + 2, bottom + plot_h + 3, f"{max_v:.0f}", fontSize=6.5, fillColor=palette["muted"]))
        return drawing

    def _weekly_ema_chart(points: list[dict[str, object]], *, width_mm: float = 174, height_mm: float = 55) -> Drawing | None:
        values = [_coerce_float(point.get("value")) for point in points]
        emas = [_coerce_float(point.get("ema")) for point in points]
        values = [value for value in values if value is not None]
        emas = [value for value in emas if value is not None]
        if len(values) < 2 or len(emas) < 2:
            return None
        width = width_mm * mm
        height = height_mm * mm
        left = 28
        bottom = 17
        plot_w = width - 46
        plot_h = height - 32
        max_v = max(values + emas)
        min_v = min([0] + values + emas)
        if min_v == max_v:
            max_v += 1
        drawing = Drawing(width, height)
        drawing.add(Rect(0, 0, width, height, fillColor=palette["card"], strokeColor=palette["line"], strokeWidth=0.5))
        drawing.add(String(left, height - 9, "sRPE semanal + EMA 6 semanas", fontSize=8, fillColor=palette["navy"]))
        drawing.add(Line(left, bottom, left + plot_w, bottom, strokeColor=palette["line"], strokeWidth=0.7))
        drawing.add(Line(left, bottom, left, bottom + plot_h, strokeColor=palette["line"], strokeWidth=0.7))
        drawing.add(String(7, bottom + plot_h - 2, "UA", fontSize=6.5, fillColor=palette["muted"]))
        bar_gap = 4
        bar_w = max(5, (plot_w - (bar_gap * max(0, len(points) - 1))) / max(1, len(points)))
        ema_coords: list[float] = []
        for idx, point in enumerate(points):
            value = _coerce_float(point.get("value"))
            ema = _coerce_float(point.get("ema"))
            x = left + idx * (bar_w + bar_gap)
            if value is not None:
                bar_h = ((value - min_v) / (max_v - min_v)) * plot_h
                drawing.add(Rect(x, bottom, bar_w, bar_h, fillColor=colors.HexColor("#DCE5EA"), strokeColor=None))
                if idx == len(points) - 1:
                    drawing.add(String(x, bottom + bar_h + 2, f"{value:.0f}", fontSize=6, fillColor=palette["gray"]))
            if ema is not None:
                ema_x = x + (bar_w / 2)
                ema_y = _scale(ema, min_v, max_v, bottom, plot_h)
                ema_coords.extend([ema_x, ema_y])
        if len(ema_coords) >= 4:
            drawing.add(PolyLine(ema_coords, strokeColor=palette["navy"], strokeWidth=2.0))
        for idx, point in enumerate(points):
            if idx in {0, len(points) - 1}:
                x = left + idx * (bar_w + bar_gap)
                drawing.add(String(x, 4, _pdf_label(point.get("label")), fontSize=6.5, fillColor=palette["gray"]))
            if point.get("evaluation"):
                x = left + idx * (bar_w + bar_gap) + (bar_w / 2)
                drawing.add(Line(x, bottom, x, bottom + plot_h, strokeColor=palette["orange"], strokeWidth=0.8))
                drawing.add(String(x + 2, bottom + plot_h - 8, "evaluación", fontSize=5.8, fillColor=palette["orange"]))
        drawing.add(String(left + 2, bottom + plot_h + 3, f"{max_v:.0f}", fontSize=6.5, fillColor=palette["muted"]))
        legend_y = height - 10
        legend_x = left + plot_w - 78
        drawing.add(Rect(legend_x, legend_y - 4, 7, 4, fillColor=colors.HexColor("#DCE5EA"), strokeColor=None))
        drawing.add(String(legend_x + 10, legend_y - 4, "sRPE semanal", fontSize=6.5, fillColor=palette["gray"]))
        drawing.add(Line(legend_x + 50, legend_y - 2, legend_x + 62, legend_y - 2, strokeColor=palette["navy"], strokeWidth=1.4))
        drawing.add(String(legend_x + 65, legend_y - 4, "EMA 6", fontSize=6.5, fillColor=palette["gray"]))
        return drawing

    def _wellness_chart(points: list[dict[str, object]], *, width_mm: float = 82, height_mm: float = 42) -> Drawing | None:
        valid_values = [_coerce_float(point.get("score")) for point in points]
        valid_values = [value for value in valid_values if value is not None]
        if len(valid_values) < 2:
            return None
        width = width_mm * mm
        height = height_mm * mm
        left = 24
        bottom = 15
        plot_w = width - 36
        plot_h = height - 28
        min_v = 1.0
        max_v = 5.0
        drawing = Drawing(width, height)
        drawing.add(Rect(0, 0, width, height, fillColor=palette["card"], strokeColor=palette["line"], strokeWidth=0.5))
        drawing.add(String(left, height - 9, "Wellness score", fontSize=7.5, fillColor=palette["navy"]))
        drawing.add(Line(left, bottom, left + plot_w, bottom, strokeColor=palette["line"], strokeWidth=0.7))
        drawing.add(Line(left, bottom, left, bottom + plot_h, strokeColor=palette["line"], strokeWidth=0.7))
        coords: list[float] = []
        valid_points = [point for point in points if _coerce_float(point.get("score")) is not None]
        for idx, point in enumerate(valid_points):
            value = float(_coerce_float(point.get("score")) or 0)
            x = left + (plot_w * idx / max(1, len(valid_points) - 1))
            y = _scale(value, min_v, max_v, bottom, plot_h)
            coords.extend([x, y])
        drawing.add(PolyLine(coords, strokeColor=palette["green"], strokeWidth=1.7))
        for idx, point in enumerate(valid_points):
            if idx not in {0, len(valid_points) - 1}:
                continue
            value = float(_coerce_float(point.get("score")) or 0)
            x = left + (plot_w * idx / max(1, len(valid_points) - 1))
            y = _scale(value, min_v, max_v, bottom, plot_h)
            drawing.add(Circle(x, y, 2.2, fillColor=palette["green"], strokeColor=colors.white, strokeWidth=0.4))
        drawing.add(String(left, 4, _pdf_label(valid_points[0].get("label")), fontSize=6.5, fillColor=palette["gray"]))
        drawing.add(String(left + plot_w - 20, 4, _pdf_label(valid_points[-1].get("label")), fontSize=6.5, fillColor=palette["gray"]))
        drawing.add(String(left + 2, bottom + plot_h + 3, f"{max_v:.1f}", fontSize=6.5, fillColor=palette["muted"]))
        drawing.add(String(left + plot_w - 62, bottom + plot_h + 3, "Wellness score", fontSize=6.5, fillColor=palette["gray"]))
        return drawing

    def _unpack_metric_row(row: dict[str, object] | tuple[str, str]) -> tuple[str, str, str]:
        if isinstance(row, dict):
            return (
                str(row.get("label", "")),
                safe_value(row.get("value"), fallback=PDF_MISSING_TEXT),
                str(row.get("note", "") or ""),
            )
        label, value = row
        return str(label), safe_value(value, fallback=PDF_MISSING_TEXT), ""

    def _key_value_table(
        rows: list[dict[str, object] | tuple[str, str]],
        *,
        label_width_mm: float = 54,
        value_width_mm: float = 120,
    ) -> Table:
        if not rows:
            return _collapsed_box("Faltan datos para construir esta tabla.")
        data = []
        for raw_row in rows:
            label, value, note = _unpack_metric_row(raw_row)
            value_flow: list[object] = [_p(value, "ProfBody")]
            if note:
                value_flow.append(_p(note, "ProfMutedItalic"))
            data.append([_p(label, "ProfCardTitle"), value_flow])
        table = Table(data, colWidths=[label_width_mm * mm, value_width_mm * mm], hAlign="LEFT")
        table.setStyle(
            TableStyle(
                [
                    ("BOX", (0, 0), (-1, -1), 0.8, palette["line_dark"]),
                    ("INNERGRID", (0, 0), (-1, -1), 0.35, palette["line"]),
                    ("BACKGROUND", (0, 0), (0, -1), palette["panel_alt"]),
                    ("ROWBACKGROUNDS", (1, 0), (-1, -1), [palette["card"], palette["panel"]]),
                    ("LEFTPADDING", (0, 0), (-1, -1), 7),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 7),
                    ("TOPPADDING", (0, 0), (-1, -1), 6),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ]
            )
        )
        return table

    def _mini_cards_table(
        rows: list[dict[str, object] | tuple[str, str]],
        *,
        columns: int = 2,
        width_mm: float = 174,
    ) -> Table:
        if not rows:
            return _collapsed_box("Faltan datos para construir estos indicadores.")
        columns = max(1, min(columns, len(rows)))
        cell_width_mm = (width_mm - (max(0, columns - 1) * 3)) / columns
        cells: list[object] = []
        for raw_row in rows:
            label, value, note = _unpack_metric_row(raw_row)
            flowables: list[object] = [_p(label, "ProfCardTitle"), _p(value, "ProfCardValue")]
            if note:
                flowables.append(_p(note, "ProfMutedItalic"))
            card = Table(
                [[flowables]],
                colWidths=[cell_width_mm * mm],
            )
            card.setStyle(
                TableStyle(
                    [
                        ("BOX", (0, 0), (-1, -1), 0.8, palette["line_dark"]),
                        ("LINEABOVE", (0, 0), (-1, 0), 2.2, palette["steel"]),
                        ("BACKGROUND", (0, 0), (-1, -1), palette["card"]),
                        ("LEFTPADDING", (0, 0), (-1, -1), 6),
                        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
                        ("TOPPADDING", (0, 0), (-1, -1), 7),
                        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
                        ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ]
                )
            )
            cells.append(card)
        table_rows = [cells[idx:idx + columns] for idx in range(0, len(cells), columns)]
        if table_rows and len(table_rows[-1]) < columns:
            table_rows[-1].extend([""] * (columns - len(table_rows[-1])))
        table = Table(table_rows, colWidths=[cell_width_mm * mm] * columns, hAlign="LEFT")
        table.setStyle(
            TableStyle(
                [
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("LEFTPADDING", (0, 0), (-1, -1), 0),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 0),
                    ("TOPPADDING", (0, 0), (-1, -1), 0),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
                ]
            )
        )
        return table

    def _load_summary_rows(payload: dict[str, object]) -> list[dict[str, object]]:
        scope = str(payload.get("analysis_scope") or "")
        change_pct = _coerce_float(payload.get("weekly_change_pct"))
        change = _coerce_float(payload.get("weekly_change"))
        if scope == "current_week_partial":
            change_text = "Semana en curso incompleta"
        elif change_pct is not None:
            change_text = f"{change_pct:+.1f}%"
        elif change is not None:
            change_text = f"{change:+.0f} UA"
        else:
            change_text = PDF_MISSING_TEXT
        total = _coerce_float(payload.get("current_week_total" if scope == "current_week_partial" else "last_week_total"))
        srpe_daily_mean = _coerce_float(payload.get("current_week_daily_mean" if scope == "current_week_partial" else "last_week_daily_mean"))
        sessions = payload.get("current_week_sessions" if scope == "current_week_partial" else "sessions_registered")
        sample_days = int(payload.get("current_week_days" if scope == "current_week_partial" else "last_week_days_with_data") or 0)
        days = payload.get("current_week_days" if scope == "current_week_partial" else "days_without_data")
        weekly_value = f"{total:.0f} UA" if total is not None else PDF_MISSING_TEXT
        daily_value = f"{srpe_daily_mean:.1f} UA" if srpe_daily_mean is not None else PDF_MISSING_TEXT
        sample_suffix = _report_sample_suffix(sample_days)
        if weekly_value != PDF_MISSING_TEXT and sample_suffix:
            weekly_value = f"{weekly_value}  {sample_suffix}"
        if daily_value != PDF_MISSING_TEXT and sample_suffix:
            daily_value = f"{daily_value}  {sample_suffix}"
        sample_note = _report_sample_warning(sample_days)
        return [
            {"label": "sRPE semanal", "value": weekly_value, "note": sample_note},
            {"label": "sRPE diario", "value": daily_value, "note": sample_note},
            {"label": "Sesiones registradas", "value": str(int(sessions or 0)) if sessions is not None else PDF_MISSING_TEXT},
            {"label": "Días con registro" if scope == "current_week_partial" else "Días sin sesión registrada", "value": str(int(days)) if days is not None else PDF_MISSING_TEXT},
            {"label": "Cambio vs semana previa", "value": change_text},
        ]

    def _wellness_summary_rows(payload: dict[str, object]) -> list[dict[str, object]]:
        summary = payload.get("last_week_summary", {}) if isinstance(payload.get("last_week_summary"), dict) else {}
        scales = payload.get("scales", {}) if isinstance(payload.get("scales"), dict) else {}
        def fmt(value: object, digits: int = 1) -> str:
            numeric = _coerce_float(value)
            return f"{numeric:.{digits}f}" if numeric is not None else PDF_MISSING_TEXT
        def with_scale(value: object, key: str) -> str:
            rendered = fmt(value)
            if rendered == PDF_MISSING_TEXT:
                return rendered
            scale = str(scales.get(key, "Escala no definida"))
            if scale == "h":
                return f"{rendered} h"
            if scale.startswith("/"):
                return f"{rendered}{scale}"
            return f"{rendered} · {scale}"
        score_value = _coerce_float(summary.get("score_mean"))
        score_meta = _report_wellness_score_label(score_value) if score_value is not None else None
        score_days = int(_coerce_float(summary.get("score_n")) or 0)
        rendered_score = (
            f"{float(score_meta['score']):.1f} / 5.0 ({score_meta['label']}) (n={score_days} días)"
            if score_meta is not None
            else PDF_MISSING_TEXT
        )
        return [
            ("Sueño promedio", with_scale(summary.get("sleep_mean"), "sleep")),
            ("Estrés promedio", with_scale(summary.get("stress_mean"), "stress")),
            ("Dolor promedio", with_scale(summary.get("pain_mean"), "pain")),
            {"label": "Wellness score", "value": rendered_score, "note": _report_sample_warning(score_days)},
            ("Días con registro", str(int(summary.get("days"))) if summary.get("days") is not None else PDF_MISSING_TEXT),
        ]

    def _quadrant_chart(section: dict[str, object], *, width_mm: float = 82, height_mm: float = 55) -> Drawing | None:
        if section.get("selected") is None:
            return None
        points = section.get("points", [])
        athlete_label = _professional_quadrant_display_name(report_athlete)
        width = width_mm * mm
        height = height_mm * mm
        left = 25
        bottom = 18
        plot_w = width - 38
        plot_h = height - 30
        min_v = -2.5
        max_v = 2.5
        drawing = Drawing(width, height)
        drawing.add(Rect(0, 0, width, height, fillColor=palette["card"], strokeColor=palette["line"], strokeWidth=0.5))
        x0 = _scale(0, min_v, max_v, left, plot_w)
        y0 = _scale(0, min_v, max_v, bottom, plot_h)
        drawing.add(Line(left, y0, left + plot_w, y0, strokeColor=palette["line"], strokeWidth=0.8))
        drawing.add(Line(x0, bottom, x0, bottom + plot_h, strokeColor=palette["line"], strokeWidth=0.8))
        for point in points:
            x_value = max(min_v, min(max_v, float(point["x"])))
            y_value = max(min_v, min(max_v, float(point["y"])))
            x = _scale(x_value, min_v, max_v, left, plot_w)
            y = _scale(y_value, min_v, max_v, bottom, plot_h)
            if point.get("selected"):
                drawing.add(Circle(x, y, 4.2, fillColor=palette["navy"], strokeColor=colors.white, strokeWidth=0.8))
                drawing.add(
                    String(
                        min(x + 5, left + plot_w - 48),
                        min(y + 5, bottom + plot_h - 8),
                        _pdf_label(athlete_label),
                        fontSize=6.4,
                        fillColor=palette["navy"],
                    )
                )
            else:
                drawing.add(Circle(x, y, 2.4, fillColor=colors.HexColor("#B8C2C9"), strokeColor=None))
        drawing.add(String(left, 4, _pdf_label(section.get("x_label", "")), fontSize=6.5, fillColor=palette["gray"]))
        drawing.add(String(2, bottom + plot_h - 5, _pdf_label(section.get("y_label", ""))[:18], fontSize=6.5, fillColor=palette["gray"]))
        return drawing

    def _explanation_lines(payload: dict[str, object]) -> list[object]:
        if payload.get("state") == "missing":
            return [_p("Faltan datos para generar una interpretación confiable.", "ProfMuted")]
        if payload.get("selected") is not None:
            return [
                _p(f"Qué estoy viendo: {payload.get('what', PDF_MISSING_TEXT)}", "ProfMuted"),
                _p(f"Dónde se ubica el atleta: {payload.get('location', PDF_MISSING_TEXT)}", "ProfMuted"),
                _p(f"Qué significa para este atleta: {payload.get('athlete_meaning', payload.get('meaning', PDF_MISSING_TEXT))}", "ProfMuted"),
                _p(f"Qué decisión sugiere: {payload.get('decision', PDF_MISSING_TEXT)}", "ProfMuted"),
            ]
        return [
            _p(f"Qué estoy viendo: {payload.get('what', PDF_MISSING_TEXT)}", "ProfMuted"),
            _p(f"Qué significa: {payload.get('meaning', PDF_MISSING_TEXT)}", "ProfMuted"),
            _p(f"Qué decisión sugiere: {payload.get('decision', PDF_MISSING_TEXT)}", "ProfMuted"),
        ]

    def _chart_explanation_panel(chart: Drawing, title: str, payload: dict[str, object]) -> Table:
        text_flow: list[object] = [_p(title, "ProfCardTitle")]
        if payload.get("selected") is not None:
            text_flow.extend(
                [
                    _p(f"Qué estoy viendo: {payload.get('what', PDF_MISSING_TEXT)}", "ProfMuted"),
                    _p(f"Dónde se ubica el atleta: {payload.get('location', PDF_MISSING_TEXT)}", "ProfMuted"),
                    _p(
                        f"Qué significa para este atleta: {payload.get('athlete_meaning', payload.get('meaning', PDF_MISSING_TEXT))}",
                        "ProfBody",
                    ),
                ]
            )
        else:
            text_flow.extend(
                [
                    _p(f"Qué estoy viendo: {payload.get('what', PDF_MISSING_TEXT)}", "ProfMuted"),
                    _p(f"Qué significa: {payload.get('meaning', PDF_MISSING_TEXT)}", "ProfBody"),
                ]
            )
        decision = str(payload.get("decision") or "").strip()
        if decision:
            text_flow.append(Spacer(1, 1.5 * mm))
            text_flow.append(
                _decision_box(
                    "Decisión sugerida",
                    decision,
                    width_mm=86,
                    note="Traducir la ubicación del cuadrante a una prioridad operativa.",
                    inverted=False,
                )
            )
        return Table(
            [[_chart_panel(chart, width_mm=84), text_flow]],
            colWidths=[84 * mm, 90 * mm],
            style=TableStyle(
                [
                    ("BOX", (0, 0), (-1, -1), 0.8, palette["line_dark"]),
                    ("BACKGROUND", (0, 0), (-1, -1), palette["card"]),
                    ("LEFTPADDING", (0, 0), (-1, -1), 6),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 6),
                    ("TOPPADDING", (0, 0), (-1, -1), 6),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ]
            ),
        )

    def _quadrant_compact_panel(section: dict[str, object], *, width_mm: float = 84) -> Table | None:
        chart = _quadrant_chart(section, width_mm=max(62, width_mm - 16), height_mm=36)
        if chart is None:
            return None
        flowables: list[object] = [
            _p(str(section.get("title", "Cuadrante")), "ProfCardTitle"),
            chart,
            _p(f"Ubicación: {section.get('location', PDF_MISSING_TEXT)}", "ProfMuted"),
            _p(f"Lectura: {section.get('athlete_meaning', section.get('meaning', PDF_MISSING_TEXT))}", "ProfBody"),
        ]
        decision = str(section.get("decision") or "").strip()
        if decision:
            flowables.extend([_p("Decisión sugerida", "ProfCardTitle"), _p(decision, "ProfMuted")])
        return _box(
            flowables,
            width_mm=width_mm,
            background=palette["card"],
            border_color=palette["line_dark"],
            padding=5,
            accent_color=palette["steel"],
        )

    def _dataframe_table(
        frame: pd.DataFrame,
        *,
        col_widths_mm: list[float] | None = None,
        body_style: str = "ProfTableCell",
    ) -> Table:
        result = frame.copy() if frame is not None else pd.DataFrame()
        if result.empty:
            return _collapsed_box("Faltan datos para construir esta tabla.")
        result = result.fillna("-").astype(str)
        headers = [_p(column, "ProfTableHeader") for column in result.columns.tolist()]
        rows: list[list[object]] = [headers]
        for _, row in result.iterrows():
            rows.append([_p(value, body_style) for value in row.tolist()])
        widths = col_widths_mm or [174 / max(1, len(result.columns))] * len(result.columns)
        table = Table(rows, colWidths=[width * mm for width in widths], hAlign="LEFT")
        table.setStyle(
            TableStyle(
                [
                    ("BOX", (0, 0), (-1, -1), 0.8, palette["line_dark"]),
                    ("INNERGRID", (0, 0), (-1, -1), 0.35, palette["line"]),
                    ("BACKGROUND", (0, 0), (-1, 0), palette["navy"]),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [palette["card"], palette["panel"]]),
                    ("LEFTPADDING", (0, 0), (-1, -1), 6),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 6),
                    ("TOPPADDING", (0, 0), (-1, -1), 5),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ]
            )
        )
        return table

    def _bullet_box(
        title: str,
        lines: list[str],
        *,
        style_name: str = "ProfBody",
        width_mm: float = 174,
        background=None,
        border_color=None,
        accent_color=None,
        title_style: str = "ProfCardTitle",
    ) -> Table:
        flowables: list[object] = [_p(title, title_style)]
        cleaned = []
        for line in lines:
            text = str(line or "").strip()
            if not text:
                continue
            normalized = _professional_normalized_text(text)
            if normalized == "estimulos bajos o ausentes: faltan datos.":
                continue
            if normalized.startswith("estimulos bajos o ausentes:"):
                detail = _professional_visible_metric_text(text.partition(":")[2]).strip()
                if detail:
                    detail = detail[:1].lower() + detail[1:]
                    text = f"Menor exposición relativa: {detail}"
                else:
                    text = "Menor exposición relativa"
            cleaned.append(text)
        if not cleaned:
            cleaned = ["Faltan datos para generar una interpretación confiable."]
        flowables.extend(
            _build_threshold_bullets(
                cleaned,
                paragraph_builder=_p,
                style_name=style_name,
            )
        )
        return _box(
            flowables,
            padding=6,
            width_mm=width_mm,
            background=background,
            border_color=border_color,
            accent_color=accent_color,
        )

    def _executive_summary_table(payload: dict[str, object]) -> Table:
        rows: list[dict[str, object]] = [
            {"label": "Atleta", "value": payload.get("athlete", PDF_MISSING_TEXT)},
            {"label": "Fecha", "value": payload.get("date", PDF_MISSING_TEXT)},
            {"label": "Ventana temporal", "value": payload.get("period", PDF_MISSING_TEXT)},
            {"label": "Cobertura / calidad", "value": payload.get("coverage", PDF_MISSING_TEXT)},
            {
                "label": "Nivel de confianza",
                "value": safe_value(payload.get("confidence")),
                "note": safe_value(payload.get("confidence_detail")),
            },
        ]
        return _mini_cards_table(rows, columns=2)

    def _composite_radar_image(payload: dict[str, object], *, width_mm: float = 174, height_mm: float = 72) -> Image | None:
        if payload.get("state") == "missing" or payload.get("profile_row") is None:
            return None
        try:
            from charts.dashboard_charts import chart_composite_profile_radar
        except Exception:
            return None
        figure = chart_composite_profile_radar(payload["profile_row"], report_athlete, theme=_build_report_chart_theme())
        return _plotly_chart_image(figure, width_mm=width_mm, height_mm=height_mm, width_px=1100, height_px=760)

    def _exposure_chart_image(*, width_mm: float = 174, height_mm: float = 68) -> Image | None:
        try:
            from charts.load_charts import chart_volume_by_tag
        except Exception:
            return None
        prepared = state.get("prepared_raw_df")
        if prepared is None:
            prepared = prepare_raw_workouts_df(state.get("raw_df"))
        if prepared is None or prepared.empty:
            return None
        figure = chart_volume_by_tag(prepared, report_athlete, theme=_build_report_chart_theme())
        return _plotly_chart_image(figure, width_mm=width_mm, height_mm=height_mm, width_px=1200, height_px=680)

    cards = _build_professional_metric_cards(state, report_athlete)
    evolution_sections = _build_professional_evolution_sections(state, report_athlete)
    quadrant_sections = _build_professional_quadrant_sections(state, report_athlete)
    training_context = _build_professional_training_context(state, report_athlete)
    internal_load = _build_professional_internal_load_context(state, report_athlete)
    wellness_context = _professional_wellness_context(state, report_athlete)
    overview = _build_professional_report_overview(state, report_athlete, cards, training_context, internal_load)
    available_cards, missing_metric_titles = _professional_metric_display_groups(cards)
    has_evaluations = bool(available_cards)
    evaluation_state = "missing" if not has_evaluations else "partial" if missing_metric_titles else "available"
    has_evolution_charts = _professional_assessment_date_count(state, report_athlete) >= 2 and any(
        len(section.get("points", [])) >= 2 for section in evolution_sections
    )
    has_quadrant_charts = _professional_any_quadrant_ready(quadrant_sections)
    assessment_interval_warning = _professional_short_assessment_interval_warning(state, report_athlete)
    integrated_lines = _build_professional_integrated_interpretation(
        internal_load,
        wellness_context,
        evaluation_state=evaluation_state,
        assessment_interval_warning=assessment_interval_warning,
    )
    next_steps = _build_professional_next_steps(evaluation_state)
    composite_profile = _build_professional_composite_profile_payload(state, report_athlete)
    change_payload = _build_professional_change_payload(state, report_athlete)
    isometric_payload = _build_professional_isometric_payload(state, report_athlete, cards)
    force_time_payload = isometric_payload.get("imtp_payload", {})
    load_tolerance_payload = _build_professional_load_tolerance_payload(state, report_athlete, internal_load)
    wellness_availability_payload = _build_professional_wellness_availability_payload(
        state,
        report_athlete,
        training_context,
        internal_load,
        wellness_context,
    )
    exposure_payload = _build_professional_exposure_payload(state, report_athlete, change_payload=change_payload)
    integrated_decision_payload = _build_professional_integrated_decision_payload(
        state,
        report_athlete,
        evaluation_state=evaluation_state,
        assessment_interval_warning=assessment_interval_warning,
        composite_payload=composite_profile,
        change_payload=change_payload,
        load_payload=load_tolerance_payload,
        wellness_payload=wellness_availability_payload,
        exposure_payload=exposure_payload,
        training_context=training_context,
    )
    action_plan_payload = _build_professional_action_plan_payload(
        state,
        report_athlete,
        evaluation_state=evaluation_state,
        composite_payload=composite_profile,
        change_payload=change_payload,
        integrated_payload=integrated_decision_payload,
    )
    executive_payload = _build_professional_executive_payload(
        state,
        report_athlete,
        evaluation_state=evaluation_state,
        assessment_interval_warning=assessment_interval_warning,
        composite_payload=composite_profile,
        change_payload=change_payload,
        load_payload=load_tolerance_payload,
        wellness_payload=wellness_availability_payload,
        exposure_payload=exposure_payload,
        training_context=training_context,
        integrated_payload=integrated_decision_payload,
    )

    def _append_metric_cards(target: list[object]) -> None:
        if not available_cards:
            target.append(_p("Evaluaciones físicas no disponibles", "ProfSection"))
            target.append(_collapsed_box(PROFESSIONAL_NO_EVALUATION_TEXT))
            return
        target.append(_p("Tarjetas de evaluación", "ProfSection"))
        target.append(_metric_cards_table(available_cards))
        if missing_metric_titles:
            target.append(Spacer(1, 3 * mm))
            target.append(_box([_p(f"Métricas no disponibles: {', '.join(missing_metric_titles)}.", "ProfMuted")]))
        if any(card.get("title") in {"RSI", "mRSI"} for card in available_cards):
            target.append(Spacer(1, 3 * mm))
            target.append(_box([_p(f"Nota metodológica: {PROFESSIONAL_RSI_METHOD_NOTE}", "ProfMuted")], padding=5))

    def _append_evolution_section(target: list[object]) -> None:
        target.extend(
            [
                _p("Evolución del perfil físico", "ProfSection"),
                _p("Evolución entre fechas de evaluación. No interpretar como readiness semanal.", "ProfMuted"),
            ]
        )
        if assessment_interval_warning:
            target.append(_box([_p(assessment_interval_warning, "ProfMuted")], padding=5))
            target.append(Spacer(1, 3 * mm))
        if _professional_assessment_date_count(state, report_athlete) < 2:
            target.append(_collapsed_box(PROFESSIONAL_NO_EVOLUTION_TEXT))
            return
        missing_evolution_titles: list[str] = []
        selected_sections, omitted_evolution_titles = _professional_prioritized_evolution_sections(evolution_sections)
        for section in selected_sections:
            chart = _line_chart(section.get("points", []), color=palette["steel"])
            if chart is None:
                missing_evolution_titles.append(str(section.get("title", "Métrica")))
                continue
            target.append(_chart_explanation_panel(chart, str(section.get("title", "Métrica")), section))
            if section.get("large_change_warning"):
                target.append(_box([_p(str(section.get("large_change_warning")), "ProfMuted")], padding=5))
            target.append(Spacer(1, 4 * mm))
        charted_titles = {str(section.get("title", "Métrica")) for section in selected_sections}
        for section in evolution_sections:
            title = str(section.get("title", "Métrica"))
            if title not in charted_titles and len(section.get("points", [])) < 2:
                missing_evolution_titles.append(title)
        if omitted_evolution_titles:
            target.append(_box([_p(f"Evolución no graficada para compactar el reporte: {', '.join(omitted_evolution_titles)}.", "ProfMuted")], padding=5))
        if missing_evolution_titles:
            target.append(_box([_p(f"Métricas sin evolución suficiente: {', '.join(missing_evolution_titles)}. Faltan datos.", "ProfMuted")]))

    def _append_quadrants_section(target: list[object]) -> None:
        target.append(_p("Cuadrantes de perfil físico", "ProfSection"))
        if not _professional_any_quadrant_ready(quadrant_sections):
            target.append(_collapsed_box(PROFESSIONAL_NO_QUADRANTS_TEXT))
            return
        for section in quadrant_sections[:3]:
            chart = _quadrant_chart(section)
            if chart is None:
                target.append(_collapsed_box(str(section.get("message") or "Faltan datos para ubicar al atleta en este cuadrante.")))
                target.append(Spacer(1, 4 * mm))
                continue
            target.append(_chart_explanation_panel(chart, str(section.get("title", "Cuadrante")), section))
            target.append(Spacer(1, 4 * mm))

    def _append_training_section(target: list[object]) -> None:
        target.append(_p("Contexto de entrenamiento", "ProfSection"))
        if training_context["state"] == "missing":
            target.append(_collapsed_box("Faltan datos de entrenamiento para este período."))
            return
        target.append(_key_value_table([(str(label), str(value)) for label, value in training_context["rows"]]))
        target.append(Spacer(1, 3 * mm))
        target.append(_box(_explanation_lines(training_context), padding=6))

    def _append_internal_last_week(target: list[object], *, include_title: bool = True, compact: bool = False) -> None:
        if include_title:
            target.append(_p(internal_load.get("analysis_title", "Carga interna - última semana completa"), "ProfSection"))
        if internal_load["state"] == "missing":
            target.append(_collapsed_box("Faltan datos suficientes para analizar la última semana completa."))
            return
        summary_rows = _load_summary_rows(internal_load)
        if compact:
            compact_parts: list[str] = []
            for row in summary_rows:
                if isinstance(row, dict):
                    compact_parts.append(f"{row.get('label', '')}: {safe_value(row.get('value'), fallback=PDF_MISSING_TEXT)}")
                else:
                    label, value = row
                    compact_parts.append(f"{label}: {value}")
            target.append(_box([_p(" | ".join(compact_parts), "ProfMuted")], padding=5))
        else:
            target.append(_key_value_table(summary_rows))
        point_key = "current_week_points" if internal_load.get("analysis_scope") == "current_week_partial" else "daily_points"
        daily_chart = _bar_chart(
            internal_load.get(point_key, []),
            height_mm=34 if compact else 42,
            title="sRPE diario",
            y_unit="UA",
        )
        if daily_chart is None:
            target.append(Spacer(1, 3 * mm))
            message = (
                str(internal_load.get("current_week_partial_message"))
                if internal_load.get("analysis_scope") == "current_week_partial"
                else "Faltan datos suficientes para analizar la última semana completa."
            )
            target.append(_collapsed_box(message))
            return
        target.append(Spacer(1, 2 * mm if compact else 3 * mm))
        if internal_load.get("analysis_scope") == "current_week_partial":
            target.append(_collapsed_box(str(internal_load.get("current_week_partial_message"))))
            target.append(Spacer(1, 2 * mm if compact else 3 * mm))
        target.append(
            _chart_explanation_panel(
                daily_chart,
                "sRPE diario",
                {"state": "available", **internal_load["daily_explanation"]},
            )
        )

    def _append_internal_16_weeks(target: list[object], *, include_title: bool = True, compact: bool = False) -> None:
        if include_title:
            target.append(_p("Carga interna - últimas 16 semanas", "ProfSection"))
        weekly_chart = _weekly_ema_chart(internal_load.get("weekly_points", []), height_mm=42 if compact else 55)
        if weekly_chart is None:
            target.append(_collapsed_box("Faltan datos para mostrar sRPE semanal + EMA 6 semanas."))
            return
        target.append(weekly_chart)
        if internal_load.get("analysis_scope") == "current_week_partial":
            total = _coerce_float(internal_load.get("current_week_total"))
            change_pct = None
            summary = (
                f"Acumulado parcial de la semana en curso: {total:.0f} UA. No comparar directamente con semanas completas."
                if total is not None
                else "Semana en curso parcial: Faltan datos."
            )
        else:
            total = _coerce_float(internal_load.get("last_week_total"))
            change_pct = _coerce_float(internal_load.get("weekly_change_pct"))
            summary = f"Total última semana completa: {total:.0f} UA." if total is not None else "Total última semana completa: Faltan datos."
        if internal_load.get("analysis_scope") != "current_week_partial" and change_pct is not None:
            summary = f"{summary} Cambio vs semana previa: {change_pct:+.1f}%."
        target.append(
            _box(
                [
                    _p(summary, "ProfMuted"),
                    _p(
                        "Permite observar tendencia de carga, acumulaciones, descargas o aumentos bruscos del bloque.",
                        "ProfMuted",
                    ),
                ],
                padding=5 if compact else 6,
            )
        )

    rpe_daily_frame_cache: pd.DataFrame | None = None

    def _rpe_daily_frame_for_pdf() -> pd.DataFrame:
        nonlocal rpe_daily_frame_cache
        if rpe_daily_frame_cache is not None:
            return rpe_daily_frame_cache.copy()
        rpe_df = state.get("rpe_df")
        if rpe_df is None or rpe_df.empty or not {"Athlete", "Date", "sRPE"}.issubset(rpe_df.columns):
            rpe_daily_frame_cache = pd.DataFrame()
            return rpe_daily_frame_cache.copy()
        daily = rpe_df.copy()
        daily["Date"] = pd.to_datetime(daily["Date"], errors="coerce").dt.normalize()
        daily["sRPE"] = pd.to_numeric(daily["sRPE"], errors="coerce")
        daily = daily.dropna(subset=["Athlete", "Date", "sRPE"])
        daily = daily[_professional_athlete_mask(daily["Athlete"], report_athlete)]
        if daily.empty:
            rpe_daily_frame_cache = pd.DataFrame()
            return rpe_daily_frame_cache.copy()
        daily = daily.groupby("Date", as_index=False)["sRPE"].sum().sort_values("Date")
        full_index = pd.date_range(daily["Date"].min(), daily["Date"].max(), freq="D")
        daily = (
            daily.set_index("Date")
            .reindex(full_index, fill_value=0.0)
            .rename_axis("Date")
            .reset_index()
            .rename(columns={"index": "Date", "sRPE": "sRPE_diario"})
        )
        rpe_daily_frame_cache = daily
        return rpe_daily_frame_cache.copy()

    def _acwr_zone_label(value: float | None) -> str:
        if value is None:
            return PDF_MISSING_TEXT
        if value < 0.80:
            return "Subcarga"
        if value <= 1.30:
            return "Óptimo"
        if value <= 1.50:
            return "Precaución"
        return "Alto riesgo"

    def _build_acwr_ewma_frame_for_pdf() -> pd.DataFrame:
        daily = _rpe_daily_frame_for_pdf()
        if daily.empty or len(daily) < 7:
            return pd.DataFrame()
        result = daily.copy()
        result["Aguda_7d"] = result["sRPE_diario"].ewm(alpha=0.25, adjust=False).mean()
        result["Cronica_28d"] = result["sRPE_diario"].ewm(alpha=(2 / 29), adjust=False).mean()
        result["ACWR_EWMA"] = result.apply(
            lambda row: (float(row["Aguda_7d"]) / float(row["Cronica_28d"])) if _coerce_float(row["Cronica_28d"]) not in [None, 0] else None,
            axis=1,
        )
        result["Zona"] = result["ACWR_EWMA"].map(lambda value: _acwr_zone_label(_coerce_float(value)))
        return result

    def _build_monotony_strain_frame_for_pdf() -> pd.DataFrame:
        daily = _rpe_daily_frame_for_pdf()
        if daily.empty:
            return pd.DataFrame()
        result = daily.copy()
        result["week_start"] = result["Date"] - pd.to_timedelta(result["Date"].dt.weekday, unit="D")
        current_week_start = _professional_week_start(_professional_report_today())
        result = result[result["week_start"] < current_week_start]
        if result.empty:
            return pd.DataFrame()
        rows: list[dict[str, object]] = []
        for week_start, week_df in result.groupby("week_start"):
            values = week_df.sort_values("Date")["sRPE_diario"].astype(float).tolist()
            if len(values) < 7:
                values.extend([0.0] * (7 - len(values)))
            series = pd.Series(values[:7], dtype=float)
            mean_value = float(series.mean())
            std_value = float(series.std(ddof=0))
            monotony_value = None if std_value == 0 else (mean_value / std_value)
            strain_value = (float(series.sum()) * monotony_value) if monotony_value is not None else None
            rows.append(
                {
                    "Semana": pd.Timestamp(week_start).normalize(),
                    "Monotonia": monotony_value,
                    "Strain": strain_value,
                    "Alerta": bool(monotony_value is not None and monotony_value > 2.0),
                }
            )
        return pd.DataFrame(rows).sort_values("Semana").tail(16).reset_index(drop=True) if rows else pd.DataFrame()

    def _plotly_chart_image(
        figure: object | None,
        *,
        width_mm: float = 174,
        height_mm: float = 76,
        width_px: int = 1200,
        height_px: int = 560,
    ) -> Image | None:
        image_bytes = export_plotly_figure_png(figure, width=width_px, height=height_px, scale=2) if figure is not None else None
        if not image_bytes:
            return None
        image = Image(BytesIO(image_bytes), width=width_mm * mm, height=height_mm * mm)
        image.hAlign = "LEFT"
        return image

    def _build_acwr_figure(acwr_frame: pd.DataFrame) -> object | None:
        if acwr_frame.empty:
            return None
        theme = _build_report_chart_theme()
        try:
            from charts.load_charts import chart_acwr
            return chart_acwr(acwr_frame, report_athlete, theme=theme)
        except Exception:
            try:
                import plotly.graph_objects as go
            except Exception:
                return None
            fig = go.Figure()
            for y0, y1, fill, label in [
                (0.00, 0.80, "#E3F2FD", "Subcarga"),
                (0.80, 1.30, "#E8F5E9", "Óptimo"),
                (1.30, 1.50, "#FFF9C4", "Precaución"),
                (1.50, max(2.5, float(pd.to_numeric(acwr_frame["ACWR_EWMA"], errors="coerce").max() or 2.5) + 0.2), "#FFEBEE", "Alto riesgo"),
            ]:
                fig.add_hrect(y0=y0, y1=y1, fillcolor=fill, line_width=0, layer="below", annotation_text=label, annotation_position="right")
            fig.add_trace(
                go.Scatter(
                    x=acwr_frame["Date"],
                    y=acwr_frame["ACWR_EWMA"],
                    mode="lines",
                    line=dict(color="#1E4D8C", width=2),
                    name="ACWR EWMA",
                )
            )
            fig.update_layout(height=420, margin=dict(l=44, r=28, t=48, b=40), xaxis=dict(title="Fecha"), yaxis=dict(title="ACWR EWMA"), showlegend=False)
            return fig

    def _build_monotony_figure(monotony_frame: pd.DataFrame) -> object | None:
        if monotony_frame.empty:
            return None
        theme = _build_report_chart_theme()
        try:
            from charts.load_charts import chart_monotony_strain
            return chart_monotony_strain(monotony_frame, theme=theme)
        except Exception:
            try:
                import plotly.graph_objects as go
                from plotly.subplots import make_subplots
            except Exception:
                return None
            fig = make_subplots(specs=[[{"secondary_y": True}]])
            bar_colors = ["#D94F4F" if bool(row["Alerta"]) else "#4FC97E" for _, row in monotony_frame.iterrows()]
            fig.add_trace(go.Bar(x=monotony_frame["Semana"], y=monotony_frame["Strain"], marker_color=bar_colors, name="Strain"), secondary_y=False)
            fig.add_trace(
                go.Scatter(
                    x=monotony_frame["Semana"],
                    y=monotony_frame["Monotonia"],
                    mode="lines+markers",
                    line=dict(color="#E8C84A", width=2, dash="dash"),
                    name="Monotonía",
                ),
                secondary_y=True,
            )
            fig.add_hline(y=2.0, line_dash="dash", line_color="#E8C84A", annotation_text="Umbral monotonía", secondary_y=True)
            fig.update_layout(height=420, margin=dict(l=44, r=28, t=48, b=40), showlegend=False)
            fig.update_yaxes(title_text="Strain", secondary_y=False)
            fig.update_yaxes(title_text="Monotonía", secondary_y=True)
            return fig

    def _append_acwr_section(target: list[object], *, compact: bool = False) -> None:
        target.append(_p("GESTIÓN DE RIESGO", "ProfMuted"))
        target.append(_p("Ratio Agudo:Crónico (ACWR EWMA)", "ProfSection"))
        target.append(_p("¿Cómo se ubica la carga reciente respecto a la carga acumulada habitual?", "ProfMuted"))
        acwr_frame = _build_acwr_ewma_frame_for_pdf()
        if acwr_frame.empty:
            target.append(_collapsed_box("Sin suficientes datos de carga para calcular ACWR (mínimo 7 días requeridos)."))
            return
        acwr_image = _plotly_chart_image(_build_acwr_figure(acwr_frame), height_mm=70 if compact else 78)
        if acwr_image is None:
            target.append(_collapsed_box("No se pudo renderizar el gráfico de ACWR para este reporte."))
            return
        target.append(acwr_image)
        latest_acwr = _coerce_float(acwr_frame.dropna(subset=["ACWR_EWMA"]).tail(1).iloc[0].get("ACWR_EWMA")) if not acwr_frame.dropna(subset=["ACWR_EWMA"]).empty else None
        zone_label = _acwr_zone_label(latest_acwr)
        caption = (
            "ACWR EWMA: rango operativo habitual 0.80–1.30. Valores >1.50 sugieren exposición aguda relativamente alta y requieren revisar contexto, calendario y tolerancia individual "
            "(Gabbett, 2016, BJSM; Hulin et al., 2016, BJSM). "
            f"Interpretación: zona actual = {zone_label}."
        )
        target.append(_box([_p(caption, "ProfMuted")], padding=5 if compact else 6))

    def _append_monotony_section(target: list[object], *, compact: bool = False) -> None:
        target.append(_p("Monotonía y Strain semanal", "ProfSection"))
        target.append(_p("¿Qué tan variado fue el estímulo? ¿Se acumuló estrés sin descarga?", "ProfMuted"))
        monotony_frame = _build_monotony_strain_frame_for_pdf()
        if monotony_frame.empty:
            target.append(_collapsed_box("Sin suficientes datos para calcular monotonía y strain semanal."))
            return
        monotony_image = _plotly_chart_image(_build_monotony_figure(monotony_frame), height_mm=70 if compact else 78)
        if monotony_image is None:
            target.append(_collapsed_box("No se pudo renderizar el gráfico de monotonía y strain para este reporte."))
            return
        target.append(monotony_image)
        target.append(
            _box(
                [
                    _p(
                        "Monotonía >2.0 sugiere distribución homogénea de la carga y menor variabilidad semanal. "
                        "Strain alto sugiere acumulación de estrés semanal y conviene leerlo junto con volumen total, calendario y wellness.",
                        "ProfMuted",
                    )
                ],
                padding=5 if compact else 6,
            )
        )

    def _professional_wellness_rows_for_pdf(payload: dict[str, object]) -> list[dict[str, object]]:
        summary = payload.get("last_week_summary", {}) if isinstance(payload.get("last_week_summary"), dict) else {}
        scales = payload.get("scales", {}) if isinstance(payload.get("scales"), dict) else {}

        def fmt(value: object, digits: int = 1) -> str:
            numeric = _coerce_float(value)
            return f"{numeric:.{digits}f}" if numeric is not None else PDF_MISSING_TEXT

        def with_scale(value: object, key: str) -> str:
            rendered = fmt(value)
            if rendered == PDF_MISSING_TEXT:
                return rendered
            scale = str(scales.get(key, "Escala no definida"))
            if scale == "h":
                return f"{rendered} h"
            if scale.startswith("/"):
                return f"{rendered}{scale}"
            return f"{rendered} · {scale}"

        def metric_row(label: str, value: str, n_value: object, *, include_count: bool = True) -> dict[str, object]:
            count = int(_coerce_float(n_value) or 0)
            if include_count and value != PDF_MISSING_TEXT and count > 0:
                value = f"{value}  {_report_sample_suffix(count)}"
            return {"label": label, "value": value, "note": _report_sample_warning(count)}

        score_value = _coerce_float(summary.get("score_mean"))
        score_meta = _report_wellness_score_label(score_value) if score_value is not None else None
        score_days = int(_coerce_float(summary.get("score_n")) or 0)
        rendered_score = (
            f"{float(score_meta['score']):.1f} / 5.0 ({score_meta['label']}) (n={score_days} días)"
            if score_meta is not None
            else PDF_MISSING_TEXT
        )

        return [
            metric_row("Sueño promedio", with_scale(summary.get("sleep_mean"), "sleep"), summary.get("sleep_n")),
            metric_row("Estrés promedio", with_scale(summary.get("stress_mean"), "stress"), summary.get("stress_n")),
            metric_row("Dolor promedio", with_scale(summary.get("pain_mean"), "pain"), summary.get("pain_n")),
            metric_row("Wellness score", rendered_score, summary.get("score_n"), include_count=False),
            {"label": "Días con registro", "value": str(int(summary.get("days"))) if summary.get("days") is not None else PDF_MISSING_TEXT},
        ]

    def _append_wellness_last_week(target: list[object], *, include_title: bool = True, compact: bool = False) -> None:
        if include_title:
            target.append(_p(wellness_context.get("analysis_title", "Wellness - última semana completa"), "ProfSection"))
        if wellness_context.get("state") == "missing":
            target.append(_collapsed_box("Faltan datos de wellness para este período."))
            return
        summary_rows = _professional_wellness_rows_for_pdf(wellness_context)
        target.append(_mini_cards_table(summary_rows))
        if wellness_context.get("state") == "partial" and not wellness_context.get("trend_allowed"):
            target.append(Spacer(1, 3 * mm))
            target.append(_collapsed_box(str(wellness_context.get("partial_message") or _professional_wellness_partial_message(wellness_context))))
            return
        daily_chart = _wellness_chart(wellness_context.get("daily_points", []), height_mm=34 if compact else 42)
        if daily_chart is None:
            target.append(Spacer(1, 3 * mm))
            target.append(_collapsed_box("Faltan registros suficientes para tendencia. La lectura debe tomarse como alerta contextual puntual."))
            return
        target.append(Spacer(1, 2 * mm if compact else 3 * mm))
        target.append(
            _chart_explanation_panel(
                daily_chart,
                "Wellness score",
                {
                    "state": wellness_context.get("state"),
                    "what": "Tendencia diaria del wellness score de la última semana completa." if wellness_context.get("analysis_scope") == "last_complete_week" else "Registro parcial de wellness de la semana en curso.",
                    "meaning": "Muestra recuperación percibida y carga contextual alrededor del entrenamiento sin mezclar escalas crudas.",
                    "decision": "Regular volumen, densidad o selección de ejercicios si empeoran sueño, estrés o dolor." if wellness_context.get("analysis_scope") == "last_complete_week" else "No leerlo como tendencia semanal cerrada; usarlo solo como alerta contextual hasta completar la semana.",
                },
            )
        )

    def _append_wellness_16_weeks(target: list[object], *, include_title: bool = True, compact: bool = False) -> None:
        if include_title:
            target.append(_p("Wellness - últimas 16 semanas", "ProfSection"))
        if wellness_context.get("state") == "missing":
            target.append(_collapsed_box("Faltan datos de wellness para este período."))
            return
        if wellness_context.get("state") == "partial" and not wellness_context.get("trend_allowed"):
            target.append(_collapsed_box("Tendencia de wellness no disponible por baja cantidad de registros."))
            return
        weekly_chart = _wellness_chart(wellness_context.get("weekly_points", []), width_mm=174, height_mm=38 if compact else 46)
        if weekly_chart is None:
            target.append(_collapsed_box("Faltan registros suficientes para tendencia. La lectura debe tomarse como alerta contextual puntual."))
            return
        target.append(weekly_chart)
        missing_variables = wellness_context.get("missing_variables") or []
        note = "Cruzar la tendencia del score con carga interna ayuda a distinguir tolerancia aparente de fatiga contextual sin mezclar escalas crudas."
        if missing_variables:
            note = f"{note} Variables no disponibles: {', '.join(map(str, missing_variables))}."
        target.append(_box([_p(note, "ProfMuted")], padding=5 if compact else 6))

    def _append_acute_interpretation(target: list[object]) -> None:
        target.append(_p("Interpretación aguda", "ProfSection"))
        lines: list[str] = []
        if internal_load.get("analysis_scope") == "current_week_partial":
            total = _coerce_float(internal_load.get("current_week_total"))
            sessions = int(_coerce_float(internal_load.get("current_week_sessions")) or 0)
            if total is not None:
                lines.append(
                    f"El acumulado parcial de la semana actual es {total:.0f} UA con {sessions} sesiones registradas. "
                    "No debe compararse directamente con una semana completa."
                )
            else:
                lines.append("La semana en curso está incompleta y faltan datos suficientes para interpretarla.")
        else:
            total = _coerce_float(internal_load.get("last_week_total"))
            weekly_change_pct = _coerce_float(internal_load.get("weekly_change_pct"))
            if total is not None:
                load_text = f"La carga interna de la última semana completa fue {total:.0f} UA."
                if weekly_change_pct is not None:
                    load_text = f"{load_text} Cambió {weekly_change_pct:+.1f}% respecto a la semana previa."
                else:
                    load_text = f"{load_text} No hay comparación suficiente contra la semana previa."
                lines.append(load_text)
            else:
                lines.append("Faltan datos para interpretar la carga interna aguda.")

        summary = wellness_context.get("last_week_summary", {}) if isinstance(wellness_context.get("last_week_summary"), dict) else {}
        scales = wellness_context.get("scales", {}) if isinstance(wellness_context.get("scales"), dict) else {}
        stress_mean = _coerce_float(summary.get("stress_mean"))
        pain_mean = _coerce_float(summary.get("pain_mean"))
        sleep_mean = _coerce_float(summary.get("sleep_mean"))
        if wellness_context.get("state") == "missing":
            lines.append("Faltan datos de wellness para cruzar la carga aguda con recuperación percibida.")
        else:
            details = []
            if sleep_mean is not None:
                details.append(f"sueño {sleep_mean:.1f} h")
            if stress_mean is not None:
                details.append(f"estrés {stress_mean:.1f}{scales.get('stress', '')}")
            if pain_mean is not None:
                details.append(f"dolor {pain_mean:.1f}{scales.get('pain', '')}")
            wellness_label = "Wellness última semana completa" if wellness_context.get("analysis_scope") == "last_complete_week" else "Wellness semana en curso"
            lines.append(f"{wellness_label}: " + (", ".join(details) if details else PDF_MISSING_TEXT) + ".")
        lines.append("Decisión sugerida: progresar con cautela si sube la carga y aparecen peor sueño, estrés elevado, dolor o caída de calidad de sesión.")
        target.append(_box([_p(" ".join(lines), "ProfMuted")], padding=5))

    def _append_integrated_interpretation(target: list[object], *, compact: bool = False) -> None:
        target.append(_p("Interpretación integrada", "ProfSection"))
        if compact:
            target.append(_box([_p(" ".join(integrated_lines), "ProfMuted")], padding=5))
            return
        target.append(_box([_p(line, "ProfBody") for line in integrated_lines], padding=6))

    def _append_next_steps(target: list[object], *, compact: bool = False) -> None:
        target.append(_p("Próximo paso recomendado", "ProfSection"))
        if compact:
            target.append(_box([_p(" ".join(f"{idx}. {line}" for idx, line in enumerate(next_steps, start=1)), "ProfMuted")], padding=5))
            return
        target.append(_box([_p(f"- {line}", "ProfBody") for line in next_steps], padding=6))

    def _append_limitations(target: list[object], *, compact: bool = False) -> None:
        target.append(_p("Limitaciones del reporte", "ProfSection"))
        if compact:
            target.append(
                _box(
                    [
                        _p(
                            "Las evaluaciones son perfilado físico cada 6-8 semanas y no deben interpretarse como readiness semanal. El sRPE es una estimación práctica de carga interna.",
                            "ProfMuted",
                        ),
                        _p(
                            "Las decisiones deben considerar contexto deportivo, dolor, wellness y criterio profesional. Si una sección muestra \"Faltan datos\", no había información suficiente para generar esa parte del análisis.",
                            "ProfMuted",
                        ),
                    ],
                    padding=5,
                )
            )
            return
        limitation_lines = [
            "- Las evaluaciones son perfilado físico cada 6-8 semanas.",
            "- No deben interpretarse como readiness semanal.",
            "- El sRPE es una estimación práctica de carga interna.",
            "- Las decisiones deben considerar contexto deportivo, dolor, wellness y criterio profesional.",
            "- Si una sección muestra \"Faltan datos\", no había información suficiente para generar esa parte del análisis.",
        ]
        target.append(_box([_p(line, "ProfBody") for line in limitation_lines], padding=6))

    def _append_full_executive_page(target: list[object]) -> None:
        brand = _brand_lockup()
        if brand is not None:
            target.append(brand)
            target.append(Spacer(1, 4 * mm))
        target.append(
            _box(
                [
                    _p("REPORTE PROFESIONAL", "ProfEyebrow"),
                    _p(report_athlete, "ProfHeroTitle"),
                    _p(
                        f"Generado el {datetime.now():%d/%m/%Y %H:%M} · Ventana {executive_payload.get('period', PDF_MISSING_TEXT)}",
                        "ProfHeroMeta",
                    ),
                    _p(
                        "Reporte madre profesional orientado a toma de decisiones para Threshold S&C.",
                        "ProfBodyWhite",
                    ),
                ],
                background=palette["navy"],
                border_color=palette["navy"],
                padding=9,
            )
        )
        target.append(Spacer(1, 5 * mm))
        target.append(
            _page_header(
                executive_payload.get("title", "Resumen ejecutivo profesional"),
                "Prioriza lectura principal, cobertura, señales clave y decisión operativa del próximo bloque.",
                eyebrow="Resumen ejecutivo",
            )
        )
        target.append(Spacer(1, 3 * mm))
        target.append(_executive_summary_table(executive_payload))
        target.append(Spacer(1, 4 * mm))
        target.append(
            _two_column(
                _bullet_box(
                    "Señales clave",
                    executive_payload.get("signals", []),
                    width_mm=104,
                    background=palette["panel"],
                    border_color=palette["line_dark"],
                    accent_color=palette["steel"],
                ),
                _decision_box(
                    "Decisión sugerida",
                    executive_payload.get("decision_suggested", PDF_MISSING_TEXT),
                    width_mm=66,
                    note=f"Confianza: {safe_value(executive_payload.get('confidence'), fallback=PDF_MISSING_TEXT)}",
                ),
                left_width_mm=106,
                right_width_mm=68,
            )
        )

    def _append_full_composite_profile_page(target: list[object]) -> None:
        target.append(
            _page_header(
                composite_profile.get("title", "Perfil actual compuesto"),
                "Lectura de perfil actual compuesta desde la misma fuente que el dashboard, sin inferir adaptación cerrada del bloque.",
                eyebrow="Perfil físico",
            )
        )
        if composite_profile.get("state") == "missing":
            target.append(Spacer(1, 3 * mm))
            missing_flowables: list[object] = [_p(str(composite_profile.get("message") or "Faltan datos para el perfil compuesto."), "ProfBody")]
            if change_payload.get("state") == "missing":
                missing_flowables.append(
                    _p(
                        f"Cambios vs evaluación anterior: {change_payload.get('message') or PROFESSIONAL_NO_EVOLUTION_TEXT}",
                        "ProfMuted",
                    )
                )
            target.append(
                _box(
                    missing_flowables,
                    background=palette["panel"],
                    border_color=palette["line_dark"],
                    padding=6,
                    accent_color=palette["line_dark"],
                )
            )
            return
        target.append(Spacer(1, 3 * mm))
        radar_image = _composite_radar_image(composite_profile, width_mm=102, height_mm=66)
        methodological_notes: list[str] = [str(composite_profile.get("note") or PROFESSIONAL_COMPOSITE_PROFILE_NOTE)]
        if str(composite_profile.get("scope_note") or "").strip():
            methodological_notes.append(str(composite_profile.get("scope_note")))
        if change_payload.get("state") == "missing":
            methodological_notes.append(
                f"Cambios vs evaluación anterior: {change_payload.get('message') or PROFESSIONAL_NO_EVOLUTION_TEXT}"
            )
        note_title = "Notas metodológicas" if len(methodological_notes) > 1 else "Nota metodológica"
        profile_summary_flow: list[object] = [
            _decision_box(
                "Lectura principal del perfil",
                composite_profile.get("summary_line", PDF_MISSING_TEXT),
                width_mm=68,
                inverted=False,
                note=f"Última evaluación incluida: {safe_value(composite_profile.get('latest_profile_date'), fallback=PDF_MISSING_TEXT)}",
            ),
            Spacer(1, 2.5 * mm),
            _box(
                [_p(note_title, "ProfCardTitle")] + [_p(note, "ProfMuted") for note in methodological_notes],
                background=palette["panel"],
                border_color=palette["line_dark"],
                width_mm=68,
                padding=6,
                accent_color=palette["steel"],
            ),
        ]
        radar_panel = (
            _chart_panel(
                radar_image,
                title="Radar del perfil compuesto",
                note="El radar resume la expresión actual; la tabla conserva el detalle métrico y sus z-scores visibles.",
                width_mm=102,
            )
            if radar_image is not None
            else _box(
                [_p("Radar del perfil compuesto no disponible para esta exportación.", "ProfMuted")],
                width_mm=102,
                background=palette["panel"],
                border_color=palette["line_dark"],
                padding=6,
                accent_color=palette["steel"],
            )
        )
        target.append(_two_column(radar_panel, profile_summary_flow, left_width_mm=104, right_width_mm=70))
        target.append(Spacer(1, 4 * mm))
        target.append(_dataframe_table(composite_profile.get("metric_table"), col_widths_mm=[35, 31, 21, 87]))
        feedback = composite_profile.get("feedback", {}) if isinstance(composite_profile.get("feedback"), dict) else {}
        target.append(Spacer(1, 4 * mm))
        target.append(
            _two_column(
                _bullet_box(
                    "Lectura fisiológica y biomecánica",
                    [
                        f"Lectura fisiológica: {feedback.get('physiological', PDF_MISSING_TEXT)}",
                        f"Lectura biomecánica: {feedback.get('biomechanical', PDF_MISSING_TEXT)}",
                    ],
                    width_mm=84,
                    background=palette["panel"],
                    border_color=palette["line_dark"],
                    accent_color=palette["steel"],
                ),
                _bullet_box(
                    "Decisión para el próximo bloque",
                    [
                        f"Variable dominante: {feedback.get('high', PDF_MISSING_TEXT)}",
                        f"Variable rezagada: {feedback.get('low', PDF_MISSING_TEXT)}",
                        f"Implicancia para próximo bloque: {feedback.get('next_block', PDF_MISSING_TEXT)}",
                    ],
                    width_mm=86,
                    background=palette["panel_alt"],
                    border_color=palette["line_dark"],
                    accent_color=palette["navy"],
                ),
            )
        )

    def _append_full_change_page(target: list[object]) -> None:
        target.append(
            _page_header(
                change_payload.get("title", "Cambios vs evaluación anterior"),
                "Comparación contra la evaluación inmediatamente anterior, sin leerla como readiness semanal.",
                eyebrow="Evolución",
            )
        )
        if change_payload.get("state") == "missing":
            target.append(Spacer(1, 3 * mm))
            target.append(_collapsed_box(str(change_payload.get("message") or PROFESSIONAL_NO_EVOLUTION_TEXT)))
            return
        target.append(Spacer(1, 3 * mm))
        target.append(_dataframe_table(change_payload.get("display_table"), col_widths_mm=[34, 18, 18, 18, 18, 32, 36]))
        target.append(Spacer(1, 3 * mm))
        target.append(
            _bullet_box(
                "Síntesis de cambios",
                change_payload.get("summary_lines", []),
                background=palette["panel"],
                border_color=palette["line_dark"],
                accent_color=palette["steel"],
            )
        )
        if not _professional_any_quadrant_ready(quadrant_sections):
            target.append(Spacer(1, 3 * mm))
            target.append(
                _box(
                    [_p("Relaciones de perfil / cuadrantes: datos insuficientes para una ubicación útil en esta exportación.", "ProfMuted")],
                    background=palette["panel"],
                    border_color=palette["line_dark"],
                    padding=5,
                    accent_color=palette["steel"],
                )
            )

    def _append_full_quadrants_page(target: list[object]) -> None:
        target.append(
            _page_header(
                "Relaciones de perfil / cuadrantes",
                "Cada relación se usa como apoyo técnico para priorizar decisiones, no como diagnóstico aislado.",
                eyebrow="Relaciones de perfil",
            )
        )
        if not _professional_any_quadrant_ready(quadrant_sections):
            target.append(Spacer(1, 3 * mm))
            target.append(_collapsed_box(PROFESSIONAL_NO_QUADRANTS_TEXT))
            return
        target.append(Spacer(1, 3 * mm))
        panels = [panel for panel in (_quadrant_compact_panel(section, width_mm=84) for section in quadrant_sections[:3]) if panel is not None]
        if not panels:
            target.append(_collapsed_box(PROFESSIONAL_NO_QUADRANTS_TEXT))
            return
        if len(panels) == 1:
            target.append(panels[0])
            return
        target.append(_two_column(panels[0], panels[1], left_width_mm=84, right_width_mm=90))
        if len(panels) > 2:
            target.append(Spacer(1, 3 * mm))
            target.append(_two_column(panels[2], "", left_width_mm=84, right_width_mm=90))

    def _append_full_isometrics_page(target: list[object]) -> None:
        target.append(
            _page_header(
                isometric_payload.get("title", "Isométricos y force-time avanzado"),
                "Anexo técnico compacto para IMTP, fuerza relativa, asimetrías y lectura descriptiva de la curva force-time.",
                eyebrow="Anexo técnico",
            )
        )
        if isometric_payload.get("state") == "missing":
            target.append(Spacer(1, 3 * mm))
            target.append(_collapsed_box(str(isometric_payload.get("message") or "Faltan datos isométricos.")))
            return
        target.append(Spacer(1, 3 * mm))
        imtp_priority = [
            row
            for row in isometric_payload.get("imtp_rows", [])
            if _professional_normalized_text(row[0]) in {"peak force", "fuerza relativa", "force avg", "time to peak", "asimetria", "lado dominante"}
        ]
        if imtp_priority:
            practical_notes = [
                note
                for note in isometric_payload.get("imtp_notes", [])
                if all(
                    marker not in _professional_normalized_text(note)
                    for marker in (
                        "rfd exportada descriptiva",
                        "perfil de fuerza por puntos exportados",
                    )
                )
            ][:2]
            target.append(
                _bullet_box(
                    "Lectura práctica",
                    practical_notes
                    or ["Usar el bloque isométrico para leer capacidad máxima, apoyo de fuerza relativa y consistencia de la expresión temprana."],
                    background=palette["panel"],
                    border_color=palette["line_dark"],
                    accent_color=palette["steel"],
                )
            )
            target.append(Spacer(1, 3 * mm))
            target.append(_mini_cards_table(imtp_priority, columns=3))
        if isometric_payload.get("iso_available"):
            target.append(Spacer(1, 3 * mm))
            iso_priority = [row for row in isometric_payload.get("iso_rows", []) if row[0] in {"Peak Force", "Force Avg", "Time to Peak", "Asimetría"}]
            iso_block = _mini_cards_table(iso_priority, columns=2, width_mm=84) if iso_priority else ""
            iso_notes = _bullet_box(
                "ISO Push Hip-Hamstring Bilateral",
                isometric_payload.get("iso_notes", [])[:3] or ["Usar este test como complemento del IMTP para seguimiento descriptivo."],
                width_mm=86,
                style_name="ProfMuted",
                background=palette["panel_alt"],
                border_color=palette["line_dark"],
                accent_color=palette["steel"],
            )
            target.append(_two_column(iso_block, iso_notes, left_width_mm=84, right_width_mm=90))
        if isometric_payload.get("force_time_available"):
            target.append(Spacer(1, 3 * mm))
            _draw_compact_professional_force_time_block(
                {
                    "story": target,
                    "p": _p,
                    "box": _box,
                    "Table": Table,
                    "TableStyle": TableStyle,
                    "Spacer": Spacer,
                    "mm": mm,
                    "palette": palette,
                },
                force_time_payload,
                report_type="professional",
            )

    def _append_full_load_page(target: list[object]) -> None:
        target.append(
            _page_header(
                load_tolerance_payload.get("title", "Carga interna y tolerancia"),
                "Responder rápido si la carga reciente luce tolerable, baja, alta o potencialmente riesgosa.",
                eyebrow="Tolerancia de carga",
            )
        )
        if load_tolerance_payload.get("state") == "missing":
            target.append(Spacer(1, 3 * mm))
            target.append(_collapsed_box(str(load_tolerance_payload.get("message") or "Faltan datos de carga interna.")))
            return
        target.append(Spacer(1, 3 * mm))
        rows = load_tolerance_payload.get("rows", [])
        primary_labels = {
            "srpe semanal",
            "cambio vs semana previa",
            "sesiones registradas",
            "acwr ewma",
            "zona acwr",
            "monotonia",
            "strain",
        }
        primary_rows = [row for row in rows if _professional_normalized_text(row[0]) in primary_labels]
        secondary_rows = [row for row in rows if _professional_normalized_text(row[0]) not in primary_labels]
        if primary_rows:
            target.append(_mini_cards_table(primary_rows, columns=3))
        if secondary_rows:
            target.append(Spacer(1, 3 * mm))
            target.append(_key_value_table(secondary_rows, label_width_mm=46, value_width_mm=128))
        target.append(Spacer(1, 3 * mm))
        weekly_chart = _weekly_ema_chart(load_tolerance_payload.get("weekly_points", []), height_mm=48)
        if weekly_chart is not None:
            primary_row_labels = {_professional_normalized_text(row[0]) for row in primary_rows}
            load_chart_note = (
                "La curva semanal se lee junto con ACWR, monotonía y strain para contextualizar tolerancia reciente."
                if primary_row_labels & {"acwr ewma", "zona acwr", "monotonia", "strain"}
                else "La curva semanal resume el acumulado visible de la semana y debe leerse con cautela si faltan métricas de tolerancia."
            )
            target.append(
                _chart_panel(
                    weekly_chart,
                    note=load_chart_note,
                )
            )
            target.append(Spacer(1, 3 * mm))
        target.append(
            _decision_box(
                "Lectura final",
                load_tolerance_payload.get("risk_line", PDF_MISSING_TEXT),
                width_mm=174,
                inverted=False,
            )
        )

    def _append_full_wellness_page(target: list[object]) -> None:
        target.append(
            _page_header(
                wellness_availability_payload.get("title", "Wellness, disponibilidad y adherencia"),
                "Usar wellness como contexto de disponibilidad y tolerancia, no como lectura aislada del rendimiento.",
                eyebrow="Disponibilidad",
            )
        )
        if wellness_availability_payload.get("state") == "missing":
            target.append(Spacer(1, 3 * mm))
            target.append(_collapsed_box(str(wellness_availability_payload.get("message") or "Faltan datos de wellness.")))
            return
        target.append(Spacer(1, 3 * mm))
        target.append(_mini_cards_table(wellness_availability_payload.get("rows", []), columns=3))
        chart_points = (
            wellness_availability_payload.get("weekly_points", [])
            if wellness_availability_payload.get("trend_allowed")
            else wellness_availability_payload.get("daily_points", [])
        )
        chart = _wellness_chart(chart_points, width_mm=174, height_mm=44)
        if chart is not None:
            target.append(Spacer(1, 3 * mm))
            target.append(
                _chart_panel(
                    chart,
                    note="La tendencia disponible ayuda a poner en contexto sueño, estrés, dolor y consistencia de registros.",
                )
            )
        target.append(Spacer(1, 3 * mm))
        note_lines = [str(wellness_availability_payload.get("compatibility") or "")]
        if str(wellness_availability_payload.get("quality_note") or "").strip():
            note_lines.append(str(wellness_availability_payload.get("quality_note")))
        target.append(
            _bullet_box(
                "Lectura de disponibilidad",
                note_lines,
                style_name="ProfMuted",
                background=palette["panel"],
                border_color=palette["line_dark"],
                accent_color=palette["green"],
            )
        )

    def _append_full_exposure_page(target: list[object]) -> None:
        target.append(
            _page_header(
                exposure_payload.get("title", "Exposición del bloque / contenido entrenado"),
                "Resume qué se entrenó, cómo se distribuyó el bloque y dónde aparece la menor exposición relativa.",
                eyebrow="Contenido entrenado",
            )
        )
        if exposure_payload.get("state") == "missing":
            target.append(Spacer(1, 3 * mm))
            target.append(_collapsed_box(str(exposure_payload.get("message") or "Faltan datos de exposición.")))
            return
        target.append(Spacer(1, 3 * mm))
        chart_image = _exposure_chart_image(width_mm=104, height_mm=56)
        summary_width = 66 if chart_image is not None else 174
        summary_box = _bullet_box(
            "Lectura del bloque",
            [
                str(exposure_payload.get("summary_line") or ""),
                str(exposure_payload.get("context_link") or ""),
                f"Estímulos bajos o ausentes: {_professional_join_labels(exposure_payload.get('low_or_absent', [])[:3])}.",
            ],
            width_mm=summary_width,
            style_name="ProfMuted",
            background=palette["panel_alt"],
            border_color=palette["line_dark"],
            accent_color=palette["navy"],
        )
        if chart_image is not None:
            target.append(
                _two_column(
                    _chart_panel(
                        chart_image,
                        title="Distribución visual del bloque",
                        note="El gráfico acompaña la tabla para localizar rápido el peso relativo de cada estímulo.",
                        width_mm=104,
                    ),
                    summary_box,
                    left_width_mm=106,
                    right_width_mm=68,
                )
            )
            target.append(Spacer(1, 4 * mm))
        else:
            target.append(summary_box)
            target.append(Spacer(1, 4 * mm))
        target.append(_dataframe_table(exposure_payload.get("table"), col_widths_mm=[40, 28, 24, 82]))

    def _append_full_integrated_page(target: list[object]) -> None:
        target.append(
            _page_header(
                integrated_decision_payload.get("title", "Interpretación integrada profesional"),
                "Integra evaluación física, carga, wellness y exposición para sostener una decisión práctica única.",
                eyebrow="Síntesis profesional",
            )
        )
        target.append(Spacer(1, 3 * mm))
        target.append(
            _two_column(
                [
                    _bullet_box(
                        "Qué sabemos con buena confianza",
                        integrated_decision_payload.get("good_confidence", []),
                        width_mm=84,
                        background=palette["panel_alt"],
                        border_color=palette["line_dark"],
                        accent_color=palette["navy"],
                    ),
                    Spacer(1, 3 * mm),
                    _bullet_box(
                        "Qué parece probable",
                        integrated_decision_payload.get("probable", []),
                        width_mm=84,
                        style_name="ProfMuted",
                        background=palette["panel"],
                        border_color=palette["line_dark"],
                        accent_color=palette["steel"],
                    ),
                ],
                [
                    _bullet_box(
                        "Qué no podemos afirmar todavía",
                        integrated_decision_payload.get("unknown", []),
                        width_mm=86,
                        style_name="ProfMuted",
                        background=palette["panel"],
                        border_color=palette["line_dark"],
                        accent_color=palette["line_dark"],
                    ),
                    Spacer(1, 3 * mm),
                    _bullet_box(
                        "Qué monitorear en el próximo bloque",
                        integrated_decision_payload.get("monitor", []),
                        width_mm=86,
                        style_name="ProfMuted",
                        background=palette["card"],
                        border_color=palette["line_dark"],
                        accent_color=palette["steel"],
                    ),
                ],
                left_width_mm=84,
                right_width_mm=90,
            )
        )
        target.append(Spacer(1, 4 * mm))
        target.append(
            _bullet_box(
                "Decisión práctica",
                integrated_decision_payload.get("decision_practical", []),
                style_name="ProfBodyWhite",
                background=palette["navy"],
                border_color=palette["navy"],
                title_style="ProfDecisionTitle",
            )
        )

    def _append_full_action_plan_page(target: list[object]) -> None:
        target.append(
            _page_header(
                action_plan_payload.get("title", "Próximos pasos y limitaciones metodológicas"),
                "Traducir la lectura a acciones simples de mantener, ajustar, monitorear y medir.",
                eyebrow="Plan de acción",
            )
        )
        target.append(Spacer(1, 3 * mm))
        actions = action_plan_payload.get("actions", {})
        target.append(
            _two_column(
                _bullet_box(
                    "Mantener",
                    actions.get("Mantener", []),
                    width_mm=84,
                    background=palette["panel"],
                    border_color=palette["line_dark"],
                    accent_color=palette["steel"],
                ),
                _bullet_box(
                    "Ajustar",
                    actions.get("Ajustar", []),
                    width_mm=86,
                    style_name="ProfBodyWhite",
                    title_style="ProfDecisionTitle",
                    background=palette["navy"],
                    border_color=palette["navy"],
                ),
                left_width_mm=84,
                right_width_mm=90,
            )
        )
        target.append(Spacer(1, 3 * mm))
        target.append(
            _two_column(
                _bullet_box(
                    "Monitorear",
                    actions.get("Monitorear", []),
                    width_mm=84,
                    background=palette["card"],
                    border_color=palette["line_dark"],
                    accent_color=palette["steel"],
                ),
                _bullet_box(
                    "Medir",
                    actions.get("Medir", []),
                    width_mm=86,
                    background=palette["card"],
                    border_color=palette["line_dark"],
                    accent_color=palette["steel"],
                ),
                left_width_mm=84,
                right_width_mm=90,
            )
        )
        target.append(Spacer(1, 4 * mm))
        target.append(
            _bullet_box(
                "Limitaciones metodológicas",
                action_plan_payload.get("limitations", []),
                style_name="ProfMuted",
                background=palette["panel"],
                border_color=palette["line_dark"],
                accent_color=palette["line_dark"],
            )
        )

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        leftMargin=18 * mm,
        rightMargin=18 * mm,
        topMargin=18 * mm,
        bottomMargin=14 * mm,
    )
    def _draw_professional_footer(canvas_obj, doc_obj) -> None:
        _draw_threshold_footer(canvas_obj, doc_obj, palette=palette, mm_unit=mm)
    story: list[object] = []
    # The professional report always uses the mother-report structure; missing data is handled inside each section.
    _append_full_executive_page(story)
    story.append(PageBreak())
    _append_full_composite_profile_page(story)
    if change_payload.get("state") != "missing":
        story.append(PageBreak())
        _append_full_change_page(story)
    if _professional_any_quadrant_ready(quadrant_sections):
        story.append(PageBreak())
        _append_full_quadrants_page(story)
    if isometric_payload.get("state") != "missing":
        story.append(PageBreak())
        _append_full_isometrics_page(story)
    story.append(PageBreak())
    _append_full_load_page(story)
    story.append(PageBreak())
    _append_full_wellness_page(story)
    story.append(PageBreak())
    _append_full_exposure_page(story)
    story.append(PageBreak())
    _append_full_integrated_page(story)
    story.append(PageBreak())
    _append_full_action_plan_page(story)
    try:
        doc.build(story, onFirstPage=_draw_professional_footer, onLaterPages=_draw_professional_footer)
    except Exception:
        return None
    return buffer.getvalue()

def _generate_visual_report_pdf_reportlab(
    state: dict[str, pd.DataFrame | None],
    report_athlete: str = "Todos",
    report_audience: str = "profe",
) -> bytes | None:
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
        from reportlab.lib.units import mm
        from reportlab.lib.utils import ImageReader
        from reportlab.platypus import Image, PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
        from xml.sax.saxutils import escape
    except Exception:
        return None

    audience = normalize_report_audience(report_audience)
    effective_athlete = resolve_report_scope(state, report_athlete, audience)
    if effective_athlete is None:
        return None
    if audience == "profe" and effective_athlete != "Todos":
        professional_pdf = _generate_professional_profile_pdf_reportlab(state, effective_athlete)
        if professional_pdf:
            return professional_pdf

    summary_df = build_executive_summary_df(state, effective_athlete, audience)
    insights = generate_module_insights(state, effective_athlete, audience)
    blocks = _audience_blocks(state, effective_athlete, summary_df, insights, audience)
    charts = collect_report_plotly_figures(state, effective_athlete, audience)

    palette = _pdf_theme_threshold(colors, variant="general")
    styles = _register_threshold_pdf_styles(
        getSampleStyleSheet(),
        ParagraphStyle,
        palette,
        variant="general",
    )

    def _p(text: object, style_name: str = "ReportBody") -> Paragraph:
        safe = escape(_ascii_text(text) or "").replace("\n", "<br/>")
        return Paragraph(safe, styles[style_name])

    def _fit_image(path: Path | None, max_width_mm: float, max_height_mm: float) -> Image | None:
        return _fit_pdf_image(
            path,
            max_width_mm=max_width_mm,
            max_height_mm=max_height_mm,
            mm_unit=mm,
            ImageClass=Image,
            ImageReaderClass=ImageReader,
        )

    def _metric_cards_table(cards: list[tuple[str, str, str]]) -> Table:
        rows = []
        row: list[object] = []
        for idx, (label, value, _) in enumerate(cards, start=1):
            cell = _build_metric_card(
                label,
                value,
                paragraph_builder=_p,
                TableClass=Table,
                TableStyleClass=TableStyle,
                mm_unit=mm,
                palette=palette,
            )
            row.append(cell)
            if idx % 3 == 0:
                rows.append(row)
                row = []
        if row:
            while len(row) < 3:
                row.append("")
            rows.append(row)
        table = Table(rows, colWidths=[56 * mm, 56 * mm, 56 * mm], hAlign="LEFT")
        table.setStyle(TableStyle([("VALIGN", (0, 0), (-1, -1), "TOP")]))
        return table

    def _summary_meta_table() -> Table:
        readiness = _readiness_payload(_report_quality_report(state))
        weekly_summaries = _report_weekly_summaries(state)
        current_team = _current_week_slice(weekly_summaries.get("weekly_team"))
        current_load = _current_week_slice(weekly_summaries.get("weekly_load"), effective_athlete)
        completion_value = _focus_completion_value(state, effective_athlete)
        completion_text = _display_metric(completion_value, digits=1, suffix="%") if completion_value is not None else "Sin dato"
        adherence_label = "Adherencia del atleta" if effective_athlete != "Todos" else "Adherencia promedio"
        if effective_athlete == "Todos":
            current_row = current_team.tail(1).iloc[0] if not current_team.empty else pd.Series(dtype=object)
            athletes_active = _coerce_float(current_row.get("athletes_active"))
            week_value = f"{int(athletes_active)} activos" if athletes_active is not None else "Sin dato"
        else:
            current_row = current_load.tail(1).iloc[0] if not current_load.empty else pd.Series(dtype=object)
            sessions_count = _coerce_float(current_row.get("sessions_count"))
            week_value = f"{int(sessions_count)} sesiones" if sessions_count is not None else "Sin dato"
        data = [
            [
                _p(f"Semana actual\n{week_value}", "ReportMuted"),
                _p(f"Readiness\n{readiness['label']}", "ReportMuted"),
                _p(f"{adherence_label}\n{completion_text}", "ReportMuted"),
            ]
        ]
        return _build_threshold_table(
            data,
            TableClass=Table,
            TableStyleClass=TableStyle,
            mm_unit=mm,
            palette=palette,
            col_widths_mm=[56, 56, 56],
            show_inner_grid=False,
            border_width=0.7,
            left_padding=8,
            right_padding=8,
            top_padding=8,
            bottom_padding=8,
        )

    def _metric_table(rows: list[tuple[str, str]], *, title: str) -> list[object]:
        data = [[_p(label, "CardLabel"), _p(value, "ReportBody")] for label, value in rows]
        table = _build_threshold_table(
            data,
            TableClass=Table,
            TableStyleClass=TableStyle,
            mm_unit=mm,
            palette=palette,
            col_widths_mm=[58, 112],
            border_width=0.7,
            inner_grid_width=0.4,
            left_padding=8,
            right_padding=8,
            top_padding=6,
            bottom_padding=6,
        )
        return [_p(title, "ReportSection"), table]

    def _snapshot_table() -> Table | None:
        if summary_df.empty:
            return None
        display_df = summary_df.copy().fillna("-")
        keep_cols = [col for col in ["Atleta", "ACWR EWMA", "Zona", "Wellness 3d", "CMJ cm", "DRI", "IMTP N", "Perfil NM"] if col in display_df.columns]
        if not keep_cols:
            return None
        display_df = display_df[keep_cols]
        display_df = display_df.rename(
            columns={
                "ACWR EWMA": "ACWR EWMA",
                "Wellness 3d": "Wellness",
                "CMJ cm": "CMJ",
                "IMTP N": "IMTP",
            }
        )
        data: list[list[object]] = [[_p(col, "CardLabel") for col in display_df.columns]]
        for _, row in display_df.iterrows():
            data.append([_p(_snapshot_value(col, row.get(col, "-")), "ReportBody") for col in display_df.columns])
        widths_map = {
            "Atleta": 38 * mm,
            "ACWR EWMA": 18 * mm,
            "Zona": 24 * mm,
            "Wellness": 24 * mm,
            "CMJ": 18 * mm,
            "DRI": 18 * mm,
            "IMTP": 22 * mm,
            "Perfil NM": 36 * mm,
        }
        col_widths = [widths_map.get(col, 24 * mm) for col in display_df.columns]
        return _build_threshold_table(
            data,
            TableClass=Table,
            TableStyleClass=TableStyle,
            mm_unit=mm,
            palette=palette,
            col_widths=col_widths,
            header_rows=1,
            repeat_rows=1,
            border_width=0.7,
            inner_grid_width=0.35,
            left_padding=6,
            right_padding=6,
            top_padding=6,
            bottom_padding=6,
        )

    def _chart_story(chart_payload: dict[str, object]) -> list[object]:
        image_bytes = export_plotly_figure_png(chart_payload.get("figure"))
        if not image_bytes:
            return []
        image = Image(BytesIO(image_bytes), width=175 * mm, height=102 * mm)
        image.hAlign = "LEFT"
        return [
            _p(chart_payload.get("title", "Gráfico"), "ReportSection"),
            image,
        ]

    def _box_flow(flowables: list[object], *, padding: int = 8) -> Table:
        return _build_threshold_box(
            flowables,
            TableClass=Table,
            TableStyleClass=TableStyle,
            mm_unit=mm,
            palette=palette,
            padding=padding,
            border_width=0.7,
        )

    def _compact_key_value_table(rows: list[dict[str, object] | tuple[str, str]]) -> Table:
        data = []
        for row in rows:
            if isinstance(row, dict):
                label = str(row.get("label", ""))
                value = safe_value(row.get("value"), fallback=PDF_MISSING_TEXT)
                note = str(row.get("note", "") or "")
            else:
                label, value = row
                label = str(label)
                value = safe_value(value, fallback=PDF_MISSING_TEXT)
                note = ""
            value_flow: list[object] = [_p(value, "ReportMuted")]
            if note:
                value_flow.append(_p(note, "ReportMutedItalic"))
            data.append([_p(label, "CardLabel"), value_flow])
        return _build_threshold_table(
            data,
            TableClass=Table,
            TableStyleClass=TableStyle,
            mm_unit=mm,
            palette=palette,
            col_widths_mm=[42, 132],
            border_width=0.7,
            inner_grid_width=0.35,
            left_padding=7,
            right_padding=7,
            top_padding=5,
            bottom_padding=5,
        )

    def _chart_image(chart_payload: dict[str, object] | None, *, width_mm: float = 174, height_mm: float = 88) -> Image | None:
        if not chart_payload:
            return None
        image_bytes = export_plotly_figure_png(chart_payload.get("figure"), width=1150, height=620, scale=2)
        if not image_bytes:
            return None
        image = Image(BytesIO(image_bytes), width=width_mm * mm, height=height_mm * mm)
        image.hAlign = "LEFT"
        return image

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        leftMargin=18 * mm,
        rightMargin=18 * mm,
        topMargin=18 * mm,
        bottomMargin=14 * mm,
    )

    story: list[object] = []
    wordmark = _resolve_brand_asset_path("wordmark")
    icon = _resolve_brand_asset_path("icon")
    wordmark_logo = _fit_image(wordmark, 150, 28) if wordmark is not None else None
    icon_logo = _fit_image(icon, 18, 18) if icon is not None else None
    if icon_logo is not None and wordmark_logo is not None:
        brand_table = Table(
            [[icon_logo, wordmark_logo]],
            colWidths=[22 * mm, 128 * mm],
            style=TableStyle(
                [
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("LEFTPADDING", (0, 0), (-1, -1), 0),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 0),
                    ("TOPPADDING", (0, 0), (-1, -1), 0),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
                ]
            ),
        )
        story.append(brand_table)
        story.append(Spacer(1, 8 * mm))
    elif wordmark_logo is not None:
        story.append(wordmark_logo)
        story.append(Spacer(1, 8 * mm))
    elif icon_logo is not None:
        story.append(icon_logo)
        story.append(Spacer(1, 8 * mm))

    if audience == "atleta" and effective_athlete != "Todos":
        focus_row = summary_df.iloc[0] if not summary_df.empty else pd.Series(dtype=object)
        athlete_neuromuscular_profile = _build_current_pdf_neuromuscular_profile_payload(
            state,
            effective_athlete,
            audience="atleta",
        )
        force_time_payload = build_force_time_report_payload(
            _latest_jump_row(state, effective_athlete),
            test_id="imtp",
            report_type="athlete",
        )
        completion_value = _focus_completion_value(state, effective_athlete)
        internal_load = _build_professional_internal_load_context(state, effective_athlete)
        wellness_context = _professional_wellness_context(state, effective_athlete)
        profile_reading = _athlete_profile_interpretation(
            focus_row,
            neuromuscular_profile=athlete_neuromuscular_profile,
        )
        final_blocks = _athlete_final_focus_blocks(
            focus_row,
            completion_value,
            neuromuscular_profile=athlete_neuromuscular_profile,
        )
        athlete_charts = _collect_athlete_pdf_chart_payloads(state, effective_athlete)
        evaluation_available = _row_has_eval_data(focus_row)
        load_available = _row_has_load_data(focus_row)
        wellness_available = _row_has_wellness_data(focus_row)
        athlete_focus = (
            _professional_visible_metric_text(athlete_neuromuscular_profile.get("training_priority_short"))
            if evaluation_available and athlete_neuromuscular_profile.get("source") == "core"
            else (_current_focus_text(focus_row, audience="atleta") if not summary_df.empty else "Nueva evaluación")
        )
        load_lines = _athlete_load_status_lines(focus_row, internal_load)
        wellness_summary = wellness_context.get("last_week_summary", {}) if isinstance(wellness_context.get("last_week_summary"), dict) else {}
        wellness_scale = wellness_context.get("scales", {}) if isinstance(wellness_context.get("scales"), dict) else {}
        athlete_score_value = _coerce_float(wellness_summary.get("score_mean"))
        athlete_score_meta = _report_wellness_score_label(athlete_score_value) if athlete_score_value is not None else None
        force_time_summary = dict(force_time_payload.get("summary", {}))
        asymmetry_summary = dict(force_time_payload.get("asymmetry", {}))
        interpretation = dict(force_time_payload.get("interpretation", {}))
        force_time_points = [
            point
            for point in list(force_time_payload.get("force_time_points", []))
            if _coerce_float(point.get("value_n")) is not None
        ]
        rfd_points = [
            point
            for point in list(force_time_payload.get("rfd_points", []))
            if _coerce_float(point.get("value_n_s")) is not None
        ]
        profile_label = _professional_visible_metric_text(
            athlete_neuromuscular_profile.get("profile_label")
            or _profile_text(focus_row, fallback="Pendiente de evaluación")
        )

        def _athlete_metric(value: object, *, digits: int = 1, suffix: str = "", fallback: str = "Dato todavía no disponible") -> str:
            numeric = _coerce_float(value)
            if numeric is None:
                return fallback
            return _display_metric(numeric, digits=digits, suffix=suffix)

        def _athlete_zone_text() -> str:
            if load_available and _has_text(focus_row.get("Zona")):
                return _display_zone(focus_row.get("Zona"))
            return "En seguimiento"

        def _athlete_completion_text() -> str:
            return _athlete_metric(completion_value, digits=1, suffix="%", fallback="Todavía no registrado")

        def _athlete_recent_wellness_text() -> str:
            recent_value = _coerce_float(focus_row.get("Wellness 3d"))
            if recent_value is not None:
                return _display_metric(recent_value, digits=1)
            if athlete_score_meta is not None:
                return f"{float(athlete_score_meta['score']):.1f} / 5.0"
            return "Todavía no registrado"

        def _athlete_focus_summary_sentence() -> str:
            if evaluation_available and athlete_neuromuscular_profile.get("source") == "core":
                focus_phrase = _sentence_fragment(athlete_focus, lowercase=True) or "sostener la calidad actual"
                return f"El foco actual es {focus_phrase}."
            normalized = _professional_normalized_text(athlete_focus)
            if normalized == "nueva evaluacion":
                return "El foco actual es completar una nueva evaluación y sostener continuidad."
            if normalized == "fuerza base":
                return "El foco actual es construir una base de fuerza más sólida antes de pedir más reactividad."
            if normalized == "recuperar cmj":
                return "El foco actual es recuperar calidad de salto sin perder continuidad."
            if normalized == "regular carga":
                return "El foco actual es ordenar la carga y sostener mejor la tolerancia."
            if normalized == "subir estimulo":
                return "El foco actual es recuperar continuidad de estímulo sin aumentar la carga de forma brusca."
            return "El foco actual es sostener lo que viene funcionando y seguir construyendo con continuidad."

        def _athlete_focus_plan_text() -> str:
            if evaluation_available and athlete_neuromuscular_profile.get("source") == "core":
                return _professional_visible_metric_text(
                    athlete_neuromuscular_profile.get("training_priority_detailed")
                    or athlete_neuromuscular_profile.get("training_priority_short")
                    or athlete_focus
                )
            normalized = _professional_normalized_text(athlete_focus)
            if normalized == "nueva evaluacion":
                return "Completar una nueva evaluación para tener una referencia más clara y seguir ajustando el trabajo."
            if normalized == "fuerza base":
                return "Construir una base de fuerza más sólida antes de pedir más reactividad."
            if normalized == "recuperar cmj":
                return "Recuperar calidad de salto y disponibilidad antes de volver a empujar el volumen."
            if normalized == "regular carga":
                return "Seguir progresando sin aumentar la carga de forma brusca y sostener mejor la tolerancia."
            if normalized == "subir estimulo":
                return "Recuperar continuidad de estímulo para volver a construir con una señal más clara."
            return "Sostener lo que viene funcionando y seguir construyendo con continuidad."

        def _athlete_overview_text() -> str:
            if summary_df.empty:
                return "Todavía falta información para una lectura más clara. Por ahora conviene completar una evaluación, registrar bienestar y sostener continuidad."
            lines: list[str] = ["Hoy tenemos una referencia útil para seguir tu proceso."]
            if evaluation_available:
                lines.append(profile_reading["meaning"])
            elif load_available or wellness_available or completion_value is not None:
                lines.append("Todavía faltan evaluaciones físicas, así que la lectura se apoya en carga, bienestar y adherencia recientes.")
            else:
                lines.append("Todavía faltan evaluaciones y registros recientes para una lectura más completa.")
            zone = _professional_normalized_text(_athlete_zone_text())
            if load_available:
                if "alto riesgo" in zone:
                    lines.append("La carga reciente viene exigente y conviene ordenarla para sostener mejor la tolerancia.")
                elif "precaucion" in zone:
                    lines.append("La carga reciente pide un poco más de cuidado para sostener continuidad.")
                elif "subcarga" in zone:
                    lines.append("La carga reciente quedó baja y conviene recuperar continuidad de estímulo.")
                else:
                    lines.append("La carga reciente está en una zona útil para seguir construyendo.")
            if athlete_score_meta is not None and athlete_score_value is not None:
                if athlete_score_value < 3.0:
                    lines.append("El bienestar reciente viene bajo y conviene cruzarlo de cerca con sueño, estrés y dolor.")
                else:
                    lines.append("El bienestar reciente acompaña y suma contexto para entender cómo venís recuperando.")
            elif wellness_context.get("state") == "partial":
                lines.append("Hay pocos registros de bienestar: sirven como primera señal, pero todavía no como una tendencia cerrada.")
            if completion_value is not None and completion_value < 80:
                lines.append("La adherencia del plan necesita más continuidad para leer mejor cómo responde el proceso.")
            lines.append(_athlete_focus_summary_sentence())
            return " ".join(lines)

        def _athlete_status_box() -> Table:
            return _build_threshold_box(
                [
                    _p("Carga actual", "CardLabel"),
                    _p(_athlete_zone_text(), "ReportSection"),
                    _p("Lectura simple de esta ventana.", "ReportMuted"),
                ],
                TableClass=Table,
                TableStyleClass=TableStyle,
                mm_unit=mm,
                palette=palette,
                width_mm=58,
                padding=8,
                background=palette["panel_alt"],
                border_color=palette["line_dark"],
                accent_color=palette["navy"],
                border_width=0.7,
            )

        def _athlete_cover_cards() -> list[tuple[str, str, str]]:
            cards: list[tuple[str, str, str]] = [
                ("Perfil actual", _short_profile_label(profile_label) if evaluation_available else "Pendiente", "#134263"),
                (
                    "CMJ",
                    _athlete_metric(focus_row.get("CMJ cm"), digits=1, suffix=" cm", fallback="Pendiente de evaluación")
                    if evaluation_available
                    else "Pendiente de evaluación",
                    "#0D3C5E",
                ),
                (
                    "ACWR EWMA",
                    _athlete_metric(focus_row.get("ACWR EWMA"), digits=2, fallback="Todavía no registrado")
                    if load_available
                    else "Todavía no registrado",
                    _zone_color(focus_row.get("Zona")) if load_available else "#708C9F",
                ),
                ("Bienestar", _athlete_recent_wellness_text(), "#2F6B52"),
                ("Adherencia", _athlete_completion_text(), "#708C9F"),
            ]
            return cards

        def _athlete_dictionary_lines() -> list[str]:
            return [f"{label}: {text}" for label, text in _athlete_metric_explanation_rows(focus_row)]

        def _athlete_metric_row(label: str, value: str, n_value: object, *, include_count: bool = True) -> dict[str, object]:
            count = int(_coerce_float(n_value) or 0)
            rendered = value
            if include_count and value != PDF_MISSING_TEXT and count > 0:
                rendered = f"{value}  {_report_sample_suffix(count)}"
            return {"label": label, "value": rendered, "note": _report_sample_warning(count)}

        def _athlete_load_rows() -> list[dict[str, object]]:
            if internal_load.get("analysis_scope") == "current_week_partial":
                week_text = (
                    _athlete_metric(internal_load.get("current_week_total"), digits=0, suffix=" UA", fallback="Semana en curso")
                    if _coerce_float(internal_load.get("current_week_total")) is not None
                    else "Semana en curso"
                )
                change_text = "Semana en curso"
                sessions_text = (
                    str(int(_coerce_float(internal_load.get("current_week_sessions")) or 0))
                    if _coerce_float(internal_load.get("current_week_sessions")) is not None
                    else PDF_MISSING_TEXT
                )
            else:
                week_text = _athlete_metric(internal_load.get("last_week_total"), digits=0, suffix=" UA")
                weekly_change_pct = _coerce_float(internal_load.get("weekly_change_pct"))
                change_text = f"{weekly_change_pct:+.1f}%" if weekly_change_pct is not None else PDF_MISSING_TEXT
                sessions_registered = _coerce_float(internal_load.get("sessions_registered"))
                sessions_text = str(int(sessions_registered)) if sessions_registered is not None else PDF_MISSING_TEXT
            return [
                {"label": "Estado de carga", "value": _athlete_zone_text()},
                {"label": "Última semana visible", "value": week_text},
                {"label": "Cambio vs semana previa", "value": change_text},
                {"label": "ACWR EWMA", "value": _athlete_metric(focus_row.get("ACWR EWMA"), digits=2)},
                {"label": "Monotonía", "value": _athlete_metric(focus_row.get("Monotonia"), digits=2)},
                {"label": "Sesiones registradas", "value": sessions_text},
            ]

        def _athlete_load_summary_line() -> str:
            zone = _professional_normalized_text(_athlete_zone_text())
            if internal_load.get("analysis_scope") == "current_week_partial":
                total = _coerce_float(internal_load.get("current_week_total"))
                sessions = int(_coerce_float(internal_load.get("current_week_sessions")) or 0)
                if total is not None:
                    return f"La semana sigue en curso: acumulás {total:.0f} UA en {sessions} sesiones y conviene esperar el cierre semanal para compararla mejor."
                return "La semana sigue en curso y todavía faltan datos suficientes para leer mejor la carga reciente."
            if not load_available:
                return "Faltan datos recientes de carga interna para entender mejor cómo venís tolerando el entrenamiento."
            if "alto riesgo" in zone:
                return "La carga reciente viene alta y conviene ordenarla para sostener mejor la tolerancia."
            if "precaucion" in zone:
                return "La carga reciente pide un poco más de cuidado para sostener continuidad."
            if "subcarga" in zone:
                return "La carga reciente quedó baja y conviene recuperar continuidad de estímulo."
            return "La carga reciente se mueve en una zona útil para seguir construyendo con continuidad."

        def _athlete_load_watch_lines() -> list[str]:
            lines: list[str] = []
            if _coerce_float(focus_row.get("Monotonia")) is not None:
                lines.append("Si la semana se parece demasiado entre días, conviene darle más variación al bloque.")
            if completion_value is not None and completion_value < 80:
                lines.append("La adherencia también importa: si baja, la lectura de carga se vuelve más ruidosa.")
            if not lines:
                lines.append("Lo importante es sostener continuidad y evitar cambios bruscos de una semana a otra.")
            return lines[:2]

        def _athlete_wellness_signal_is_concerning() -> bool:
            stress_mean = _coerce_float(wellness_summary.get("stress_mean"))
            pain_mean = _coerce_float(wellness_summary.get("pain_mean"))
            recent_value = _coerce_float(focus_row.get("Wellness 3d"))
            mean_low = athlete_score_value is not None and athlete_score_value < 3.0
            stress_high = stress_mean is not None and _professional_wellness_high(stress_mean, wellness_scale.get("stress"))
            pain_high = pain_mean is not None and _professional_wellness_high(pain_mean, wellness_scale.get("pain"))
            recent_low = recent_value is not None and recent_value < 3.0
            return bool(
                mean_low
                or stress_high
                or pain_high
                or (wellness_context.get("state") == "partial" and recent_low)
                or (athlete_score_value is None and recent_low)
            )

        def _athlete_wellness_rows() -> list[dict[str, object]]:
            athlete_score_days = int(_coerce_float(wellness_summary.get("score_n")) or 0)
            return [
                _athlete_metric_row(
                    "Bienestar",
                    (
                        f"{float(athlete_score_meta['score']):.1f} / 5.0 ({athlete_score_meta['label']})"
                        if athlete_score_meta is not None
                        else "Todavía no registrado"
                    ),
                    wellness_summary.get("score_n"),
                ),
                _athlete_metric_row(
                    "Sueño",
                    _athlete_metric(wellness_summary.get("sleep_mean"), digits=1, suffix=" h", fallback="Todavía no registrado"),
                    wellness_summary.get("sleep_n"),
                ),
                _athlete_metric_row(
                    "Estrés",
                    (
                        f"{_display_metric(wellness_summary.get('stress_mean'), digits=1)}{wellness_scale.get('stress', '')}"
                        if _coerce_float(wellness_summary.get("stress_mean")) is not None
                        else "Todavía no registrado"
                    ),
                    wellness_summary.get("stress_n"),
                ),
                _athlete_metric_row(
                    "Dolor",
                    (
                        f"{_display_metric(wellness_summary.get('pain_mean'), digits=1)}{wellness_scale.get('pain', '')}"
                        if _coerce_float(wellness_summary.get("pain_mean")) is not None
                        else "Todavía no registrado"
                    ),
                    wellness_summary.get("pain_n"),
                ),
                {"label": "Adherencia", "value": _athlete_completion_text(), "note": ""},
                {"label": "Días con registro", "value": str(athlete_score_days) if athlete_score_days > 0 else "Todavía no registrado", "note": ""},
            ]

        def _athlete_wellness_summary_line() -> str:
            if wellness_context.get("state") == "missing":
                return "Todavía faltan registros recientes de bienestar para entender mejor cómo venís recuperando."
            if wellness_context.get("state") == "partial":
                if _athlete_wellness_signal_is_concerning():
                    return "Hay pocos registros, pero lo disponible sugiere cuidar recuperación, estrés y dolor de cerca."
                return "Hay pocos registros de bienestar: sirven como una primera señal, pero todavía no como una tendencia cerrada."
            if athlete_score_value is not None and athlete_score_value < 3.0:
                return "El bienestar reciente viene bajo y conviene cruzarlo de cerca con la carga, el sueño, el estrés y el dolor."
            return "El bienestar reciente acompaña y suma contexto para entender mejor cómo venís recuperando."

        def _athlete_wellness_watch_lines() -> list[str]:
            if wellness_context.get("state") == "missing":
                return [
                    "Sueño, estrés y dolor todavía necesitan más registros para leerse mejor.",
                    "La carga se interpreta mejor cuando se cruza con sueño, estrés, dolor y adherencia.",
                ]
            if wellness_context.get("state") == "partial":
                if _athlete_wellness_signal_is_concerning():
                    return [
                        "Los registros disponibles sugieren cuidar recuperación, estrés y dolor de cerca.",
                        "Con más registros vamos a poder leer mejor la tendencia sin sobrerreaccionar a un solo día.",
                    ]
                return [
                    "Con más registros vamos a poder leer mejor la tendencia de recuperación.",
                    "La carga se interpreta mejor cuando se cruza con sueño, estrés, dolor y adherencia.",
                ]
            if _athlete_wellness_signal_is_concerning():
                return [
                    "Si sueño, estrés o dolor empeoran, conviene ajustar la carga antes de forzar.",
                    "La carga se interpreta mejor cuando se cruza con sueño, estrés, dolor y adherencia.",
                ]
            return [
                "El bienestar reciente acompaña, pero conviene seguir registrándolo para confirmar la tendencia.",
                "La carga se interpreta mejor cuando se cruza con sueño, estrés, dolor y adherencia.",
            ]

        def _athlete_imtp_peak_force() -> object:
            return force_time_summary.get("peak_force_n") if _coerce_float(force_time_summary.get("peak_force_n")) is not None else focus_row.get("IMTP N")

        def _athlete_imtp_cards() -> list[tuple[str, str, str]]:
            stronger_side = _professional_force_side_label(asymmetry_summary.get("stronger_side"))
            time_to_peak = _coerce_float(force_time_summary.get("time_to_peak_s"))
            return [
                ("Fuerza máxima", _athlete_metric(_athlete_imtp_peak_force(), digits=0, suffix=" N"), "#134263"),
                ("Asimetría", _athlete_metric(force_time_summary.get("absolute_asymmetry_pct"), digits=1, suffix="%", fallback=_athlete_metric(focus_row.get("IMTP_asym_pct"), digits=1, suffix="%")), "#B87445"),
                ("Lado dominante", stronger_side if stronger_side != PDF_MISSING_TEXT else "Sin dato", "#708C9F"),
                ("Tiempo al pico", _athlete_metric(time_to_peak, digits=2, suffix=" s"), "#0D3C5E"),
            ]

        def _athlete_force_time_copy(text: object) -> str:
            clean = _repair_mojibake_text(str(text or "")).strip()
            replacements = {
                "evaluacion": "evaluación",
                "adquisicion": "adquisición",
                "Asimetria": "Asimetría",
                "asimetria": "asimetría",
                "produccion": "producción",
                "maxima": "máxima",
                "isometrica": "isométrica",
                "posicion": "posición",
                "medicion": "medición",
                "descripcion": "descripción",
                "especifico": "específico",
                "expresion": "expresión",
            }
            for source, target in replacements.items():
                clean = re.sub(rf"\b{re.escape(source)}\b", target, clean)
            clean = clean.replace("describe como expresa", "describe cómo expresa")
            clean = clean.replace("como descripción", "como descripción")
            clean = clean.replace("cómo descripción", "como descripción")
            return clean

        def _athlete_two_col_cards(cards: list[tuple[str, str, str]]) -> Table:
            rows: list[list[object]] = []
            row_cells: list[object] = []
            for idx, (label, value, _) in enumerate(cards, start=1):
                row_cells.append(
                    _build_metric_card(
                        label,
                        value,
                        paragraph_builder=_p,
                        TableClass=Table,
                        TableStyleClass=TableStyle,
                        mm_unit=mm,
                        palette=palette,
                        width_mm=86,
                    )
                )
                if idx % 2 == 0:
                    rows.append(row_cells)
                    row_cells = []
            if row_cells:
                while len(row_cells) < 2:
                    row_cells.append("")
                rows.append(row_cells)
            table = Table(rows, colWidths=[86 * mm, 86 * mm], hAlign="LEFT")
            table.setStyle(TableStyle([("VALIGN", (0, 0), (-1, -1), "TOP")]))
            return table

        def _athlete_force_time_table() -> Table | None:
            if not force_time_points:
                return None
            data: list[list[object]] = [[_p("Ventana", "ReportTableHeader"), _p("Fuerza", "ReportTableHeader")]]
            for point in force_time_points[:5]:
                label = str(point.get("label") or "-")
                if label.casefold() == "peak":
                    label = "Pico"
                data.append(
                    [
                        _p(label, "ReportTableCell"),
                        _p(_athlete_metric(point.get("value_n"), digits=0, suffix=" N"), "ReportTableCell"),
                    ]
                )
            return _build_threshold_table(
                data,
                TableClass=Table,
                TableStyleClass=TableStyle,
                mm_unit=mm,
                palette=palette,
                col_widths_mm=[44, 130],
                header_rows=1,
                repeat_rows=1,
                border_width=0.7,
                inner_grid_width=0.35,
                left_padding=6,
                right_padding=6,
                top_padding=5,
                bottom_padding=5,
            )

        def _athlete_imtp_summary_line() -> str:
            peak_force = _coerce_float(_athlete_imtp_peak_force())
            asymmetry_value = _coerce_float(force_time_summary.get("absolute_asymmetry_pct"))
            if peak_force is None:
                return "Todavía no hay un IMTP reciente para leer tu base de fuerza con más claridad."
            line = "El IMTP te da una referencia de fuerza máxima registrada en una posición fija."
            if asymmetry_value is not None:
                line += " La diferencia entre lados ayuda a ver si uno está empujando más que el otro."
            return line

        def _athlete_imtp_meaning_lines() -> list[str]:
            peak_force = _coerce_float(_athlete_imtp_peak_force())
            asymmetry_value = _coerce_float(force_time_summary.get("absolute_asymmetry_pct"))
            lines: list[str] = []
            if peak_force is not None:
                lines.append(f"Tu fuerza máxima registrada en esta medición fue {_display_metric(peak_force, digits=0, suffix=' N')}.")
            if asymmetry_value is not None:
                stronger_side = _professional_force_side_label(asymmetry_summary.get("stronger_side"))
                if stronger_side != PDF_MISSING_TEXT:
                    lines.append(
                        f"La diferencia entre lados fue de {_display_metric(asymmetry_value, digits=1, suffix='%')} y el lado que más empujó fue {stronger_side}."
                    )
                else:
                    lines.append(f"La diferencia entre lados fue de {_display_metric(asymmetry_value, digits=1, suffix='%')}.")
            elif peak_force is not None:
                lines.append("En esta medición no aparece una diferencia clara entre lados para seguir de cerca.")
            if force_time_points:
                lines.append("Los puntos force-time muestran cómo aparece la fuerza al inicio y cómo sigue creciendo antes del pico.")
            elif peak_force is not None:
                lines.append("Aunque falten puntos force-time, igual podemos usar el IMTP como referencia simple de fuerza máxima.")
            return lines[:3] or ["Usamos esta referencia para entender tu base de fuerza y qué conviene cuidar en el próximo bloque."]

        def _athlete_imtp_care_line() -> str:
            if _coerce_float(_athlete_imtp_peak_force()) is None:
                return "Conviene completar un IMTP para tener una referencia más clara de fuerza y simetría."
            return "Lo usamos como referencia, no como conclusión aislada: conviene cruzarlo con salto, carga y cómo venís recuperando."

        def _athlete_watch_block_text() -> str:
            if completion_value is not None and completion_value < 80:
                return "Sostener más continuidad del plan va a ayudar a leer mejor cómo responde el proceso."
            if wellness_context.get("state") == "missing":
                return "Mantener registros de bienestar va a ayudar a confirmar mejor la tendencia."
            if wellness_context.get("state") == "partial":
                return _athlete_wellness_watch_lines()[0]
            if _athlete_wellness_signal_is_concerning():
                return "La recuperación reciente merece seguimiento cercano junto con sueño, estrés y dolor."
            if _coerce_float(focus_row.get("Monotonia")) is not None and _coerce_float(focus_row.get("Monotonia")) > 2.0:
                return "Conviene que la semana no se repita demasiado para sostener mejor la tolerancia."
            if load_available and _professional_normalized_text(_athlete_zone_text()) in {"alto riesgo", "precaucion", "subcarga"}:
                return "Conviene seguir de cerca la carga reciente para sostener continuidad sin cambios bruscos."
            return "El bienestar reciente acompaña, pero conviene seguir registrándolo para confirmar la tendencia."

        def _athlete_chrome(canvas_obj, doc_obj) -> None:
            _draw_threshold_header(canvas_obj, doc_obj, palette=palette, mm_unit=mm, label="")
            _draw_threshold_footer(canvas_obj, doc_obj, palette=palette, mm_unit=mm)

        cover_intro = Table(
            [[
                _build_note_box(
                    "Lectura simple",
                    _athlete_overview_text(),
                    paragraph_builder=_p,
                    TableClass=Table,
                    TableStyleClass=TableStyle,
                    mm_unit=mm,
                    palette=palette,
                    width_mm=112,
                    title_style="BlockTitle",
                    body_style="ReportBody",
                    background=palette["panel"],
                    border_color=palette["line_dark"],
                    accent_color=palette["navy"],
                ),
                _athlete_status_box(),
            ]],
            colWidths=[112 * mm, 58 * mm],
            hAlign="LEFT",
        )
        cover_intro.setStyle(
            TableStyle(
                [
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("LEFTPADDING", (0, 0), (-1, -1), 0),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 0),
                    ("TOPPADDING", (0, 0), (-1, -1), 0),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
                ]
            )
        )

        story.extend(
            [
                _p("Reporte individual para atleta", "ReportMuted"),
                _p(effective_athlete, "ReportTitle"),
                _p(f"Fecha: {datetime.now():%d/%m/%Y} | Ventana visible: últimas 6 semanas", "ReportMuted"),
                Spacer(1, 4 * mm),
                cover_intro,
                Spacer(1, 5 * mm),
                _metric_cards_table(_athlete_cover_cards()),
                Spacer(1, 5 * mm),
                _build_decision_box(
                    "Foco del próximo bloque",
                    _athlete_focus_plan_text(),
                    paragraph_builder=_p,
                    TableClass=Table,
                    TableStyleClass=TableStyle,
                    mm_unit=mm,
                    palette=palette,
                    title_style="ReportDecisionTitle",
                    body_style="ReportBodyWhite",
                    note_style="ReportMutedWhite",
                    note="Lectura pensada para entender qué sigue y qué conviene cuidar ahora.",
                ),
                Spacer(1, 4 * mm),
                _build_note_box(
                    "Diccionario rápido",
                    _athlete_dictionary_lines(),
                    paragraph_builder=_p,
                    TableClass=Table,
                    TableStyleClass=TableStyle,
                    mm_unit=mm,
                    palette=palette,
                    title_style="BlockTitle",
                    body_style="ReportMuted",
                    background=palette["card"],
                    border_color=palette["line_dark"],
                    accent_color=palette["steel"],
                ),
            ]
        )

        story.append(PageBreak())
        story.append(
            _build_pdf_page_title(
                "Perfil neuromuscular",
                paragraph_builder=_p,
                TableClass=Table,
                TableStyleClass=TableStyle,
                mm_unit=mm,
                palette=palette,
                subtitle="Una lectura simple de tu perfil actual de salto, reactividad y fuerza para orientar el próximo bloque.",
                eyebrow="",
                title_style="ReportSection",
                subtitle_style="ReportMuted",
                eyebrow_style="ReportMuted",
            )
        )
        story.append(Spacer(1, 4 * mm))
        story.append(
            _build_note_box(
                "Qué muestra",
                profile_reading["what"],
                paragraph_builder=_p,
                TableClass=Table,
                TableStyleClass=TableStyle,
                mm_unit=mm,
                palette=palette,
                title_style="BlockTitle",
                body_style="ReportBody",
                background=palette["card"],
                border_color=palette["line_dark"],
                accent_color=palette["navy"],
            )
        )
        story.append(Spacer(1, 4 * mm))
        radar_payload = athlete_charts.get("radar_perfil")
        radar_image = _chart_image(radar_payload, height_mm=70)
        story.append(
            _build_chart_container(
                radar_image,
                paragraph_builder=_p,
                TableClass=Table,
                TableStyleClass=TableStyle,
                mm_unit=mm,
                palette=palette,
                title="Radar de perfil",
                note="El radar acompaña la lectura, pero la decisión se apoya sobre todo en qué significa hoy para vos.",
                empty_text=(
                    "Hoy no se pudo renderizar el radar, pero la lectura del perfil sigue disponible."
                    if radar_payload
                    else "Faltan datos de evaluación para mostrar el radar neuromuscular."
                ),
                title_style="BlockTitle",
                note_style="ReportMuted",
                accent_color=palette["steel"],
            )
        )
        story.append(Spacer(1, 4 * mm))
        story.append(
            _build_note_box(
                "Qué significa para vos",
                profile_reading["meaning"],
                paragraph_builder=_p,
                TableClass=Table,
                TableStyleClass=TableStyle,
                mm_unit=mm,
                palette=palette,
                title_style="BlockTitle",
                body_style="ReportBody",
                background=palette["panel"],
                border_color=palette["line_dark"],
                accent_color=palette["steel"],
            )
        )
        story.append(Spacer(1, 4 * mm))
        story.append(
            _build_decision_box(
                "Qué vamos a priorizar",
                profile_reading["priority"],
                paragraph_builder=_p,
                TableClass=Table,
                TableStyleClass=TableStyle,
                mm_unit=mm,
                palette=palette,
                title_style="ReportDecisionTitle",
                body_style="ReportBodyWhite",
                note_style="ReportMutedWhite",
                note="El foco no cambia la etiqueta del perfil: traduce esa foto a una prioridad de trabajo.",
            )
        )

        story.append(PageBreak())
        story.append(
            _build_pdf_page_title(
                "Fuerza e IMTP",
                paragraph_builder=_p,
                TableClass=Table,
                TableStyleClass=TableStyle,
                mm_unit=mm,
                palette=palette,
                subtitle="Una lectura simple de tu referencia de fuerza isométrica y, si existe, de cómo aparece esa fuerza en el tiempo.",
                eyebrow="",
                title_style="ReportSection",
                subtitle_style="ReportMuted",
                eyebrow_style="ReportMuted",
            )
        )
        story.append(Spacer(1, 4 * mm))
        story.append(
            _build_note_box(
                "Qué muestra",
                _athlete_imtp_summary_line(),
                paragraph_builder=_p,
                TableClass=Table,
                TableStyleClass=TableStyle,
                mm_unit=mm,
                palette=palette,
                title_style="BlockTitle",
                body_style="ReportBody",
                background=palette["card"],
                border_color=palette["line_dark"],
                accent_color=palette["navy"],
            )
        )
        story.append(Spacer(1, 4 * mm))
        story.append(_athlete_two_col_cards(_athlete_imtp_cards()))
        force_time_table = _athlete_force_time_table()
        if force_time_table is not None:
            story.append(Spacer(1, 4 * mm))
            story.append(
                _build_note_box(
                    "Cómo aparece la fuerza al inicio",
                    "Estos puntos muestran cómo empieza a aparecer la fuerza y cómo sigue creciendo antes del pico.",
                    paragraph_builder=_p,
                    TableClass=Table,
                    TableStyleClass=TableStyle,
                    mm_unit=mm,
                    palette=palette,
                    title_style="BlockTitle",
                    body_style="ReportMuted",
                    background=palette["panel"],
                    border_color=palette["line_dark"],
                    accent_color=palette["steel"],
                )
            )
            story.append(Spacer(1, 3 * mm))
            story.append(force_time_table)
        story.append(Spacer(1, 4 * mm))
        story.append(
            _build_note_box(
                "Qué significa para vos",
                _athlete_imtp_meaning_lines(),
                paragraph_builder=_p,
                TableClass=Table,
                TableStyleClass=TableStyle,
                mm_unit=mm,
                palette=palette,
                title_style="BlockTitle",
                body_style="ReportBody",
                background=palette["panel"],
                border_color=palette["line_dark"],
                accent_color=palette["steel"],
            )
        )
        story.append(Spacer(1, 4 * mm))
        story.append(
            _build_note_box(
                "Qué vamos a cuidar",
                [
                    _athlete_imtp_care_line(),
                    "La RFD ayuda a ver qué tan rápido aparece la fuerza, pero la usamos con cautela cuando todavía no tenemos una referencia propia de confiabilidad."
                    if rfd_points
                    else "Si todavía faltan datos force-time, igual podemos usar el IMTP como referencia simple de fuerza.",
                ],
                paragraph_builder=_p,
                TableClass=Table,
                TableStyleClass=TableStyle,
                mm_unit=mm,
                palette=palette,
                title_style="BlockTitle",
                body_style="ReportBody",
                background=palette["card"],
                border_color=palette["line_dark"],
                accent_color=palette["navy"],
            )
        )

        story.append(PageBreak())
        story.append(
            _build_pdf_page_title(
                "Carga reciente",
                paragraph_builder=_p,
                TableClass=Table,
                TableStyleClass=TableStyle,
                mm_unit=mm,
                palette=palette,
                subtitle="Una lectura simple de cómo venís tolerando el entrenamiento y de si la carga se está moviendo con continuidad.",
                eyebrow="",
                title_style="ReportSection",
                subtitle_style="ReportMuted",
                eyebrow_style="ReportMuted",
            )
        )
        story.append(Spacer(1, 4 * mm))
        story.append(
            _build_note_box(
                "Qué significa",
                _athlete_load_summary_line(),
                paragraph_builder=_p,
                TableClass=Table,
                TableStyleClass=TableStyle,
                mm_unit=mm,
                palette=palette,
                title_style="BlockTitle",
                body_style="ReportBody",
                background=palette["card"],
                border_color=palette["line_dark"],
                accent_color=palette["navy"],
            )
        )
        story.append(Spacer(1, 4 * mm))
        load_chart_payload = athlete_charts.get("acwr")
        load_image = _chart_image(load_chart_payload, height_mm=56)
        story.append(
            _build_chart_container(
                load_image,
                paragraph_builder=_p,
                TableClass=Table,
                TableStyleClass=TableStyle,
                mm_unit=mm,
                palette=palette,
                title="Cómo venís tolerando el entrenamiento",
                note=(
                    "El gráfico acompaña la lectura de continuidad, tolerancia y ritmo reciente del entrenamiento."
                    if load_image is not None
                    else "Con la información disponible, la lectura sigue siendo útil aunque hoy no haya gráfico para mostrar."
                ),
                empty_text=(
                    "Hoy no se pudo renderizar el gráfico de carga, pero la lectura reciente sigue disponible."
                    if load_chart_payload
                    else "Faltan datos de carga interna para mostrar este gráfico reciente."
                ),
                title_style="BlockTitle",
                note_style="ReportMuted",
                accent_color=palette["steel"],
            )
        )
        story.append(Spacer(1, 4 * mm))
        story.append(_compact_key_value_table(_athlete_load_rows()))
        story.append(Spacer(1, 4 * mm))
        story.append(
            _build_note_box(
                "Qué cuidar",
                _athlete_load_watch_lines(),
                paragraph_builder=_p,
                TableClass=Table,
                TableStyleClass=TableStyle,
                mm_unit=mm,
                palette=palette,
                title_style="BlockTitle",
                body_style="ReportBody",
                background=palette["panel"],
                border_color=palette["line_dark"],
                accent_color=palette["steel"],
            )
        )

        story.append(PageBreak())
        story.append(
            _build_pdf_page_title(
                "Bienestar y adherencia",
                paragraph_builder=_p,
                TableClass=Table,
                TableStyleClass=TableStyle,
                mm_unit=mm,
                palette=palette,
                subtitle="La carga se entiende mejor cuando se cruza con sueño, estrés, dolor y continuidad del plan.",
                eyebrow="",
                title_style="ReportSection",
                subtitle_style="ReportMuted",
                eyebrow_style="ReportMuted",
            )
        )
        story.append(Spacer(1, 4 * mm))
        story.append(
            _build_note_box(
                "Qué significa",
                _athlete_wellness_summary_line(),
                paragraph_builder=_p,
                TableClass=Table,
                TableStyleClass=TableStyle,
                mm_unit=mm,
                palette=palette,
                title_style="BlockTitle",
                body_style="ReportBody",
                background=palette["card"],
                border_color=palette["line_dark"],
                accent_color=palette["navy"],
            )
        )
        story.append(Spacer(1, 4 * mm))
        wellness_chart_payload = athlete_charts.get("wellness") or athlete_charts.get("completion")
        wellness_image = _chart_image(wellness_chart_payload, height_mm=52)
        story.append(
            _build_chart_container(
                wellness_image,
                paragraph_builder=_p,
                TableClass=Table,
                TableStyleClass=TableStyle,
                mm_unit=mm,
                palette=palette,
                title=str(wellness_chart_payload.get("title") or "Bienestar reciente") if wellness_chart_payload else "Bienestar y adherencia",
                note=(
                    "Sirve para sumar contexto sobre recuperación y continuidad, no para juzgar un día aislado."
                    if wellness_image is not None
                    else "Con más registros esta vista va a ayudar a seguir mejor cómo venís recuperando y sosteniendo el plan."
                ),
                empty_text=(
                    "Hoy no se pudo renderizar este gráfico, pero la lectura de bienestar y adherencia sigue disponible."
                    if wellness_chart_payload
                    else "Todavía no hay registros suficientes para mostrar este gráfico de bienestar o adherencia."
                ),
                title_style="BlockTitle",
                note_style="ReportMuted",
                accent_color=palette["steel"],
            )
        )
        story.append(Spacer(1, 4 * mm))
        story.append(_compact_key_value_table(_athlete_wellness_rows()))
        story.append(Spacer(1, 4 * mm))
        story.append(
            _build_note_box(
                "Qué cuidar",
                _athlete_wellness_watch_lines(),
                paragraph_builder=_p,
                TableClass=Table,
                TableStyleClass=TableStyle,
                mm_unit=mm,
                palette=palette,
                title_style="BlockTitle",
                body_style="ReportBody",
                background=palette["panel"],
                border_color=palette["line_dark"],
                accent_color=palette["steel"],
            )
        )

        story.append(PageBreak())
        story.append(
            _build_pdf_page_title(
                "Fortalezas y próximos pasos",
                paragraph_builder=_p,
                TableClass=Table,
                TableStyleClass=TableStyle,
                mm_unit=mm,
                palette=palette,
                subtitle="Lo más importante de esta ventana, lo que conviene vigilar y qué sigue para el próximo bloque.",
                eyebrow="",
                title_style="ReportSection",
                subtitle_style="ReportMuted",
                eyebrow_style="ReportMuted",
            )
        )
        story.append(Spacer(1, 4 * mm))
        strength_block = next((block for block in final_blocks if block.get("title") == "Fortaleza principal"), {})
        review_block = next((block for block in final_blocks if block.get("title") == "Próxima medición o revisión"), {})
        story.append(
            _build_decision_box(
                "Foco del próximo bloque",
                _athlete_focus_plan_text(),
                paragraph_builder=_p,
                TableClass=Table,
                TableStyleClass=TableStyle,
                mm_unit=mm,
                palette=palette,
                title_style="ReportDecisionTitle",
                body_style="ReportBodyWhite",
                note_style="ReportMutedWhite",
                note="El foco principal traduce la foto actual del atleta a una prioridad concreta de trabajo.",
            )
        )
        story.append(Spacer(1, 4 * mm))
        story.append(
            _build_note_box(
                "Fortaleza principal",
                str(strength_block.get("body") or PDF_MISSING_TEXT),
                paragraph_builder=_p,
                TableClass=Table,
                TableStyleClass=TableStyle,
                mm_unit=mm,
                palette=palette,
                title_style="BlockTitle",
                body_style="ReportBody",
                background=palette["card"],
                border_color=palette["line_dark"],
                accent_color=palette["navy"],
            )
        )
        story.append(Spacer(1, 4 * mm))
        story.append(
            _build_note_box(
                "Punto a vigilar",
                _athlete_watch_block_text(),
                paragraph_builder=_p,
                TableClass=Table,
                TableStyleClass=TableStyle,
                mm_unit=mm,
                palette=palette,
                title_style="BlockTitle",
                body_style="ReportBody",
                background=palette["panel"],
                border_color=palette["line_dark"],
                accent_color=palette["steel"],
            )
        )
        story.append(Spacer(1, 4 * mm))
        story.append(
            _build_note_box(
                "Próxima medición o revisión",
                str(review_block.get("body") or PDF_MISSING_TEXT),
                paragraph_builder=_p,
                TableClass=Table,
                TableStyleClass=TableStyle,
                mm_unit=mm,
                palette=palette,
                title_style="BlockTitle",
                body_style="ReportBody",
                background=palette["card"],
                border_color=palette["line_dark"],
                accent_color=palette["navy"],
            )
        )
        story.append(Spacer(1, 4 * mm))
        story.append(
            _build_note_box(
                "Nota de lectura",
                "Este informe no promete rendimiento ni reemplaza el criterio del entrenador. Sirve para entender qué se ve hoy, qué significa para vos y qué conviene hacer después con la información disponible.",
                paragraph_builder=_p,
                TableClass=Table,
                TableStyleClass=TableStyle,
                mm_unit=mm,
                palette=palette,
                title_style="BlockTitle",
                body_style="ReportMuted",
                background=palette["panel"],
                border_color=palette["line_dark"],
                accent_color=palette["steel"],
            )
        )
        try:
            doc.build(story, onFirstPage=_athlete_chrome, onLaterPages=_athlete_chrome)
        except Exception:
            return None
        return buffer.getvalue()

    if audience == "cliente" and effective_athlete != "Todos":
        focus_row = summary_df.iloc[0] if not summary_df.empty else pd.Series(dtype=object)
        client_neuromuscular_profile = _build_current_pdf_neuromuscular_profile_payload(
            state,
            effective_athlete,
            audience="cliente",
        )
        completion_value = _focus_completion_value(state, effective_athlete)
        internal_load = _build_professional_internal_load_context(state, effective_athlete)
        wellness_context = _professional_wellness_context(state, effective_athlete)
        client_charts = _collect_client_pdf_chart_payloads(state, effective_athlete)

        def _client_missing_text(kind: str = "generic") -> str:
            mapping = {
                "generic": "Dato todavía no disponible",
                "evaluation": "Pendiente de evaluación",
                "registration": "Todavía no registrado",
                "reference": "Falta completar esta referencia",
            }
            return mapping.get(kind, mapping["generic"])

        def _client_text(value: object, fallback: str = "Dato todavía no disponible") -> str:
            text = _ascii_text(value).strip()
            normalized = text.casefold()
            if not text or normalized in {"-", "sin dato", "sin datos", "nan", "none"}:
                return fallback
            return text

        def _client_metric(value: object, *, digits: int = 1, suffix: str = "", fallback: str = "Dato todavía no disponible") -> str:
            return _client_text(_display_metric(value, digits=digits, suffix=suffix), fallback=fallback)

        def _client_focus_label() -> str:
            if _row_has_eval_data(focus_row) and client_neuromuscular_profile.get("source") == "core":
                return _client_text(
                    client_neuromuscular_profile.get("training_priority_short"),
                    fallback="Seguir construyendo",
                )
            return _client_text(_current_focus_text(focus_row, audience="cliente"), fallback="Seguir construyendo")

        def _client_focus_phrase() -> str:
            focus = _sentence_fragment(_client_focus_label())
            normalized = _professional_normalized_text(focus)
            if normalized == "hacer nueva evaluacion":
                return "hacer una nueva evaluación"
            return focus.lower()

        def _client_focus_summary_sentence() -> str:
            normalized = _professional_normalized_text(_client_focus_label())
            if normalized == "hacer nueva evaluacion":
                return "El foco actual es completar una nueva evaluación y sostener continuidad."
            if normalized == "sostener progreso":
                return "El foco actual es sostener continuidad y seguir construyendo."
            if normalized == "ordenar carga":
                return "El foco actual es ordenar la carga de la semana y sostener continuidad."
            if normalized == "ganar continuidad":
                return "El foco actual es ganar continuidad para leer mejor el proceso."
            if normalized == "recuperar potencia":
                return "El foco actual es recuperar potencia sin perder continuidad."
            if normalized == "construir fuerza base":
                return "El foco actual es construir fuerza base y sostener continuidad."
            return f"El foco actual es {_client_focus_phrase()}."

        def _client_focus_plan_sentence() -> str:
            normalized = _professional_normalized_text(_client_focus_label())
            if normalized == "hacer nueva evaluacion":
                return "El foco ahora es completar una nueva evaluación para tener una referencia más clara y seguir ajustando el proceso."
            if normalized == "sostener progreso":
                return "El foco ahora es sostener lo que viene funcionando y seguir construyendo con continuidad."
            if normalized == "ordenar carga":
                return "El foco ahora es ordenar la carga de la semana para que el cuerpo la tolere mejor y podamos sostener continuidad."
            if normalized == "ganar continuidad":
                return "El foco ahora es ganar continuidad en el plan para que la lectura del proceso sea más clara."
            if normalized == "recuperar potencia":
                return "El foco ahora es recuperar potencia con calma, sin perder continuidad en el proceso."
            if normalized == "construir fuerza base":
                return "El foco ahora es construir una base más sólida para que el progreso se sostenga mejor."
            return f"El foco ahora es {_client_focus_phrase()}."

        def _client_completion_needs_follow_up() -> bool:
            return completion_value is not None and completion_value < 80

        def _client_recent_wellness_value() -> float | None:
            return _coerce_float(focus_row.get("Wellness 3d"))

        def _client_recent_wellness_is_low(value: float | None) -> bool:
            if value is None:
                return False
            return value < 15 if value > 5 else value < 3.0

        def _client_recent_wellness_label() -> str:
            return (
                _client_metric(_client_recent_wellness_value(), digits=1, fallback=_client_missing_text("registration"))
                if _client_recent_wellness_value() is not None
                else _client_missing_text("registration")
            )

        def _client_wellness_mean_value() -> float | None:
            summary = wellness_context.get("last_week_summary", {}) if isinstance(wellness_context.get("last_week_summary"), dict) else {}
            return _coerce_float(summary.get("score_mean"))

        def _client_wellness_days_count() -> int:
            summary = wellness_context.get("last_week_summary", {}) if isinstance(wellness_context.get("last_week_summary"), dict) else {}
            day_candidates = [
                _coerce_float(summary.get("days")),
                _coerce_float(summary.get("score_n")),
                _coerce_float(summary.get("sleep_n")),
                _coerce_float(summary.get("stress_n")),
                _coerce_float(summary.get("pain_n")),
            ]
            return max([int(value) for value in day_candidates if value is not None] or [0])

        def _client_wellness_signal_is_concerning() -> bool:
            mean_value = _client_wellness_mean_value()
            recent_value = _client_recent_wellness_value()
            summary = wellness_context.get("last_week_summary", {}) if isinstance(wellness_context.get("last_week_summary"), dict) else {}
            scales = wellness_context.get("scales", {}) if isinstance(wellness_context.get("scales"), dict) else {}
            stress_mean = _coerce_float(summary.get("stress_mean"))
            pain_mean = _coerce_float(summary.get("pain_mean"))
            return bool(
                (mean_value is not None and mean_value < 3.0)
                or _client_recent_wellness_is_low(recent_value)
                or (stress_mean is not None and _professional_wellness_high(stress_mean, scales.get("stress")))
                or (pain_mean is not None and _professional_wellness_high(pain_mean, scales.get("pain")))
            )

        def _client_wellness_summary_line() -> str:
            if wellness_context.get("state") == "missing":
                return "Falta información reciente para entender cómo venís recuperando."
            if wellness_context.get("state") == "partial":
                if _client_wellness_signal_is_concerning():
                    return "El registro es bajo, pero los valores disponibles sugieren cuidar recuperación, estrés y dolor de cerca."
                return "Todavía hay pocos registros para sacar una conclusión firme sobre recuperación."
            mean_value = _client_wellness_mean_value()
            recent_value = _client_recent_wellness_value()
            if _client_recent_wellness_is_low(recent_value) and mean_value is not None and mean_value >= 3.0:
                return "El promedio reciente es aceptable, pero conviene seguir de cerca la recuperación percibida."
            if mean_value is not None and mean_value < 3.0:
                return "El promedio reciente viene algo bajo y conviene seguir de cerca la recuperación."
            if _client_recent_wellness_is_low(recent_value):
                return "La recuperación percibida viene baja y conviene seguirla de cerca."
            return "El bienestar reciente acompaña sin señales grandes de alerta."

        def _client_strengths_lines() -> list[str]:
            lines: list[str] = []
            zone = _professional_normalized_text(_client_zone_text())
            if _row_has_eval_data(focus_row):
                lines.append("Ya tenemos una referencia útil para seguir tu proceso.")
            if _row_has_load_data(focus_row) and all(flag not in zone for flag in ["alto riesgo", "precaucion", "subcarga"]):
                lines.append("La carga reciente está en una zona útil para seguir construyendo.")
            mean_value = _client_wellness_mean_value()
            if mean_value is not None and mean_value >= 3.0:
                lines.append("El promedio reciente de bienestar se mantiene en una zona aceptable.")
            if completion_value is not None and completion_value >= 80:
                lines.append("La constancia del plan ayuda a leer mejor cómo viene el proceso.")
            return list(dict.fromkeys(lines))[:3] or ["Ya empezamos a juntar información útil para seguir tu proceso."]

        def _client_watchouts_lines() -> list[str]:
            lines: list[str] = []
            zone = _professional_normalized_text(_client_zone_text())
            summary_parts: list[str] = []
            if _row_has_eval_data(focus_row) and client_neuromuscular_profile.get("source") == "core":
                translated_flags = list(client_neuromuscular_profile.get("flag_messages_client", []))
                if translated_flags:
                    return translated_flags[:1] + ["Conviene usar el próximo control para confirmar si esta señal se sostiene."]
            if "alto riesgo" in zone:
                summary_parts.append("carga reciente")
                lines.append("La carga reciente viene exigente y conviene ordenarla con un poco más de cuidado.")
            elif "precaucion" in zone:
                summary_parts.append("carga reciente")
                lines.append("La carga reciente pide un poco más de cuidado para sostener continuidad.")
            elif "subcarga" in zone:
                summary_parts.append("carga reciente")
                lines.append("La carga reciente quedó baja y conviene recuperar continuidad.")
            wellness_line = _client_wellness_summary_line()
            if "conviene seguir" in wellness_line or "viene algo bajo" in wellness_line:
                summary_parts.append("recuperación percibida")
                lines.append(wellness_line)
            elif "registro es bajo" in wellness_line:
                summary_parts.extend(["recuperación", "estrés", "dolor"])
                lines.append(wellness_line)
            if _client_completion_needs_follow_up():
                summary_parts.append("constancia")
                lines.append("La constancia todavía merece seguimiento para leer mejor el proceso.")
            if summary_parts:
                labels = list(dict.fromkeys(summary_parts))
                if labels == ["recuperación percibida"]:
                    summary_text = "La recuperación percibida necesita seguimiento."
                elif labels == ["constancia"]:
                    summary_text = "La constancia necesita seguimiento."
                elif labels == ["carga reciente"]:
                    summary_text = "La carga reciente necesita seguimiento."
                elif len(labels) == 1:
                    summary_text = labels[0].capitalize() + " requiere seguimiento."
                else:
                    summary_text = (", ".join(labels[:-1]) + f" y {labels[-1]}").capitalize() + " requieren seguimiento."
                return [summary_text] + list(dict.fromkeys(lines))[:2]
            return ["No aparece una alarma grande; el foco está en sostener continuidad."]

        def _client_next_steps_list() -> list[str]:
            lines: list[str] = []
            zone = _professional_normalized_text(_client_zone_text())
            if _row_has_eval_data(focus_row) and client_neuromuscular_profile.get("source") == "core":
                lines.append(_professional_visible_metric_text(f"Sostener el foco en {_client_focus_phrase()}."))
            if _professional_normalized_text(_client_focus_label()) == "hacer nueva evaluacion":
                lines.append("Programar una nueva evaluación para sumar una referencia más clara del proceso.")
            if _client_completion_needs_follow_up():
                lines.append("Sostener más continuidad en el plan para que la lectura del proceso sea más clara.")
            if any(flag in zone for flag in ["alto riesgo", "precaucion", "subcarga"]):
                lines.append("Lo importante es evitar cambios bruscos y sostener continuidad de una semana a otra.")
            wellness_line = _client_wellness_summary_line()
            if "conviene seguir" in wellness_line or "viene algo bajo" in wellness_line or "registro es bajo" in wellness_line:
                lines.append("Si sueño, estrés, dolor o recuperación empeoran, ajustamos la carga antes de forzar.")
            if len(lines) < 2 and _professional_normalized_text(_client_focus_label()) != "hacer nueva evaluacion":
                lines.append("Volver a medir nos va a ayudar a confirmar si el proceso se sostiene.")
            if not lines:
                lines.append("Mantener la línea actual y volver a medir para confirmar si el proceso se sostiene.")
            return list(dict.fromkeys([line for line in lines if str(line).strip()]))[:3]

        def _client_load_watch_lines() -> list[str]:
            return [
                "Lo importante no es un día aislado, sino sostener continuidad y evitar cambios bruscos.",
                "Si la semana se vuelve demasiado exigente o demasiado liviana, ajustamos antes de que el proceso se vuelva confuso.",
            ]

        def _client_wellness_watch_lines() -> list[str]:
            if wellness_context.get("state") == "missing":
                return [
                    "Sueño, estrés y dolor todavía necesitan más registros para leerse mejor.",
                    "Cuando aparezcan más registros, vamos a ajustar con más precisión.",
                ]
            if wellness_context.get("state") == "partial":
                if _client_wellness_signal_is_concerning():
                    return [
                        "Recuperación, estrés y dolor merecen seguimiento cercano mientras sumamos más registros.",
                        "Si los próximos registros siguen bajos, ajustamos la carga antes de forzar.",
                    ]
                return [
                    "Sueño, estrés y dolor merecen seguimiento mientras sumamos más registros.",
                    "Con más registros vamos a poder leer mejor la tendencia.",
                ]
            if _client_wellness_signal_is_concerning():
                return [
                    "Recuperación percibida, estrés y dolor merecen seguimiento cercano.",
                    "Si sueño, estrés o dolor empeoran, ajustamos la carga antes de forzar.",
                ]
            return [
                "Seguir registrando bienestar nos va a ayudar a confirmar la tendencia.",
                "Si sueño, estrés o dolor empeoran, ajustamos la carga antes de forzar.",
            ]

        def _client_zone_text() -> str:
            if not _row_has_load_data(focus_row) or not _has_text(focus_row.get("Zona")):
                return "En seguimiento"
            return _display_zone(focus_row.get("Zona"))

        def _client_cards() -> list[tuple[str, str, str]]:
            return [
                (
                    "Salto vertical",
                    _client_metric(focus_row.get("CMJ cm"), digits=1, suffix=" cm", fallback=_client_missing_text("evaluation"))
                    if _row_has_eval_data(focus_row)
                    else _client_missing_text("evaluation"),
                    "#134263",
                ),
                (
                    "Bienestar reciente",
                    _client_recent_wellness_label(),
                    "#2F6B52",
                ),
                (
                    "Constancia",
                    _client_metric(completion_value, digits=1, suffix="%", fallback=_client_missing_text("registration"))
                    if completion_value is not None
                    else _client_missing_text("registration"),
                    "#708C9F",
                ),
                ("Foco actual", _client_focus_label(), "#134263"),
            ]

        def _client_overview_text() -> str:
            if summary_df.empty:
                return "Todavía falta información para una lectura más clara. Por ahora el foco es completar registros, sumar una evaluación inicial y sostener continuidad."
            if _row_has_eval_data(focus_row) or _row_has_load_data(focus_row) or _row_has_wellness_data(focus_row):
                lines = ["Hoy tenemos una referencia útil para seguir el proceso."]
                if _row_has_eval_data(focus_row) and client_neuromuscular_profile.get("source") == "core":
                    lines.append(
                        _professional_visible_metric_text(
                            client_neuromuscular_profile.get("summary_client")
                            or client_neuromuscular_profile.get("summary_short")
                            or ""
                        )
                    )
                zone = _professional_normalized_text(_client_zone_text())
                if _row_has_load_data(focus_row):
                    if "alto riesgo" in zone:
                        lines.append("La carga reciente viene exigente y conviene ordenarla con un poco más de cuidado.")
                    elif "precaucion" in zone:
                        lines.append("La carga reciente pide un poco más de cuidado para sostener continuidad.")
                    elif "subcarga" in zone:
                        lines.append("La carga reciente quedó baja y conviene recuperar continuidad.")
                    else:
                        lines.append("La carga reciente está en una zona útil para construir.")
                if _row_has_wellness_data(focus_row) or wellness_context.get("state") == "partial":
                    lines.append(_client_wellness_summary_line())
                if _client_completion_needs_follow_up():
                    if any(key in _professional_normalized_text(" ".join(lines)) for key in ["recuperacion", "registro es bajo", "pocos registros"]):
                        lines.append("La recuperación y la constancia necesitan seguimiento.")
                    else:
                        lines.append("La constancia necesita seguimiento para leer mejor el proceso.")
                lines.append(_client_focus_summary_sentence())
                return " ".join(lines)
            return "Ya empezamos a juntar información útil, pero todavía falta completar algunas referencias para leer mejor tu proceso. El foco actual es completar una nueva evaluación y sostener continuidad."

        def _client_state_rows() -> list[dict[str, object]]:
            rows: list[dict[str, object]] = [
                {"label": "Estado de carga", "value": _client_zone_text()},
                {
                    "label": "Tu salto hoy",
                    "value": _client_metric(focus_row.get("CMJ cm"), digits=1, suffix=" cm", fallback=_client_missing_text("evaluation"))
                    if _row_has_eval_data(focus_row)
                    else _client_missing_text("evaluation"),
                },
                {
                    "label": "Cambio contra tu primera referencia",
                    "value": _client_metric(focus_row.get("CMJ vs BL %"), digits=1, suffix="%", fallback=_client_missing_text("reference"))
                    if _coerce_float(focus_row.get("CMJ vs BL %")) is not None
                    else _client_missing_text("reference"),
                },
                {
                    "label": "Bienestar reciente",
                    "value": _client_metric(focus_row.get("Wellness 3d"), digits=1, fallback=_client_missing_text("registration"))
                    if _row_has_wellness_data(focus_row)
                    else _client_missing_text("registration"),
                },
                {
                    "label": "Constancia del plan",
                    "value": _client_metric(completion_value, digits=1, suffix="%", fallback=_client_missing_text("registration"))
                    if completion_value is not None
                    else _client_missing_text("registration"),
                },
                {
                    "label": "Lectura actual",
                    "value": (
                        _professional_visible_metric_text(
                            client_neuromuscular_profile.get("profile_label")
                            or "Ya tenemos una referencia útil para seguir tu proceso."
                        )
                        if _row_has_eval_data(focus_row) and client_neuromuscular_profile.get("source") == "core"
                        else (
                            "Ya tenemos una referencia útil para seguir tu proceso."
                            if _row_has_eval_data(focus_row)
                            else "Falta completar esta referencia para definir mejor la lectura actual."
                        )
                    ),
                },
            ]
            return rows

        def _client_load_rows() -> list[dict[str, object]]:
            change_text = _client_missing_text("reference")
            weekly_change_pct = _coerce_float(internal_load.get("weekly_change_pct"))
            sessions_registered = _coerce_float(internal_load.get("sessions_registered"))
            if str(internal_load.get("analysis_scope") or "") == "current_week_partial":
                change_text = "Semana en curso"
            elif weekly_change_pct is not None:
                change_text = f"{weekly_change_pct:+.1f}%"
            elif _coerce_float(internal_load.get("weekly_change")) is not None:
                change_text = _client_metric(internal_load.get("weekly_change"), digits=0, suffix=" UA")
            return [
                {"label": "Estado de la carga", "value": _client_zone_text()},
                {
                    "label": "Última semana visible",
                    "value": _client_metric(internal_load.get("last_week_total"), digits=0, suffix=" UA", fallback=_client_missing_text("registration"))
                    if _coerce_float(internal_load.get("last_week_total")) is not None
                    else ("Semana en curso" if str(internal_load.get("analysis_scope") or "") == "current_week_partial" else _client_missing_text("registration")),
                },
                {"label": "Cambio reciente", "value": change_text},
                {
                    "label": "Sesiones registradas",
                    "value": str(int(sessions_registered))
                    if sessions_registered is not None and (_row_has_load_data(focus_row) or str(internal_load.get("analysis_scope") or "") == "current_week_partial")
                    else _client_missing_text("registration"),
                },
                {
                    "label": "Constancia",
                    "value": _client_metric(completion_value, digits=1, suffix="%", fallback=_client_missing_text("registration"))
                    if completion_value is not None
                    else _client_missing_text("registration"),
                },
            ]

        def _client_load_lines() -> list[str]:
            lines: list[str] = []
            zone = _client_zone_text().casefold()
            scope = str(internal_load.get("analysis_scope") or "")
            if scope == "current_week_partial":
                total = _coerce_float(internal_load.get("current_week_total"))
                sessions = int(_coerce_float(internal_load.get("current_week_sessions")) or 0)
                if total is not None:
                    lines.append(f"La semana sigue en curso. Por ahora acumulaste {total:.0f} UA en {sessions} sesiones y conviene esperar el cierre semanal para compararla mejor.")
                else:
                    lines.append("La semana sigue en curso y todavía falta información para leer bien la carga reciente.")
            elif "alto riesgo" in zone:
                lines.append("La carga reciente viene exigente y conviene ordenar la semana para que el cuerpo la tolere mejor.")
            elif "precaucion" in zone:
                lines.append("La carga reciente pide un poco más de cuidado para sostener el progreso sin apurarse.")
            elif "subcarga" in zone:
                lines.append("La carga reciente quedó baja y puede faltar continuidad para seguir construyendo.")
            elif _row_has_load_data(focus_row):
                lines.append("La carga reciente se mueve en una zona útil para seguir construyendo con continuidad.")
            else:
                lines.append("Falta información reciente para entender bien cómo venís tolerando el entrenamiento.")
            if completion_value is not None and completion_value < 70:
                lines.append(f"La constancia actual ({completion_value:.1f}%) merece seguimiento porque puede volver más ruidosa la lectura del proceso.")
            elif completion_value is not None:
                lines.append("La constancia del plan ayuda a leer mejor si el proceso se está sosteniendo.")
            else:
                lines.append("La constancia del plan todavía no está registrada con suficiente claridad.")
            return lines[:3]

        def _client_wellness_rows() -> list[dict[str, object]]:
            summary = wellness_context.get("last_week_summary", {}) if isinstance(wellness_context.get("last_week_summary"), dict) else {}
            scales = wellness_context.get("scales", {}) if isinstance(wellness_context.get("scales"), dict) else {}
            score_value = _coerce_float(summary.get("score_mean"))
            score_label = _client_missing_text("registration")
            if score_value is not None:
                score_meta = _report_wellness_score_label(score_value)
                score_label = f"{float(score_meta['score']):.1f} / 5.0 ({score_meta['label']})"
            days_count = _client_wellness_days_count()
            return [
                {"label": "Promedio reciente", "value": score_label, "note": "Últimos registros visibles"},
                {"label": "Sueño", "value": _client_metric(summary.get("sleep_mean"), digits=1, suffix=" h", fallback=_client_missing_text("registration"))},
                {"label": "Estrés", "value": _client_metric(summary.get("stress_mean"), digits=1, suffix=str(scales.get("stress", "")), fallback=_client_missing_text("registration"))},
                {"label": "Dolor", "value": _client_metric(summary.get("pain_mean"), digits=1, suffix=str(scales.get("pain", "")), fallback=_client_missing_text("registration"))},
                {"label": "Días con registro", "value": str(days_count) if days_count > 0 else _client_missing_text("registration")},
            ]

        def _client_wellness_lines() -> list[str]:
            summary = wellness_context.get("last_week_summary", {}) if isinstance(wellness_context.get("last_week_summary"), dict) else {}
            scales = wellness_context.get("scales", {}) if isinstance(wellness_context.get("scales"), dict) else {}
            sleep_mean = _coerce_float(summary.get("sleep_mean"))
            stress_mean = _coerce_float(summary.get("stress_mean"))
            pain_mean = _coerce_float(summary.get("pain_mean"))
            if wellness_context.get("state") == "missing":
                return [
                    "Con más registros vamos a poder leer mejor la tendencia.",
                    "Sueño, estrés y dolor van a ayudarnos a entender mejor cómo venís recuperando.",
                ]
            if wellness_context.get("state") == "partial":
                if _client_wellness_signal_is_concerning():
                    return [
                        "Igual sirve como primera señal para mirar sueño, estrés y dolor.",
                        "Con más registros vamos a poder leer mejor la tendencia.",
                    ]
                return [
                    "Igual sirve como primera señal para mirar sueño, estrés y dolor.",
                    "Con más registros vamos a poder leer mejor la tendencia.",
                ]
            lines: list[str] = []
            if sleep_mean is not None and sleep_mean < 6.5:
                lines.append("El sueño reciente merece seguimiento para que la recuperación acompañe mejor el trabajo.")
            if stress_mean is not None and _professional_wellness_high(stress_mean, scales.get("stress")):
                lines.append("El estrés reciente aparece alto y conviene mirarlo junto con la carga de la semana.")
            if pain_mean is not None and _professional_wellness_high(pain_mean, scales.get("pain")):
                lines.append("El dolor reciente merece seguimiento antes de subir demasiado la exigencia.")
            if not lines:
                lines.append("Sueño, estrés y dolor no muestran desvíos grandes en esta ventana.")
            lines.append("La idea es usar esta información para cuidar recuperación y continuidad, no para juzgar un día aislado.")
            return list(dict.fromkeys(lines))[:3]

        def _client_review_text() -> str:
            if _row_has_eval_data(focus_row):
                return "Volver a medir en 6-8 semanas ayuda a confirmar si el progreso se sostiene."
            return "Cuando completemos una nueva evaluación, vamos a tener una referencia más clara para seguir ajustando el proceso."

        def _client_status_box() -> Table:
            return _build_threshold_box(
                [
                    _p("Carga actual", "CardLabel"),
                    _p(_client_zone_text(), "ReportSection"),
                    _p("Lectura general de esta ventana.", "ReportMuted"),
                ],
                TableClass=Table,
                TableStyleClass=TableStyle,
                mm_unit=mm,
                palette=palette,
                width_mm=58,
                padding=8,
                background=palette["panel_alt"],
                border_color=palette["line_dark"],
                accent_color=palette["navy"],
                border_width=0.7,
            )

        def _client_cover_cards_table(cards: list[tuple[str, str, str]]) -> Table:
            rows: list[list[object]] = []
            row: list[object] = []
            for idx, (label, value, _) in enumerate(cards, start=1):
                cell = _build_metric_card(
                    label,
                    value,
                    paragraph_builder=_p,
                    TableClass=Table,
                    TableStyleClass=TableStyle,
                    mm_unit=mm,
                    palette=palette,
                    width_mm=86,
                )
                row.append(cell)
                if idx % 2 == 0:
                    rows.append(row)
                    row = []
            if row:
                while len(row) < 2:
                    row.append("")
                rows.append(row)
            table = Table(rows, colWidths=[86 * mm, 86 * mm], hAlign="LEFT")
            table.setStyle(TableStyle([("VALIGN", (0, 0), (-1, -1), "TOP")]))
            return table

        def _client_chrome(canvas_obj, doc_obj) -> None:
            _draw_threshold_header(canvas_obj, doc_obj, palette=palette, mm_unit=mm, label="")
            _draw_threshold_footer(canvas_obj, doc_obj, palette=palette, mm_unit=mm)

        client_cards = _client_cards()
        client_has_profile_signal = (
            _row_has_eval_data(focus_row)
            or _row_has_load_data(focus_row)
            or _row_has_wellness_data(focus_row)
        )
        client_strengths = _client_strengths_lines()
        client_gaps = _client_watchouts_lines()
        client_next_steps = _client_next_steps_list()
        if not client_has_profile_signal:
            client_strengths = ["Ya empezamos a juntar información útil para seguir tu proceso."]
            client_gaps = ["Falta información para definir con más claridad qué conviene priorizar."]
            client_next_steps = [
                "Programar una nueva evaluación nos va a dar una referencia más clara para seguir ajustando el proceso.",
                "Sostener más continuidad en el plan va a ayudar a leer mejor cómo viene el proceso.",
            ]
        client_focus = _client_focus_label()

        cover_intro = Table(
            [[
                _build_note_box(
                    "Lectura breve",
                    _client_overview_text(),
                    paragraph_builder=_p,
                    TableClass=Table,
                    TableStyleClass=TableStyle,
                    mm_unit=mm,
                    palette=palette,
                    width_mm=112,
                    title_style="BlockTitle",
                    body_style="ReportBody",
                    background=palette["panel"],
                    border_color=palette["line_dark"],
                    accent_color=palette["navy"],
                ),
                _client_status_box(),
            ]],
            colWidths=[112 * mm, 58 * mm],
            hAlign="LEFT",
        )
        cover_intro.setStyle(
            TableStyle(
                [
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("LEFTPADDING", (0, 0), (-1, -1), 0),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 0),
                    ("TOPPADDING", (0, 0), (-1, -1), 0),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
                ]
            )
        )

        story.extend(
            [
                _p("Reporte de progreso", "ReportMuted"),
                _p(effective_athlete, "ReportTitle"),
                _p(f"Fecha: {datetime.now():%d/%m/%Y} | Ventana visible: últimas 6 semanas", "ReportMuted"),
                Spacer(1, 4 * mm),
                cover_intro,
                Spacer(1, 5 * mm),
                _client_cover_cards_table(client_cards),
                Spacer(1, 5 * mm),
                _build_decision_box(
                    "Qué vamos a priorizar",
                    f"Hoy la prioridad es {_client_focus_phrase()}.",
                    paragraph_builder=_p,
                    TableClass=Table,
                    TableStyleClass=TableStyle,
                    mm_unit=mm,
                    palette=palette,
                    title_style="ReportDecisionTitle",
                    body_style="ReportBodyWhite",
                    note_style="ReportMutedWhite",
                    note="Lectura simple para entender cómo venís y qué vamos a cuidar ahora.",
                ),
            ]
        )

        story.append(PageBreak())
        story.append(_build_pdf_page_title(
            "Estado actual y progreso",
            paragraph_builder=_p,
            TableClass=Table,
            TableStyleClass=TableStyle,
            mm_unit=mm,
            palette=palette,
            subtitle="Una foto simple de cómo estás hoy y de si ya hay una referencia para comparar el proceso.",
            eyebrow="",
            title_style="ReportSection",
            subtitle_style="ReportMuted",
            eyebrow_style="ReportMuted",
        ))
        story.append(Spacer(1, 4 * mm))
        state_chart_payload = client_charts.get("cmj_trend") or client_charts.get("completion")
        state_chart = _chart_image(state_chart_payload, height_mm=66) if state_chart_payload else None
        story.append(
            _build_chart_container(
                state_chart,
                paragraph_builder=_p,
                TableClass=Table,
                TableStyleClass=TableStyle,
                mm_unit=mm,
                palette=palette,
                title=state_chart_payload.get("title", "Progreso reciente") if state_chart_payload else "Progreso reciente",
                note=(
                    "Sirve para ver si ya hay una referencia útil para comparar tu proceso con más claridad."
                    if state_chart is not None
                    else "Cuando tengamos más registros, esta vista va a ayudar a seguir mejor tu progreso."
                ),
                empty_text="Todavía no hay historial suficiente para mostrar una tendencia clara.",
                title_style="BlockTitle",
                note_style="ReportMuted",
                accent_color=palette["steel"],
            )
        )
        story.append(Spacer(1, 4 * mm))
        story.append(_compact_key_value_table(_client_state_rows()))
        story.append(Spacer(1, 4 * mm))
        story.append(
            _build_note_box(
                "Lo que estamos viendo",
                [client_strengths[0], client_gaps[0]],
                paragraph_builder=_p,
                TableClass=Table,
                TableStyleClass=TableStyle,
                mm_unit=mm,
                palette=palette,
                title_style="BlockTitle",
                body_style="ReportBody",
                background=palette["card"],
                border_color=palette["line_dark"],
                accent_color=palette["steel"],
            )
        )

        story.append(PageBreak())
        story.append(_build_pdf_page_title(
            "Carga y constancia",
            paragraph_builder=_p,
            TableClass=Table,
            TableStyleClass=TableStyle,
            mm_unit=mm,
            palette=palette,
            subtitle="Una lectura simple para ver si la semana viene estable, baja o exigente, y si hay continuidad del plan.",
            eyebrow="",
            title_style="ReportSection",
            subtitle_style="ReportMuted",
            eyebrow_style="ReportMuted",
        ))
        story.append(Spacer(1, 4 * mm))
        load_chart_payload = client_charts.get("acwr") or client_charts.get("completion")
        load_chart = _chart_image(load_chart_payload, height_mm=56) if load_chart_payload else None
        story.append(
            _build_note_box(
                "Qué significa",
                _client_load_lines()[0],
                paragraph_builder=_p,
                TableClass=Table,
                TableStyleClass=TableStyle,
                mm_unit=mm,
                palette=palette,
                title_style="BlockTitle",
                body_style="ReportBody",
                background=palette["card"],
                border_color=palette["line_dark"],
                accent_color=palette["navy"],
            )
        )
        story.append(Spacer(1, 4 * mm))
        story.append(
            _build_chart_container(
                load_chart,
                paragraph_builder=_p,
                TableClass=Table,
                TableStyleClass=TableStyle,
                mm_unit=mm,
                palette=palette,
                title=load_chart_payload.get("title", "Cómo venís tolerando el entrenamiento") if load_chart_payload else "Cómo venís tolerando el entrenamiento",
                note=(
                    "El gráfico acompaña la lectura de continuidad, tolerancia y ritmo reciente del entrenamiento."
                    if load_chart is not None
                    else "La lectura resume continuidad, tolerancia y ritmo reciente del entrenamiento con la información disponible."
                ),
                empty_text="Todavía no hay información suficiente para mostrar este gráfico de carga reciente.",
                title_style="BlockTitle",
                note_style="ReportMuted",
                accent_color=palette["steel"],
            )
        )
        story.append(Spacer(1, 4 * mm))
        story.append(_compact_key_value_table(_client_load_rows()))
        story.append(Spacer(1, 4 * mm))
        story.append(
            _build_note_box(
                "Lectura de la semana",
                _client_load_lines(),
                paragraph_builder=_p,
                TableClass=Table,
                TableStyleClass=TableStyle,
                mm_unit=mm,
                palette=palette,
                title_style="BlockTitle",
                body_style="ReportBody",
                background=palette["panel"],
                border_color=palette["line_dark"],
                accent_color=palette["steel"],
            )
        )
        story.append(Spacer(1, 4 * mm))
        story.append(
            _build_note_box(
                "Qué mirar",
                _client_load_watch_lines(),
                paragraph_builder=_p,
                TableClass=Table,
                TableStyleClass=TableStyle,
                mm_unit=mm,
                palette=palette,
                title_style="BlockTitle",
                body_style="ReportBody",
                background=palette["card"],
                border_color=palette["line_dark"],
                accent_color=palette["navy"],
            )
        )

        story.append(PageBreak())
        story.append(_build_pdf_page_title(
            "Bienestar y recuperación",
            paragraph_builder=_p,
            TableClass=Table,
            TableStyleClass=TableStyle,
            mm_unit=mm,
            palette=palette,
            subtitle="Una referencia simple para ver si sueño, estrés y dolor están acompañando el trabajo.",
            eyebrow="",
            title_style="ReportSection",
            subtitle_style="ReportMuted",
            eyebrow_style="ReportMuted",
        ))
        story.append(Spacer(1, 4 * mm))
        wellness_chart_payload = client_charts.get("wellness")
        wellness_chart = _chart_image(wellness_chart_payload, height_mm=56) if wellness_chart_payload else None
        story.append(
            _build_note_box(
                "Qué significa",
                _client_wellness_summary_line(),
                paragraph_builder=_p,
                TableClass=Table,
                TableStyleClass=TableStyle,
                mm_unit=mm,
                palette=palette,
                title_style="BlockTitle",
                body_style="ReportBody",
                background=palette["card"],
                border_color=palette["line_dark"],
                accent_color=palette["navy"],
            )
        )
        story.append(Spacer(1, 4 * mm))
        story.append(
            _build_chart_container(
                wellness_chart,
                paragraph_builder=_p,
                TableClass=Table,
                TableStyleClass=TableStyle,
                mm_unit=mm,
                palette=palette,
                title=wellness_chart_payload.get("title", "Bienestar reciente") if wellness_chart_payload else "Bienestar reciente",
                note=(
                    "La idea es entender cómo venís recuperando y si hace falta cuidar algo más de cerca."
                    if wellness_chart is not None
                    else "Con más registros esta vista va a ayudar a seguir mejor cómo venís recuperando."
                ),
                empty_text="Todavía no hay registros suficientes para mostrar un gráfico de bienestar reciente.",
                title_style="BlockTitle",
                note_style="ReportMuted",
                accent_color=palette["steel"],
            )
        )
        story.append(Spacer(1, 4 * mm))
        story.append(_compact_key_value_table(_client_wellness_rows()))
        story.append(Spacer(1, 4 * mm))
        story.append(
            _build_note_box(
                "Lectura de recuperación",
                _client_wellness_lines(),
                paragraph_builder=_p,
                TableClass=Table,
                TableStyleClass=TableStyle,
                mm_unit=mm,
                palette=palette,
                title_style="BlockTitle",
                body_style="ReportBody",
                background=palette["panel"],
                border_color=palette["line_dark"],
                accent_color=palette["steel"],
            )
        )
        story.append(Spacer(1, 4 * mm))
        story.append(
            _build_note_box(
                "Qué vamos a cuidar",
                _client_wellness_watch_lines(),
                paragraph_builder=_p,
                TableClass=Table,
                TableStyleClass=TableStyle,
                mm_unit=mm,
                palette=palette,
                title_style="BlockTitle",
                body_style="ReportBody",
                background=palette["card"],
                border_color=palette["line_dark"],
                accent_color=palette["navy"],
            )
        )

        story.append(PageBreak())
        story.append(_build_pdf_page_title(
            "Próximos pasos",
            paragraph_builder=_p,
            TableClass=Table,
            TableStyleClass=TableStyle,
            mm_unit=mm,
            palette=palette,
            subtitle="Lo que está funcionando, lo que vamos a cuidar y la prioridad de esta etapa.",
            eyebrow="",
            title_style="ReportSection",
            subtitle_style="ReportMuted",
            eyebrow_style="ReportMuted",
        ))
        story.append(Spacer(1, 4 * mm))
        next_focus_box = _build_decision_box(
            "Qué vamos a priorizar",
            _client_focus_plan_sentence(),
            paragraph_builder=_p,
            TableClass=Table,
            TableStyleClass=TableStyle,
            mm_unit=mm,
            palette=palette,
            title_style="ReportDecisionTitle",
            body_style="ReportBodyWhite",
            note_style="ReportMutedWhite",
            note="La decisión se apoya en la información disponible de esta ventana.",
        )
        what_is_working = _box_flow(
            [_p("Qué está funcionando bien", "BlockTitle"), _p(client_strengths[0], "ReportBody")],
            padding=8,
        )
        what_to_watch = _box_flow(
            [_p("Qué vamos a cuidar", "BlockTitle"), _p(client_gaps[0], "ReportBody")],
            padding=8,
        )
        next_steps_box = _box_flow(
            [_p("Próximos pasos", "BlockTitle")] + _build_threshold_bullets(client_next_steps[:3], paragraph_builder=_p, style_name="ReportBody"),
            padding=8,
        )
        story.append(next_focus_box)
        story.append(Spacer(1, 4 * mm))
        story.append(what_is_working)
        story.append(Spacer(1, 4 * mm))
        story.append(what_to_watch)
        story.append(Spacer(1, 4 * mm))
        story.append(next_steps_box)
        story.append(Spacer(1, 5 * mm))
        story.append(
            _build_note_box(
                "Volver a medir",
                _client_review_text(),
                paragraph_builder=_p,
                TableClass=Table,
                TableStyleClass=TableStyle,
                mm_unit=mm,
                palette=palette,
                title_style="BlockTitle",
                body_style="ReportBody",
                background=palette["panel"],
                border_color=palette["line_dark"],
                accent_color=palette["navy"],
            )
        )
        try:
            doc.build(story, onFirstPage=_client_chrome, onLaterPages=_client_chrome)
        except Exception:
            return None
        return buffer.getvalue()

    subtitle = {
        "atleta": "Reporte individual para atleta",
        "profe": "Reporte técnico para profesional",
        "cliente": "Reporte de progreso",
    }[audience]
    target_name = effective_athlete if effective_athlete != "Todos" else "Resumen general"
    story.extend(
        [
            _p(subtitle, "ReportMuted"),
            _p(target_name, "ReportTitle"),
            _p(f"Generado el {datetime.now():%d/%m/%Y %H:%M} | Ventana visible: últimas 6 semanas", "ReportMuted"),
            Spacer(1, 4 * mm),
            Table(
                [[_p(insights.get("report", {}).get("summary", "Reporte listo para revisión."), "ReportBody")]],
                colWidths=[174 * mm],
                style=TableStyle(
                    [
                        ("BOX", (0, 0), (-1, -1), 0.7, palette["line"]),
                        ("BACKGROUND", (0, 0), (-1, -1), palette["card"]),
                        ("LEFTPADDING", (0, 0), (-1, -1), 10),
                        ("RIGHTPADDING", (0, 0), (-1, -1), 10),
                        ("TOPPADDING", (0, 0), (-1, -1), 8),
                        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
                    ]
                ),
            ),
            Spacer(1, 6 * mm),
            _p("Focos prioritarios", "ReportSection"),
        ]
    )
    for focus in insights.get("report", {}).get("focuses", []):
        story.append(_p(f"- {focus}", "ReportBody"))
    story.append(PageBreak())

    if not summary_df.empty:
        focus_row = summary_df.iloc[0]
        story.extend(
            [
                _p("Resumen ejecutivo", "ReportSection"),
                _p(
                    {
                        "atleta": "Lectura cuasi técnica de tu estado actual, perfil y prioridades inmediatas.",
                        "profe": "Última foto integrada para revisar disponibilidad, perfil y toma de decisiones.",
                        "cliente": "Lectura clara del estado actual y del progreso reciente.",
                    }[audience],
                    "ReportMuted",
                ),
                _metric_cards_table(_audience_dashboard_cards(state, focus_row, effective_athlete, audience)),
                Spacer(1, 5 * mm),
                _summary_meta_table(),
                Spacer(1, 5 * mm),
            ]
        )

        if effective_athlete != "Todos":
            story.extend(_metric_table(_audience_metric_rows(state, focus_row, effective_athlete, audience), title="Indicadores principales"))
        else:
            snapshot = _snapshot_table()
            if snapshot is not None:
                story.append(_p("Tabla ejecutiva integrada", "ReportSection"))
                story.append(snapshot)

    for chart in charts:
        chart_story = _chart_story(chart)
        if chart_story:
            story.append(PageBreak())
            story.extend(chart_story)

    if blocks:
        story.append(PageBreak())
        header = {
            "profe": "Interpretación y focos",
            "atleta": "Fortalezas y próximos pasos",
            "cliente": "Lectura simple y próximos pasos",
        }[audience]
        story.append(_p(header, "ReportSection"))
        for block in blocks:
            box_story = [
                _p(block.get("title", "Bloque"), "BlockTitle"),
                _p(block.get("summary", ""), "ReportBody"),
            ]
            for item in block.get("focuses", []):
                box_story.append(_p(f"- {item}", "ReportBody"))
            table = Table(
                [[box_story]],
                colWidths=[174 * mm],
                style=TableStyle(
                    [
                        ("BOX", (0, 0), (-1, -1), 0.7, palette["line"]),
                        ("BACKGROUND", (0, 0), (-1, -1), palette["card"]),
                        ("LEFTPADDING", (0, 0), (-1, -1), 10),
                        ("RIGHTPADDING", (0, 0), (-1, -1), 10),
                        ("TOPPADDING", (0, 0), (-1, -1), 8),
                        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
                    ]
                ),
            )
            story.append(table)
            story.append(Spacer(1, 4 * mm))

    try:
        doc.build(story)
    except Exception:
        return None
    return buffer.getvalue()


def _build_pdf_document(page_contents: list[str]) -> bytes:
    objects: list[bytes] = []
    objects.append(b"<< /Type /Catalog /Pages 2 0 R >>")

    font_objects = [
        b"<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica >>",
        b"<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica-Bold >>",
        b"<< /Type /Font /Subtype /Type1 /BaseFont /Courier >>",
    ]
    pages_object_num = 2
    first_dynamic_obj = 3 + len(font_objects)
    page_obj_nums = []
    content_obj_nums = []

    for idx in range(len(page_contents)):
        page_obj_nums.append(first_dynamic_obj + idx * 2)
        content_obj_nums.append(first_dynamic_obj + idx * 2 + 1)

    kids = " ".join(f"{obj} 0 R" for obj in page_obj_nums)
    objects.append(f"<< /Type /Pages /Count {len(page_contents)} /Kids [{kids}] >>".encode("ascii"))
    objects.extend(font_objects)

    for page_obj_num, content_obj_num, content in zip(page_obj_nums, content_obj_nums, page_contents):
        page_obj = (
            f"<< /Type /Page /Parent {pages_object_num} 0 R /MediaBox [0 0 595 842] "
            f"/Resources << /Font << /F1 3 0 R /F2 4 0 R /F3 5 0 R >> >> "
            f"/Contents {content_obj_num} 0 R >>"
        )
        objects.append(page_obj.encode("ascii"))
        content_bytes = content.encode("latin-1", "ignore")
        stream = b"<< /Length " + str(len(content_bytes)).encode("ascii") + b" >>\nstream\n" + content_bytes + b"\nendstream"
        objects.append(stream)

    output = bytearray(b"%PDF-1.4\n")
    offsets = [0]
    for idx, obj in enumerate(objects, start=1):
        offsets.append(len(output))
        output.extend(f"{idx} 0 obj\n".encode("ascii"))
        output.extend(obj)
        output.extend(b"\nendobj\n")

    xref_offset = len(output)
    output.extend(f"xref\n0 {len(objects) + 1}\n".encode("ascii"))
    output.extend(b"0000000000 65535 f \n")
    for offset in offsets[1:]:
        output.extend(f"{offset:010d} 00000 n \n".encode("ascii"))
    output.extend(f"trailer\n<< /Size {len(objects) + 1} /Root 1 0 R >>\nstartxref\n{xref_offset}\n%%EOF".encode("ascii"))
    return bytes(output)


def generate_visual_report_pdf(
    state: dict[str, pd.DataFrame | None],
    report_athlete: str = "Todos",
    report_audience: str = "profe",
) -> bytes | None:
    audience = normalize_report_audience(report_audience)
    effective_athlete = resolve_report_scope(state, report_athlete, audience)
    if effective_athlete is None:
        return None
    premium_pdf = _generate_visual_report_pdf_reportlab(state, effective_athlete, audience)
    if premium_pdf:
        return premium_pdf

    summary_df = build_executive_summary_df(state, effective_athlete, audience)
    insights = generate_module_insights(state, effective_athlete, audience)

    page_contents = [_build_cover_page(effective_athlete, summary_df, insights, audience)]
    dashboard_page = _build_dashboard_page(state, effective_athlete, summary_df, audience)
    if dashboard_page:
        page_contents.append(dashboard_page)
    if audience == "profe" or effective_athlete == "Todos":
        page_contents.extend(_build_snapshot_pages(summary_df))
    else:
        metric_page = _build_metric_profile_page(state, effective_athlete, summary_df, audience)
        if metric_page:
            page_contents.append(metric_page)
    trend_page = _build_trend_page(state, effective_athlete, audience)
    if trend_page:
        page_contents.append(trend_page)
    page_contents.extend(_build_insight_pages(state, effective_athlete, summary_df, insights, audience))
    return _build_pdf_document(page_contents)
