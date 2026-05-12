from __future__ import annotations

import os
import re
import unicodedata
from io import BytesIO
from pathlib import Path
from typing import BinaryIO

from openpyxl import load_workbook

from modules.evaluation_registry import get_evaluation_spec


EXPECTED_METRIC_FIELDS = (
    "force_max_n",
    "force_avg_n",
    "force_left_max_n",
    "force_right_max_n",
    "asymmetry_pct",
    "pre_tension_n",
    "time_to_peak_s",
    "time_pull_s",
    "force_50_n",
    "force_100_n",
    "force_150_n",
    "force_200_n",
    "force_250_n",
    "rfd_50_n_s",
    "rfd_100_n_s",
    "rfd_150_n_s",
    "rfd_250_n_s",
)

_SIGNATURE_TO_FIELD: dict[str, str] = {}
_REP_COLUMN_RE = re.compile(r"^rep\s*\d+$")


def _register_signatures(field_name: str, *signatures: str) -> None:
    for signature in signatures:
        _SIGNATURE_TO_FIELD[signature] = field_name


_register_signatures("time_to_peak_s", "time max force", "time max force s")
_register_signatures("time_pull_s", "time pull", "time pull s")
_register_signatures("force_max_n", "force max", "force max n")
_register_signatures("force_avg_n", "force avg", "force avg n")
_register_signatures("force_left_max_n", "force left max", "force left max n")
_register_signatures("force_right_max_n", "force right max", "force right max n")
_register_signatures("asymmetry_pct", "asimmetry", "asimmetry pct", "asymmetry", "asymmetry pct")
_register_signatures("pre_tension_n", "pre tension", "pre tension n")

for _window in (50, 100, 150, 200, 250):
    _register_signatures(
        f"force_{_window}_n",
        f"force at {_window}",
        f"force at {_window} n",
        f"force {_window}",
        f"force {_window} n",
    )

for _window in (50, 100, 150, 250):
    _register_signatures(
        f"rfd_{_window}_n_s",
        f"rfd at {_window}",
        f"rfd at {_window} n s",
        f"rfd {_window}",
        f"rfd {_window} n s",
    )


def _normalize_label(value: object) -> str:
    if value is None:
        return ""
    text = unicodedata.normalize("NFKD", str(value))
    text = "".join(char for char in text if not unicodedata.combining(char))
    text = text.lower().strip()
    text = text.replace("%", " pct ")
    text = text.replace("/", " ")
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def _coerce_number(value: object) -> int | float | None:
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return int(value) if value.is_integer() else value

    text = str(value).strip()
    if not text:
        return None

    text = text.replace(" ", "")
    if "," in text and "." not in text:
        left, right = text.split(",", 1)
        if len(right) == 3 and right.isdigit() and left.replace("-", "").isdigit():
            text = f"{left}{right}"
        else:
            text = f"{left}.{right}"
    elif "," in text and "." in text:
        if text.rfind(",") > text.rfind("."):
            text = text.replace(".", "").replace(",", ".")
        else:
            text = text.replace(",", "")
    else:
        text = text.replace(",", "")

    try:
        number = float(text)
    except ValueError:
        return None
    return int(number) if number.is_integer() else number


def _resolve_metric_field(label: object) -> str | None:
    signature = _normalize_label(label)
    return _SIGNATURE_TO_FIELD.get(signature)


def _open_workbook_source(file_like: object) -> str | Path | BytesIO | BinaryIO:
    if isinstance(file_like, (str, Path, os.PathLike)):
        return Path(file_like)
    if isinstance(file_like, (bytes, bytearray)):
        return BytesIO(file_like)
    if isinstance(file_like, BytesIO):
        file_like.seek(0)
        return file_like
    if hasattr(file_like, "read"):
        try:
            file_like.seek(0)
        except Exception:
            pass
        payload = file_like.read()
        if isinstance(payload, str):
            payload = payload.encode("utf-8")
        return BytesIO(payload)
    raise TypeError("Unsupported workbook source type.")


def _find_summary_table(workbook) -> tuple[str, dict[str, object]] | None:
    for worksheet in workbook.worksheets:
        rows = list(worksheet.iter_rows(values_only=True))
        for row_index, row in enumerate(rows):
            normalized = [_normalize_label(value) for value in row]
            if "references" not in normalized:
                continue

            label_col = normalized.index("references")
            rep_col = next((idx for idx, cell in enumerate(normalized) if _REP_COLUMN_RE.match(cell)), None)
            if rep_col is None:
                continue

            rep_label = str(row[rep_col]).strip() if row[rep_col] is not None else "Rep 1"
            metric_rows: dict[str, object] = {}
            for metric_row in rows[row_index + 1 :]:
                if not metric_row:
                    continue
                label_value = metric_row[label_col] if label_col < len(metric_row) else None
                if label_value is None or str(label_value).strip() == "":
                    continue
                metric_rows[str(label_value)] = metric_row[rep_col] if rep_col < len(metric_row) else None

            if any(_resolve_metric_field(label) for label in metric_rows):
                return rep_label, metric_rows
    return None


def parse_involution_summary_excel(
    file_like: object,
    test_id: str = "imtp",
    athlete_name: object = None,
    test_date: object = None,
) -> dict[str, object]:
    if get_evaluation_spec(test_id) is None:
        raise ValueError(f"Unsupported evaluation test_id: {test_id}")

    workbook_source = _open_workbook_source(file_like)
    workbook = load_workbook(workbook_source, data_only=True, read_only=True)
    try:
        summary_table = _find_summary_table(workbook)
    finally:
        workbook.close()

    if summary_table is None:
        raise ValueError("No Involution summary table found in workbook.")

    rep_column, metric_rows = summary_table
    metrics = {field_name: None for field_name in EXPECTED_METRIC_FIELDS}

    for label, raw_value in metric_rows.items():
        field_name = _resolve_metric_field(label)
        if field_name is None:
            continue
        metrics[field_name] = _coerce_number(raw_value)

    missing_metrics = [field_name for field_name, value in metrics.items() if value is None]

    return {
        "test_id": test_id,
        "athlete_name": athlete_name,
        "test_date": test_date,
        "rep_column": rep_column,
        "metrics": metrics,
        "missing_metrics": missing_metrics,
        "source_format": "involution_summary",
    }
