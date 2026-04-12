from __future__ import annotations

import unittest
from io import BytesIO
from unittest.mock import patch

import pandas as pd
from openpyxl import load_workbook

from modules.report_generator import (
    build_executive_summary_df,
    build_report_sheets,
    collect_report_athletes,
    export_excel,
    export_plotly_figure_png,
)


class Phase1ReportingTest(unittest.TestCase):
    def test_export_plotly_figure_png_avoids_deprecated_engine_argument(self):
        figure = object()

        with patch("plotly.io.to_image", side_effect=[RuntimeError("first try"), b"png-bytes"]) as mocked_to_image:
            with patch("kaleido.get_chrome_sync", return_value="chrome.exe"):
                exported = export_plotly_figure_png(figure, width=640, height=360, scale=1)

        self.assertEqual(exported, b"png-bytes")
        self.assertEqual(mocked_to_image.call_count, 2)
        first_call = mocked_to_image.call_args_list[0]
        second_call = mocked_to_image.call_args_list[1]
        self.assertNotIn("engine", first_call.kwargs)
        self.assertNotIn("engine", second_call.kwargs)
        self.assertEqual(first_call.kwargs["width"], 640)
        self.assertEqual(first_call.kwargs["height"], 360)
        self.assertEqual(first_call.kwargs["scale"], 1)

    def test_collect_report_athletes_uses_all_supported_datasets(self):
        state = {
            "rpe_df": None,
            "wellness_df": pd.DataFrame(
                [{"Athlete": "Carla Diaz", "Date": "2026-04-03", "Wellness_Score": 21}]
            ),
            "completion_df": pd.DataFrame(
                [{"Athlete": "Ana Lopez", "Date": "2026-04-02", "Pct": 90}]
            ),
            "rep_load_df": None,
            "raw_df": pd.DataFrame(
                [{"Athlete": "Bruno Rey", "Assigned Date": "2026-04-04", "Volume_Load": 520}]
            ),
            "maxes_df": pd.DataFrame(
                [{"Athlete": "Dario Paz", "Exercise Name": "Back Squat", "Max Value": 140}]
            ),
            "jump_df": None,
        }

        self.assertEqual(
            collect_report_athletes(state),
            ["Ana Lopez", "Bruno Rey", "Carla Diaz", "Dario Paz"],
        )

    def test_completion_exports_include_summary_and_expected_sheet_names(self):
        completion_df = pd.DataFrame(
            [
                {"Athlete": "Ana Lopez", "Date": "2026-04-02", "Assigned": 10, "Completed": 9, "Pct": 90},
                {"Athlete": "Ana Lopez", "Date": "2026-04-03", "Assigned": 8, "Completed": 8, "Pct": 100},
                {"Athlete": "Bruno Rey", "Date": "2026-04-02", "Assigned": 12, "Completed": 9, "Pct": 75},
            ]
        )
        state = {
            "rpe_df": None,
            "wellness_df": None,
            "completion_df": completion_df,
            "rep_load_df": None,
            "raw_df": None,
            "maxes_df": None,
            "jump_df": None,
            "acwr_dict": {},
            "mono_dict": {},
        }

        sheets = build_report_sheets(
            state,
            report_athlete="Todos",
            include_acwr=False,
            include_mono=False,
            include_wellness=False,
            include_jumps=False,
            include_maxes=False,
            include_volume=False,
            include_completion=True,
        )

        self.assertIn("Completion_Resumen", sheets)
        self.assertIn("Completion_Rate", sheets)
        self.assertIn("Reporte_Meta", sheets)

        workbook = load_workbook(BytesIO(export_excel(sheets)))
        sheetnames = workbook.sheetnames

        for sheet_name in ["01_Resumen", "02_Interpretacion", "03_Completion_Resumen", "10_Completion_Detalle", "99_Meta"]:
            self.assertIn(sheet_name, sheetnames)

        ordered_indexes = [sheetnames.index(name) for name in ["01_Resumen", "02_Interpretacion", "03_Completion_Resumen", "10_Completion_Detalle", "99_Meta"]]
        self.assertEqual(ordered_indexes, sorted(ordered_indexes))

        summary_sheet = workbook["03_Completion_Resumen"]
        self.assertEqual(summary_sheet["A2"].value, "Equipo")
        self.assertEqual(summary_sheet["C2"].number_format, '0.0"%"')

    def test_jump_reports_label_eur_as_ratio(self):
        jump_df = pd.DataFrame(
            [
                {
                    "Athlete": "Ana Lopez",
                    "Date": "2026-04-03",
                    "CMJ_cm": 35,
                    "SJ_cm": 30,
                    "EUR": 1.167,
                    "DRI": 1.6,
                    "IMTP_N": 1820,
                    "NM_Profile": "Reactivo / Poca Base",
                }
            ]
        )
        state = {
            "rpe_df": None,
            "wellness_df": None,
            "completion_df": None,
            "rep_load_df": None,
            "raw_df": None,
            "maxes_df": None,
            "jump_df": jump_df,
            "acwr_dict": {},
            "mono_dict": {},
        }

        summary_df = build_executive_summary_df(state, report_athlete="Ana Lopez")
        self.assertIn("EUR (ratio)", summary_df.columns)
        self.assertNotIn("EUR", summary_df.columns)
        self.assertEqual(summary_df.loc[0, "EUR (ratio)"], "1.167")

        sheets = build_report_sheets(
            state,
            report_athlete="Ana Lopez",
            include_acwr=False,
            include_mono=False,
            include_wellness=False,
            include_jumps=True,
            include_maxes=False,
            include_volume=False,
            include_completion=False,
        )

        workbook = load_workbook(BytesIO(export_excel(sheets)))
        eval_sheet = workbook["07_Evaluaciones"]
        headers = [cell.value for cell in eval_sheet[1]]
        self.assertIn("EUR (ratio)", headers)
        self.assertNotIn("EUR", headers)
        eur_column = headers.index("EUR (ratio)") + 1
        self.assertEqual(eval_sheet.cell(row=2, column=eur_column).value, 1.17)


if __name__ == "__main__":
    unittest.main()
